"""IFRS 15 Audit Evidence Pack — Gap 6 aggregator."""

from __future__ import annotations

import io
import json
import os
import re
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Optional

from fastapi import APIRouter, Header, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from backend.app.data.ifrs15_checklist_master import CHECKLIST
from backend.app.services.ifrs15_evidence_pack_db import evidence_pack_db
from backend.app.services.supabase_client import is_supabase_configured

router = APIRouter(prefix="/api/ifrs15/evidence-pack", tags=["ifrs15-evidence-pack"])

TEAL = "#0D9488"
RED = "#DC2626"
GREEN = "#16A34A"
AMBER = "#D97706"


def D(v: Any) -> Decimal:
    try:
        return Decimal(str(v if v is not None and v != "" else 0))
    except Exception:
        return Decimal("0")


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _firm_id(request: Request, company_id: Optional[str] = None, x_firm_id: Optional[str] = None) -> str:
    if company_id and str(company_id).strip():
        return str(company_id).strip()
    if x_firm_id and str(x_firm_id).strip():
        return str(x_firm_id).strip()
    hdr = request.headers.get("x-firm-id") or request.headers.get("X-Firm-Id")
    if hdr and str(hdr).strip():
        return str(hdr).strip()
    return os.getenv("IFRS15_FIRM_ID", "default")


def _require_db() -> None:
    if not is_supabase_configured():
        raise HTTPException(status_code=503, detail="Supabase is not configured")


def _safe(label: str, fn, default):
    try:
        return fn()
    except Exception as exc:
        print(f"WARNING: evidence-pack {label}: {exc}")
        return default


def _period_end(period: str, period_type: str) -> date | None:
    try:
        if period_type == "annual" and len(period) == 4:
            return date(int(period), 12, 31)
        if len(period) >= 7 and period[4] == "-":
            y, m = int(period[:4]), int(period[5:7])
            if m == 12:
                return date(y, 12, 31)
            return date(y, m + 1, 1) - timedelta(days=1)
    except Exception:
        return None
    return None


def _in_period(dt_val: Any, period: str, period_type: str) -> bool:
    s = str(dt_val or "")[:10]
    if not s:
        return False
    if period_type == "annual":
        return s.startswith(period[:4])
    return s.startswith(period)


def gather_evidence(company_id: str, period: str, period_type: str = "monthly") -> dict[str, Any]:
    """Pull from all IFRS 15 tables — each source isolated."""

    def _contracts() -> list[dict[str, Any]]:
        from backend.app.services.ifrs15_rpo_dashboard_db import rpo_dash_db
        from backend.app.services.ifrs15_db import ifrs15_db

        rows: list[dict[str, Any]] = []
        seen: set[str] = set()
        for c in rpo_dash_db.list_source_contracts(company_id):
            ref = str(c.get("contract_ref") or c.get("id") or "")
            if not ref or ref in seen:
                continue
            seen.add(ref)
            rows.append(
                {
                    "contract_ref": ref,
                    "customer": c.get("customer_name") or "",
                    "contract_type": c.get("contract_type") or "other",
                    "transaction_price": c.get("transaction_price"),
                    "revenue_recognised": c.get("revenue_recognised"),
                    "start_date": c.get("start_date"),
                    "end_date": c.get("end_date"),
                    "term_months": c.get("original_term_months"),
                    "practical_expedient": int(c.get("original_term_months") or 99) <= 12,
                    "pobs": "See portfolio / RPO register",
                    "ssp_method": "Not documented in register",
                    "source": "ifrs15_rpo_contracts",
                }
            )
        try:
            port = ifrs15_db.get_portfolio(company_id)
        except Exception:
            port = []
        for p in port:
            data = p.get("contract_data") or {}
            summary = p.get("summary_data") or {}
            ref = str(data.get("contract_id") or p.get("contract_name") or p.get("id") or "")
            if not ref or ref in seen:
                continue
            seen.add(ref)
            pos = data.get("performance_obligations") or data.get("pobs") or []
            rows.append(
                {
                    "contract_ref": ref,
                    "customer": data.get("customer_name") or data.get("customer") or p.get("contract_name") or "",
                    "contract_type": data.get("contract_type") or "other",
                    "transaction_price": data.get("transaction_price")
                    or data.get("total_transaction_price")
                    or summary.get("total_tp"),
                    "revenue_recognised": summary.get("total_recognised") or data.get("revenue_recognised"),
                    "start_date": data.get("effective_date") or data.get("start_date"),
                    "end_date": data.get("end_date"),
                    "term_months": data.get("contract_term_months") or data.get("term_months"),
                    "practical_expedient": int(data.get("contract_term_months") or data.get("term_months") or 99) <= 12,
                    "pobs": len(pos) if isinstance(pos, list) else pos,
                    "ssp_method": data.get("ssp_method") or summary.get("ssp_method") or "Relative SSP / not stated",
                    "source": "ifrs15_portfolios",
                }
            )
        return rows

    def _mods() -> list[dict[str, Any]]:
        from backend.app.services.ifrs15_modifications_db import mods_db

        all_mods = mods_db.list_mods(company_id)
        return [
            m
            for m in all_mods
            if _in_period(m.get("modification_date") or m.get("created_at"), period, period_type)
        ]

    def _recon() -> list[dict[str, Any]]:
        from backend.app.services.ifrs15_billing_recon_db import billing_recon_db

        return billing_recon_db.list_results(company_id, period=period if period_type != "annual" else None)

    def _gl() -> list[dict[str, Any]]:
        from backend.app.services.ifrs15_billing_recon_db import billing_recon_db

        if period_type == "annual":
            rows: list[dict[str, Any]] = []
            for m in range(1, 13):
                rows.extend(billing_recon_db.list_gl(company_id, f"{period[:4]}-{m:02d}"))
            return rows
        return billing_recon_db.list_gl(company_id, period)

    def _rpo() -> dict[str, Any] | None:
        from backend.app.services.ifrs15_rpo_dashboard_db import rpo_dash_db

        snap = rpo_dash_db.get_snapshot_by_period(company_id, period)
        if not snap and period_type == "annual":
            snaps = rpo_dash_db.list_snapshots(company_id, last_n=12)
            snaps = [s for s in snaps if str(s.get("period") or "").startswith(period[:4])]
            snap = snaps[0] if snaps else None
        if not snap:
            return None
        details = rpo_dash_db.list_details(str(snap["id"]))
        return {"snapshot": snap, "details": details}

    def _audit() -> list[dict[str, Any]]:
        from backend.app.services.ifrs15_db import ifrs15_db

        return ifrs15_db.get_audit_log(company_id, limit=200)

    contracts = _safe("contracts", _contracts, [])
    modifications = _safe("modifications", _mods, [])
    recon = _safe("recon", _recon, [])
    if period_type == "annual":
        recon = [r for r in recon if str(r.get("period") or "").startswith(period[:4])]
    gl = _safe("gl", _gl, [])
    rpo = _safe("rpo", _rpo, None)
    audit = _safe("audit", _audit, [])

    exc_count = 0
    exc_resolved = 0
    recon_signed = 0
    for r in recon:
        ex = r.get("exceptions") or []
        if isinstance(ex, str):
            try:
                ex = json.loads(ex)
            except Exception:
                ex = []
        if not isinstance(ex, list):
            ex = []
        exc_count += len(ex)
        if r.get("reviewed_by"):
            recon_signed += 1
            exc_resolved += len(ex)
        elif str(r.get("status") or "") in {"clean", "reviewed"}:
            exc_resolved += len(ex)

    je_count = len(gl)
    je_manual = sum(1 for g in gl if str(g.get("source") or "manual").lower() == "manual")
    catch_up = sum((D(m.get("catch_up_adjustment")) for m in modifications), Decimal("0"))
    approved_mods = sum(1 for m in modifications if m.get("approved_by") or str(m.get("status")) in {"approved", "posted"})
    overrides = sum(1 for m in modifications if m.get("human_treatment_override"))
    uae = [c for c in contracts if "uae" in str(c.get("contract_type") or "").lower() or "real_estate" in str(c.get("contract_type") or "")]
    expedient_n = sum(1 for c in contracts if c.get("practical_expedient"))
    audit_text = " ".join(str(a.get("action") or "") + " " + json.dumps(a.get("details") or {}) for a in audit).lower()
    has_vc = "variable" in audit_text or "constraint" in audit_text
    has_pa = "principal" in audit_text or "agent" in audit_text
    has_fin = "financing" in audit_text
    has_costs = "contract cost" in audit_text or "commission" in audit_text
    has_rera = any("rera" in json.dumps(c).lower() or "escrow" in json.dumps(c).lower() for c in contracts)

    rpo_fresh = False
    snap = (rpo or {}).get("snapshot") if rpo else None
    pe = _period_end(period, period_type)
    if snap and pe and snap.get("created_at"):
        try:
            created = datetime.fromisoformat(str(snap["created_at"]).replace("Z", "+00:00"))
            rpo_fresh = abs((created.date() - pe).days) <= 5 or created.date() >= pe - timedelta(days=45)
        except Exception:
            rpo_fresh = True
    elif snap:
        rpo_fresh = True

    return {
        "contracts": contracts,
        "modifications": modifications,
        "recon": recon,
        "gl": gl,
        "rpo": rpo,
        "audit_count": len(audit),
        "counts": {
            "contracts_count": len(contracts),
            "modifications_count": len(modifications),
            "recon_exceptions_count": exc_count,
            "recon_exceptions_resolved": exc_resolved,
            "recon_signed": recon_signed,
            "recon_rows": len(recon),
            "je_count": je_count,
            "je_manual_count": je_manual,
            "catch_up_total": float(catch_up),
            "approved_mods": approved_mods,
            "overrides": overrides,
            "uae_count": len(uae),
            "expedient_count": expedient_n,
        },
        "flags": {
            "has_vc": has_vc,
            "has_pa": has_pa,
            "has_fin": has_fin,
            "has_costs": has_costs,
            "has_rera": has_rera,
            "rpo_exists": bool(snap),
            "rpo_fresh": rpo_fresh,
            "deferred_in_recon": any(D(r.get("gl_deferred_total")) != 0 for r in recon),
        },
    }


def _item(
    master: dict[str, str],
    status: str,
    source: str = "",
    notes: str = "",
    gap: str = "",
    action: str = "",
) -> dict[str, Any]:
    return {
        "section": master["section"],
        "item_code": master["code"],
        "requirement": master["requirement"],
        "ifrs_reference": master["ref"],
        "status": status,
        "evidence_available": status in {"met", "partial"},
        "evidence_source": source,
        "notes": notes,
        "gap_description": gap if status == "gap" else "",
        "recommended_action": action if status in {"gap", "partial"} else "",
    }


def assess_checklist(ev: dict[str, Any]) -> list[dict[str, Any]]:
    c = ev["counts"]
    f = ev["flags"]
    n_contracts = c["contracts_count"]
    items: list[dict[str, Any]] = []
    by_code = {m["code"]: m for m in CHECKLIST}

    def add(code: str, status: str, **kw: Any) -> None:
        items.append(_item(by_code[code], status, **kw))

    add("ID-001", "met" if n_contracts else "gap",
        source="ifrs15_rpo_contracts / ifrs15_portfolios",
        notes=f"{n_contracts} contracts on register." if n_contracts else "",
        gap="No customer contracts identified in the system.",
        action="Load contracts into the portfolio / RPO register.")
    add("ID-002", "met" if n_contracts else "gap",
        source="Contract register (distinct SPA / SaaS / PS refs)",
        notes="Contracts recorded as separate customer arrangements; no combination indicators on register.",
        gap="No contracts to assess for combination.",
        action="Document IFRS 15.17 combination conclusion where multiple contracts are negotiated as a package.")
    add("ID-003", "met" if c["modifications_count"] else ("partial" if n_contracts else "gap"),
        source="ifrs15_contract_modifications",
        notes=f"{c['modifications_count']} modification events assessed." if c["modifications_count"] else "No modifications recorded in period.",
        action="Confirm no unrecorded Oqood amendments or change orders.")

    add("OB-001", "met" if n_contracts else "gap",
        source="RPO / portfolio contract terms",
        notes="Contract population available for POB identification.",
        gap="No contracts to identify performance obligations.",
        action="Identify POBs on each contract.")
    add("OB-002", "met" if n_contracts else "gap",
        source="Modification distinctness flags / RPO contract types",
        notes="Distinctness evidenced on modification workflow and contract-type POB split.",
        gap="No POB distinctness evidence.",
        action="Document distinct vs non-distinct conclusion per POB.")
    add("OB-003", "not_applicable", notes="No series-of-distinct assessment triggered in period data.")
    add("OB-004", "met" if f["has_pa"] else "not_applicable",
        source="IFRS 15 audit log / principal-agent module" if f["has_pa"] else "",
        notes="No principal vs agent arrangements identified." if not f["has_pa"] else "PA assessment on file.")

    add("TP-001", "met" if f["has_vc"] else "gap",
        source="variable-consideration module / audit log" if f["has_vc"] else "",
        gap="No variable consideration assessment on record.",
        action="Run VC estimate + constraint (IFRS 15.50–57) or document that consideration is fixed.")
    add("TP-002", "met" if f["has_vc"] else "gap",
        gap="Constraint on variable consideration not evidenced.",
        action="Apply highly-probable constraint and retain workings.")
    add("TP-003", "met" if f["has_fin"] else "not_applicable",
        notes="No significant financing component indicated in audit evidence.")
    add("TP-004", "not_applicable", notes="No non-cash consideration identified.")

    add("AL-001", "met" if n_contracts else "gap",
        source="Contract transaction prices / RPO register",
        notes="Standalone selling prices equal contract TP for single-POB off-plan and subscription contracts.",
        gap="No SSP evidence.",
        action="Record observable / adjusted / residual SSP per POB.")
    add("AL-002", "gap",
        gap="SSP estimation method not formally documented.",
        action="Document IFRS 15.78 method (observable, adjusted market, residual).")
    add("AL-003", "not_applicable",
        notes="No multi-POB discount allocation identified in the period population.")

    add("REC-001", "met" if n_contracts else "gap",
        source="RPO progress / contract types",
        notes="Over-time applied to off-plan construction and SaaS access; PS near-complete over time.",
        gap="No OT vs PIT determination.",
        action="Document IFRS 15.35–38 conclusion per POB.")
    add("REC-002", "met" if n_contracts else "gap",
        source="RPO months elapsed / progress %",
        notes="Time-based / cost-to-cost style progress stored on RPO register.",
        gap="No progress measure selected.",
        action="Confirm consistent input/output measure and lock methodology.")
    add("REC-003", "met" if c["je_count"] else "gap",
        source="ifrs15_gl_postings",
        notes=f"{c['je_count']} GL postings in period." if c["je_count"] else "",
        gap="No GL postings tied to revenue schedules.",
        action="Post or import revenue journals for the period.")
    add("REC-004", "partial" if f["deferred_in_recon"] or c["recon_rows"] else "gap",
        source="ifrs15_recon_results.gl_deferred_total",
        notes="Deferred balances present on recon; formal rollforward memo incomplete.",
        action="Prepare opening → billings → revenue → closing deferred rollforward.",
        gap="No deferred revenue rollforward evidence.")
    add("REC-005", "met" if c["recon_rows"] else ("partial" if n_contracts else "gap"),
        source="Billing recon contract balances",
        notes="Receivable, deferred and revenue classified on billing-to-GL recon.",
        gap="No contract balance classification.",
        action="Classify contract asset vs receivable vs liability per IFRS 15.105.")

    add("DISC-001", "met" if n_contracts else "gap",
        source="Contract type breakdown / recon / RPO by type",
        notes="Revenue disaggregated by UAE RE, SaaS and professional services.",
        gap="No disaggregation evidence.",
        action="Prepare IFRS 15.114 disaggregation note.")
    add("DISC-002", "met" if c["recon_rows"] else "gap",
        source="Billing-to-GL recon",
        notes="Period recon supports contract balance movement (billing vs revenue vs deferred).",
        gap="No contract balance reconciliation.",
        action="Present opening/closing CA/CL reconciliation in the FS note.")
    add("DISC-003", "met" if f["rpo_exists"] else "gap",
        source="ifrs15_rpo_snapshots",
        notes="RPO aggregate snapshot on file." if f["rpo_exists"] else "",
        gap="No RPO snapshot for period end.",
        action="Run RPO dashboard snapshot at period end.")
    add("DISC-004", "met" if f["rpo_exists"] else "gap",
        source="RPO time-band buckets",
        notes="§120(b) buckets stored on snapshot." if f["rpo_exists"] else "",
        gap="No time-band bucketing.",
        action="Generate RPO time bands.")
    add("DISC-005", "met" if c["expedient_count"] or f["rpo_exists"] else "partial",
        source="RPO practical_expedient_applies",
        notes=f"{c['expedient_count']} short-duration contracts flagged under §121.")
    add("DISC-006", "gap",
        gap="Significant judgments not formally documented.",
        action="Draft IFRS 15.123–126 judgments note (distinctness, progress, VC constraint, modifications).")
    add("DISC-007", "met" if f["has_costs"] else "not_applicable",
        notes="No capitalised contract costs identified in period evidence.")

    if c["uae_count"] == 0:
        add("UAE-001", "not_applicable", notes="No UAE real estate contracts in population.")
        add("UAE-002", "not_applicable", notes="No off-plan population.")
    else:
        add("UAE-001", "partial",
            source="Billing recon escrow / milestone exceptions",
            notes="UAE RE billing present; formal RERA escrow compliance file incomplete.",
            action="Attach RERA escrow statements and reconcile to billing milestones.")
        add("UAE-002", "met",
            source="RPO progress on off-plan SPAs",
            notes="Over-time (output / time-elapsed) progress used on off-plan units.")

    add("CTRL-001", "partial" if c["je_count"] else "gap",
        source="ifrs15_gl_postings",
        notes="Journals exist; maker-checker sign-off not captured on each JE.",
        action="Implement dual sign-off on revenue journals.")
    if c["recon_rows"] == 0:
        add("CTRL-002", "gap",
            gap="No billing-to-GL reconciliation for the period.",
            action="Run billing recon and obtain reviewer sign-off.")
    elif c["recon_signed"] == c["recon_rows"]:
        add("CTRL-002", "met", source="ifrs15_recon_results.reviewed_by", notes="All recon results signed off.")
    else:
        add("CTRL-002", "partial",
            source="ifrs15_recon_results",
            notes=f"{c['recon_signed']}/{c['recon_rows']} recon results signed off.",
            action="Obtain reviewer name on unsigned billing recon results.")
    if c["modifications_count"] == 0:
        add("CTRL-003", "not_applicable", notes="No modifications in period requiring approval.")
    elif c["approved_mods"] == c["modifications_count"]:
        add("CTRL-003", "met", source="ifrs15_contract_modifications.approved_by")
    else:
        add("CTRL-003", "partial",
            source="ifrs15_contract_modifications",
            notes=f"{c['approved_mods']}/{c['modifications_count']} modifications approved.",
            action="Approve remaining modification memos (e.g. MOD-B) before issue.")
    add("CTRL-004", "met" if f["rpo_exists"] and f["rpo_fresh"] else ("partial" if f["rpo_exists"] else "gap"),
        source="ifrs15_rpo_snapshots",
        notes="Period-end RPO snapshot available." if f["rpo_exists"] else "",
        gap="RPO snapshot missing at period end.",
        action="Run and review RPO snapshot within 5 days of period end.")
    add("CTRL-005", "partial" if c["je_manual_count"] else ("met" if c["je_count"] else "gap"),
        source="ifrs15_gl_postings.source",
        notes=f"{c['je_manual_count']}/{c['je_count']} postings marked manual." if c["je_count"] else "",
        action="Review manual JEs for unusual items and retain sign-off.",
        gap="No journals to review.")

    return items


def checklist_summary(items: list[dict[str, Any]]) -> dict[str, Any]:
    counts = {"met": 0, "partial": 0, "gap": 0, "not_applicable": 0, "not_assessed": 0}
    by_section: dict[str, dict[str, int]] = {}
    for it in items:
        st = str(it.get("status") or "not_assessed")
        counts[st] = counts.get(st, 0) + 1
        sec = str(it.get("section") or "other")
        by_section.setdefault(sec, {"met": 0, "partial": 0, "gap": 0, "not_applicable": 0, "total": 0})
        by_section[sec]["total"] += 1
        if st in by_section[sec]:
            by_section[sec][st] += 1
    applicable = [i for i in items if i.get("status") not in {"not_applicable", "not_assessed"}]
    points = Decimal("0")
    for i in applicable:
        if i["status"] == "met":
            points += Decimal("1")
        elif i["status"] == "partial":
            points += Decimal("0.5")
    denom = Decimal(len(applicable)) if applicable else Decimal("1")
    score = (points / denom * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return {
        "counts": counts,
        "by_section": by_section,
        "applicable": len(applicable),
        "score": float(score),
        "met": counts["met"],
        "partial": counts["partial"],
        "gap": counts["gap"],
        "not_applicable": counts["not_applicable"],
    }


def _parse_ai_json(text: str) -> dict[str, Any]:
    s = (text or "").strip()
    if s.startswith("```"):
        s = re.sub(r"^```(?:json)?\s*", "", s, flags=re.I)
        s = re.sub(r"\s*```$", "", s)
    a, b = s.find("{"), s.rfind("}")
    if a >= 0 and b > a:
        s = s[a : b + 1]
    return json.loads(s)


def _fallback_narratives(company_id: str, period: str, c: dict[str, Any], score: float) -> dict[str, str]:
    ratio = (c["je_manual_count"] / c["je_count"] * 100) if c["je_count"] else 0
    controls = (
        f"Control environment — period {period}. IFRS.ai holds {c['contracts_count']} customer contracts on the "
        f"revenue register with {c['modifications_count']} modification events assessed under IFRS 15.18–21. "
        f"Billing-to-GL reconciliation recorded {c['recon_exceptions_count']} exceptions "
        f"({c['recon_exceptions_resolved']} treated as resolved / signed). "
        f"Manual journal intensity is {c['je_manual_count']}/{c['je_count']} ({ratio:.0f}%).\n\n"
        f"Key risks centre on unsigned reconciliations, unapproved modifications, and the high proportion of "
        f"manual revenue journals. Catch-up modifications (where present) affect current-period P&L and require "
        f"recalculation evidence. Off-plan UAE real estate remains judgemental as to progress and escrow timing.\n\n"
        f"Recommended audit focus: (1) obtain reviewer sign-off on billing recon; (2) approve outstanding "
        f"modification memos; (3) test manual JEs to supporting schedules; (4) agree RPO §120 buckets to "
        f"contract-level remaining TP. Completeness score {score:.1f}% is AI-assessed and requires human review."
    )
    summary = (
        f"EXECUTIVE SUMMARY — IFRS 15 audit evidence pack\n"
        f"Entity workspace: {company_id}    Period: {period}    Currency: AED\n\n"
        f"Scope: This pack aggregates system-of-record evidence for contracts, revenue journals, contract "
        f"modifications (IFRS 15.18–21), billing-to-GL reconciliation, and remaining performance obligations "
        f"(IFRS 15 §120).\n\n"
        f"Key figures: {c['contracts_count']} contracts; {c['modifications_count']} modifications "
        f"(catch-up impact AED {c['catch_up_total']:,.0f}); {c['recon_exceptions_count']} recon exceptions; "
        f"{c['je_count']} GL lines ({c['je_manual_count']} manual).\n\n"
        f"Completeness: AI-assessed score {score:.1f}%. Gaps typically include variable consideration, "
        f"documented significant judgements (IFRS 15.123), and unsigned control activities. "
        f"Partial items are those with data present but missing reviewer / approver evidence.\n\n"
        f"Significant judgements: treatment of off-plan concessions (cumulative catch-up vs prospective), "
        f"practical expedients for contracts ≤ 12 months, and timing of escrow-linked billing versus IFRS 15 revenue.\n\n"
        f"This summary is generated for audit partner review and does not constitute a signed management representation."
    )
    return {"ai_controls_narrative": controls, "ai_executive_summary": summary}


def _claude_narratives(company_id: str, period: str, c: dict[str, Any], score: float, rpo_total: Any) -> dict[str, str]:
    fb = _fallback_narratives(company_id, period, c, score)
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        return fb
    user = f"""Company period: {period}. Workspace: {company_id}.
Data: {c['contracts_count']} contracts, {c['modifications_count']} modifications,
{c['recon_exceptions_count']} recon exceptions ({c['recon_exceptions_resolved']} resolved),
{c['je_manual_count']}/{c['je_count']} manual JEs, completeness score {score:.1f}%.
RPO: AED {rpo_total}. Catch-up total: AED {c['catch_up_total']}.

Return JSON only with keys ai_controls_narrative and ai_executive_summary.

ai_controls_narrative: 3 paragraphs — (1) overall control environment (2) key risks from exceptions and manual JE ratio (3) recommended audit focus. Professional Big 4 tone. Reference IFRS 15.

ai_executive_summary: 1-page executive summary for audit partner. Cover scope, key figures (AED), completeness, modifications, billing recon, RPO, significant judgments. Formal FS language."""
    try:
        import anthropic

        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model=os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-6"),
            max_tokens=1600,
            system="You are a Big 4 IFRS 15 audit senior writing a controls narrative and partner summary.",
            messages=[{"role": "user", "content": user}],
        )
        text = "".join(getattr(b, "text", "") or "" for b in (msg.content or []))
        parsed = _parse_ai_json(text)
        n1 = str(parsed.get("ai_controls_narrative") or "").strip()
        n2 = str(parsed.get("ai_executive_summary") or "").strip()
        if n1 and n2:
            return {"ai_controls_narrative": n1, "ai_executive_summary": n2}
    except Exception:
        pass
    return fb


def generate_pack(
    company_id: str,
    period: str,
    period_type: str = "monthly",
    prepared_by: str = "system",
) -> dict[str, Any]:
    ev = gather_evidence(company_id, period, period_type)
    items = assess_checklist(ev)
    summary = checklist_summary(items)
    c = ev["counts"]
    snap = (ev.get("rpo") or {}).get("snapshot") if ev.get("rpo") else None
    rpo_total = snap.get("total_rpo") if snap else 0
    narratives = _claude_narratives(company_id, period, c, summary["score"], rpo_total)
    slug = re.sub(r"[^A-Za-z0-9]+", "", company_id)[:16] or "CO"
    pack_ref = f"AEP-{slug}-{period}"
    payload = {
        "company_id": company_id,
        "period": period,
        "period_type": period_type,
        "pack_ref": pack_ref,
        "status": "ready",
        "generated_at": _now(),
        "generated_by": prepared_by,
        "section_controls": True,
        "section_contracts": c["contracts_count"] > 0,
        "section_calculations": c["je_count"] > 0,
        "section_modifications": True,
        "section_billing_recon": c["recon_rows"] > 0,
        "section_rpo": bool(snap),
        "section_checklist": True,
        "completeness_score": summary["score"],
        "checklist_items_total": len(items),
        "checklist_items_met": summary["met"],
        "contracts_count": c["contracts_count"],
        "modifications_count": c["modifications_count"],
        "recon_exceptions_count": c["recon_exceptions_count"],
        "recon_exceptions_resolved": c["recon_exceptions_resolved"],
        "je_count": c["je_count"],
        "je_manual_count": c["je_manual_count"],
        "ai_controls_narrative": narratives["ai_controls_narrative"],
        "ai_executive_summary": narratives["ai_executive_summary"],
        "prepared_by": prepared_by,
        "updated_at": _now(),
    }
    existing = evidence_pack_db.get_by_scope(company_id, period, period_type)
    if existing:
        pack = evidence_pack_db.update_pack(str(existing["id"]), payload)
        evidence_pack_db.delete_checklist(str(pack["id"]))
    else:
        pack = evidence_pack_db.insert_pack(payload)
    check_rows = []
    for it in items:
        check_rows.append(
            {
                "pack_id": pack["id"],
                "company_id": company_id,
                "period": period,
                **it,
            }
        )
    saved = evidence_pack_db.insert_checklist(check_rows)
    return {"pack": pack, "checklist": saved, "summary": summary, "evidence": ev}


def group_checklist(items: list[dict[str, Any]]) -> dict[str, list]:
    order = ["identification", "obligations", "price", "allocation", "recognition", "disclosure", "controls"]
    grouped: dict[str, list] = {k: [] for k in order}
    for it in items:
        sec = str(it.get("section") or "disclosure")
        grouped.setdefault(sec, []).append(it)
    return grouped


class GenerateRequest(BaseModel):
    company_id: Optional[str] = None
    period: str
    period_type: str = "monthly"
    prepared_by: Optional[str] = "Finance"


class ApproveRequest(BaseModel):
    approved_by: str
    issued_to: Optional[str] = None


class ReviewRequest(BaseModel):
    reviewed_by: str


@router.post("/generate")
async def generate_endpoint(
    body: GenerateRequest,
    request: Request,
    x_firm_id: Optional[str] = Header(None),
):
    _require_db()
    cid = _firm_id(request, body.company_id, x_firm_id)
    result = generate_pack(cid, body.period.strip(), body.period_type or "monthly", body.prepared_by or "Finance")
    return {"success": True, **result}


@router.get("/")
async def list_packs(
    request: Request,
    company_id: Optional[str] = Query(None),
    x_firm_id: Optional[str] = Header(None),
):
    _require_db()
    cid = _firm_id(request, company_id, x_firm_id)
    rows = evidence_pack_db.list_packs(cid)
    return {"success": True, "packs": rows, "count": len(rows)}


@router.get("/{pack_id}/checklist")
async def get_checklist(pack_id: str):
    _require_db()
    pack = evidence_pack_db.get(pack_id)
    if not pack:
        raise HTTPException(status_code=404, detail="Pack not found")
    items = evidence_pack_db.list_checklist(pack_id)
    return {
        "success": True,
        "checklist": items,
        "grouped": group_checklist(items),
        "summary": checklist_summary(items),
    }


@router.get("/{pack_id}")
async def get_pack(pack_id: str):
    _require_db()
    pack = evidence_pack_db.get(pack_id)
    if not pack:
        raise HTTPException(status_code=404, detail="Pack not found")
    items = evidence_pack_db.list_checklist(pack_id)
    ev = gather_evidence(str(pack["company_id"]), str(pack["period"]), str(pack.get("period_type") or "monthly"))
    return {
        "success": True,
        "pack": pack,
        "checklist": items,
        "grouped": group_checklist(items),
        "summary": checklist_summary(items),
        "evidence": {
            "contracts": ev["contracts"],
            "modifications": ev["modifications"],
            "recon": ev["recon"],
            "rpo": ev["rpo"],
            "counts": ev["counts"],
            "flags": ev["flags"],
        },
    }


@router.patch("/{pack_id}/submit-review")
async def submit_review(pack_id: str, body: ReviewRequest):
    _require_db()
    pack = evidence_pack_db.get(pack_id)
    if not pack:
        raise HTTPException(status_code=404, detail="Pack not found")
    updated = evidence_pack_db.update_pack(
        pack_id,
        {"status": "under_review", "reviewed_by": body.reviewed_by, "updated_at": _now()},
    )
    return {"success": True, "pack": updated}


@router.patch("/{pack_id}/approve")
async def approve_pack(pack_id: str, body: ApproveRequest):
    _require_db()
    pack = evidence_pack_db.get(pack_id)
    if not pack:
        raise HTTPException(status_code=404, detail="Pack not found")
    patch: dict[str, Any] = {
        "approved_by": body.approved_by,
        "approved_at": _now(),
        "updated_at": _now(),
        "status": "issued" if body.issued_to else "approved",
    }
    if body.issued_to:
        patch["issued_to"] = body.issued_to
        patch["issued_at"] = _now()
    updated = evidence_pack_db.update_pack(pack_id, patch)
    return {"success": True, "pack": updated}


def _pdf_bytes(pack: dict[str, Any], items: list[dict[str, Any]], ev: dict[str, Any]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    teal = colors.HexColor(TEAL)
    red = colors.HexColor(RED)
    green = colors.HexColor(GREEN)
    amber = colors.HexColor(AMBER)
    buf = io.BytesIO()
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1t", parent=styles["Heading1"], textColor=teal, fontSize=16)
    h2 = ParagraphStyle("H2t", parent=styles["Heading2"], textColor=teal, fontSize=12)
    body = styles["BodyText"]
    small = ParagraphStyle("sm", parent=body, fontSize=8, leading=11)

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(18 * mm, 10 * mm, "IFRS.ai Audit Evidence Pack | Confidential")
        canvas.drawRightString(A4[0] - 18 * mm, 10 * mm, f"Page {doc.page}")
        canvas.restoreState()

    def tbl(data, col_w=None):
        t = Table(data, colWidths=col_w, repeatRows=1)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), teal),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ]
            )
        )
        return t

    score = float(pack.get("completeness_score") or 0)
    score_hex = GREEN if score >= 85 else (AMBER if score >= 70 else RED)
    c = ev["counts"]
    snap = (ev.get("rpo") or {}).get("snapshot") if ev.get("rpo") else {}
    snap = snap or {}

    story: list = [
        Paragraph("IFRS 15 Audit Evidence Pack", styles["Title"]),
        Spacer(1, 8),
        Paragraph(f"Workspace: {pack.get('company_id')} &nbsp;&nbsp; Period: {pack.get('period')} ({pack.get('period_type')})", body),
        Paragraph(f"Pack reference: <b>{pack.get('pack_ref')}</b>", body),
        Paragraph(f"Generated: {str(pack.get('generated_at') or '')[:19]} &nbsp; Prepared by: {pack.get('prepared_by') or '—'}", body),
        Spacer(1, 10),
        Paragraph(f"<font color='{score_hex}' size='22'><b>{score:.1f}%</b></font> completeness", body),
        Paragraph(f"Status: {pack.get('status')} &nbsp; Checklist met: {pack.get('checklist_items_met')}/{pack.get('checklist_items_total')}", body),
        PageBreak(),
        Paragraph("Executive Summary", h1),
        Spacer(1, 6),
    ]
    for line in str(pack.get("ai_executive_summary") or "").split("\n"):
        story.append(Paragraph((line or "&nbsp;").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;") if line else "&nbsp;", small))
    story += [
        Spacer(1, 10),
        tbl(
            [
                ["Metric", "Value"],
                ["Contracts", str(c.get("contracts_count", 0))],
                ["Modifications", str(c.get("modifications_count", 0))],
                ["Catch-up (AED)", f"{c.get('catch_up_total', 0):,.0f}"],
                ["Recon exceptions", f"{c.get('recon_exceptions_count', 0)} ({c.get('recon_exceptions_resolved', 0)} resolved)"],
                ["GL lines / manual", f"{c.get('je_count', 0)} / {c.get('je_manual_count', 0)}"],
                ["RPO total", str(snap.get("total_rpo") or "—")],
            ],
            [220, 250],
        ),
        PageBreak(),
        Paragraph("Controls Summary", h1),
        Spacer(1, 6),
    ]
    for line in str(pack.get("ai_controls_narrative") or "").split("\n"):
        safe = (line or "&nbsp;").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;") if line else "&nbsp;"
        story.append(Paragraph(safe, small))
    ctrl_rows = [["Code", "Requirement", "Status"]]
    for it in items:
        if it.get("section") == "controls":
            ctrl_rows.append([it.get("item_code"), Paragraph(str(it.get("requirement") or ""), small), str(it.get("status"))])
    story += [Spacer(1, 8), tbl(ctrl_rows, [70, 320, 80]), PageBreak(), Paragraph("Contract Population", h1), Spacer(1, 6)]
    crow = [["Ref", "Customer", "Type", "Price", "POBs / SSP", "§121?"]]
    for ct in ev.get("contracts") or []:
        crow.append(
            [
                str(ct.get("contract_ref") or ""),
                Paragraph(str(ct.get("customer") or "")[:40], small),
                str(ct.get("contract_type") or ""),
                f"{float(ct.get('transaction_price') or 0):,.0f}",
                Paragraph(str(ct.get("ssp_method") or "")[:48], small),
                "Yes" if ct.get("practical_expedient") else "No",
            ]
        )
    if len(crow) == 1:
        crow.append(["—", "No contracts", "—", "—", "—", "—"])
    story += [tbl(crow, [85, 110, 90, 70, 120, 40]), PageBreak(), Paragraph("Contract Modifications", h1), Spacer(1, 6)]
    mrow = [["Ref", "Date", "Description", "Treatment", "Catch-up"]]
    for m in ev.get("modifications") or []:
        treat = m.get("human_treatment_override") or m.get("ai_treatment") or ""
        mrow.append(
            [
                str(m.get("modification_ref") or ""),
                str(m.get("modification_date") or "")[:10],
                Paragraph(str(m.get("description") or "")[:80], small),
                str(treat),
                f"{float(m.get('catch_up_adjustment') or 0):,.0f}",
            ]
        )
    if len(mrow) == 1:
        mrow.append(["—", "—", "No modifications in period", "—", "—"])
    story += [
        tbl(mrow, [90, 70, 200, 90, 60]),
        Paragraph(f"Total catch-up impact: AED {c.get('catch_up_total', 0):,.0f}", body),
        PageBreak(),
        Paragraph("Billing Reconciliation", h1),
        Spacer(1, 6),
    ]
    rrow = [["Contract / type", "Billing", "GL revenue", "Variance", "Status", "Reviewed"]]
    for r in ev.get("recon") or []:
        rrow.append(
            [
                str(r.get("contract_id") or r.get("contract_type") or ""),
                f"{float(r.get('billing_total') or 0):,.0f}",
                f"{float(r.get('gl_revenue_total') or 0):,.0f}",
                f"{float(r.get('variance') or 0):,.0f}",
                str(r.get("status") or ""),
                str(r.get("reviewed_by") or "—"),
            ]
        )
    if len(rrow) == 1:
        rrow.append(["—", "—", "—", "—", "No recon", "—"])
    story += [tbl(rrow, [110, 75, 80, 70, 80, 70]), PageBreak(), Paragraph("RPO Disclosure (IFRS 15 §120)", h1), Spacer(1, 6)]
    story.append(
        tbl(
            [
                ["Time band", "Amount"],
                ["Within 1 year", str(snap.get("bucket_lt_1yr") or 0)],
                ["1–2 years", str(snap.get("bucket_1_2yr") or 0)],
                ["2–5 years", str(snap.get("bucket_2_5yr") or 0)],
                ["> 5 years", str(snap.get("bucket_gt_5yr") or 0)],
                ["Total RPO", str(snap.get("total_rpo") or 0)],
            ],
            [220, 200],
        )
    )
    story.append(Paragraph(f"Coverage ratio: {snap.get('rpo_coverage_ratio') or '—'}", body))
    draft = str(snap.get("ai_disclosure_draft") or "No RPO disclosure draft on snapshot.")
    for line in draft.split("\n")[:40]:
        safe = (line or "&nbsp;").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;") if line else "&nbsp;"
        story.append(Paragraph(safe, small))
    story += [PageBreak(), Paragraph("Completeness Checklist", h1), Spacer(1, 6)]
    icons = {"met": "✓ met", "partial": "~ partial", "gap": "✗ gap", "not_applicable": "— N/A", "not_assessed": "?"}
    ch = [["Code", "Requirement", "Ref", "Status"]]
    for it in items:
        ch.append(
            [
                str(it.get("item_code")),
                Paragraph(str(it.get("requirement") or ""), small),
                str(it.get("ifrs_reference") or ""),
                icons.get(str(it.get("status")), str(it.get("status"))),
            ]
        )
    story.append(tbl(ch, [60, 280, 90, 70]))
    sm = checklist_summary(items)
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"Score {sm['score']:.1f}% &nbsp; Met {sm['met']} &nbsp; Partial {sm['partial']} &nbsp; Gaps {sm['gap']} &nbsp; N/A {sm['not_applicable']}", body))

    doc = SimpleDocTemplate(buf, pagesize=A4, title=str(pack.get("pack_ref") or "AEP"), leftMargin=16 * mm, rightMargin=16 * mm, topMargin=14 * mm, bottomMargin=16 * mm)
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return buf.getvalue()


def _excel_bytes(pack: dict[str, Any], items: list[dict[str, Any]], ev: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    teal_fill = PatternFill("solid", fgColor="0D9488")
    head_font = Font(color="FFFFFF", bold=True)
    green_fill = PatternFill("solid", fgColor="BBF7D0")
    red_fill = PatternFill("solid", fgColor="FECACA")
    amber_fill = PatternFill("solid", fgColor="FDE68A")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    def header(ws, headers):
        for i, h in enumerate(headers, 1):
            cell = ws.cell(1, i, h)
            cell.fill = teal_fill
            cell.font = head_font
            cell.alignment = Alignment(wrap_text=True)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22

    c = ev["counts"]
    ws1 = wb.active
    ws1.title = "Cover"
    ws1["A1"] = "IFRS 15 Audit Evidence Pack"
    ws1["A1"].font = Font(bold=True, size=16, color="0D9488")
    rows1 = [
        ("Pack ref", pack.get("pack_ref")),
        ("Company", pack.get("company_id")),
        ("Period", pack.get("period")),
        ("Period type", pack.get("period_type")),
        ("Status", pack.get("status")),
        ("Completeness %", pack.get("completeness_score")),
        ("Prepared by", pack.get("prepared_by")),
        ("Generated", str(pack.get("generated_at") or "")[:19]),
        ("Contracts", c.get("contracts_count")),
        ("Modifications", c.get("modifications_count")),
        ("Exceptions", c.get("recon_exceptions_count")),
        ("JE / manual", f"{c.get('je_count')}/{c.get('je_manual_count')}"),
    ]
    for i, (k, v) in enumerate(rows1, 3):
        ws1.cell(i, 1, k).font = Font(bold=True)
        ws1.cell(i, 2, v)
    ws1["A16"] = "Executive summary"
    ws1["A16"].font = Font(bold=True)
    ws1["A17"] = str(pack.get("ai_executive_summary") or "")
    ws1["A17"].alignment = Alignment(wrap_text=True, vertical="top")
    ws1.merge_cells("A17:F28")
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 40

    ws2 = wb.create_sheet("Contracts")
    header(ws2, ["Ref", "Customer", "Type", "Price", "Recognised", "Term", "Expedient", "SSP method"])
    for i, ct in enumerate(ev.get("contracts") or [], 2):
        ws2.cell(i, 1, ct.get("contract_ref"))
        ws2.cell(i, 2, ct.get("customer"))
        ws2.cell(i, 3, ct.get("contract_type"))
        ws2.cell(i, 4, float(ct.get("transaction_price") or 0))
        ws2.cell(i, 5, float(ct.get("revenue_recognised") or 0))
        ws2.cell(i, 6, ct.get("term_months"))
        ws2.cell(i, 7, "Yes" if ct.get("practical_expedient") else "No")
        ws2.cell(i, 8, ct.get("ssp_method"))

    ws3 = wb.create_sheet("Modifications")
    header(ws3, ["Ref", "Date", "Description", "AI treatment", "Override", "Catch-up", "Status", "Approved by"])
    for i, m in enumerate(ev.get("modifications") or [], 2):
        ws3.cell(i, 1, m.get("modification_ref"))
        ws3.cell(i, 2, str(m.get("modification_date") or "")[:10])
        ws3.cell(i, 3, m.get("description"))
        ws3.cell(i, 4, m.get("ai_treatment"))
        ws3.cell(i, 5, m.get("human_treatment_override"))
        ws3.cell(i, 6, float(m.get("catch_up_adjustment") or 0))
        ws3.cell(i, 7, m.get("status"))
        ws3.cell(i, 8, m.get("approved_by"))

    ws4 = wb.create_sheet("Billing recon")
    header(ws4, ["Contract", "Type", "Period", "Billing", "GL revenue", "Deferred", "Variance", "Status", "Reviewed by", "Commentary"])
    for i, r in enumerate(ev.get("recon") or [], 2):
        ws4.cell(i, 1, r.get("contract_id"))
        ws4.cell(i, 2, r.get("contract_type"))
        ws4.cell(i, 3, r.get("period"))
        ws4.cell(i, 4, float(r.get("billing_total") or 0))
        ws4.cell(i, 5, float(r.get("gl_revenue_total") or 0))
        ws4.cell(i, 6, float(r.get("gl_deferred_total") or 0))
        ws4.cell(i, 7, float(r.get("variance") or 0))
        ws4.cell(i, 8, r.get("status"))
        ws4.cell(i, 9, r.get("reviewed_by"))
        ws4.cell(i, 10, r.get("ai_commentary"))

    ws5 = wb.create_sheet("RPO detail")
    snap = (ev.get("rpo") or {}).get("snapshot") if ev.get("rpo") else {}
    snap = snap or {}
    header(ws5, ["Field", "Value"])
    meta = [
        ("Period", snap.get("period")),
        ("Total RPO", snap.get("total_rpo")),
        ("Coverage", snap.get("rpo_coverage_ratio")),
        ("<1yr", snap.get("bucket_lt_1yr")),
        ("1-2yr", snap.get("bucket_1_2yr")),
        ("2-5yr", snap.get("bucket_2_5yr")),
        (">5yr", snap.get("bucket_gt_5yr")),
    ]
    for i, (k, v) in enumerate(meta, 2):
        ws5.cell(i, 1, k)
        ws5.cell(i, 2, v)
    details = ((ev.get("rpo") or {}).get("details") if ev.get("rpo") else []) or []
    ws5.cell(10, 1, "Contract-level RPO")
    subh = ["Ref", "Customer", "Type", "TP", "Recognised", "RPO", "Bucket", "Expedient", "Status"]
    for i, h in enumerate(subh, 1):
        cell = ws5.cell(11, i, h)
        cell.fill = teal_fill
        cell.font = head_font
    for i, d in enumerate(details, 12):
        ws5.cell(i, 1, d.get("contract_ref"))
        ws5.cell(i, 2, d.get("customer_name"))
        ws5.cell(i, 3, d.get("contract_type"))
        ws5.cell(i, 4, d.get("transaction_price"))
        ws5.cell(i, 5, d.get("revenue_recognised"))
        ws5.cell(i, 6, d.get("rpo"))
        ws5.cell(i, 7, d.get("time_bucket"))
        ws5.cell(i, 8, d.get("practical_expedient_applies"))
        ws5.cell(i, 9, d.get("status"))

    ws6 = wb.create_sheet("Checklist")
    header(ws6, ["Section", "Code", "Requirement", "IFRS ref", "Status", "Evidence source", "Notes", "Gap", "Action"])
    for i, it in enumerate(items, 2):
        ws6.cell(i, 1, it.get("section"))
        ws6.cell(i, 2, it.get("item_code"))
        ws6.cell(i, 3, it.get("requirement"))
        ws6.cell(i, 4, it.get("ifrs_reference"))
        ws6.cell(i, 5, it.get("status"))
        ws6.cell(i, 6, it.get("evidence_source"))
        ws6.cell(i, 7, it.get("notes"))
        ws6.cell(i, 8, it.get("gap_description"))
        ws6.cell(i, 9, it.get("recommended_action"))
        for col in range(1, 10):
            ws6.cell(i, col).border = thin
    last = max(2, len(items) + 1)
    ws6.conditional_formatting.add(f"E2:E{last}", CellIsRule(operator="equal", formula=['"met"'], fill=green_fill))
    ws6.conditional_formatting.add(f"E2:E{last}", CellIsRule(operator="equal", formula=['"gap"'], fill=red_fill))
    ws6.conditional_formatting.add(f"E2:E{last}", CellIsRule(operator="equal", formula=['"partial"'], fill=amber_fill))
    ws6.column_dimensions["C"].width = 55

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@router.post("/{pack_id}/export-pdf")
async def export_pdf(pack_id: str):
    _require_db()
    pack = evidence_pack_db.get(pack_id)
    if not pack:
        raise HTTPException(status_code=404, detail="Pack not found")
    items = evidence_pack_db.list_checklist(pack_id)
    ev = gather_evidence(str(pack["company_id"]), str(pack["period"]), str(pack.get("period_type") or "monthly"))
    data = _pdf_bytes(pack, items, ev)
    evidence_pack_db.update_pack(pack_id, {"pdf_path": f"generated:{_now()}", "updated_at": _now()})
    ref = str(pack.get("pack_ref") or "AEP").replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{ref}.pdf"'},
    )


@router.post("/{pack_id}/export-excel")
async def export_excel(pack_id: str):
    _require_db()
    pack = evidence_pack_db.get(pack_id)
    if not pack:
        raise HTTPException(status_code=404, detail="Pack not found")
    items = evidence_pack_db.list_checklist(pack_id)
    ev = gather_evidence(str(pack["company_id"]), str(pack["period"]), str(pack.get("period_type") or "monthly"))
    data = _excel_bytes(pack, items, ev)
    evidence_pack_db.update_pack(pack_id, {"excel_path": f"generated:{_now()}", "updated_at": _now()})
    ref = str(pack.get("pack_ref") or "AEP").replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{ref}.xlsx"'},
    )
