"""
IFRS 16 Excel Export Module
Professional Excel workbook generation with multiple sheets
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart
import pandas as pd
from typing import Any, Dict
from datetime import datetime
from pathlib import Path

from currency_format import format_currency_value, excel_money_number_format, currency_display_symbol


def normalize_results_for_excel_export(results: Dict[str, Any]) -> Dict[str, Any]:
    """
    API/JSON payloads use amortization_schedule as list[dict]; exporter expects a DataFrame.
    """
    out = dict(results)
    am = out.get("amortization_schedule")
    if am is not None and not isinstance(am, pd.DataFrame):
        if isinstance(am, list):
            out["amortization_schedule"] = pd.DataFrame(am)
        elif isinstance(am, dict):
            out["amortization_schedule"] = pd.DataFrame(am)
    return out


class IFRS16ExcelExporter:
    """Export IFRS 16 calculations to professional Excel format"""
    
    def __init__(self):
        # Styling
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=14, color="1F4E78")
        self.subtitle_font = Font(bold=True, size=12, color="1F4E78")
        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        self.thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

    def _money_fmt(self, results: Dict[str, Any]) -> str:
        return excel_money_number_format(
            (results.get("disclosure_data") or {}).get("currency") or "INR"
        )
    
    def _build_workbook(self, results: Dict[str, Any]) -> Workbook:
        results = normalize_results_for_excel_export(results)
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        self._create_summary_sheet(wb, results)
        self._create_amortization_sheet(wb, results)
        self._create_journal_entries_sheet(wb, results)
        self._create_maturity_sheet(wb, results)
        self._create_disclosure_sheet(wb, results)
        return wb

    def export_ifrs16_workbook(self, results: Dict, filename: str):
        """
        Create complete IFRS 16 Excel workbook with multiple sheets
        
        Sheets:
        1. Summary
        2. Amortization Schedule
        3. Journal Entries
        4. Maturity Analysis
        5. Disclosure Notes
        
        Args:
            results: Dictionary from IFRS16Calculator.calculate_full_ifrs16()
            filename: Output filename (path)
        """
        output_path = Path(filename)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = self._build_workbook(results)
        wb.save(filename)
        print(f"Excel workbook saved: {filename}")

    def export_ifrs16_workbook_bytes(self, results: Dict[str, Any]) -> bytes:
        """Build workbook in memory (no disk). For direct HTTP download after bulk-calculate JSON."""
        wb = self._build_workbook(dict(results))
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
    
    def _create_summary_sheet(self, wb: Workbook, results: Dict):
        """Summary sheet with key metrics and overview"""
        money_fmt = self._money_fmt(results)

        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "IFRS 16 LEASE ACCOUNTING SUMMARY"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Lease details section
        row = 3
        ws[f'A{row}'] = "LEASE DETAILS"
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        lease_details = [
            ("Lease ID:", results['disclosure_data']['lease_id']),
            ("Asset:", results['disclosure_data']['asset']),
            ("Lessee:", results['disclosure_data']['lessee']),
            ("Lessor:", results['disclosure_data']['lessor']),
            ("Commencement Date:", results['disclosure_data']['commencement']),
            ("End Date:", results['disclosure_data']['end_date']),
            ("Lease Term:", f"{results['disclosure_data']['term_months']} months"),
            ("Discount Rate:", f"{results['disclosure_data']['discount_rate_pct']:.2f}%"),
            ("Currency:", results['disclosure_data']['currency'])
        ]
        
        for label, value in lease_details:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True, size=10)
            row += 1
        
        # Initial measurement section
        row += 1
        ws[f'A{row}'] = "INITIAL MEASUREMENT"
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Amount"
        ws[f'A{row}'].font = self.header_font
        ws[f'B{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws[f'B{row}'].fill = self.header_fill

        metrics = [
            ("Lease Liability", results['lease_liability']),
            ("Right-of-Use Asset", results['rou_asset']),
            ("Monthly Depreciation", results['monthly_depreciation']),
            ("Total Interest Expense", results['total_interest'])
        ]
        
        for metric_name, value in metrics:
            row += 1
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = money_fmt
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border

        comp = results.get("component_analysis") or {}
        if comp.get("non_lease_component", 0) or comp.get("lease_component"):
            row += 2
            ws[f'A{row}'] = "LEASE vs NON-LEASE COMPONENTS"
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            component_rows = [
                ("Total monthly contract payment", comp.get("total_monthly_payment", 0)),
                ("Lease component (in liability schedule)", comp.get("lease_component", 0)),
                ("Non-lease component (P&L straight-line)", comp.get("non_lease_component", 0)),
            ]
            for label, value in component_rows:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = money_fmt
                row += 1
            if comp.get("non_lease_description"):
                ws[f'A{row}'] = "Non-lease description"
                ws[f'B{row}'] = comp.get("non_lease_description")
                row += 1

        if results.get("incentive_disclosure_note"):
            row += 1
            ws[f'A{row}'] = "Rent-free / incentives"
            ws[f'B{row}'] = results["incentive_disclosure_note"]
            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        # Lease liability breakdown (if RVG present)
        if results.get('liability_breakdown') and results['liability_breakdown'].get('pv_residual_value_guarantee', 0) != 0:
            row += 2
            ws[f'A{row}'] = "LEASE LIABILITY CALCULATION"
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            lb = results['liability_breakdown']
            liability_items = [
                ("PV of lease payments", lb.get('pv_regular_payments', 0)),
                ("PV of residual value guarantee", lb.get('pv_residual_value_guarantee', 0)),
                ("Total lease liability", lb.get('total_lease_liability', 0)),
            ]
            for label, val in liability_items:
                row += 1
                ws[f'A{row}'] = label
                ws[f'B{row}'] = val
                ws[f'B{row}'].number_format = money_fmt
                ws[f'A{row}'].border = self.border
                ws[f'B{row}'].border = self.border

        # ROU Asset build-up (if lease incentives or IDC present)
        rb = results.get('rou_build_up')
        if rb and (rb.get('less_lease_incentives', 0) != 0 or rb.get('add_initial_direct_costs', 0) != 0):
            row += 2
            ws[f'A{row}'] = "ROU ASSET CALCULATION"
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            build_up = [("PV of lease payments", rb.get('pv_lease_payments', 0))]
            if rb.get('add_initial_direct_costs', 0) != 0:
                if rb.get('legal_fees', 0) != 0:
                    build_up.append(("  Legal fees", rb.get('legal_fees', 0)))
                if rb.get('brokerage_fees', 0) != 0:
                    build_up.append(("  Brokerage / agent fees", rb.get('brokerage_fees', 0)))
                if rb.get('other_initial_direct_costs', 0) != 0:
                    build_up.append(("  Other initial direct costs", rb.get('other_initial_direct_costs', 0)))
                build_up.append(("Add: Initial direct costs", rb.get('add_initial_direct_costs', 0)))
            if rb.get('less_lease_incentives', 0) != 0:
                build_up.append(("Less: Lease incentives", -float(rb.get('less_lease_incentives', 0))))
            build_up.append(("ROU Asset at commencement", rb.get('rou_asset_at_commencement', 0)))
            for label, val in build_up:
                row += 1
                ws[f'A{row}'] = label
                ws[f'B{row}'] = val
                ws[f'B{row}'].number_format = money_fmt
                ws[f'A{row}'].border = self.border
                ws[f'B{row}'].border = self.border
        
        # Balance sheet classification
        row += 2
        ws[f'A{row}'] = "BALANCE SHEET CLASSIFICATION"
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = "Classification"
        ws[f'B{row}'] = "Amount"
        ws[f'A{row}'].font = self.header_font
        ws[f'B{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws[f'B{row}'].fill = self.header_fill
        
        bs_items = [
            ("Current Portion", results['liability_split']['current_portion']),
            ("Non-Current Portion", results['liability_split']['non_current_portion']),
            ("Total Lease Liability", results['liability_split']['total_liability'])
        ]
        
        for item_name, value in bs_items:
            row += 1
            ws[f'A{row}'] = item_name
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = money_fmt
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            
            if item_name == "Total Lease Liability":
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
        
        # Year 1 P&L Impact
        row += 2
        ws[f'A{row}'] = "YEAR 1 P&L IMPACT"
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = "Item"
        ws[f'B{row}'] = "Amount"
        ws[f'A{row}'].font = self.header_font
        ws[f'B{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws[f'B{row}'].fill = self.header_fill
        
        y1_items = [
            ("Interest Expense", results['year_1_impact']['interest_expense']),
            ("Depreciation Expense", results['year_1_impact']['depreciation_expense']),
            ("Total P&L Expense", results['year_1_impact']['total_p_l_expense']),
            ("Cash Outflow", results['year_1_impact']['cash_outflow']),
            ("EBITDA Improvement", results['year_1_impact']['ebitda_improvement'])
        ]
        
        for item_name, value in y1_items:
            row += 1
            ws[f'A{row}'] = item_name
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = money_fmt
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            
            if item_name == "Total P&L Expense":
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
        
        # Total lease cost
        row += 2
        ws[f'A{row}'] = "TOTAL LEASE COST"
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = "Component"
        ws[f'B{row}'] = "Amount"
        ws[f'A{row}'].font = self.header_font
        ws[f'B{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws[f'B{row}'].fill = self.header_fill
        
        cost_items = [
            ("Total Lease Payments", results['total_lease_cost']['total_payments']),
            ("Initial Direct Costs", results['total_lease_cost']['initial_costs']),
            ("Total Interest Component", results['total_lease_cost']['total_interest']),
            ("Grand Total", results['total_lease_cost']['grand_total'])
        ]
        
        for item_name, value in cost_items:
            row += 1
            ws[f'A{row}'] = item_name
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = money_fmt
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            
            if item_name == "Grand Total":
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'A{row}'].border = self.thick_border
                ws[f'B{row}'].border = self.thick_border
        
        # Metadata
        row += 3
        ws[f'A{row}'] = f"Generated: {results['calculation_metadata']['calculation_date']}"
        ws[f'A{row}'].font = Font(size=9, italic=True, color="666666")
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
    
    def _create_amortization_sheet(self, wb: Workbook, results: Dict):
        """Amortization schedule sheet with full payment table"""
        mf = excel_money_number_format(results.get('disclosure_data', {}).get('currency') or 'INR')

        ws = wb.create_sheet("Amortization Schedule")
        
        # Title
        ws['A1'] = "LEASE LIABILITY AMORTIZATION SCHEDULE"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25
        
        # Subtitle
        ws['A2'] = f"Lease ID: {results['disclosure_data']['lease_id']}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:H2')
        
        # Add DataFrame to sheet
        df = results['amortization_schedule']
        
        row_offset = 4
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), row_offset):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Header row
                if r_idx == row_offset:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    # Format numbers
                    if c_idx >= 4 and c_idx <= 8:  # Amount columns
                        cell.number_format = mf
                
                cell.border = self.border
        
        # Auto-size columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
    
    def _create_journal_entries_sheet(self, wb: Workbook, results: Dict):
        """Journal entries sheet with all accounting entries"""
        ccy = results.get('disclosure_data', {}).get('currency') or 'INR'
        currency_symbol = currency_display_symbol(ccy)
        mf = excel_money_number_format(ccy)

        ws = wb.create_sheet("Journal Entries")
        
        ws['A1'] = "ACCOUNTING JOURNAL ENTRIES"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25
        
        row = 3
        
        # Entry 1: Initial Recognition
        ws[f'A{row}'] = "1. INITIAL RECOGNITION"
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = "Date: " + results['journal_entries']['initial_recognition']['date']
        ws[f'A{row}'].font = Font(italic=True)
        row += 1
        ws[f'A{row}'] = "Description: " + results['journal_entries']['initial_recognition']['description']
        ws[f'A{row}'].font = Font(italic=True)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = "Account"
        ws[f'B{row}'] = f"Debit ({currency_symbol})"
        ws[f'C{row}'] = f"Credit ({currency_symbol})"
        ws[f'D{row}'] = "Narration"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = self.header_font
            ws[f'{col}{row}'].fill = self.header_fill
            ws[f'{col}{row}'].border = self.border
        
        for entry in results['journal_entries']['initial_recognition']['entries']:
            row += 1
            ws[f'A{row}'] = entry['account']
            ws[f'B{row}'] = entry['dr'] if entry['dr'] > 0 else ""
            ws[f'C{row}'] = entry['cr'] if entry['cr'] > 0 else ""
            ws[f'D{row}'] = entry['narration']
            
            if entry['dr'] > 0:
                ws[f'B{row}'].number_format = mf
            if entry['cr'] > 0:
                ws[f'C{row}'].number_format = mf
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = self.border
        
        # Entry 2: Monthly Depreciation
        row += 2
        ws[f'A{row}'] = "2. MONTHLY DEPRECIATION (Recurring)"
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = "Frequency: " + results['journal_entries']['monthly_depreciation']['frequency']
        ws[f'A{row}'].font = Font(italic=True)
        row += 1
        
        ws[f'A{row}'] = "Account"
        ws[f'B{row}'] = f"Debit ({currency_symbol})"
        ws[f'C{row}'] = f"Credit ({currency_symbol})"
        ws[f'D{row}'] = "Narration"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = self.header_font
            ws[f'{col}{row}'].fill = self.header_fill
            ws[f'{col}{row}'].border = self.border
        
        for entry in results['journal_entries']['monthly_depreciation']['entries']:
            row += 1
            ws[f'A{row}'] = entry['account']
            ws[f'B{row}'] = entry['dr'] if entry['dr'] > 0 else ""
            ws[f'C{row}'] = entry['cr'] if entry['cr'] > 0 else ""
            ws[f'D{row}'] = entry['narration']
            
            if entry['dr'] > 0:
                ws[f'B{row}'].number_format = mf
            if entry['cr'] > 0:
                ws[f'C{row}'].number_format = mf
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = self.border
        
        # Entry 3: Monthly Payment
        row += 2
        ws[f'A{row}'] = "3. MONTHLY LEASE PAYMENT (Example - Month 1)"
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws[f'A{row}'] = "Frequency: " + results['journal_entries']['monthly_payment_example']['frequency']
        ws[f'A{row}'].font = Font(italic=True)
        row += 1
        
        ws[f'A{row}'] = "Account"
        ws[f'B{row}'] = f"Debit ({currency_symbol})"
        ws[f'C{row}'] = f"Credit ({currency_symbol})"
        ws[f'D{row}'] = "Narration"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = self.header_font
            ws[f'{col}{row}'].fill = self.header_fill
            ws[f'{col}{row}'].border = self.border
        
        for entry in results['journal_entries']['monthly_payment_example']['entries']:
            row += 1
            ws[f'A{row}'] = entry['account']
            ws[f'B{row}'] = entry['dr'] if entry['dr'] > 0 else ""
            ws[f'C{row}'] = entry['cr'] if entry['cr'] > 0 else ""
            ws[f'D{row}'] = entry['narration']
            
            if entry['dr'] > 0:
                ws[f'B{row}'].number_format = mf
            if entry['cr'] > 0:
                ws[f'C{row}'].number_format = mf
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = self.border
        
        # Column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 45
    
    def _create_maturity_sheet(self, wb: Workbook, results: Dict):
        """Maturity analysis sheet with payment breakdown by year"""
        mf = self._money_fmt(results)

        ws = wb.create_sheet("Maturity Analysis")
        
        ws['A1'] = "LEASE PAYMENT MATURITY ANALYSIS"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25
        
        ws['A2'] = "Future lease payments by period"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:B2')
        
        row = 4
        ws[f'A{row}'] = "Period"
        ws[f'B{row}'] = "Future Payments"
        ws[f'A{row}'].font = self.header_font
        ws[f'B{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws[f'B{row}'].fill = self.header_fill
        ws[f'A{row}'].border = self.border
        ws[f'B{row}'].border = self.border

        for period, amount in results['maturity_analysis'].items():
            row += 1
            ws[f'A{row}'] = period.replace('_', ' ')
            ws[f'B{row}'] = amount
            ws[f'B{row}'].number_format = mf
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            
            if period == 'Total':
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'A{row}'].border = self.thick_border
                ws[f'B{row}'].border = self.thick_border
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
    
    def _create_disclosure_sheet(self, wb: Workbook, results: Dict):
        """Disclosure notes sheet for financial statements"""
        
        ws = wb.create_sheet("Disclosure Notes")
        
        ws['A1'] = "IFRS 16 DISCLOSURE NOTES"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25
        
        row = 3
        ws[f'A{row}'] = "Note: Leases (IFRS 16)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 2
        disc_ccy = results.get('disclosure_data', {}).get('currency') or 'INR'
        y1 = results.get('year_1_impact') or {}
        ls = results.get('liability_split') or {}
        disclosure_text = f"""The Company has adopted IFRS 16 'Leases' with effect from {results['disclosure_data']['commencement']}.

The Company has lease contracts for office premises. The Company's obligations under its leases are secured by the lessor's title to the leased assets.

The Company has recognized right-of-use assets and lease liabilities for these leases. The details are as follows:

Right-of-Use Assets:
- Asset description: {results['disclosure_data']['asset']}
- Carrying amount: {format_currency_value(float(results['rou_asset']), disc_ccy)}
- Accumulated depreciation: Depreciated on a straight-line basis over lease term

Lease Liabilities:
- Total lease liability: {format_currency_value(float(results['lease_liability']), disc_ccy)}
- Current portion: {format_currency_value(float(ls.get('current_portion', 0)), disc_ccy)}
- Non-current portion: {format_currency_value(float(ls.get('non_current_portion', 0)), disc_ccy)}

The Company has used incremental borrowing rate of {results['disclosure_data']['discount_rate_pct']:.2f}% to calculate the present value of lease payments.

Amounts recognized in statement of profit and loss:
- Depreciation expense: {format_currency_value(float(y1.get('depreciation_expense', 0)), disc_ccy)} (Year 1)
- Interest expense: {format_currency_value(float(y1.get('interest_expense', 0)), disc_ccy)} (Year 1)
- Total expense: {format_currency_value(float(y1.get('total_p_l_expense', 0)), disc_ccy)} (Year 1)

Cash outflow for leases: {format_currency_value(float(y1.get('cash_outflow', 0)), disc_ccy)} (Year 1)

The maturity analysis of lease liabilities is presented in the 'Maturity Analysis' sheet.

Report generated on: {results['calculation_metadata']['calculation_date']}
"""
        
        # Write disclosure text with wrapping
        ws[f'A{row}'] = disclosure_text
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:B{row+50}')
        
        ws.column_dimensions['A'].width = 100
        ws.column_dimensions['B'].width = 20


# Example usage
if __name__ == "__main__":
    print("IFRS 16 Excel Exporter")
    print("="*70)
    print("\nThis module exports IFRS 16 calculations to Excel format.")
    print("\nUsage:")
    print("  from ifrs16_calculator import IFRS16Calculator, LeaseInput")
    print("  from ifrs16_excel_export import IFRS16ExcelExporter")
    print("  from datetime import datetime")
    print("  from decimal import Decimal")
    print()
    print("  # Calculate lease")
    print("  lease = LeaseInput(...)")
    print("  calc = IFRS16Calculator()")
    print("  results = calc.calculate_full_ifrs16(lease)")
    print()
    print("  # Export to Excel")
    print("  exporter = IFRS16ExcelExporter()")
    print("  exporter.export_ifrs16_workbook(results, 'output.xlsx')")
    print()
    print("See example_usage.py for a complete working example.")

