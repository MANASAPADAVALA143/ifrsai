#!/usr/bin/env python3
"""Throwaway IFRS 15 known-answer verification — diagnosis only, no engine patches."""

from __future__ import annotations

import math
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "backend"))

from ifrs15_calculator import (  # noqa: E402
    BillAndHoldInput,
    ContractCostInput,
    ContractModification,
    FinancingComponentInput,
    IFRS15Calculator,
    IFRS15VariableConsiderationEngine,
    LicenseIPInput,
    NonCashConsiderationInput,
    PrincipalAgentInput,
    WarrantyInput,
)

try:
    from app.services.ifrs15_realestate import OffPlanSalesEngine  # noqa: E402
except ImportError:
    from backend.app.services.ifrs15_realestate import OffPlanSalesEngine  # noqa: E402


def approx(a: float, b: float, tol: float = 1000.0) -> bool:
    return abs(a - b) <= tol


def row(num: str, case: str, expected: str, actual: str, status: str, notes: str = "") -> dict:
    return {
        "#": num,
        "Test case": case,
        "Expected": expected,
        "Engine output": actual,
        "Status": status,
        "Notes": notes,
    }


def main() -> None:
    calc = IFRS15Calculator()
    rows: list[dict] = []

    # 1 — Bill and hold
    bh = calc.assess_bill_and_hold(
        BillAndHoldInput(
            arrangement_id="BH-2025-FURN-009",
            contract_id="BH-2025-FURN-009",
            customer_name="Buyer",
            product_description="Furniture order",
            contract_value=3_200_000,
            expected_delivery_date="2026-06-01",
            billing_date="2025-12-01",
            reason_is_substantive=True,
            product_separately_identified=True,
            product_ready_for_transfer=True,
            entity_cannot_redirect=True,
            currency="AED",
        )
    )
    exp_bh = "REVENUE_RECOGNISABLE @ AED 3,200,000 on billing date"
    act_bh = f"{bh['conclusion']} @ AED {bh['revenue_recognisable_now']:,.0f}"
    rows.append(
        row(
            "1",
            "Bill-and-hold (BH-2025-FURN-009)",
            exp_bh,
            act_bh,
            "PASS" if bh["conclusion"] == "REVENUE_RECOGNISABLE" and bh["revenue_recognisable_now"] == 3_200_000 else "FAIL",
        )
    )

    # 2 — Contract costs
    cc = calc.calculate_contract_costs(
        [
            ContractCostInput(
                cost_id="COMM-001",
                contract_id="SPA-2025-VILLA-022",
                description="Broker commission",
                cost_type="incremental_obtaining",
                cost_amount=68_000,
                incurred_date="2025-01-15",
                contract_start="2025-01-01",
                contract_end="2027-06-30",  # ~30 months
                expected_renewal=False,
                currency="AED",
            )
        ]
    )
    item = cc["costs"][0]
    exp_monthly = round(68_000 / 30, 2)
    act_monthly = item.get("monthly_amortisation", 0)
    rows.append(
        row(
            "2",
            "Contract costs (SPA-2025-VILLA-022)",
            f"CAPITALISE AED 68,000; monthly ~ AED {exp_monthly:,.2f}",
            f"{item['treatment']}; {item.get('amortisation_period_months', 'n/a')} mo; monthly AED {act_monthly:,.2f}",
            "PASS"
            if item["treatment"] == "CAPITALISE"
            and item.get("amortisation_period_months") == 30
            and abs(act_monthly - exp_monthly) < 0.02
            else "MISMATCH-NEEDS-REVIEW",
            f"Engine uses contract_end-start months={item.get('amortisation_period_months')}",
        )
    )

    # 3 — Financing component
    fin = calc.calculate_financing_component(
        FinancingComponentInput(
            contract_id="LAND-2025-DEFER-005",
            description="Land sale deferred payment",
            contract_value=5_000_000,
            transfer_date="2025-01-01",
            payment_date="2028-01-01",
            payment_timing="deferred",
            discount_rate=7.0,
            currency="AED",
        )
    )
    exp_pv = round(5_000_000 / (1.07**3), 0)
    act_pv = fin["revenue_amount"]
    act_fin = fin["financing_amount"]
    rows.append(
        row(
            "3",
            "Financing component (LAND-2025-DEFER-005)",
            f"PV revenue ~ AED {exp_pv:,.0f}; interest ~ AED {5_000_000 - exp_pv:,.0f}",
            f"PV AED {act_pv:,.2f}; financing AED {act_fin:,.2f}; period_months={fin.get('period_months')}",
            "PASS" if approx(act_pv, exp_pv, 2000) else "MISMATCH-NEEDS-REVIEW",
            f"period_years uses months/12={fin.get('period_months', 0)/12:.2f}",
        )
    )

    # 4 — License IP
    lic = calc.assess_license_ip(
        LicenseIPInput(
            license_id="LIC-2025-PMSOFT-003",
            product_name="Property Management Software",
            license_description="SaaS hosted access",
            license_fee=540_000,
            license_start="2025-01-01",
            license_end="2027-12-31",
            is_perpetual=False,
            entity_activities_affect_ip=True,
            customer_exposed_to_effect=True,
            no_separate_functional_utility=True,
            currency="AED",
        )
    )
    monthly_lic = lic["recognition_schedule"][0]["amount"] if lic["recognition_schedule"] else 0
    rows.append(
        row(
            "4",
            "License/IP (LIC-2025-PMSOFT-003)",
            "RIGHT_TO_ACCESS; OVER_TIME; AED 15,000/month",
            f"{lic['license_type']}; {lic['recognition_basis']}; AED {monthly_lic:,.2f}/mo ({len(lic['recognition_schedule'])} periods)",
            "PASS"
            if lic["license_type"] == "RIGHT_TO_ACCESS"
            and lic["recognition_basis"] == "OVER_TIME"
            and abs(monthly_lic - 15_000) < 1
            else "MISMATCH-NEEDS-REVIEW",
        )
    )

    # 5 — Contract modification
    mod = calc.assess_modification(
        ContractModification(
            original_contract_id="FITOUT-2024-077",
            modification_date="2025-06-01",
            modification_description="Add Floor 9",
            new_goods_services=["Floor 9 fit-out"],
            price_change=1_150_000,
            remaining_transaction_price=720_000,
            remaining_performance_obligations=["Floor 8 fit-out"],
            original_ssps={"Floor 9 fit-out": 1_150_000, "Floor 8 fit-out": 1_200_000},
        )
    )
    rows.append(
        row(
            "5",
            "Modification (FITOUT-2024-077-MOD1)",
            "TYPE_1 New Separate Contract; no catch-up on original",
            f"{mod['modification_type']} — {mod['modification_type_name']}",
            "PASS" if mod["modification_type"] == "TYPE_1" else "MISMATCH-NEEDS-REVIEW",
            mod.get("explanation", "")[:120],
        )
    )

    # 6 — Principal vs agent
    pa = calc.assess_principal_agent(
        PrincipalAgentInput(
            arrangement_id="MKT-2025-AGENT-014",
            description="Property marketplace referral",
            third_party_involved=True,
            gross_contract_value=1_800_000,
            third_party_cost=1_764_000,  # net commission 36,000
            controls_before_transfer=False,
            primary_obligor=False,
            inventory_risk=False,
            pricing_discretion=False,
            credit_risk=False,
        )
    )
    act_pa = f"{pa['conclusion']}; {pa['revenue_treatment']}; AED {pa.get('revenue_recognised', pa.get('net_margin', 0)):,.0f}"
    rows.append(
        row(
            "6",
            "Principal vs agent (MKT-2025-AGENT-014)",
            "AGENT; NET revenue AED 36,000",
            act_pa,
            "PASS"
            if pa["conclusion"] == "AGENT" and pa.get("net_margin") == 36_000
            else "FAIL",
            "CRITICAL: must not return 1,800,000 gross",
        )
    )

    # 7 — Non-cash consideration
    nc = calc.calculate_non_cash_consideration(
        [
            NonCashConsiderationInput(
                item_id="LAND-PLOT",
                contract_id="SWAP-2025-LAND-CONST-002",
                description="Land plot received",
                consideration_type="goods",
                fair_value_determinable=True,
                fair_value=6_000_000,
                fallback_ssp=6_200_000,
                currency="AED",
            )
        ]
    )
    tp = nc["total_tp_from_non_cash"]
    rows.append(
        row(
            "7",
            "Non-cash / TP (SWAP-2025-LAND-CONST-002)",
            "Transaction price AED 6,000,000 (land FV)",
            f"AED {tp:,.0f} via {nc['items'][0]['measurement_method']}",
            "PASS" if tp == 6_000_000 else "FAIL",
            "Service cost-plus estimate 6.2M is fallback only",
        )
    )

    # 8 — Variable consideration + source inconsistency flag
    vc_engine = IFRS15VariableConsiderationEngine()
    ev_cap_480 = 0.30 * 600_000 - 0.20 * 480_000
    ev_cap_240 = 0.30 * 600_000 - 0.20 * 240_000
    scenarios_doc_style = [
        {"label": "bonus", "amount": 600_000, "probability": 0.30},
        {"label": "penalty", "amount": -240_000, "probability": 0.20},
        {"label": "base", "amount": 0, "probability": 0.50},
    ]
    vc = vc_engine.estimate(
        {
            "method": "expected_value",
            "scenarios": scenarios_doc_style,
            "constraint_factors": [True, True, False, False, False],
            "total_contract_value": 4_800_000,
        }
    )
    inconsistency = (
        "SOURCE INCONSISTENCY: contract states penalty cap AED 480,000 but EV formula uses "
        "AED 240,000 (20% × 240k). Engine EV with 240k penalty = "
        f"AED {ev_cap_240:,.0f}; with 480k cap would be AED {ev_cap_480:,.0f}."
    )
    stated_constrained_bonus = 300_000
    rows.append(
        row(
            "8",
            "Variable consideration (CON-ARM-2025-088)",
            f"Stated constrained variable inclusion ~ AED {stated_constrained_bonus:,.0f} (per contract); EV per doc math",
            f"EV AED {vc['expected_amount']:,.2f}; constrained AED {vc['constrained_amount']:,.2f}",
            "MISMATCH-NEEDS-REVIEW"
            if abs(vc["constrained_amount"] - stated_constrained_bonus) > 5000
            else "PASS",
            inconsistency
            + f" Stated '300k of bonus' vs engine constrained {vc['constrained_amount']:,.0f}.",
        )
    )

    # 9 — Warranties (two components)
    std = calc.classify_warranty(
        WarrantyInput(
            warranty_id="STD-12M",
            contract_id="HVAC-2025-SUPPLY-031",
            product_description="HVAC equipment",
            warranty_description="12-month standard warranty",
            warranty_period_months=12,
            warranty_value=0,
            required_by_law=False,
            covers_specs_only=True,
            customer_can_purchase_separately=False,
            provides_additional_service=False,
            currency="AED",
        )
    )
    ext = calc.classify_warranty(
        WarrantyInput(
            warranty_id="EXT-3Y",
            contract_id="HVAC-2025-SUPPLY-031",
            product_description="HVAC extended service plan",
            warranty_description="3-year extended service plan",
            warranty_period_months=36,
            warranty_value=95_000,
            required_by_law=False,
            covers_specs_only=False,
            customer_can_purchase_separately=True,
            provides_additional_service=True,
            allocated_fee=95_000,
            currency="AED",
        )
    )
    rows.append(
        row(
            "9a",
            "Warranty standard 12m (HVAC-2025-SUPPLY-031)",
            "ASSURANCE-TYPE; not separate PO",
            f"{std['warranty_type']}; {std.get('treatment', '')[:40]}",
            "PASS" if std["warranty_type"] == "ASSURANCE" else "FAIL",
        )
    )
    rows.append(
        row(
            "9b",
            "Warranty extended 3y (HVAC-2025-SUPPLY-031)",
            "SERVICE-TYPE; separate PO AED 95,000",
            f"{ext['warranty_type']}; allocated AED {ext.get('allocated_fee', ext.get('warranty_value', 0)):,.0f}",
            "PASS" if ext["warranty_type"] == "SERVICE" else "FAIL",
        )
    )

    # 10 — Excel / real estate (Emaar demo vs OffPlanSalesEngine)
    excel_candidates = [
        Path(r"C:\Users\HCSUSER\Downloads\Emaar_IFRS15_UAE_RealEstate_Demo (1).xlsx"),
        *ROOT.rglob("Emaar_IFRS15_UAE_RealEstate_Demo*.xlsx"),
    ]
    excel_path = next((p for p in excel_candidates if p.exists()), None)
    engine = OffPlanSalesEngine()

    def _compare_realestate_row(
        portfolio: dict,
        schedule: dict,
    ) -> dict:
        contract_value = float(portfolio["total_contract_value_aed"])
        completion = float(schedule["completion_pct"])
        if completion <= 1:
            completion *= 100
        cash = float(schedule["cash_received_to_date_aed"])
        rev_td_excel = float(schedule["revenue_recognised_to_date_aed"])
        rev_period = float(schedule["revenue_recognised_this_period_aed"])
        ca_excel = float(schedule["contract_asset_aed"] or 0)
        cl_excel = float(schedule["contract_liability_aed"] or 0)

        data = {
            "contract_value": contract_value,
            "rera_certificate_verified_pct": completion,
            "construction_start": str(portfolio.get("construction_start_date", ""))[:10],
            "expected_handover": str(portfolio.get("expected_handover_date", ""))[:10],
            "current_date": "2024-12-31",
            "revenue_prior_period": round(rev_td_excel - rev_period, 2),
            "escrow_receipts": [{"date": "2024-12-31", "amount": cash}],
        }
        out = engine.calculate(data)
        simple_rev = round(contract_value * completion / 100, 2)
        eng_rev = out["revenue_recognised_to_date"]
        eng_ca = out["contract_asset"]
        eng_cl = out["contract_liability"]

        rev_match = abs(eng_rev - rev_td_excel) < 1
        ca_match = abs(eng_ca - ca_excel) < 1
        cl_match = abs(eng_cl - cl_excel) < 1
        balances_match = ca_match and cl_match

        if rev_match and balances_match:
            status = "PASS"
            notes = "Revenue and contract balances match Excel schedule."
        elif rev_match and not balances_match:
            status = "MISMATCH-NEEDS-REVIEW"
            notes = (
                f"Revenue matches simple % (engine {eng_rev:,.0f} = Excel {rev_td_excel:,.0f}) "
                f"but CA/CL differ (engine CA {eng_ca:,.0f} vs Excel {ca_excel:,.0f}; "
                f"CL {eng_cl:,.0f} vs {cl_excel:,.0f}). "
                "Excel may use different billings basis than cash_received_to_date."
            )
        else:
            status = "MISMATCH-NEEDS-REVIEW"
            notes = (
                f"Revenue engine {eng_rev:,.0f} vs Excel {rev_td_excel:,.0f}; "
                f"simple % would be {simple_rev:,.0f}."
            )
        return {
            "expected": (
                f"Rev TD AED {rev_td_excel:,.0f}; CA {ca_excel:,.0f}; CL {cl_excel:,.0f} "
                f"({completion:.0f}% × {contract_value:,.0f})"
            ),
            "actual": f"Rev TD AED {eng_rev:,.0f}; CA {eng_ca:,.0f}; CL {eng_cl:,.0f}",
            "status": status,
            "notes": notes,
        }

    if excel_path:
        import openpyxl

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        port_ws = wb["Off-Plan Sales Portfolio"]
        rev_ws = wb["IFRS 15 Revenue Schedule"]
        portfolio_by_id: dict[str, dict] = {}
        for r in range(4, port_ws.max_row + 1):
            cid = port_ws.cell(r, 1).value
            if not cid or not str(cid).startswith("SPA-"):
                continue
            if str(cid) == "TOTAL":
                break
            portfolio_by_id[str(cid)] = {
                "contract_id": str(cid),
                "project_name": port_ws.cell(r, 2).value,
                "spa_date": port_ws.cell(r, 6).value,
                "total_contract_value_aed": port_ws.cell(r, 7).value,
                "construction_start_date": port_ws.cell(r, 10).value,
                "expected_handover_date": port_ws.cell(r, 11).value,
                "current_completion_pct": port_ws.cell(r, 12).value,
            }
        schedule_by_id: dict[str, dict] = {}
        for r in range(3, rev_ws.max_row + 1):
            cid = rev_ws.cell(r, 1).value
            if not cid or str(cid) == "TOTAL":
                continue
            schedule_by_id[str(cid)] = {
                "contract_id": str(cid),
                "completion_pct": rev_ws.cell(r, 4).value,
                "revenue_recognised_to_date_aed": rev_ws.cell(r, 5).value,
                "revenue_recognised_this_period_aed": rev_ws.cell(r, 6).value,
                "contract_asset_aed": rev_ws.cell(r, 7).value,
                "contract_liability_aed": rev_ws.cell(r, 8).value,
                "cash_received_to_date_aed": rev_ws.cell(r, 9).value,
            }

        pick = ["SPA-UAE-001", "SPA-UAE-004", "SPA-UAE-003"]
        for i, cid in enumerate(pick, 1):
            comp = _compare_realestate_row(portfolio_by_id[cid], schedule_by_id[cid])
            rows.append(
                row(
                    f"10{i}",
                    f"Emaar Excel — {cid} ({portfolio_by_id[cid]['project_name']})",
                    comp["expected"],
                    comp["actual"],
                    comp["status"],
                    comp["notes"],
                )
            )
    else:
        samples = [
            {
                "name": "Marina Heights (65% POC)",
                "contract_value": 2_000_000,
                "completion_pct": 65.0,
                "costs_incurred_to_date": 1_300_000,
                "total_estimated_costs": 2_000_000,
                "construction_start": "2023-01-01",
                "expected_handover": "2025-09-30",
                "current_date": "2024-12-31",
                "revenue_prior_period": 1_026_000,
                "escrow_receipts": [{"date": "2024-12-31", "amount": 1_200_000}],
            },
        ]
        for i, s in enumerate(samples, 1):
            data = {k: v for k, v in s.items() if k not in ("name", "completion_pct")}
            data["rera_certificate_verified_pct"] = s["completion_pct"]
            out = engine.calculate(data)
            simple_rev = round(s["contract_value"] * s["completion_pct"] / 100, 2)
            eng_rev = out["revenue_recognised_to_date"]
            match = abs(simple_rev - eng_rev) < 1
            rows.append(
                row(
                    f"10{i}",
                    f"Real estate proxy — {s['name']}",
                    f"Simple % rev TD AED {simple_rev:,.0f}",
                    f"Engine rev TD AED {eng_rev:,.0f}; CA {out['contract_asset']:,.0f}; CL {out['contract_liability']:,.0f}",
                    "PASS" if match else "MISMATCH-NEEDS-REVIEW",
                    "Emaar xlsx not found; Marina Heights fixture only.",
                )
            )

    # Print table
    headers = ["#", "Test case", "Expected", "Engine output", "Status", "Notes"]
    widths = [4, 42, 36, 40, 22, 50]
    print("| " + " | ".join(headers) + " |")
    print("|" + "|".join("-" * (w + 2) for w in widths) + "|")
    for r in rows:
        print(
            "| "
            + " | ".join(
                str(r.get(h, "")).replace("\n", " ")[: widths[i]]
                for i, h in enumerate(headers)
            )
            + " |"
        )

    fails = sum(1 for r in rows if r["Status"] in ("FAIL", "MISMATCH-NEEDS-REVIEW"))
    print(f"\nSummary: {len(rows)} checks, {len(rows)-fails} PASS, {fails} FAIL/REVIEW")


if __name__ == "__main__":
    main()
