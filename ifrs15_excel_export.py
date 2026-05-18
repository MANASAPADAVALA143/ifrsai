"""
IFRS 15 Excel Export Module
Professional Excel workbook generation with audit pack (master summary + core sheets + optional deferred revenue)
"""

import io
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from currency_format import format_currency_value, excel_money_number_format


class IFRS15ExcelExporter:
    """Export IFRS 15 calculation results to a professional workbook (optional master summary + five core sheets)."""

    def __init__(self) -> None:
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=14, color="1F4E78")
        self.subtitle_font = Font(bold=True, size=12, color="1F4E78")
        self.alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        self.total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        self.border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

    def _write_master_summary_placeholder(self, ws: Any) -> None:
        """Sheet 1 when no master report payload is supplied with the audit pack download."""
        dark_blue = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells("A1:D8")
        cell = ws["A1"]
        cell.value = (
            "IFRS 15 Master Summary\n\n"
            "Run all modules and Generate Master Report to populate this sheet."
        )
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = dark_blue
        for i in range(1, 9):
            ws.row_dimensions[i].height = 18
        ws.column_dimensions["A"].width = 44
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 44
        ws.column_dimensions["D"].width = 36

    def _write_master_summary_sheet(self, ws: Any, master_report: Dict[str, Any]) -> None:
        ccy = (master_report.get("contract_overview") or {}).get("currency") or "USD"
        ccy = str(ccy).upper()
        money_fmt = excel_money_number_format(ccy)

        dark_blue = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        orange_hdr = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        hdr_f = Font(bold=True, size=11, color="1F4E78")

        cust = master_report.get("customer_name") or "—"
        gen = master_report.get("generated_at") or ""
        ws.merge_cells("A1:D3")
        cell = ws["A1"]
        cell.value = f"IFRS 15 Revenue Recognition\nMaster Compliance Report\nContract: {cust} | Generated: {gen}"
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = dark_blue
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 18

        r = 5

        def section_header(title: str) -> None:
            nonlocal r
            ws.merge_cells(f"A{r}:D{r}")
            c = ws[f"A{r}"]
            c.value = title
            c.font = hdr_f
            c.fill = orange_hdr
            c.alignment = Alignment(horizontal="left", vertical="center")
            r += 1

        def row_pair(left: str, right: Any, money: bool = False) -> None:
            nonlocal r
            ws[f"A{r}"] = left
            ws[f"A{r}"].font = Font(bold=True, size=10)
            ws[f"A{r}"].border = self.border
            ws[f"B{r}"] = right
            ws[f"B{r}"].border = self.border
            if money and isinstance(right, (int, float)):
                ws[f"B{r}"].number_format = money_fmt
            r += 1

        co = master_report.get("contract_overview") or {}
        section_header("CONTRACT OVERVIEW")
        row_pair("Contract Value", float(co.get("total_contract_value", 0) or 0), True)
        row_pair("Term (months)", int(co.get("contract_term_months", 0) or 0))
        row_pair("Obligations", int(co.get("number_of_obligations", 0) or 0))
        row_pair("Recognition Pattern", co.get("recognition_pattern", "—"))
        row_pair("Currency", co.get("currency", "USD"))
        r += 1

        fs = master_report.get("financial_summary") or {}
        section_header("FINANCIAL SUMMARY")
        row_pair("Gross Revenue", float(fs.get("gross_revenue", 0) or 0), True)
        row_pair("Net Revenue", float(fs.get("net_revenue", 0) or 0), True)
        row_pair("Revenue Basis", fs.get("revenue_basis", "—"))
        row_pair("Deferred Revenue", float(fs.get("deferred_revenue", 0) or 0), True)
        row_pair("Contract Assets", float(fs.get("contract_assets", 0) or 0), True)
        row_pair("Variable Consideration (included)", float(fs.get("variable_consideration_included", 0) or 0), True)
        row_pair("Variable Consideration (constrained amt)", float(fs.get("variable_consideration_constrained", 0) or 0), True)
        row_pair("Commission Asset", float(fs.get("commission_asset_recognised", 0) or 0), True)
        row_pair("Commission Monthly Amortisation", float(fs.get("commission_monthly_amortisation", 0) or 0), True)
        row_pair("Total RPO", float(fs.get("total_rpo", 0) or 0), True)
        r += 1

        f5 = master_report.get("five_step_status") or {}
        section_header("5-STEP STATUS")
        row_pair("Step 1 — Contract identified", "✓" if f5.get("step1_contract_identified") else "✗")
        row_pair("Step 2 — Obligations count", int(f5.get("step2_obligations_identified", 0) or 0))
        row_pair("Step 3 — Transaction price", float(f5.get("step3_transaction_price", 0) or 0), True)
        row_pair("Step 3 — VC included", "✓" if f5.get("step3_variable_consideration_included") else "✗")
        row_pair("Step 4 — Allocation", f5.get("step4_allocation_method", "SSP"))
        row_pair("Step 5 — Revenue recognised", float(f5.get("step5_revenue_recognised", 0) or 0), True)
        row_pair("Step 5 — Deferred", float(f5.get("step5_deferred", 0) or 0), True)
        row_pair("All steps + modules complete", "✓" if f5.get("all_steps_complete") else "✗")
        r += 1

        asm = master_report.get("assessments") or {}
        section_header("ASSESSMENTS")
        ws[f"A{r}"] = "Module"
        ws[f"B{r}"] = "Assessed"
        ws[f"C{r}"] = "Key Finding"
        ws[f"D{r}"] = "Risk"
        for col in ("A", "B", "C", "D"):
            c = ws[f"{col}{r}"]
            c.font = self.header_font
            c.fill = self.header_fill
            c.border = self.border
        r += 1
        rows_asm = [
            ("Modifications", asm.get("modifications"), lambda x: f"Catch-up {x.get('total_catch_up_amount', 0)}", lambda b: "Yes" if b.get("risk_flag") else "No"),
            ("Variable consideration", asm.get("variable_consideration"), lambda x: f"{x.get('method', '')} / {x.get('constraint_level', '')}", lambda b: "Yes" if b.get("risk_flag") else "No"),
            ("RPO", asm.get("rpo"), lambda x: f"Total RPO {x.get('total_rpo', 0)}", lambda b: "—"),
            ("Contract costs", asm.get("contract_costs"), lambda x: f"Asset {x.get('asset_amount', 0)}", lambda b: "Yes" if b.get("impairment_flag") else "No"),
            ("Principal / Agent", asm.get("principal_agent"), lambda x: str(x.get("conclusion", "")), lambda b: "Yes" if b.get("borderline") else "No"),
            ("Licence of IP", asm.get("license"), lambda x: str(x.get("license_type", "")), lambda b: "Yes" if b.get("royalty_exception") else "No"),
        ]
        for name, block, fn, rfn in rows_asm:
            b = block or {}
            assessed = "Yes" if b.get("assessed") else "No"
            kf = fn(b) if b.get("assessed") else "—"
            rf = rfn(b) if b.get("assessed") else "—"
            ws[f"A{r}"] = name
            ws[f"B{r}"] = assessed
            ws[f"C{r}"] = kf
            ws[f"D{r}"] = rf
            for col in ("A", "B", "C", "D"):
                ws[f"{col}{r}"].border = self.border
            r += 1
        r += 1

        section_header("RISK FLAGS")
        risks = master_report.get("risk_flags") or []
        if not risks:
            ws.merge_cells(f"A{r}:D{r}")
            c = ws[f"A{r}"]
            c.value = "No risk flags identified"
            c.font = Font(bold=True, color="006100")
            c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            r += 1
        else:
            ws[f"A{r}"] = "Severity"
            ws[f"B{r}"] = "Module"
            ws[f"C{r}"] = "Message"
            ws[f"D{r}"] = "Action"
            for col in ("A", "B", "C", "D"):
                ws[f"{col}{r}"].font = self.header_font
                ws[f"{col}{r}"].fill = self.header_fill
                ws[f"{col}{r}"].border = self.border
            r += 1
            for rf in risks:
                sev = str(rf.get("severity", "LOW"))
                fill = (
                    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    if sev == "HIGH"
                    else PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                    if sev == "MEDIUM"
                    else PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                )
                ws[f"A{r}"] = sev
                ws[f"B{r}"] = rf.get("module", "")
                ws[f"C{r}"] = rf.get("message", "")
                ws[f"D{r}"] = rf.get("action_required", "")
                for col in ("A", "B", "C", "D"):
                    c = ws[f"{col}{r}"]
                    c.border = self.border
                    c.fill = fill
                r += 1
        r += 1

        ar = master_report.get("audit_readiness") or {}
        section_header("AUDIT READINESS")
        score = int(ar.get("score", 0) or 0)
        lvl = str(ar.get("level", ""))
        sc_fill = (
            PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            if lvl == "Ready"
            else PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            if lvl == "Needs Review"
            else PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        )
        ws.merge_cells(f"A{r}:D{r}")
        c = ws[f"A{r}"]
        c.value = f"Audit Readiness Score: {score}/100 — {lvl}"
        c.font = Font(bold=True, size=14)
        c.fill = sc_fill
        c.alignment = Alignment(horizontal="center")
        r += 1
        ws[f"A{r}"] = "Item"
        ws[f"B{r}"] = "Status"
        ws[f"C{r}"] = "Note"
        for col in ("A", "B", "C"):
            ws[f"{col}{r}"].font = self.header_font
            ws[f"{col}{r}"].fill = self.header_fill
            ws[f"{col}{r}"].border = self.border
        r += 1
        for it in ar.get("checklist") or []:
            st = str(it.get("status", ""))
            colf = Font(color="006100") if st == "complete" else Font(color="9CA3AF") if st == "not_applicable" else Font(color="C00000")
            ws[f"A{r}"] = it.get("item", "")
            ws[f"B{r}"] = st
            ws[f"C{r}"] = it.get("note", "")
            ws[f"A{r}"].font = colf
            ws[f"B{r}"].font = colf
            for col in ("A", "B", "C"):
                ws[f"{col}{r}"].border = self.border
            r += 1
        r += 1

        section_header("AI NARRATIVE")
        ws.merge_cells(f"A{r}:D{r + 20}")
        nar = ws[f"A{r}"]
        nar.value = master_report.get("ai_narrative") or "—"
        nar.font = Font(name="Times New Roman", size=11)
        nar.alignment = Alignment(wrap_text=True, vertical="top")
        nar.border = self.border

        ws.column_dimensions["A"].width = 44
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 44
        ws.column_dimensions["D"].width = 36

    def export_ifrs15_workbook(
        self,
        results: Dict[str, Any],
        contract_meta: Dict[str, Any],
        filename: str,
        master_report: Optional[Dict[str, Any]] = None,
    ) -> None:
        output_path = Path(filename)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        currency = (contract_meta.get("currency") or "USD").upper()
        money_fmt = excel_money_number_format(currency)

        ws_master = wb.create_sheet("IFRS 15 Master Summary", 0)
        if master_report is not None:
            self._write_master_summary_sheet(ws_master, master_report)
        else:
            self._write_master_summary_placeholder(ws_master)

        self._create_summary_sheet(wb, results, contract_meta, money_fmt)
        self._create_schedule_sheet(wb, results, contract_meta, money_fmt)
        self._create_journal_entries_sheet(wb, results, money_fmt)
        self._create_obligations_sheet(wb, results, contract_meta, money_fmt)
        self._create_disclosures_sheet(wb, results, contract_meta, money_fmt, currency)
        drf = contract_meta.get("deferred_revenue_rollforward") or results.get("deferred_revenue_rollforward")
        if drf and isinstance(drf, dict):
            self._create_deferred_revenue_sheet(wb, drf, money_fmt)

        rpo120 = contract_meta.get("rpo_disclosure_ifrs120") or results.get("rpo_disclosure_ifrs120")
        if rpo120 and isinstance(rpo120, dict):
            self._create_rpo_disclosure_sheet(wb, rpo120, money_fmt)

        pa_hist = contract_meta.get("principal_agent_history") or results.get("principal_agent_history")
        if pa_hist and isinstance(pa_hist, list) and len(pa_hist) > 0:
            self._create_principal_agent_sheet(wb, pa_hist, money_fmt)

        cc_batch = contract_meta.get("contract_costs_ifrs9194") or results.get("contract_costs_ifrs9194")
        if cc_batch and isinstance(cc_batch, dict):
            self._create_contract_costs_sheet(wb, cc_batch, money_fmt)

        lip = contract_meta.get("licenses_ip_export") or results.get("licenses_ip_export")
        if lip and isinstance(lip, dict):
            self._create_licenses_ip_sheet(wb, lip, money_fmt)

        mr = contract_meta.get("material_rights_ifrs1540") or results.get("material_rights_ifrs1540")
        if mr and isinstance(mr, dict):
            self._create_material_rights_sheet(wb, mr, money_fmt)

        warr_pkg = contract_meta.get("warranties_ifrs1528") or results.get("warranties_ifrs1528")
        if warr_pkg and isinstance(warr_pkg, dict):
            wlist = warr_pkg.get("warranties") or []
            if isinstance(wlist, list) and len(wlist) > 0:
                self._create_warranties_sheet(wb, warr_pkg, money_fmt)

        bah_pkg = contract_meta.get("bill_and_hold_ifrs1579") or results.get("bill_and_hold_ifrs1579")
        if bah_pkg and isinstance(bah_pkg, dict):
            alist = bah_pkg.get("arrangements") or []
            if isinstance(alist, list) and len(alist) > 0:
                self._create_bill_and_hold_sheet(wb, bah_pkg, money_fmt)

        fc_pkg = contract_meta.get("financing_component_ifrs1560") or results.get("financing_component_ifrs1560")
        if fc_pkg and isinstance(fc_pkg, dict):
            clist = fc_pkg.get("contracts") or []
            if isinstance(clist, list) and len(clist) > 0:
                self._create_financing_component_sheet(wb, fc_pkg, money_fmt)

        tp_pkg = contract_meta.get("tp_adjustments_ifrs1566") or results.get("tp_adjustments_ifrs1566")
        if tp_pkg and isinstance(tp_pkg, dict):
            nc_items = (tp_pkg.get("non_cash") or {}).get("items") or []
            cp_items = (tp_pkg.get("consideration_payable") or {}).get("items") or []
            if (isinstance(nc_items, list) and len(nc_items) > 0) or (isinstance(cp_items, list) and len(cp_items) > 0):
                self._create_tp_adjustments_sheet(wb, tp_pkg, money_fmt)

        audit_rows = contract_meta.get("audit_entries") or results.get("audit_entries")
        if not isinstance(audit_rows, list):
            audit_rows = []
        self._create_audit_trail_sheet(wb, audit_rows, money_fmt)

        wb.save(filename)

    def export_ifrs15_master_report(self, master_report: Dict[str, Any], filename: str) -> None:
        """Standalone workbook: IFRS 15 Master Summary as first (and only) sheet."""
        output_path = Path(filename)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        ws = wb.create_sheet("IFRS 15 Master Summary", 0)
        self._write_master_summary_sheet(ws, master_report)
        wb.save(filename)

    def export_portfolio_saas_to_bytes(self, summary: Dict[str, Any], contracts: List[Dict[str, Any]]) -> io.BytesIO:
        """In-memory workbook: portfolio KPIs + contract rows (IFRS 15 portfolio dashboard)."""
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        money_fmt = excel_money_number_format("USD")
        orange = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)

        ws1 = wb.active
        ws1.title = "Portfolio Summary"
        ws1.append(["Metric", "Value"])
        for c in ws1[1]:
            c.fill = orange
            c.font = hdr_font
            kpis = [
            ("Total Contracts", summary.get("total_contracts")),
            ("Active Contracts", summary.get("active_contracts")),
            ("At-Risk Contracts", summary.get("at_risk_contracts")),
            ("Total ARR", summary.get("total_arr")),
            ("Total MRR", summary.get("total_mrr")),
            ("Total Deferred Revenue", summary.get("total_deferred_revenue")),
            ("Total RPO", summary.get("total_rpo")),
            ("Revenue Backlog", summary.get("revenue_backlog")),
            ("Churn Rate %", summary.get("churn_rate_pct")),
            ("At-Risk Rate %", summary.get("at_risk_rate_pct")),
            ("Total Recognised to Date", summary.get("total_recognised_to_date")),
        ]
        for k, v in kpis:
            ws1.append([k, v])

        ws2 = wb.create_sheet("Contract Detail")
        headers = [
            "Contract ID",
            "Customer",
            "Type",
            "ARR",
            "MRR",
            "Total TP",
            "Recognised",
            "Deferred",
            "RPO",
            "Status",
            "Start",
            "End",
            "Currency",
        ]
        ws2.append(headers)
        for c in ws2[1]:
            c.fill = orange
            c.font = hdr_font
        green = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        amber = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        blue = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        status_fills = {"active": green, "at_risk": amber, "churned": red, "renewed": blue}
        for contr in contracts:
            if not isinstance(contr, dict):
                continue
            row = [
                contr.get("contract_id", ""),
                contr.get("customer_name", ""),
                contr.get("contract_type", ""),
                float(contr.get("arr") or 0),
                float(contr.get("mrr") or 0),
                float(contr.get("total_tp") or 0),
                float(contr.get("recognised_to_date") or 0),
                float(contr.get("deferred_balance") or 0),
                float(contr.get("rpo_amount") or 0),
                contr.get("status", ""),
                contr.get("start_date", ""),
                contr.get("end_date", ""),
                contr.get("currency", "USD"),
            ]
            ws2.append(row)
            fill = status_fills.get(str(contr.get("status") or ""))
            if fill:
                for cell in ws2[ws2.max_row]:
                    cell.fill = fill
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=4, max_col=9):
            for cell in row:
                cell.number_format = money_fmt

        for ws in (ws1, ws2):
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 18
            ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _create_summary_sheet(self, wb: Workbook, results: Dict[str, Any], contract_meta: Dict[str, Any], money_fmt: str) -> None:
        ccy = (contract_meta.get("currency") or "USD").upper()
        ws = wb.create_sheet("Summary", 1)
        ws["A1"] = "IFRS 15 Revenue Recognition — Summary"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:B1")
        ws["A1"].alignment = Alignment(horizontal="center")

        balances = results.get("contract_balances", {}) or {}
        perf_obs = (results.get("disclosure_data", {}) or {}).get("performance_obligations", []) or []
        rec_method = "Point in Time" if all((x.get("recognition_method") == "point_in_time" for x in perf_obs)) else "Over Time"

        rows = [
            ("Customer Name", contract_meta.get("customer_name") or "—"),
            ("Contract Date", contract_meta.get("effective_date") or "—"),
            ("Contract Duration (months)", contract_meta.get("contract_term_months") or 0),
            ("Total Contract Value", float(results.get("transaction_price", 0) or 0)),
            ("Total Revenue Recognised (this period)", float(balances.get("revenue_recognized_to_date", 0) or 0)),
            ("Deferred Revenue — Contract Liability", float(balances.get("contract_liability_amount", 0) or 0)),
            ("Contract Assets", float(balances.get("contract_asset_amount", 0) or 0)),
            ("No. of Performance Obligations", len(perf_obs)),
            ("Recognition Method (Over Time / Point in Time)", rec_method),
            ("Report Generated Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        row = 3
        for key, value in rows:
            ws[f"A{row}"] = key
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"A{row}"].border = self.border
            ws[f"B{row}"] = value
            ws[f"B{row}"].border = self.border
            if isinstance(value, (int, float)) and "obligations" not in key.lower() and "months" not in key.lower():
                ws[f"B{row}"].number_format = money_fmt
            row += 1

        # --- Variable Consideration (IFRS 15.50–58) ---
        vc_shade = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        row += 1
        ws[f"A{row}"] = "Variable Consideration"
        ws[f"A{row}"].font = self.subtitle_font
        ws[f"A{row}"].fill = vc_shade
        ws[f"B{row}"].fill = vc_shade
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        vca = (results.get("variable_consideration_assessment") or contract_meta.get("variable_consideration_assessment") or None)
        if vca and isinstance(vca, dict) and (vca.get("expected_amount") is not None or vca.get("constrained_amount") is not None):
            meth = (vca.get("method") or "").replace("_", " ").title()
            if meth.strip().title() in ("Most Likely", "Most_Likely"):
                meth_line = "Most Likely"
            else:
                meth_line = "Expected Value" if "expected" in (vca.get("method") or "").lower() else meth
            vc_rows = [
                ("Estimation Method", meth_line),
                ("Unconstrained Amount", float(vca.get("expected_amount", 0) or 0)),
                ("Constraint Level", (vca.get("constraint_level") or "none").title()),
                ("Constraint Reduction", f"{format_currency_value(float(vca.get('reduction_amount', 0) or 0), ccy)} ({float(vca.get('reduction_pct', 0) or 0):.1f}%)"),
                ("Amount Included in TP", float(vca.get("include_in_transaction_price", 0) or vca.get("constrained_amount", 0) or 0)),
                ("Active Constraint Factors", ", ".join(vca.get("active_factors") or []) or "—"),
            ]
            for vlabel, vval in vc_rows:
                ws[f"A{row}"] = vlabel
                ws[f"A{row}"].font = Font(bold=True, size=10)
                ws[f"A{row}"].border = self.border
                ws[f"B{row}"] = vval
                ws[f"B{row}"].border = self.border
                if isinstance(vval, (int, float)) and vlabel not in ("Constraint Level",):
                    ws[f"B{row}"].number_format = money_fmt
                row += 1
        else:
            ws[f"A{row}"] = "—"
            ws[f"B{row}"] = "Not applicable — fixed price contract"
            ws[f"A{row}"].font = Font(italic=True, size=10, color="6B7280")
            ws[f"B{row}"].font = Font(italic=True, size=10, color="6B7280")
            row += 1

        ws.column_dimensions["A"].width = 48
        ws.column_dimensions["B"].width = 30

    def _create_schedule_sheet(self, wb: Workbook, results: Dict[str, Any], contract_meta: Dict[str, Any], money_fmt: str) -> None:
        ws = wb.create_sheet("Revenue Schedule", 2)
        ws["A1"] = "REVENUE RECOGNITION SCHEDULE"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:G1")
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "Period",
            "Performance Obligation",
            "Opening Balance",
            "Revenue Recognised",
            "Closing Balance",
            "Cumulative Recognised",
            "Status",
        ]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        raw_schedule = results.get("revenue_schedule") or results.get("recognition_schedule") or []
        total_rev = 0.0
        row_no = 4
        for i, item in enumerate(raw_schedule):
            revenue = float(item.get("Revenue", item.get("revenue", item.get("revenue_amount", 0))) or 0)
            cumulative = float(item.get("Cumulative", item.get("cumulative_recognized", item.get("closing_balance", revenue))) or 0)
            opening = cumulative - revenue
            status = "Recognised" if revenue > 0 else "Deferred"
            values = [
                item.get("Period", item.get("Month", item.get("Date", i + 1))),
                item.get("Obligation", item.get("Obligation_ID", item.get("obligation", "—"))),
                opening,
                revenue,
                cumulative,
                cumulative,
                status,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_no, column=col, value=val)
                cell.border = self.border
                if col in (3, 4, 5, 6):
                    cell.number_format = money_fmt
                if i % 2 == 1:
                    cell.fill = self.alt_fill
            total_rev += revenue
            row_no += 1

        ws.cell(row=row_no, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row_no, column=4, value=total_rev).number_format = money_fmt
        ws.cell(row=row_no, column=4).font = Font(bold=True)
        for col in range(1, 8):
            c = ws.cell(row=row_no, column=col)
            c.fill = self.total_fill
            c.border = self.border

        ws.freeze_panes = "A4"
        widths = [12, 34, 18, 20, 18, 22, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

    def _create_journal_entries_sheet(self, wb: Workbook, results: Dict[str, Any], money_fmt: str) -> None:
        ws = wb.create_sheet("Journal Entries", 3)
        ws["A1"] = "JOURNAL ENTRIES"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:F1")
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Period", "Date", "Account", "Dr", "Cr", "Description"]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        balances = results.get("contract_balances", {}) or {}
        rec = float(balances.get("revenue_recognized_to_date", 0) or 0)
        def_rev = float(balances.get("contract_liability_amount", 0) or 0)
        contract_asset = float(balances.get("contract_asset_amount", 0) or 0)
        base_date = datetime.now().strftime("%Y-%m-%d")

        rows = [
            ["Initial", base_date, "Accounts Receivable / Contract Asset", contract_asset or def_rev, "", "Initial recognition"],
            ["Initial", base_date, "Contract Liability (Deferred Revenue)", "", def_rev, "Initial recognition"],
            ["Current", base_date, "Contract Liability", rec, "", "Revenue recognition"],
            ["Current", base_date, "Revenue", "", rec, "Revenue recognition"],
        ]
        journal_entries = (((results.get("journal_entries") or {}).get("revenue_recognition") or {}).get("entries") or [])
        for j in journal_entries:
            rows.append(["Current", base_date, j.get("account", "—"), float(j.get("dr", 0) or 0), float(j.get("cr", 0) or 0), j.get("narration", "Generated entry")])

        row_no = 4
        for r in rows:
            for col, val in enumerate(r, start=1):
                cell = ws.cell(row=row_no, column=col, value=val)
                cell.border = self.border
                if col in (4, 5) and isinstance(val, (int, float)) and val != 0:
                    cell.number_format = money_fmt
            row_no += 1

        widths = [12, 14, 40, 16, 16, 40]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

    def _create_obligations_sheet(self, wb: Workbook, results: Dict[str, Any], contract_meta: Dict[str, Any], money_fmt: str) -> None:
        ws = wb.create_sheet("Obligations Detail", 4)
        ws["A1"] = "PERFORMANCE OBLIGATIONS DETAIL"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:J1")
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "Obligation No.",
            "Description",
            "SSP",
            "Allocated Amount",
            "% of Total",
            "Recognition Pattern",
            "Start Date",
            "End Date",
            "Revenue to Date",
            "Revenue Remaining",
        ]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        perf_obs: List[Dict[str, Any]] = ((results.get("disclosure_data") or {}).get("performance_obligations") or [])
        term_months = int(contract_meta.get("contract_term_months") or 0)
        start_date = contract_meta.get("effective_date") or "—"
        total_tp = float(results.get("transaction_price", 0) or 0)

        row_no = 4
        for idx, ob in enumerate(perf_obs, start=1):
            allocated = float(ob.get("allocated_amount", 0) or 0)
            recognized = float(ob.get("revenue_recognized", 0) or 0)
            remaining = float(ob.get("remaining", max(allocated - recognized, 0)) or 0)
            pct = (allocated / total_tp * 100) if total_tp else 0
            values = [
                ob.get("obligation", f"PO-{idx}"),
                ob.get("description", ob.get("obligation", f"PO-{idx}")),
                allocated,
                allocated,
                pct,
                (ob.get("recognition_method", "over_time") or "over_time").replace("_", " ").title(),
                start_date,
                f"+{term_months} months",
                recognized,
                remaining,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_no, column=col, value=val)
                cell.border = self.border
                if col in (3, 4, 9, 10):
                    cell.number_format = money_fmt
                if col == 5:
                    cell.number_format = "0.00"
            row_no += 1

        widths = [16, 28, 14, 16, 12, 20, 14, 14, 16, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        # Contract Modifications Log (IFRS 15.18–21)
        row_no += 2
        orange_mod = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        mod_hdr_font = Font(bold=True, size=11, color="1F4E78")

        ws.merge_cells(f"A{row_no}:E{row_no}")
        c_title = ws[f"A{row_no}"]
        c_title.value = "CONTRACT MODIFICATIONS LOG"
        c_title.font = self.subtitle_font
        c_title.alignment = Alignment(horizontal="left", vertical="center")
        row_no += 1

        mod_headers = ["Date", "Type", "Description", "Catch-Up Amount", "IFRS Reference"]
        for idx, header in enumerate(mod_headers, start=1):
            cell = ws.cell(row=row_no, column=idx, value=header)
            cell.font = mod_hdr_font
            cell.fill = orange_mod
            cell.border = self.border
        row_no += 1

        modifications = results.get("modifications_log") or []
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        if not modifications:
            ws.merge_cells(f"A{row_no}:E{row_no}")
            c = ws[f"A{row_no}"]
            c.value = "No modifications recorded"
            c.font = Font(italic=True, color="6B7280")
            row_no += 1
        else:
            for i, mod in enumerate(modifications):
                if not isinstance(mod, dict):
                    continue
                d = str(mod.get("modification_date") or mod.get("date") or "")
                typ = str(
                    mod.get("modification_type_name")
                    or mod.get("type")
                    or mod.get("modification_type")
                    or ""
                )
                desc = str(mod.get("description") or mod.get("notes") or mod.get("modification_description") or "")
                catch = float(mod.get("catch_up_amount") or mod.get("catch_up") or 0)
                ref = str(mod.get("ifrs_reference") or mod.get("reference") or "")
                row_fill = white_fill if i % 2 == 0 else grey_fill
                vals = [d, typ, desc, catch, ref]
                for col, val in enumerate(vals, start=1):
                    c = ws.cell(row=row_no, column=col, value=val)
                    c.border = self.border
                    c.fill = row_fill
                    if col == 4 and isinstance(val, (int, float)):
                        c.number_format = money_fmt
                row_no += 1
        row_no += 2

        section_dark = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        section_font = Font(bold=True, color="FFFFFF", size=11)

        ws[f"A{row_no}"] = "Principal vs Agent Assessment (IFRS 15.B34)"
        ws[f"A{row_no}"].font = section_font
        ws[f"A{row_no}"].fill = section_dark
        ws[f"B{row_no}"].fill = section_dark
        ws.merge_cells(f"A{row_no}:B{row_no}")
        row_no += 1

        pa = (results.get("principal_agent_assessment") or contract_meta.get("principal_agent_assessment") or None)
        if pa and isinstance(pa, dict):
            pa_rows = [
                ("Conclusion", pa.get("conclusion", "—")),
                ("Revenue Basis", pa.get("revenue_recognition", "—")),
                ("Gross Revenue", float(pa.get("gross_revenue", 0) or 0)),
                ("Net Revenue (Fee)", float(pa.get("net_revenue", 0) or 0)),
                ("Revenue Recognised", float(pa.get("revenue_to_recognise", 0) or 0)),
                ("Commission Rate", f"{float(pa.get('commission_rate_pct', 0) or 0):.2f}%"),
                ("Borderline Flag", "Yes" if pa.get("borderline") else "No"),
            ]
            for lbl, val in pa_rows:
                ws[f"A{row_no}"] = lbl
                ws[f"A{row_no}"].font = Font(bold=True, size=10)
                ws[f"A{row_no}"].border = self.border
                ws[f"B{row_no}"] = val
                ws[f"B{row_no}"].border = self.border
                if isinstance(val, (int, float)):
                    ws[f"B{row_no}"].number_format = money_fmt
                row_no += 1
        else:
            ws[f"A{row_no}"] = "Principal/agent"
            ws[f"B{row_no}"] = "Principal/agent assessment not applicable"
            ws[f"A{row_no}"].font = Font(italic=True, size=10, color="6B7280")
            ws[f"B{row_no}"].font = Font(italic=True, size=10, color="6B7280")
            ws.merge_cells(f"B{row_no}:J{row_no}")
            row_no += 1

        row_no += 2
        ws[f"A{row_no}"] = "Licence of IP Classification (IFRS 15.B52)"
        ws[f"A{row_no}"].font = section_font
        ws[f"A{row_no}"].fill = section_dark
        ws[f"B{row_no}"].fill = section_dark
        ws.merge_cells(f"A{row_no}:B{row_no}")
        row_no += 1

        lic = (results.get("license_classification") or contract_meta.get("license_classification") or None)
        if lic and isinstance(lic, dict):
            lt = str(lic.get("license_type", "")).replace("_", " ").title()
            pat = "Over Time" if lic.get("pattern") == "OVER_TIME" else "Point in Time"
            monthly = lic.get("revenue_per_period")
            monthly_s = "N/A" if lic.get("pattern") != "OVER_TIME" else float(monthly or 0)
            lic_rows = [
                ("Licence Type", lt),
                ("Recognition Pattern", pat),
                ("Transaction Price", float(lic.get("transaction_price", 0) or 0)),
                ("Licence Term", f"{int(lic.get('licence_term_months', 0) or 0)} months"),
                ("Monthly Revenue", monthly_s if isinstance(monthly_s, str) else float(monthly_s)),
                ("Royalty Exception", "Yes" if lic.get("royalty_exception") else "No"),
            ]
            for lbl, val in lic_rows:
                ws[f"A{row_no}"] = lbl
                ws[f"A{row_no}"].font = Font(bold=True, size=10)
                ws[f"A{row_no}"].border = self.border
                ws[f"B{row_no}"] = val
                ws[f"B{row_no}"].border = self.border
                if isinstance(val, (int, float)):
                    ws[f"B{row_no}"].number_format = money_fmt
                row_no += 1
        else:
            ws[f"A{row_no}"] = "Licence"
            ws[f"B{row_no}"] = "Licence classification not applicable"
            ws[f"A{row_no}"].font = Font(italic=True, size=10, color="6B7280")
            ws[f"B{row_no}"].font = Font(italic=True, size=10, color="6B7280")
            ws.merge_cells(f"B{row_no}:J{row_no}")
            row_no += 1

    def _create_disclosures_sheet(
        self,
        wb: Workbook,
        results: Dict[str, Any],
        contract_meta: Dict[str, Any],
        money_fmt: str,
        currency: str,
    ) -> None:
        ws = wb.create_sheet("Disclosure Notes", 5)
        ws["A1"] = "IFRS 15 DISCLOSURE NOTES"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:B1")
        ws["A1"].alignment = Alignment(horizontal="center")

        balances = results.get("contract_balances", {}) or {}
        disclosure = results.get("disclosure_data", {}) or {}
        perf_obs = disclosure.get("performance_obligations", []) or []
        tp = float(results.get("transaction_price", 0) or 0)
        rec = float(balances.get("revenue_recognized_to_date", 0) or 0)

        sections = [
            ("1. Accounting Policy", "Revenue is recognised when performance obligations are satisfied and control transfers to the customer."),
            ("2. Disaggregation of Revenue", f"Total recognised revenue: {format_currency_value(rec, currency)}."),
            ("3. Contract Balances", f"Contract Assets: {format_currency_value(float(balances.get('contract_asset_amount', 0) or 0), currency)}; Contract Liabilities: {format_currency_value(float(balances.get('contract_liability_amount', 0) or 0), currency)}."),
            ("4. Performance Obligations", f"Number of obligations: {len(perf_obs)}."),
            ("5. Transaction Price Allocated to RPO", f"Unrecognised (RPO): {format_currency_value(max(tp - rec, 0), currency)}."),
            ("6. Significant Judgements", "Judgement is applied in identifying distinct obligations, estimating variable consideration and determining recognition pattern."),
        ]

        row = 3
        for title, body in sections:
            ws[f"A{row}"] = title
            ws[f"A{row}"].font = self.subtitle_font
            row += 1
            ws[f"A{row}"] = body
            ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(f"A{row}:B{row}")
            row += 2

        # Contract balances table
        ws[f"A{row}"] = "Contract Balance Type"
        ws[f"B{row}"] = "Amount"
        ws[f"A{row}"].font = self.header_font
        ws[f"B{row}"].font = self.header_font
        ws[f"A{row}"].fill = self.header_fill
        ws[f"B{row}"].fill = self.header_fill
        row += 1
        for label, amount in [
            ("Contract Asset", float(balances.get("contract_asset_amount", 0) or 0)),
            ("Contract Liability", float(balances.get("contract_liability_amount", 0) or 0)),
            ("Revenue Recognised to Date", rec),
        ]:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = amount
            ws[f"B{row}"].number_format = money_fmt
            ws[f"A{row}"].border = self.border
            ws[f"B{row}"].border = self.border
            row += 1

        section_dark = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        section_font = Font(bold=True, color="FFFFFF", size=11)

        row += 1
        ws[f"A{row}"] = "Remaining Performance Obligations (IFRS 15.120)"
        ws[f"A{row}"].font = section_font
        ws[f"A{row}"].fill = section_dark
        ws[f"B{row}"].fill = section_dark
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        rpo = (results.get("rpo_disclosure") or contract_meta.get("rpo_disclosure") or None)
        if rpo and isinstance(rpo, dict):
            rpo_rows = [
                ("Total RPO", float(rpo.get("total_rpo", 0) or 0)),
                ("— Within 1 Year", float(rpo.get("within_1_year", 0) or 0)),
                ("— 1 to 2 Years", float(rpo.get("one_to_two_years", 0) or 0)),
                ("— Beyond 2 Years", float(rpo.get("beyond_2_years", 0) or 0)),
                ("Disclosure Required", "Yes" if rpo.get("disclosure_required") else "No"),
                (
                    "Practical Expedient",
                    "Available" if rpo.get("practical_expedient_available") else "Not Available",
                ),
            ]
            for lbl, val in rpo_rows:
                ws[f"A{row}"] = lbl
                ws[f"A{row}"].font = Font(bold=True, size=10)
                ws[f"A{row}"].border = self.border
                ws[f"B{row}"] = val
                ws[f"B{row}"].border = self.border
                if isinstance(val, (int, float)):
                    ws[f"B{row}"].number_format = money_fmt
                row += 1
            row += 1
            ws[f"A{row}"] = "Obligation"
            ws[f"B{row}"] = "RPO Amount"
            ws[f"C{row}"] = "Bucket"
            ws[f"D{row}"] = "Expected Date"
            for col in ("A", "B", "C", "D"):
                c = ws[f"{col}{row}"]
                c.font = self.header_font
                c.fill = self.header_fill
                c.border = self.border
            row += 1
            for ob in rpo.get("by_obligation") or []:
                if not isinstance(ob, dict):
                    continue
                ws[f"A{row}"] = ob.get("obligation_name", "")
                ws[f"B{row}"] = float(ob.get("rpo_amount", 0) or 0)
                ws[f"C{row}"] = str(ob.get("bucket", ""))
                ws[f"D{row}"] = ob.get("expected_completion_date", "")
                ws[f"B{row}"].number_format = money_fmt
                for col in ("A", "B", "C", "D"):
                    ws[f"{col}{row}"].border = self.border
                row += 1
        else:
            ws[f"A{row}"] = "RPO"
            ws[f"B{row}"] = "Not assessed — run RPO disclosure in the app"
            ws[f"A{row}"].font = Font(italic=True, size=10, color="6B7280")
            ws[f"B{row}"].font = Font(italic=True, size=10, color="6B7280")
            ws.merge_cells(f"B{row}:D{row}")
            row += 1

        row += 1
        ws[f"A{row}"] = "Costs to Obtain Contracts (IFRS 15.91)"
        ws[f"A{row}"].font = section_font
        ws[f"A{row}"].fill = section_dark
        ws[f"B{row}"].fill = section_dark
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        cc = (results.get("contract_costs_assessment") or contract_meta.get("contract_costs_assessment") or None)
        if cc and isinstance(cc, dict):
            cc_rows = [
                ("Commission Amount", float(cc.get("commission_amount", 0) or 0)),
                ("Practical Expedient Used", "Yes" if cc.get("use_practical_expedient") else "No"),
                ("Asset Recognised", float(cc.get("total_asset_recognised", 0) or 0)),
                ("Monthly Amortisation", float(cc.get("monthly_amortisation", 0) or 0)),
                ("Contract Term", f"{int(cc.get('contract_term_months', 0) or 0)} months"),
                ("Impairment Flag", "Yes" if cc.get("impairment_flag") else "No"),
            ]
            for lbl, val in cc_rows:
                ws[f"A{row}"] = lbl
                ws[f"A{row}"].font = Font(bold=True, size=10)
                ws[f"A{row}"].border = self.border
                ws[f"B{row}"] = val
                ws[f"B{row}"].border = self.border
                if isinstance(val, (int, float)):
                    ws[f"B{row}"].number_format = money_fmt
                row += 1
            row += 1
            ws[f"A{row}"] = "Month"
            ws[f"B{row}"] = "Opening"
            ws[f"C{row}"] = "Amortisation"
            ws[f"D{row}"] = "Closing"
            for col in ("A", "B", "C", "D"):
                c = ws[f"{col}{row}"]
                c.font = self.header_font
                c.fill = self.header_fill
                c.border = self.border
            row += 1
            sched = list(cc.get("amortisation_schedule") or [])[:12]
            for line in sched:
                if not isinstance(line, dict):
                    continue
                ws[f"A{row}"] = line.get("month", "")
                ws[f"B{row}"] = float(line.get("opening_balance", 0) or 0)
                ws[f"C{row}"] = float(line.get("amortisation", 0) or 0)
                ws[f"D{row}"] = float(line.get("closing_balance", 0) or 0)
                for col in ("B", "C", "D"):
                    ws[f"{col}{row}"].number_format = money_fmt
                for col in ("A", "B", "C", "D"):
                    ws[f"{col}{row}"].border = self.border
                row += 1
        else:
            ws[f"A{row}"] = "Contract costs"
            ws[f"B{row}"] = "No contract costs assessed for this contract"
            ws[f"A{row}"].font = Font(italic=True, size=10, color="6B7280")
            ws[f"B{row}"].font = Font(italic=True, size=10, color="6B7280")
            ws.merge_cells(f"B{row}:D{row}")
            row += 1

        ws.column_dimensions["A"].width = 52
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 18

    def _create_deferred_revenue_sheet(self, wb: Workbook, rollforward: Dict[str, Any], money_fmt: str) -> None:
        """Sheet: Deferred revenue roll-forward reconciliation (IFRS 15.116)."""
        ws = wb.create_sheet("Deferred Rev Reconciliation", 6)
        orange_hdr = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        hdr_f = Font(bold=True, size=11, color="1F4E78")

        ws["A1"] = "DEFERRED REVENUE ROLL-FORWARD"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:C1")
        per = str(rollforward.get("period") or "—")
        ccy = str(rollforward.get("currency") or "USD").upper()
        ws["A2"] = f"Reconciliation — IFRS 15.116 | Period: {per} | {ccy}"
        ws["A2"].font = Font(size=10, color="6B7280")
        ws.merge_cells("A2:C2")

        r = 4
        ws[f"A{r}"] = "Line Item"
        ws[f"B{r}"] = "Amount"
        ws[f"C{r}"] = "Notes"
        for col in ("A", "B", "C"):
            c = ws[f"{col}{r}"]
            c.font = hdr_f
            c.fill = orange_hdr
            c.border = self.border
        r += 1

        lines = rollforward.get("rollforward_lines") or []
        var_amt = float(rollforward.get("variance") or 0)
        green_v = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_v = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for i, ln in enumerate(lines):
            if not isinstance(ln, dict):
                continue
            line_type = str(ln.get("type") or "")
            amt = float(ln.get("amount") or 0)
            ws[f"A{r}"] = ln.get("line", "")
            ws[f"B{r}"] = amt
            ws[f"B{r}"].number_format = money_fmt
            note = ""
            if line_type == "variance":
                note = "Must be zero when reconciled"
            ws[f"C{r}"] = note
            for col in ("A", "B", "C"):
                ws[f"{col}{r}"].border = self.border
            if line_type == "opening":
                ws[f"A{r}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                ws[f"B{r}"].fill = ws[f"A{r}"].fill
            elif line_type == "addition":
                for col in ("A", "B", "C"):
                    ws[f"{col}{r}"].border = Border(
                        left=Side(style="medium", color="92D050"),
                        right=Side(style="thin", color="000000"),
                        top=Side(style="thin", color="000000"),
                        bottom=Side(style="thin", color="000000"),
                    )
            elif line_type == "deduction":
                for col in ("A", "B", "C"):
                    ws[f"{col}{r}"].border = Border(
                        left=Side(style="medium", color="C00000"),
                        right=Side(style="thin", color="000000"),
                        top=Side(style="thin", color="000000"),
                        bottom=Side(style="thin", color="000000"),
                    )
            elif line_type == "subtotal":
                for col in ("A", "B", "C"):
                    ws[f"{col}{r}"].font = Font(bold=True)
                    ws[f"{col}{r}"].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            elif line_type == "control":
                ws[f"B{r}"].font = Font(bold=True, color="1F4E78")
            elif line_type == "variance":
                vf = green_v if abs(var_amt) < 0.01 else red_v
                ws[f"A{r}"].fill = vf
                ws[f"B{r}"].fill = vf
                ws[f"C{r}"].fill = vf
            if i % 2 == 1 and line_type not in ("subtotal", "variance", "opening"):
                alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                if line_type not in ("addition", "deduction"):
                    ws[f"A{r}"].fill = alt
                    ws[f"B{r}"].fill = alt
                    ws[f"C{r}"].fill = alt
            r += 1

        r += 2
        ws[f"A{r}"] = "Exceptions"
        ws[f"A{r}"].font = self.subtitle_font
        r += 1
        excs = rollforward.get("exceptions") or []
        if not excs:
            ws[f"A{r}"] = "No exceptions"
            ws[f"A{r}"].font = Font(italic=True, color="6B7280")
            r += 1
        else:
            for ex in excs:
                if not isinstance(ex, dict):
                    continue
                ws[f"A{r}"] = f"{ex.get('severity', '')} — {ex.get('type', '')}"
                ws[f"A{r}"].font = Font(bold=True, size=10)
                r += 1
                ws[f"A{r}"] = ex.get("description", "")
                ws[f"A{r}"].alignment = Alignment(wrap_text=True)
                ws.merge_cells(f"A{r}:C{r}")
                r += 1
                ws[f"A{r}"] = ex.get("action", "")
                ws[f"A{r}"].font = Font(italic=True, color="1F4E78")
                ws[f"A{r}"].alignment = Alignment(wrap_text=True)
                ws.merge_cells(f"A{r}:C{r}")
                r += 1

        ws.column_dimensions["A"].width = 48
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 40

    def _create_rpo_disclosure_sheet(self, wb: Workbook, rpo: Dict[str, Any], money_fmt: str) -> None:
        """Sheet: RPO disclosure (IFRS 15.120–122) — mandatory remaining performance obligations note."""
        ws = wb.create_sheet("RPO Disclosure")
        orange_hdr = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        hdr_f = Font(bold=True, size=11, color="1F4E78")
        blue_total = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        ws["A1"] = "REMAINING PERFORMANCE OBLIGATIONS"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:B1")
        ws["A2"] = "IFRS 15.120-122"
        ws["A2"].font = self.subtitle_font
        ws.merge_cells("A2:B2")

        r = 4
        ws[f"A{r}"] = "Metric"
        ws[f"B{r}"] = "Amount"
        for col in ("A", "B"):
            c = ws[f"{col}{r}"]
            c.font = hdr_f
            c.fill = orange_hdr
            c.border = self.border
        r += 1

        buckets = rpo.get("buckets") or {}
        total_rpo = float(rpo.get("total_rpo") or 0)

        summary_rows = [
            ("Total RPO", total_rpo, True),
            ("Within 1 Year", float(buckets.get("within_1_year", 0) or 0), True),
            ("1 to 2 Years", float(buckets.get("1_to_2_years", 0) or 0), True),
            ("2 to 5 Years", float(buckets.get("2_to_5_years", 0) or 0), True),
            ("Beyond 5 Years", float(buckets.get("beyond_5_years", 0) or 0), True),
        ]
        for i, (lbl, amt, is_money) in enumerate(summary_rows):
            ws[f"A{r}"] = lbl
            ws[f"B{r}"] = amt
            ws[f"A{r}"].border = self.border
            ws[f"B{r}"].border = self.border
            if is_money:
                ws[f"B{r}"].number_format = money_fmt
            if i == 0:
                ws[f"A{r}"].font = Font(bold=True, size=11)
                ws[f"B{r}"].font = Font(bold=True, size=11)
                ws[f"A{r}"].fill = blue_total
                ws[f"B{r}"].fill = blue_total
            r += 1

        r += 2
        ws[f"A{r}"] = "Contract ID"
        ws[f"B{r}"] = "Customer"
        ws[f"C{r}"] = "End Date"
        ws[f"D{r}"] = "Total TP"
        ws[f"E{r}"] = "Recognised"
        ws[f"F{r}"] = "RPO Amount"
        ws[f"G{r}"] = "Expedient Applied"
        for col in ("A", "B", "C", "D", "E", "F", "G"):
            c = ws[f"{col}{r}"]
            c.font = hdr_f
            c.fill = orange_hdr
            c.border = self.border
        r += 1

        def write_contract_row(row: int, cd: Dict[str, Any], expedient: bool) -> int:
            ws[f"A{row}"] = cd.get("contract_id", "")
            ws[f"B{row}"] = cd.get("customer_name", "")
            ws[f"C{row}"] = cd.get("contract_end", "")
            ws[f"D{row}"] = float(cd.get("total_transaction_price", 0) or 0)
            ws[f"E{row}"] = float(cd.get("revenue_recognised_to_date", 0) or 0)
            ws[f"F{row}"] = float(cd.get("rpo_amount", 0) or 0) if not expedient else 0.0
            ws[f"G{row}"] = "Yes" if expedient else "No"
            for col in ("D", "E", "F"):
                ws[f"{col}{row}"].number_format = money_fmt
            for col in ("A", "B", "C", "D", "E", "F", "G"):
                ws[f"{col}{row}"].border = self.border
            if expedient:
                grey = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                for col in ("A", "B", "C", "D", "E", "F", "G"):
                    ws[f"{col}{row}"].fill = grey
            return row + 1

        row_idx = r
        for cd in rpo.get("contract_details") or []:
            if not isinstance(cd, dict):
                continue
            row_idx = write_contract_row(row_idx, cd, False)

        for cd in rpo.get("expedient_contracts") or []:
            if not isinstance(cd, dict):
                continue
            row_idx = write_contract_row(row_idx, cd, True)

        r = row_idx + 1
        note = rpo.get("disclosure_note") or {}
        full_text = str(note.get("full_text") or "")
        ws.merge_cells(f"A{r}:G{r}")
        ws[f"A{r}"] = "IFRS 15 Annual Report Disclosure Note"
        ws[f"A{r}"].font = Font(bold=True, size=12, color="1F4E78")
        ws[f"A{r}"].fill = orange_hdr
        ws[f"A{r}"].alignment = Alignment(horizontal="center")
        r += 1
        ws.merge_cells(f"A{r}:G{r + 12}")
        cell = ws[f"A{r}"]
        cell.value = full_text
        cell.font = Font(italic=True, size=10, name="Calibri")
        cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        thick = Border(
            left=Side(style="medium", color="1F4E78"),
            right=Side(style="medium", color="1F4E78"),
            top=Side(style="medium", color="1F4E78"),
            bottom=Side(style="medium", color="1F4E78"),
        )
        cell.border = thick
        r += 14

        for col, w in zip(("A", "B", "C", "D", "E", "F", "G"), (18, 28, 14, 16, 16, 16, 18)):
            ws.column_dimensions[col].width = w

    def _create_principal_agent_sheet(self, wb: Workbook, history: List[Dict[str, Any]], money_fmt: str) -> None:
        """Sheet: Principal vs agent assessments (IFRS 15.B34–B38)."""
        ws = wb.create_sheet("Principal vs Agent")
        orange_hdr = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        hdr_f = Font(bold=True, size=11, color="1F4E78")
        green_c = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        blue_c = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        amber_c = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        sep_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        ws["A1"] = "PRINCIPAL VS AGENT ASSESSMENTS"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:H1")
        ws["A2"] = "IFRS 15.B34-B38"
        ws["A2"].font = self.subtitle_font
        ws.merge_cells("A2:H2")

        r = 4
        headers = [
            "Arrangement",
            "Description",
            "Gross Value",
            "Third-Party Cost",
            "Net Margin",
            "Conclusion",
            "Revenue Treatment",
            "Confidence",
        ]
        for i, h in enumerate(headers):
            col = chr(ord("A") + i)
            ws[f"{col}{r}"] = h
            ws[f"{col}{r}"].font = hdr_f
            ws[f"{col}{r}"].fill = orange_hdr
            ws[f"{col}{r}"].border = self.border
        r += 1

        for entry in history:
            if not isinstance(entry, dict):
                continue
            a = entry.get("assessment") if isinstance(entry.get("assessment"), dict) else entry
            if not isinstance(a, dict):
                continue
            arr = entry.get("arrangement_id") or a.get("arrangement_id", "")
            desc = entry.get("description") or a.get("description", "")
            gross = float(entry.get("gross_contract_value") or a.get("gross_contract_value", 0) or 0)
            cost = float(entry.get("third_party_cost") or a.get("third_party_cost", 0) or 0)
            net_m = float(a.get("net_margin", gross - cost) or 0)
            conc = str(a.get("conclusion", "") or "")
            rev_t = str(a.get("revenue_treatment", "") or "")
            conf = str(a.get("confidence", "") or "")

            ws[f"A{r}"] = arr
            ws[f"B{r}"] = desc
            ws[f"C{r}"] = gross
            ws[f"D{r}"] = cost
            ws[f"E{r}"] = net_m
            ws[f"F{r}"] = conc
            ws[f"G{r}"] = rev_t
            ws[f"H{r}"] = conf
            for col in ("C", "D", "E"):
                ws[f"{col}{r}"].number_format = money_fmt
            for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
                ws[f"{col}{r}"].border = self.border
            conc_fill = None
            if conc == "PRINCIPAL":
                conc_fill = green_c
            elif conc == "AGENT":
                conc_fill = blue_c
            elif "JUDGEMENT" in conc.upper():
                conc_fill = amber_c
            if conc_fill:
                ws[f"F{r}"].fill = conc_fill
            r += 1

        r += 2
        ws[f"A{r}"] = "INDICATOR DETAIL (by arrangement)"
        ws[f"A{r}"].font = self.subtitle_font
        ws.merge_cells(f"A{r}:D{r}")
        r += 1

        for entry in history:
            if not isinstance(entry, dict):
                continue
            a = entry.get("assessment") if isinstance(entry.get("assessment"), dict) else entry
            if not isinstance(a, dict):
                continue
            arr = entry.get("arrangement_id") or a.get("arrangement_id", "")
            ws.merge_cells(f"A{r}:D{r}")
            ws[f"A{r}"] = f"Arrangement: {arr}"
            ws[f"A{r}"].font = Font(bold=True, size=10)
            ws[f"A{r}"].fill = sep_fill
            r += 1

            detail = a.get("indicators_detail") or {}
            for key, info in detail.items():
                if not isinstance(info, dict):
                    continue
                lbl = str(info.get("label", key))
                present = "Yes" if info.get("present") else "No"
                sup = str(info.get("supports", ""))
                ws[f"A{r}"] = lbl
                ws[f"B{r}"] = present
                ws[f"C{r}"] = sup
                for col in ("A", "B", "C"):
                    ws[f"{col}{r}"].border = self.border
                sup_fill = green_c if sup == "PRINCIPAL" else blue_c
                ws[f"C{r}"].fill = sup_fill
                r += 1
            r += 1

        r += 1
        ws[f"A{r}"] = "REVENUE IMPACT (per arrangement)"
        ws[f"A{r}"].font = self.subtitle_font
        ws.merge_cells(f"A{r}:D{r}")
        r += 1

        for entry in history:
            if not isinstance(entry, dict):
                continue
            a = entry.get("assessment") if isinstance(entry.get("assessment"), dict) else entry
            if not isinstance(a, dict):
                continue
            arr = entry.get("arrangement_id") or a.get("arrangement_id", "")
            gross = float(entry.get("gross_contract_value") or a.get("gross_contract_value", 0) or 0)
            cost = float(entry.get("third_party_cost") or a.get("third_party_cost", 0) or 0)
            margin = gross - cost
            ws[f"A{r}"] = f"{arr} — If PRINCIPAL: Revenue"
            ws[f"B{r}"] = gross
            ws[f"B{r}"].number_format = money_fmt
            r += 1
            ws[f"A{r}"] = f"{arr} — If AGENT: Revenue"
            ws[f"B{r}"] = margin
            ws[f"B{r}"].number_format = money_fmt
            r += 1
            ws[f"A{r}"] = f"{arr} — Difference (gross − agent revenue)"
            ws[f"B{r}"] = cost
            ws[f"B{r}"].number_format = money_fmt
            r += 2

        ws.column_dimensions["A"].width = 52
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 22
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 14

    def _create_contract_costs_sheet(self, wb: Workbook, payload: Dict[str, Any], money_fmt: str) -> None:
        """Sheet: Contract costs (IFRS 15.91–94) — batch calculator output."""
        ws = wb.create_sheet("Contract Costs")
        orange = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        hdr_f = Font(bold=True, size=11, color="1F4E78")
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        amber = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        ws["A1"] = "CONTRACT COSTS — IFRS 15.91-94"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:I1")
        summ = payload.get("summary") or {}
        r = 3
        for lbl, key in [
            ("Total Capitalised", "total_capitalised"),
            ("Net Asset Balance", "net_asset_balance"),
            ("Amortised to Date", "total_amortised_to_date"),
            ("Expensed Immediately", "total_expensed_immediately"),
        ]:
            ws[f"A{r}"] = lbl
            ws[f"B{r}"] = float(summ.get(key, 0) or 0)
            ws[f"B{r}"].number_format = money_fmt
            ws[f"A{r}"].border = self.border
            ws[f"B{r}"].border = self.border
            r += 1
        r += 1
        headers = [
            "Cost ID",
            "Contract",
            "Description",
            "Type",
            "Amount",
            "Treatment",
            "Monthly Amort",
            "Asset Balance",
            "IFRS Ref",
        ]
        for i, h in enumerate(headers):
            col = chr(ord("A") + i)
            ws[f"{col}{r}"] = h
            ws[f"{col}{r}"].font = hdr_f
            ws[f"{col}{r}"].fill = orange
            ws[f"{col}{r}"].border = self.border
        r += 1
        for it in payload.get("costs") or []:
            if not isinstance(it, dict):
                continue
            tr = str(it.get("treatment", ""))
            ws[f"A{r}"] = it.get("cost_id", "")
            ws[f"B{r}"] = it.get("contract_id", "")
            ws[f"C{r}"] = it.get("description", "")
            ws[f"D{r}"] = it.get("cost_type", "")
            ws[f"E{r}"] = float(it.get("cost_amount", 0) or 0)
            ws[f"F{r}"] = tr
            ws[f"G{r}"] = float(it.get("monthly_amortisation", 0) or 0) if tr == "CAPITALISE" else 0.0
            ws[f"H{r}"] = float(it.get("asset_balance", 0) or 0)
            ws[f"I{r}"] = str(payload.get("ifrs_reference", "IFRS 15.91-94"))
            for col, money in (("E", True), ("G", True), ("H", True)):
                ws[f"{col}{r}"].number_format = money_fmt
            for col in ("A", "B", "C", "D", "E", "F", "G", "H", "I"):
                ws[f"{col}{r}"].border = self.border
            if tr == "CAPITALISE":
                ws[f"F{r}"].fill = green
            else:
                ws[f"F{r}"].fill = amber
            r += 1
        r += 2
        ws[f"A{r}"] = "Amortisation schedules (capitalised items)"
        ws[f"A{r}"].font = self.subtitle_font
        r += 1
        for it in payload.get("costs") or []:
            if not isinstance(it, dict) or it.get("treatment") != "CAPITALISE":
                continue
            ws.merge_cells(f"A{r}:E{r}")
            ws[f"A{r}"] = f"{it.get('cost_id')} — {it.get('description')}"
            ws[f"A{r}"].font = hdr_f
            ws[f"A{r}"].fill = orange
            r += 1
            ws[f"A{r}"] = "Period"
            ws[f"B{r}"] = "Amortisation"
            ws[f"C{r}"] = "Cumulative"
            ws[f"D{r}"] = "Asset Balance"
            ws[f"E{r}"] = "Status"
            for col in ("A", "B", "C", "D", "E"):
                ws[f"{col}{r}"].font = hdr_f
                ws[f"{col}{r}"].fill = orange
                ws[f"{col}{r}"].border = self.border
            r += 1
            for row in (it.get("amortisation_schedule") or [])[:60]:
                if not isinstance(row, dict):
                    continue
                ws[f"A{r}"] = row.get("period", "")
                ws[f"B{r}"] = float(row.get("amortisation", 0) or 0)
                ws[f"C{r}"] = float(row.get("cumulative_amortised", 0) or 0)
                ws[f"D{r}"] = float(row.get("asset_balance", 0) or 0)
                ws[f"E{r}"] = row.get("status", "")
                ws[f"B{r}"].number_format = money_fmt
                ws[f"C{r}"].number_format = money_fmt
                ws[f"D{r}"].number_format = money_fmt
                for col in ("A", "B", "C", "D", "E"):
                    ws[f"{col}{r}"].border = self.border
                r += 1
            r += 1
        r += 1
        ws[f"A{r}"] = "Contract Cost Asset — Roll-Forward"
        ws[f"A{r}"].font = self.subtitle_font
        r += 1
        opening = 0.0
        add = float(summ.get("total_capitalised", 0) or 0)
        amort = float(summ.get("total_amortised_to_date", 0) or 0)
        closing = float(summ.get("net_asset_balance", 0) or 0)
        for lbl, val in [
            ("Opening Balance", opening),
            ("Add: New Costs Capitalised", add),
            ("Less: Amortised in Period", amort),
            ("Closing Balance", closing),
        ]:
            ws[f"A{r}"] = lbl
            ws[f"B{r}"] = val
            ws[f"B{r}"].number_format = money_fmt
            r += 1
        ws[f"A{r}"] = "Reconcile to Contract Cost Asset GL monthly (SOX control)."
        ws[f"A{r}"].font = Font(italic=True, size=9, color="6B7280")
        ws.merge_cells(f"A{r}:D{r}")
        for col in "ABCDEFGHI":
            ws.column_dimensions[col].width = 14
        ws.column_dimensions["C"].width = 28

    def _create_licenses_ip_sheet(self, wb: Workbook, payload: Dict[str, Any], money_fmt: str) -> None:
        """Sheet: Licences of IP (IFRS 15.B52–B63)."""
        ws = wb.create_sheet("Licenses of IP")
        orange = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        hdr_f = Font(bold=True, size=11, color="1F4E78")
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        amb = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        ws["A1"] = "LICENSES OF IP — IFRS 15.B52-B63"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:H1")
        sm = payload.get("summary") or {}
        r = 3
        ws[f"A{r}"] = "Total licences"
        ws[f"B{r}"] = int(sm.get("total_licenses", 0) or 0)
        r += 1
        ws[f"A{r}"] = "Right to Access"
        ws[f"B{r}"] = int(sm.get("right_to_access", 0) or 0)
        r += 1
        ws[f"A{r}"] = "Right to Use"
        ws[f"B{r}"] = int(sm.get("right_to_use", 0) or 0)
        r += 1
        ws[f"A{r}"] = "Judgement required"
        ws[f"B{r}"] = int(sm.get("judgement_required", 0) or 0)
        r += 2
        heads = ["License ID", "Product", "Type", "Fee", "Classification", "Recognition Basis", "Revenue Timing", "IFRS Ref"]
        for i, h in enumerate(heads):
            c = chr(ord("A") + i)
            ws[f"{c}{r}"] = h
            ws[f"{c}{r}"].font = hdr_f
            ws[f"{c}{r}"].fill = orange
            ws[f"{c}{r}"].border = self.border
        r += 1
        for lic in payload.get("licenses") or []:
            if not isinstance(lic, dict):
                continue
            lt = str(lic.get("license_type", ""))
            ws[f"A{r}"] = lic.get("license_id", "")
            ws[f"B{r}"] = lic.get("product_name", "")
            ws[f"C{r}"] = lt
            ws[f"D{r}"] = float(lic.get("license_fee", 0) or 0)
            ws[f"E{r}"] = lt
            ws[f"F{r}"] = str(lic.get("recognition_basis", ""))
            rp = lic.get("revenue_per_period")
            ws[f"G{r}"] = (
                f"${float(rp or 0):,.2f}/period" if str(lic.get("recognition_basis")) == "OVER_TIME" else "Point in time"
            )
            ws[f"H{r}"] = str(lic.get("ifrs_reference", ""))
            ws[f"D{r}"].number_format = money_fmt
            fill = green if lt == "RIGHT_TO_USE" else blue if lt == "RIGHT_TO_ACCESS" else amb
            ws[f"E{r}"].fill = fill
            for col in "ABCDEFGH":
                ws[f"{col}{r}"].border = self.border
            r += 1
        r += 2
        ws[f"A{r}"] = "Recognition schedules (Right to Access)"
        ws[f"A{r}"].font = self.subtitle_font
        r += 1
        for lic in payload.get("licenses") or []:
            if not isinstance(lic, dict) or lic.get("license_type") != "RIGHT_TO_ACCESS":
                continue
            ws.merge_cells(f"A{r}:E{r}")
            ws[f"A{r}"] = f"{lic.get('license_id')} — {lic.get('product_name')}"
            ws[f"A{r}"].fill = orange
            ws[f"A{r}"].font = hdr_f
            r += 1
            for col, h in zip("ABCDE", ["Period", "Revenue", "Cumulative", "Balance", "Note"]):
                ws[f"{col}{r}"] = h
                ws[f"{col}{r}"].font = hdr_f
                ws[f"{col}{r}"].border = self.border
            r += 1
            for row in lic.get("recognition_schedule") or []:
                if not isinstance(row, dict):
                    continue
                ws[f"A{r}"] = row.get("period", "")
                ws[f"B{r}"] = float(row.get("amount", 0) or 0)
                ws[f"C{r}"] = float(row.get("cumulative", 0) or 0)
                ws[f"D{r}"] = float(row.get("balance", 0) or 0)
                ws[f"E{r}"] = row.get("note", "")
                for col in ("B", "C", "D"):
                    ws[f"{col}{r}"].number_format = money_fmt
                for col in "ABCDE":
                    ws[f"{col}{r}"].border = self.border
                r += 1
            r += 1
        for col in "ABCDEFGH":
            ws.column_dimensions[col].width = 16

    def _create_material_rights_sheet(self, wb: Workbook, payload: Dict[str, Any], money_fmt: str) -> None:
        """Sheet: Customer options / material rights (IFRS 15.B40–B43)."""
        ws = wb.create_sheet("Material Rights")
        orange = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        hdr_f = Font(bold=True, size=11, color="1F4E78")
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        grey = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        ws["A1"] = "CUSTOMER OPTIONS & MATERIAL RIGHTS"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:K1")
        ws["A2"] = "IFRS 15.B40-B43"
        ws["A2"].font = self.subtitle_font
        summ = payload.get("summary") or {}
        r = 4
        for lbl, key in [
            ("Options assessed", "total_options_assessed"),
            ("Material rights identified", "material_rights_found"),
            ("Total deferred to options", "total_deferred_to_options"),
        ]:
            ws[f"A{r}"] = lbl
            val = summ.get(key, 0)
            ws[f"B{r}"] = float(val) if key == "total_deferred_to_options" else int(val or 0)
            if key == "total_deferred_to_options":
                ws[f"B{r}"].number_format = money_fmt
            ws[f"A{r}"].border = self.border
            ws[f"B{r}"].border = self.border
            r += 1
        r += 1
        headers = [
            "Option ID",
            "Contract",
            "Type",
            "Description",
            "Option Price",
            "Option SSP",
            "Discount %",
            "Exercise Prob",
            "Material Right",
            "Allocated to Option",
            "IFRS Ref",
        ]
        for i, h in enumerate(headers):
            col = chr(ord("A") + i)
            ws[f"{col}{r}"] = h
            ws[f"{col}{r}"].font = hdr_f
            ws[f"{col}{r}"].fill = orange
            ws[f"{col}{r}"].border = self.border
        r += 1
        for it in payload.get("options") or []:
            if not isinstance(it, dict):
                continue
            tr = str(it.get("treatment", ""))
            mat = bool(it.get("material_right_exists"))
            disc_pct = float(it.get("discount_pct", 0) or 0)
            prob_pct = float(it.get("exercise_probability_pct", 0) or 0)
            ws[f"A{r}"] = it.get("option_id", "")
            ws[f"B{r}"] = it.get("contract_id", "")
            ws[f"C{r}"] = it.get("option_type", "")
            ws[f"D{r}"] = it.get("description", "")
            ws[f"E{r}"] = float(it.get("option_price", 0) or 0)
            ws[f"F{r}"] = float(it.get("option_ssp", 0) or 0)
            ws[f"G{r}"] = disc_pct
            ws[f"H{r}"] = prob_pct
            ws[f"I{r}"] = "YES" if mat else "NO"
            ws[f"J{r}"] = float(it.get("allocated_to_option", 0) or 0)
            ws[f"K{r}"] = str(it.get("ifrs_reference", "IFRS 15.B40"))
            for col in ("E", "F", "J"):
                ws[f"{col}{r}"].number_format = money_fmt
            for col in "ABCDEFGHIJK":
                ws[f"{col}{r}"].border = self.border
            fill_row = green if tr == "SEPARATE_PERFORMANCE_OBLIGATION" else grey
            for col in "ABCDEFGHIJK":
                ws[f"{col}{r}"].fill = fill_row
            r += 1

        r += 2
        ws[f"A{r}"] = "Allocation detail (material rights only)"
        ws[f"A{r}"].font = self.subtitle_font
        r += 1
        for it in payload.get("options") or []:
            if not isinstance(it, dict) or not it.get("material_right_exists"):
                continue
            ws.merge_cells(f"A{r}:E{r}")
            ws[f"A{r}"] = f"{it.get('option_id')} — {it.get('description')}"
            ws[f"A{r}"].font = hdr_f
            ws[f"A{r}"].fill = orange
            r += 1
            orig_ssp = float(it.get("original_ssp", 0) or 0)
            opt_est = float(it.get("option_ssp_estimated", 0) or 0)
            tot = float(it.get("total_ssp_pool", 0) or 0)
            alloc_o = float(it.get("allocated_to_original", 0) or 0)
            alloc_opt = float(it.get("allocated_to_option", 0) or 0)
            tp = float(it.get("original_contract_value", 0) or 0)
            for lbl, val, is_money in [
                ("Original SSP", orig_ssp, True),
                ("Option SSP (estimated)", opt_est, True),
                ("Total SSP pool", tot, True),
                ("Allocated to original POs", alloc_o, True),
                ("Allocated to option (contract liability)", alloc_opt, True),
                ("Transaction price (original contract)", tp, True),
            ]:
                ws[f"A{r}"] = lbl
                ws[f"B{r}"] = val
                if is_money:
                    ws[f"B{r}"].number_format = money_fmt
                ws[f"A{r}"].border = self.border
                ws[f"B{r}"].border = self.border
                r += 1
            r += 1

        for col in "ABCDEFGHIJK":
            ws.column_dimensions[col].width = 14
        ws.column_dimensions["D"].width = 28

    def _create_warranties_sheet(self, wb: Workbook, payload: Dict[str, Any], money_fmt: str) -> None:
        """Sheet: Warranties (IFRS 15.B28–B33 / IAS 37) — optional when warranties export payload is present."""
        ws = wb.create_sheet("Warranties")
        orange = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        hdr_f = Font(bold=True, size=11, color="1F4E78")
        grey_r = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        blue_r = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        amb_r = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        ws["A1"] = "WARRANTIES — IFRS 15.B28-B33 / IAS 37"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:I1")
        summ = payload.get("summary") or {}
        r = 3
        ws[f"A{r}"] = "Total warranties"
        ws[f"B{r}"] = int(summ.get("total", 0) or 0)
        r += 1
        ws[f"A{r}"] = "Assurance-type (IAS 37)"
        ws[f"B{r}"] = int(summ.get("assurance_type", 0) or 0)
        r += 1
        ws[f"A{r}"] = "Service-type (IFRS 15)"
        ws[f"B{r}"] = int(summ.get("service_type", 0) or 0)
        r += 1
        ws[f"A{r}"] = "Judgement required"
        ws[f"B{r}"] = int(summ.get("judgement_required", 0) or 0)
        r += 1
        ws[f"A{r}"] = "Total IAS 37 provision (assurance values)"
        ws[f"B{r}"] = float(summ.get("total_ias37_provision", 0) or 0)
        ws[f"B{r}"].number_format = money_fmt
        r += 1
        ws[f"A{r}"] = "Total IFRS 15 deferred (service)"
        ws[f"B{r}"] = float(summ.get("total_ifrs15_deferred", 0) or 0)
        ws[f"B{r}"].number_format = money_fmt
        r += 2

        ws[f"A{r}"] = "WARRANTY CLASSIFICATION"
        ws[f"A{r}"].font = self.subtitle_font
        r += 1
        heads = [
            "ID",
            "Contract",
            "Description",
            "Period (mo)",
            "Value",
            "Classification",
            "Standard",
            "Provision Req.",
            "Deferred Revenue",
        ]
        for i, h in enumerate(heads):
            col = chr(ord("A") + i)
            ws[f"{col}{r}"] = h
            ws[f"{col}{r}"].font = hdr_f
            ws[f"{col}{r}"].fill = orange
            ws[f"{col}{r}"].border = self.border
        r += 1

        for w in payload.get("warranties") or []:
            if not isinstance(w, dict):
                continue
            wt = str(w.get("warranty_type", ""))
            row_fill = grey_r if wt == "ASSURANCE" else blue_r if wt == "SERVICE" else amb_r
            desc = str(w.get("warranty_description") or "")
            ws[f"A{r}"] = w.get("warranty_id", "")
            ws[f"B{r}"] = w.get("contract_id", "")
            ws[f"C{r}"] = desc
            ws[f"D{r}"] = int(w.get("warranty_period_months", 0) or 0)
            ws[f"E{r}"] = float(w.get("warranty_value", 0) or 0)
            ws[f"F{r}"] = wt
            ws[f"G{r}"] = str(w.get("accounting_standard", ""))
            ws[f"H{r}"] = "Yes" if w.get("provision_required") else "No"
            ws[f"I{r}"] = float(w.get("deferred_revenue_amount", 0) or 0)
            ws[f"E{r}"].number_format = money_fmt
            ws[f"I{r}"].number_format = money_fmt
            for col in "ABCDEFGHI":
                ws[f"{col}{r}"].border = self.border
                ws[f"{col}{r}"].fill = row_fill
            r += 1

        r += 2
        ws[f"A{r}"] = "INDICATOR DETAIL (per warranty)"
        ws[f"A{r}"].font = self.subtitle_font
        r += 1

        for w in payload.get("warranties") or []:
            if not isinstance(w, dict):
                continue
            wid = str(w.get("warranty_id", ""))
            ws.merge_cells(f"A{r}:E{r}")
            ws[f"A{r}"] = f"Warranty: {wid} — {w.get('warranty_description', '')}"
            ws[f"A{r}"].font = hdr_f
            ws[f"A{r}"].fill = orange
            r += 1
            ws[f"A{r}"] = "Indicator"
            ws[f"B{r}"] = "Value"
            ws[f"C{r}"] = "Supports"
            ws[f"D{r}"] = "Label"
            for col in ("A", "B", "C", "D"):
                ws[f"{col}{r}"].font = hdr_f
                ws[f"{col}{r}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                ws[f"{col}{r}"].border = self.border
            r += 1
            detail = w.get("indicators_detail") or {}
            for _key, info in detail.items():
                if not isinstance(info, dict):
                    continue
                val = info.get("value")
                val_s = "Yes" if val is True else "No" if val is False else str(val)
                ws[f"A{r}"] = str(_key)
                ws[f"B{r}"] = val_s
                ws[f"C{r}"] = str(info.get("supports", ""))
                ws[f"D{r}"] = str(info.get("label", ""))
                for col in ("A", "B", "C", "D"):
                    ws[f"{col}{r}"].border = self.border
                sup = str(info.get("supports", ""))
                if sup == "ASSURANCE":
                    ws[f"C{r}"].fill = grey_r
                elif sup == "SERVICE":
                    ws[f"C{r}"].fill = blue_r
                r += 1
            r += 1

        for col, w in zip("ABCDEFGHI", (14, 18, 40, 12, 14, 18, 22, 14, 16)):
            ws.column_dimensions[col].width = w

    def _create_bill_and_hold_sheet(self, wb: Workbook, payload: Dict[str, Any], money_fmt: str) -> None:
        """Sheet: Bill-and-hold (IFRS 15.B79–B82) — optional when export payload is present."""
        ws = wb.create_sheet("Bill-and-Hold")
        orange = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        hdr_f = Font(bold=True, size=11, color="1F4E78")
        green_r = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_r = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_lbl = Font(bold=True, size=11, color="C00000")

        ws["A1"] = "BILL-AND-HOLD — IFRS 15.B79-B82"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:I1")
        r = 3
        ws.merge_cells(f"A{r}:C{r}")
        c = ws[f"A{r}"]
        c.value = "HIGH AUDIT RISK — bill-and-hold is rigorously tested by auditors"
        c.font = red_lbl
        c.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        r += 2
        summ = payload.get("summary") or {}
        ws[f"A{r}"] = "Total arrangements"
        ws[f"B{r}"] = int(summ.get("total", 0) or 0)
        r += 1
        ws[f"A{r}"] = "Revenue recognisable now (all criteria met)"
        ws[f"B{r}"] = int(summ.get("recognisable_now", 0) or 0)
        r += 1
        ws[f"A{r}"] = "Defer until delivery"
        ws[f"B{r}"] = int(summ.get("deferred_until_delivery", 0) or 0)
        r += 1
        ws[f"A{r}"] = "Total revenue deferred"
        ws[f"B{r}"] = float(summ.get("total_revenue_deferred", 0) or 0)
        ws[f"B{r}"].number_format = money_fmt
        r += 2

        ws[f"A{r}"] = "ARRANGEMENTS"
        ws[f"A{r}"].font = self.subtitle_font
        r += 1
        heads = [
            "ID",
            "Customer",
            "Product",
            "Value",
            "Billing Date",
            "Delivery Date",
            "Criteria Met",
            "Conclusion",
            "Revenue Deferred",
        ]
        for i, h in enumerate(heads):
            col = chr(ord("A") + i)
            ws[f"{col}{r}"] = h
            ws[f"{col}{r}"].font = hdr_f
            ws[f"{col}{r}"].fill = orange
            ws[f"{col}{r}"].border = self.border
        r += 1

        for a in payload.get("arrangements") or []:
            if not isinstance(a, dict):
                continue
            conc = str(a.get("conclusion", ""))
            row_fill = green_r if conc == "REVENUE_RECOGNISABLE" else red_r
            ws[f"A{r}"] = a.get("arrangement_id", "")
            ws[f"B{r}"] = a.get("customer_name", "")
            ws[f"C{r}"] = a.get("product_description", "")
            ws[f"D{r}"] = float(a.get("contract_value", 0) or 0)
            ws[f"E{r}"] = a.get("billing_date", "")
            ws[f"F{r}"] = a.get("expected_delivery_date", "")
            ws[f"G{r}"] = f"{int(a.get('criteria_met_count', 0) or 0)}/4"
            ws[f"H{r}"] = conc
            ws[f"I{r}"] = float(a.get("revenue_deferred", 0) or 0)
            ws[f"D{r}"].number_format = money_fmt
            ws[f"I{r}"].number_format = money_fmt
            for col in "ABCDEFGHI":
                ws[f"{col}{r}"].border = self.border
                ws[f"{col}{r}"].fill = row_fill
            r += 1

        r += 2
        ws[f"A{r}"] = "CRITERIA CHECKLIST (per arrangement)"
        ws[f"A{r}"].font = self.subtitle_font
        r += 1

        for a in payload.get("arrangements") or []:
            if not isinstance(a, dict):
                continue
            aid = str(a.get("arrangement_id", ""))
            ws.merge_cells(f"A{r}:E{r}")
            ws[f"A{r}"] = f"Arrangement: {aid} — {a.get('product_description', '')}"
            ws[f"A{r}"].font = hdr_f
            ws[f"A{r}"].fill = orange
            r += 1
            ws[f"A{r}"] = "Criterion"
            ws[f"B{r}"] = "MET?"
            ws[f"C{r}"] = "Explanation"
            for col in ("A", "B", "C"):
                ws[f"{col}{r}"].font = hdr_f
                ws[f"{col}{r}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                ws[f"{col}{r}"].border = self.border
            r += 1
            detail = a.get("criteria_detail") or {}
            for _key, info in detail.items():
                if not isinstance(info, dict):
                    continue
                met = bool(info.get("met"))
                ws[f"A{r}"] = str(info.get("label", _key))
                ws[f"B{r}"] = "MET" if met else "FAILED"
                ws[f"C{r}"] = str(info.get("explanation", ""))
                for col in ("A", "B", "C"):
                    ws[f"{col}{r}"].border = self.border
                ws[f"B{r}"].fill = green_r if met else red_r
                r += 1
            r += 1

        for col, w in zip("ABCDEFGHI", (14, 22, 36, 14, 14, 14, 12, 22, 16)):
            ws.column_dimensions[col].width = w

    def _create_financing_component_sheet(self, wb: Workbook, payload: Dict[str, Any], money_fmt: str) -> None:
        """Sheet: Significant financing component (IFRS 15.60–65) — optional when export payload is present."""
        ws = wb.create_sheet("Financing Component")
        orange = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        hdr_f = Font(bold=True, size=11, color="1F4E78")
        blue_r = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        amb_r = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        grey_r = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        ws["A1"] = "SIGNIFICANT FINANCING COMPONENT — IFRS 15.60-65"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:J1")
        summ = payload.get("summary") or {}
        r = 3
        for lbl, key, is_money in [
            ("Total contracts", "total", False),
            ("Practical expedient applied", "expedient_applied", False),
            ("Financing adjusted", "financing_adjusted", False),
            ("Total financing amount", "total_financing_amount", True),
            ("Interest income (total)", "interest_income_total", True),
            ("Interest expense (total)", "interest_expense_total", True),
        ]:
            ws[f"A{r}"] = lbl
            val = summ.get(key, 0)
            ws[f"B{r}"] = val
            if is_money and isinstance(val, (int, float)):
                ws[f"B{r}"].number_format = money_fmt
            ws[f"A{r}"].border = self.border
            ws[f"B{r}"].border = self.border
            r += 1
        r += 1

        ws[f"A{r}"] = "CONTRACTS"
        ws[f"A{r}"].font = self.subtitle_font
        r += 1
        heads = [
            "Contract ID",
            "Nominal",
            "PV",
            "Revenue",
            "Financing",
            "Type",
            "Expedient",
            "Transfer",
            "Payment",
            "IFRS Ref",
        ]
        for i, h in enumerate(heads):
            col = chr(ord("A") + i)
            ws[f"{col}{r}"] = h
            ws[f"{col}{r}"].font = hdr_f
            ws[f"{col}{r}"].fill = orange
            ws[f"{col}{r}"].border = self.border
        r += 1

        for c in payload.get("contracts") or []:
            if not isinstance(c, dict):
                continue
            ft = str(c.get("financing_type", ""))
            exp = bool(c.get("practical_expedient_applied"))
            row_fill = grey_r if exp else blue_r if ft == "INTEREST_INCOME" else amb_r if ft == "INTEREST_EXPENSE" else grey_r
            ws[f"A{r}"] = c.get("contract_id", "")
            ws[f"B{r}"] = float(c.get("nominal_payment", c.get("contract_value", 0)) or 0)
            ws[f"C{r}"] = float(c.get("pv_of_payment", 0) or 0)
            ws[f"D{r}"] = float(c.get("revenue_amount", 0) or 0)
            ws[f"E{r}"] = float(c.get("financing_amount", 0) or 0)
            ws[f"F{r}"] = ft
            ws[f"G{r}"] = "Yes" if exp else "No"
            ws[f"H{r}"] = str(c.get("transfer_date", ""))
            ws[f"I{r}"] = str(c.get("payment_date", ""))
            ws[f"J{r}"] = str(c.get("ifrs_reference", ""))
            for col in "ABCDEFGHIJ":
                ws[f"{col}{r}"].border = self.border
                ws[f"{col}{r}"].fill = row_fill
            for col in ("B", "C", "D", "E"):
                ws[f"{col}{r}"].number_format = money_fmt
            r += 1

        r += 2
        ws[f"A{r}"] = "INTEREST AMORTISATION (per contract)"
        ws[f"A{r}"].font = self.subtitle_font
        r += 1

        for c in payload.get("contracts") or []:
            if not isinstance(c, dict) or bool(c.get("practical_expedient_applied")):
                continue
            cid = str(c.get("contract_id", ""))
            ws.merge_cells(f"A{r}:E{r}")
            ws[f"A{r}"] = f"Contract: {cid} — {c.get('description', '')}"
            ws[f"A{r}"].font = hdr_f
            ws[f"A{r}"].fill = orange
            r += 1
            ws[f"A{r}"] = "Period"
            ws[f"B{r}"] = "Opening"
            ws[f"C{r}"] = "Interest"
            ws[f"D{r}"] = "Closing"
            for col in ("A", "B", "C", "D"):
                ws[f"{col}{r}"].font = hdr_f
                ws[f"{col}{r}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                ws[f"{col}{r}"].border = self.border
            r += 1
            for row in c.get("amortisation_schedule") or []:
                if not isinstance(row, dict):
                    continue
                ws[f"A{r}"] = row.get("period", "")
                ws[f"B{r}"] = float(row.get("opening_balance", 0) or 0)
                ws[f"C{r}"] = float(row.get("interest", 0) or 0)
                ws[f"D{r}"] = float(row.get("closing_balance", 0) or 0)
                for col in ("B", "C", "D"):
                    ws[f"{col}{r}"].number_format = money_fmt
                for col in ("A", "B", "C", "D"):
                    ws[f"{col}{r}"].border = self.border
                r += 1
            r += 1

        for col, w in zip("ABCDEFGHIJ", (16, 14, 14, 14, 14, 18, 12, 12, 12, 14)):
            ws.column_dimensions[col].width = w

    def _create_tp_adjustments_sheet(self, wb: Workbook, payload: Dict[str, Any], money_fmt: str) -> None:
        """Sheet 16 — IFRS 15.66-72 transaction price adjustments (non-cash + consideration payable)."""
        ws = wb.create_sheet("TP Adjustments")
        orange = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        hdr_f = Font(bold=True, size=11, color="1F4E78")
        red_r = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        amb_r = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        org_r = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")

        ws["A1"] = "TRANSACTION PRICE ADJUSTMENTS — IFRS 15.66-72"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:H1")
        r = 3

        nc = payload.get("non_cash") or {}
        nc_items = nc.get("items") or []
        ws[f"A{r}"] = "A — NON-CASH CONSIDERATION (IFRS 15.66-69)"
        ws[f"A{r}"].font = self.subtitle_font
        r += 1
        heads_nc = [
            "Item ID",
            "Contract ID",
            "Type",
            "Method",
            "TP addition",
            "Description",
            "IFRS Ref",
        ]
        for i, h in enumerate(heads_nc):
            col = chr(ord("A") + i)
            ws[f"{col}{r}"] = h
            ws[f"{col}{r}"].font = hdr_f
            ws[f"{col}{r}"].fill = orange
            ws[f"{col}{r}"].border = self.border
        r += 1
        for it in nc_items:
            if not isinstance(it, dict):
                continue
            ws[f"A{r}"] = it.get("item_id", "")
            ws[f"B{r}"] = it.get("contract_id", "")
            ws[f"C{r}"] = it.get("consideration_type", "")
            ws[f"D{r}"] = it.get("measurement_method", "")
            ws[f"E{r}"] = float(it.get("transaction_price_addition", 0) or 0)
            ws[f"F{r}"] = str(it.get("description", ""))[:200]
            ws[f"G{r}"] = str(it.get("ifrs_reference", ""))
            for col in "ABCDEFG":
                ws[f"{col}{r}"].border = self.border
            ws[f"E{r}"].number_format = money_fmt
            r += 1
        ws[f"A{r}"] = "Total TP from non-cash"
        ws[f"B{r}"] = float(nc.get("total_tp_from_non_cash", 0) or 0)
        ws[f"B{r}"].number_format = money_fmt
        ws[f"A{r}"].font = Font(bold=True)
        r += 3

        cp = payload.get("consideration_payable") or {}
        cp_items = cp.get("items") or []
        ws[f"A{r}"] = "B — CONSIDERATION PAYABLE TO CUSTOMER (IFRS 15.70-72)"
        ws[f"A{r}"].font = self.subtitle_font
        r += 1
        heads_cp = [
            "Item ID",
            "Contract ID",
            "Payment type",
            "Amount",
            "Treatment",
            "Revenue reduction",
            "Cost recognition",
            "Description",
        ]
        for i, h in enumerate(heads_cp):
            col = chr(ord("A") + i)
            ws[f"{col}{r}"] = h
            ws[f"{col}{r}"].font = hdr_f
            ws[f"{col}{r}"].fill = orange
            ws[f"{col}{r}"].border = self.border
        r += 1
        for it in cp_items:
            if not isinstance(it, dict):
                continue
            treat = str(it.get("treatment", ""))
            row_fill = red_r if treat == "REVENUE_REDUCTION" else amb_r if treat == "COST_FULL" else org_r
            ws[f"A{r}"] = it.get("item_id", "")
            ws[f"B{r}"] = it.get("contract_id", "")
            ws[f"C{r}"] = it.get("payment_type", "")
            ws[f"D{r}"] = float(it.get("amount", 0) or 0)
            ws[f"E{r}"] = treat
            ws[f"F{r}"] = float(it.get("revenue_reduction", 0) or 0)
            ws[f"G{r}"] = float(it.get("cost_recognition", 0) or 0)
            ws[f"H{r}"] = str(it.get("description", ""))[:200]
            for col in "ABCDEFGH":
                ws[f"{col}{r}"].border = self.border
                ws[f"{col}{r}"].fill = row_fill
            for col in ("D", "F", "G"):
                ws[f"{col}{r}"].number_format = money_fmt
            r += 1

        for col, w in zip("ABCDEFGH", (14, 16, 14, 14, 28, 14, 14, 36)):
            ws.column_dimensions[col].width = w

    def _create_audit_trail_sheet(self, wb: Workbook, entries: List[Dict[str, Any]], money_fmt: str) -> None:
        """Sheet: IFRS 15 audit trail (last sheet in workbook)."""
        _ = money_fmt
        ws = wb.create_sheet("Audit Trail")
        hdr_f = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        amber = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        heads = [
            "Entry ID",
            "Timestamp",
            "Action",
            "Contract ID",
            "Description",
            "IFRS Reference",
            "Sign-Off Required",
            "Signed Off By",
            "Signed Off At",
            "Notes",
        ]
        r = 1
        letters = [chr(ord("A") + j) for j in range(len(heads))]
        for i, h in enumerate(heads):
            c = letters[i]
            ws[f"{c}{r}"] = h
            ws[f"{c}{r}"].font = hdr_f
            ws[f"{c}{r}"].fill = hdr_fill
            ws[f"{c}{r}"].border = self.border
        r = 2
        for e in entries:
            if not isinstance(e, dict):
                continue
            req = bool(e.get("sign_off_required"))
            signed = bool((e.get("signed_off_by") or "").strip())
            letters = [chr(ord("A") + j) for j in range(10)]
            vals = [
                e.get("entry_id", ""),
                e.get("timestamp", ""),
                e.get("action", ""),
                e.get("contract_id", ""),
                e.get("description", ""),
                e.get("ifrs_reference", ""),
                "TRUE" if req else "FALSE",
                e.get("signed_off_by", ""),
                e.get("signed_off_at", ""),
                e.get("notes", ""),
            ]
            for i, c in enumerate(letters):
                ws[f"{c}{r}"] = vals[i]
                ws[f"{c}{r}"].border = self.border
            if req and not signed:
                ws[f"A{r}"].fill = amber
            elif req and signed:
                ws[f"A{r}"].fill = green
            r += 1
        last = max(r - 1, 2)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:J{last}"
        for j in range(10):
            ws.column_dimensions[chr(ord("A") + j)].width = 18
        ws.column_dimensions["E"].width = 40
