"""
RERA milestone & regulatory deadline tracker — UAE Law No. 8 of 2007, FTA VAT, IFRS 15.
"""

from __future__ import annotations

import calendar
import io
import re
from datetime import date, datetime, timedelta, timezone
from enum import Enum
from typing import Any, Dict, List, Literal, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field

DEFAULT_COMPLETION_RATE_PER_MONTH = 3.5
DAYS_PER_MONTH = 30.44
DUE_SOON_DAYS = 30
FTA_DUE_DAYS_AFTER_QUARTER = 28

RERA_MILESTONE_TYPES = frozenset(
    {
        "rera_30pct",
        "rera_50pct",
        "rera_60pct",
        "rera_80pct",
        "rera_100pct",
    }
)
FTA_MILESTONE_TYPES = frozenset(
    {"fta_vat_q1", "fta_vat_q2", "fta_vat_q3", "fta_vat_q4"}
)


class MilestoneType(str, Enum):
    RERA_30 = "rera_30pct"
    RERA_50 = "rera_50pct"
    RERA_60 = "rera_60pct"
    RERA_80 = "rera_80pct"
    RERA_100 = "rera_100pct"
    FTA_VAT_Q1 = "fta_vat_q1"
    FTA_VAT_Q2 = "fta_vat_q2"
    FTA_VAT_Q3 = "fta_vat_q3"
    FTA_VAT_Q4 = "fta_vat_q4"
    IFRS15_QUARTER_END = "ifrs15_quarter_end"
    OQOOD_AMENDMENT = "oqood_amendment"
    RERA_CERT_RENEWAL = "rera_cert_renewal"


class MilestoneStatus(str, Enum):
    UPCOMING = "upcoming"
    DUE_SOON = "due_soon"
    OVERDUE = "overdue"
    COMPLETED = "completed"
    NOT_YET_TRIGGERED = "not_triggered"


class RERAMilestone(BaseModel):
    milestone_id: str
    milestone_type: str
    title: str
    description: str
    threshold_pct: Optional[float] = None
    triggered_at_pct: Optional[float] = None
    due_date: Optional[str] = None
    projected_date: Optional[str] = None
    completed_date: Optional[str] = None
    status: str
    days_until_due: Optional[int] = None
    action_required: str
    law_reference: str
    escrow_release_permitted_pct: Optional[float] = None
    escrow_release_amount_aed: Optional[float] = None
    is_overdue: bool = False
    priority: Literal["critical", "high", "medium", "low"] = "medium"
    completed_by: Optional[str] = None
    notes: Optional[str] = None


class DeadlineTrackerReport(BaseModel):
    rera_registration_number: str
    project_name: str = ""
    current_completion_pct: float = 0.0
    projected_completion_rate_pct_per_month: float = DEFAULT_COMPLETION_RATE_PER_MONTH
    contract_price_aed: float = 0.0
    milestones: List[RERAMilestone] = Field(default_factory=list)
    overdue_count: int = 0
    due_soon_count: int = 0
    completed_count: int = 0
    upcoming_count: int = 0
    next_milestone: Optional[RERAMilestone] = None
    critical_alerts: List[str] = Field(default_factory=list)
    generated_at: str = ""
    currency: str = "AED"
    rate_assumption_note: str = (
        f"Completion rate default {DEFAULT_COMPLETION_RATE_PER_MONTH}%/month "
        "used when not supplied (typical UAE mid-rise construction)."
    )


def _parse_date(raw: Optional[str]) -> Optional[date]:
    if not raw:
        return None
    try:
        return datetime.strptime(str(raw).strip()[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _days_until(due: Optional[date], today: date) -> Optional[int]:
    if due is None:
        return None
    return (due - today).days


def calculate_projected_date(
    current_pct: float,
    target_pct: float,
    completion_rate_per_month: float,
    reference_date: str,
) -> Optional[str]:
    if completion_rate_per_month <= 0:
        return None
    months_needed = (target_pct - current_pct) / completion_rate_per_month
    if months_needed < 0:
        return None
    ref = _parse_date(reference_date)
    if not ref:
        return None
    projected = ref + timedelta(days=months_needed * DAYS_PER_MONTH)
    return projected.isoformat()


def _quarter_end(year: int, quarter: int) -> date:
    end_month = quarter * 3
    last_day = calendar.monthrange(year, end_month)[1]
    return date(year, end_month, last_day)


def _fta_vat_due(year: int, quarter: int) -> date:
    q_end = _quarter_end(year, quarter)
    due = q_end + timedelta(days=FTA_DUE_DAYS_AFTER_QUARTER)
    if quarter == 4:
        due = date(year + 1, 1, 28)
    else:
        month = quarter * 3 + 1
        due = date(year, month, 28)
    return due


def _quarter_period_bounds(year: int, quarter: int) -> Tuple[str, str]:
    start_month = (quarter - 1) * 3 + 1
    end_month = quarter * 3
    start = date(year, start_month, 1)
    end_day = calendar.monthrange(year, end_month)[1]
    end = date(year, end_month, end_day)
    return start.isoformat(), end.isoformat()


def _upcoming_quarter_ends(ref: date, count: int = 4) -> List[Tuple[date, int, int]]:
    """Return (quarter_end_date, quarter_num, year) for current and next quarters."""
    results: List[Tuple[date, int, int]] = []
    y, m = ref.year, ref.month
    q = (m - 1) // 3 + 1
    while len(results) < count:
        qe = _quarter_end(y, q)
        if qe >= ref or (qe.year == ref.year and qe.month >= ref.month - 2):
            results.append((qe, q, y))
        q += 1
        if q > 4:
            q = 1
            y += 1
    return results[:count]


def _apply_time_status(
    milestone: RERAMilestone,
    today: date,
    existing_completions: Dict[str, Any],
    completion_key: str,
) -> RERAMilestone:
    completed = existing_completions.get(completion_key)
    if completed:
        milestone.status = MilestoneStatus.COMPLETED.value
        milestone.completed_date = str(completed)[:10]
        milestone.is_overdue = False
        return milestone

    due = _parse_date(milestone.due_date)
    if due is None:
        milestone.status = MilestoneStatus.UPCOMING.value
        return milestone

    days = _days_until(due, today)
    milestone.days_until_due = days
    if days is not None and days < 0:
        milestone.status = MilestoneStatus.OVERDUE.value
        milestone.is_overdue = True
        if milestone.priority != "critical":
            milestone.priority = "high"
    elif days is not None and days <= DUE_SOON_DAYS:
        milestone.status = MilestoneStatus.DUE_SOON.value
    else:
        milestone.status = MilestoneStatus.UPCOMING.value
    return milestone


def _apply_rera_status(
    milestone: RERAMilestone,
    current_pct: float,
    today: date,
    completion_rate: float,
    existing_completions: Dict[str, Any],
) -> RERAMilestone:
    mtype = milestone.milestone_type
    threshold = milestone.threshold_pct or 0.0
    completed = existing_completions.get(mtype)

    if completed:
        milestone.status = MilestoneStatus.COMPLETED.value
        milestone.completed_date = str(completed)[:10]
        milestone.is_overdue = False
        milestone.triggered_at_pct = threshold
        return milestone

    if current_pct < threshold:
        projected = calculate_projected_date(
            current_pct, threshold, completion_rate, today.isoformat()
        )
        milestone.projected_date = projected
        proj_d = _parse_date(projected)
        days = _days_until(proj_d, today) if proj_d else None
        milestone.days_until_due = days
        if proj_d and days is not None and days <= DUE_SOON_DAYS:
            milestone.status = MilestoneStatus.DUE_SOON.value
        else:
            milestone.status = MilestoneStatus.NOT_YET_TRIGGERED.value
        return milestone

    milestone.triggered_at_pct = current_pct
    if milestone.due_date:
        due = _parse_date(milestone.due_date)
        milestone.days_until_due = _days_until(due, today)
    else:
        milestone.days_until_due = 0

    milestone.status = MilestoneStatus.OVERDUE.value
    milestone.is_overdue = True
    milestone.priority = "critical"
    return milestone


def build_rera_milestones(
    current_completion_pct: float,
    contract_price_aed: float,
    project_start_date: str,
    expected_completion_date: str,
    completion_rate_per_month: float,
    today: str,
    existing_completions: Dict[str, Any],
) -> List[RERAMilestone]:
    today_d = _parse_date(today) or date.today()
    rera_defs = [
        (
            MilestoneType.RERA_30,
            "RERA Escrow Release — 30% Completion",
            (
                "Construction has reached 30%. Developer may submit "
                "escrow release request to escrow bank for up to 30% "
                "of total contract value."
            ),
            30.0,
            (
                "1. Obtain RERA inspection report confirming 30% completion.\n"
                "2. Submit escrow release request to escrow bank.\n"
                "3. Attach completion certificate to release request.\n"
                "4. Update IFRS 15 revenue recognition schedule."
            ),
            "UAE Law No. 8 of 2007, Article 8; DLD Escrow Instructions",
            "high",
        ),
        (
            MilestoneType.RERA_50,
            "RERA Escrow Release — 50% Completion",
            (
                "Construction has reached 50%. Mid-point escrow release "
                "permissible up to cumulative 50% of contract value."
            ),
            50.0,
            (
                "1. Obtain RERA inspection report confirming 50% completion.\n"
                "2. Submit updated escrow release request.\n"
                "3. Reconcile with previous 30% release.\n"
                "4. Update IFRS 15 quarterly revenue schedule."
            ),
            "UAE Law No. 8 of 2007, Article 8",
            "high",
        ),
        (
            MilestoneType.RERA_60,
            "RERA Escrow Release — 60% Completion",
            (
                "Construction has reached 60%. Second tranche escrow "
                "release permissible up to cumulative 60%."
            ),
            60.0,
            (
                "1. Obtain RERA inspection report confirming 60%.\n"
                "2. Submit escrow release request for incremental amount.\n"
                "3. Notify off-plan buyers of progress per RERA rules.\n"
                "4. Review IFRS 15 contract asset / deferred revenue balance."
            ),
            "UAE Law No. 8 of 2007, Article 8",
            "high",
        ),
        (
            MilestoneType.RERA_80,
            "RERA Escrow Release — 80% Completion",
            (
                "Construction has reached 80%. Major escrow release "
                "permissible. Begin handover preparation."
            ),
            80.0,
            (
                "1. Obtain RERA inspection confirming 80%.\n"
                "2. Submit escrow release request.\n"
                "3. Issue snag list to buyers if contractually required.\n"
                "4. Begin handover documentation preparation.\n"
                "5. Review SPA handover date — approaching."
            ),
            "UAE Law No. 8 of 2007, Article 8",
            "critical",
        ),
        (
            MilestoneType.RERA_100,
            "Project Completion — Handover & Final Escrow Release",
            (
                "Construction 100% complete. Full escrow release on "
                "handover. RERA completion certificate required."
            ),
            100.0,
            (
                "1. Obtain RERA final completion certificate from DLD.\n"
                "2. Issue NOC to buyers.\n"
                "3. Execute handover per SPA terms.\n"
                "4. Release remaining escrow balance.\n"
                "5. Recognise final revenue — IFRS 15 performance obligation complete.\n"
                "6. File final VAT invoice to buyers."
            ),
            "UAE Law No. 8 of 2007, Articles 8 & 11; IFRS 15 para 38",
            "critical",
        ),
    ]

    milestones: List[RERAMilestone] = []
    for mtype, title, desc, thresh, action, law, priority in rera_defs:
        escrow_pct = thresh
        mid = RERAMilestone(
            milestone_id=f"{mtype.value}_{int(thresh)}",
            milestone_type=mtype.value,
            title=title,
            description=desc,
            threshold_pct=thresh,
            action_required=action,
            law_reference=law,
            escrow_release_permitted_pct=escrow_pct,
            escrow_release_amount_aed=round(contract_price_aed * (escrow_pct / 100.0), 2),
            priority=priority,  # type: ignore[arg-type]
            status=MilestoneStatus.NOT_YET_TRIGGERED.value,
        )
        if mtype == MilestoneType.RERA_100:
            mid.due_date = str(expected_completion_date)[:10] if expected_completion_date else None
        milestones.append(
            _apply_rera_status(
                mid,
                current_completion_pct,
                today_d,
                completion_rate_per_month,
                existing_completions,
            )
        )

    year = today_d.year
    quarter_names = {1: "Q1", 2: "Q2", 3: "Q3", 4: "Q4"}
    fta_types = [
        MilestoneType.FTA_VAT_Q1,
        MilestoneType.FTA_VAT_Q2,
        MilestoneType.FTA_VAT_Q3,
        MilestoneType.FTA_VAT_Q4,
    ]
    for qi, mtype in enumerate(fta_types, start=1):
        p_start, p_end = _quarter_period_bounds(year, qi)
        due = _fta_vat_due(year, qi)
        days_to_due = _days_until(due, today_d) or 999
        priority: Literal["critical", "high", "medium", "low"] = (
            "high" if days_to_due <= 14 else "medium"
        )
        fta = RERAMilestone(
            milestone_id=f"{mtype.value}_{year}",
            milestone_type=mtype.value,
            title=f"FTA VAT Return — {quarter_names[qi]} {year}",
            description=(
                f"Quarterly VAT return due for {p_start} to {p_end}. "
                "Box 1a must include all taxable supplies for the period."
            ),
            action_required=(
                "1. Reconcile IFRS 15 recognised revenue vs FTA taxable supply.\n"
                "2. Complete VAT reconciliation in FinReportAI.\n"
                "3. File VAT return on FTA e-Services portal.\n"
                "4. Retain filed return reference number."
            ),
            law_reference="Federal Decree-Law No. 8 of 2017, Article 64",
            due_date=due.isoformat(),
            priority=priority,
            status=MilestoneStatus.UPCOMING.value,
        )
        milestones.append(
            _apply_time_status(fta, today_d, existing_completions, mtype.value)
        )

    for qe, qnum, y in _upcoming_quarter_ends(today_d, 4):
        p_start, p_end = _quarter_period_bounds(y, qnum)
        key = f"ifrs15_q{qnum}_{y}"
        days_to_due = _days_until(qe, today_d) or 999
        priority = "high" if days_to_due <= 14 else "medium"
        ifrs = RERAMilestone(
            milestone_id=key,
            milestone_type=f"{MilestoneType.IFRS15_QUARTER_END.value}_{qnum}_{y}",
            title=f"IFRS 15 Revenue Recognition — Q{qnum} {y}",
            description=(
                "Quarter-end revenue recognition calculation required. "
                "Update completion %, run recognition, update disclosures."
            ),
            action_required=(
                "1. Obtain updated construction completion % from site team.\n"
                "2. Run recognition in FinReportAI Real Estate module.\n"
                "3. Update quarterly revenue schedule.\n"
                "4. Review IFRS 15.120-122 disclosure text.\n"
                "5. Post journal entries to GL."
            ),
            law_reference="IFRS 15 para 38; IAS 34 interim reporting",
            due_date=qe.isoformat(),
            priority=priority,
            status=MilestoneStatus.UPCOMING.value,
        )
        milestones.append(
            _apply_time_status(ifrs, today_d, existing_completions, key)
        )

    return milestones


def generate_critical_alerts(milestones: List[RERAMilestone]) -> List[str]:
    alerts: List[str] = []
    for m in milestones:
        if m.status != MilestoneStatus.OVERDUE.value:
            continue
        if m.milestone_type in RERA_MILESTONE_TYPES:
            alerts.append(
                f"OVERDUE: {m.title} — construction passed "
                f"{m.threshold_pct}% but escrow release not requested. "
                "Submit to escrow bank immediately."
            )
        elif m.milestone_type in FTA_MILESTONE_TYPES:
            alerts.append(
                f"OVERDUE: {m.title} — VAT return not filed. "
                "Late filing penalty applies. File on FTA portal immediately."
            )

    for m in milestones:
        if m.status == MilestoneStatus.DUE_SOON.value and m.priority == "critical":
            alerts.append(
                f"DUE SOON: {m.title} due in {m.days_until_due} days."
            )
    return alerts


def _sort_key_for_next(m: RERAMilestone) -> Tuple[int, str]:
    order = {
        MilestoneStatus.OVERDUE.value: 0,
        MilestoneStatus.DUE_SOON.value: 1,
        MilestoneStatus.NOT_YET_TRIGGERED.value: 2,
        MilestoneStatus.UPCOMING.value: 3,
    }
    date_key = m.due_date or m.projected_date or "9999-12-31"
    return (order.get(m.status, 9), date_key)


def build_deadline_tracker(input_data: Dict[str, Any]) -> DeadlineTrackerReport:
    today = str(input_data.get("today") or date.today().isoformat())[:10]
    rate = float(
        input_data.get("completion_rate_per_month") or DEFAULT_COMPLETION_RATE_PER_MONTH
    )
    if rate <= 0:
        rate = DEFAULT_COMPLETION_RATE_PER_MONTH

    existing = dict(input_data.get("existing_completions") or {})
    current_pct = float(input_data.get("current_completion_pct") or 0)
    contract_price = float(input_data.get("contract_price_aed") or 0)
    project_start = str(
        input_data.get("project_start_date")
        or input_data.get("construction_start")
        or today
    )[:10]
    expected_end = str(
        input_data.get("expected_completion_date")
        or input_data.get("expected_handover")
        or today
    )[:10]

    milestones = build_rera_milestones(
        current_pct,
        contract_price,
        project_start,
        expected_end,
        rate,
        today,
        existing,
    )
    critical_alerts = generate_critical_alerts(milestones)

    overdue_count = sum(1 for m in milestones if m.status == MilestoneStatus.OVERDUE.value)
    due_soon_count = sum(1 for m in milestones if m.status == MilestoneStatus.DUE_SOON.value)
    completed_count = sum(1 for m in milestones if m.status == MilestoneStatus.COMPLETED.value)
    upcoming_count = sum(
        1
        for m in milestones
        if m.status in (MilestoneStatus.UPCOMING.value, MilestoneStatus.NOT_YET_TRIGGERED.value)
    )

    candidates = [
        m
        for m in milestones
        if m.status
        in (
            MilestoneStatus.DUE_SOON.value,
            MilestoneStatus.NOT_YET_TRIGGERED.value,
            MilestoneStatus.OVERDUE.value,
        )
    ]
    candidates.sort(key=_sort_key_for_next)
    next_milestone = candidates[0] if candidates else None

    return DeadlineTrackerReport(
        rera_registration_number=str(input_data.get("rera_registration_number") or ""),
        project_name=str(input_data.get("project_name") or ""),
        current_completion_pct=current_pct,
        projected_completion_rate_pct_per_month=rate,
        contract_price_aed=contract_price,
        milestones=milestones,
        overdue_count=overdue_count,
        due_soon_count=due_soon_count,
        completed_count=completed_count,
        upcoming_count=upcoming_count,
        next_milestone=next_milestone,
        critical_alerts=critical_alerts,
        generated_at=datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        currency=str(input_data.get("currency") or "AED"),
    )


def _status_sort_rank(status: str) -> int:
    return {
        MilestoneStatus.OVERDUE.value: 0,
        MilestoneStatus.DUE_SOON.value: 1,
        MilestoneStatus.NOT_YET_TRIGGERED.value: 2,
        MilestoneStatus.UPCOMING.value: 3,
        MilestoneStatus.COMPLETED.value: 4,
    }.get(status, 5)


def _status_fill(status: str) -> PatternFill:
    colors_map = {
        MilestoneStatus.OVERDUE.value: "FEE2E2",
        MilestoneStatus.DUE_SOON.value: "FFF3CD",
        MilestoneStatus.COMPLETED.value: "D1FAE5",
    }
    hex_c = colors_map.get(status, "FFFFFF")
    return PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")


def summarize_deadlines_for_contract(
    contract: Dict[str, Any],
    today: Optional[str] = None,
) -> Dict[str, Any]:
    """Lightweight deadline summary for portfolio analytics."""
    snap = contract.get("realestate_snapshot") or {}
    rera = str(snap.get("rera_registration_number") or contract.get("risk") or "").strip()
    if not rera or rera == "—":
        return {
            "deadline_overdue": 0,
            "deadline_due_soon": 0,
            "deadline_milestones": [],
        }
    try:
        report = build_deadline_tracker(
            {
                "rera_registration_number": rera,
                "project_name": snap.get("project_name") or contract.get("project_name") or "",
                "current_completion_pct": float(
                    contract.get("construction_completion_pct")
                    or snap.get("completion_pct")
                    or 0
                ),
                "contract_price_aed": float(contract.get("total_tp") or 0),
                "project_start_date": str(contract.get("start_date") or "")[:10],
                "expected_completion_date": str(contract.get("end_date") or "")[:10],
                "existing_completions": dict(snap.get("deadline_completions") or {}),
                "today": today or date.today().isoformat(),
                "currency": str(contract.get("currency") or "AED"),
            }
        )
        return {
            "deadline_overdue": report.overdue_count,
            "deadline_due_soon": report.due_soon_count,
            "deadline_milestones": [m.model_dump() for m in report.milestones],
        }
    except Exception:
        return {
            "deadline_overdue": 0,
            "deadline_due_soon": 0,
            "deadline_milestones": [],
        }


def export_deadline_tracker_excel(report: DeadlineTrackerReport) -> Tuple[bytes, str]:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Milestone Tracker"
    headers = [
        "Milestone",
        "Threshold %",
        "Status",
        "Due Date",
        "Projected Date",
        "Days Until Due",
        "Escrow Release (AED)",
        "Action Required",
        "Law Reference",
        "Completed Date",
        "Notes",
    ]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

    sorted_ms = sorted(report.milestones, key=lambda m: _status_sort_rank(m.status))
    for m in sorted_ms:
        row_idx = ws1.max_row + 1
        ws1.append(
            [
                m.title,
                m.threshold_pct if m.threshold_pct is not None else "",
                m.status,
                m.due_date or "",
                m.projected_date or "",
                m.days_until_due if m.days_until_due is not None else "",
                m.escrow_release_amount_aed or "",
                m.action_required.replace("\n", " | "),
                m.law_reference,
                m.completed_date or "",
                m.notes or "",
            ]
        )
        for cell in ws1[row_idx]:
            cell.fill = _status_fill(m.status)

    ws2 = wb.create_sheet("Critical Alerts")
    if report.critical_alerts:
        for alert in report.critical_alerts:
            ws2.append([alert])
    else:
        ws2.append(["No critical alerts at this time."])

    buf = io.BytesIO()
    wb.save(buf)
    safe_rera = re.sub(r"[^\w\-]", "_", report.rera_registration_number or "RE")[:30]
    date_part = datetime.now(timezone.utc).strftime("%Y%m%d")
    filename = f"RERA_Deadlines_{safe_rera}_{date_part}.xlsx"
    return buf.getvalue(), filename
