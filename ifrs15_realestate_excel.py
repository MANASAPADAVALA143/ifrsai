"""
Excel export for UAE Real Estate IFRS 15 module.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


class IFRS15RealEstateExcelExporter:
    def __init__(self) -> None:
        self.header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=14, color="1F4E78")
        self.red_fill = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid")
        self.amber_fill = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
        self.green_fill = PatternFill(start_color="86EFAC", end_color="86EFAC", fill_type="solid")

    def _ccy_suffix(self, report: Dict[str, Any]) -> str:
        cur = str(report.get("currency") or "AED").upper()
        return f"({cur})"

    def _footer_note(self, ws: Any, report: Dict[str, Any], row: int) -> None:
        note = str(
            report.get("currency_note") or "All amounts in AED"
        )
        ws.cell(row=row, column=1, value=note)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    def _write_headers(self, ws: Any, headers: List[str], row: int = 1) -> None:
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = self.header_fill
            c.font = self.header_font
            c.alignment = Alignment(horizontal="center")

    def _disclosure_band(self, score: float) -> tuple:
        if score >= 90:
            return "Excellent", self.green_fill
        if score >= 80:
            return "Good", self.green_fill
        if score >= 60:
            return "Adequate", self.amber_fill
        return "Needs Improvement", self.red_fill

    def export_workbook(self, report: Dict[str, Any], filename: str) -> str:
        wb = Workbook()
        ccy = self._ccy_suffix(report)
        ws_sum = wb.active
        ws_sum.title = "Summary"

        off = report.get("off_plan") or {}
        vat = report.get("vat") or {}
        unit = (
            (report.get("ifrs15_calculate_payload") or {}).get("contract_id")
            or "Real Estate Contract"
        )
        rera = report.get("rera_registration_number") or "—"

        ws_sum["A1"] = "UAE Real Estate — IFRS 15 Report"
        ws_sum["A1"].font = self.title_font
        ws_sum["A2"] = "RERA Reg No:"
        ws_sum["B2"] = rera
        ws_sum["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws_sum["A4"] = f"Contract: {unit}"
        ws_sum["A5"] = str(report.get("recognition_trigger_summary") or "")

        rows = [
            (f"Completion %", off.get("completion_pct")),
            (f"Revenue recognised to date {ccy}", off.get("revenue_recognised_to_date")),
            (f"Revenue current period {ccy}", off.get("revenue_current_period")),
            (f"Contract asset {ccy}", off.get("contract_asset")),
            (f"Contract liability {ccy}", off.get("contract_liability")),
            (f"Escrow balance {ccy}", off.get("escrow_balance")),
            (f"Remaining revenue {ccy}", off.get("remaining_revenue")),
            ("Estimated handover", off.get("estimated_handover")),
            (f"Total VAT 5% {ccy}", vat.get("total_vat")),
            ("Disclosure score", report.get("disclosure_score")),
        ]
        r = 7
        for label, val in rows:
            ws_sum.cell(row=r, column=1, value=label)
            ws_sum.cell(row=r, column=2, value=val)
            r += 1
        self._footer_note(ws_sum, report, r + 1)

        ws_sched = wb.create_sheet("Revenue Schedule")
        self._write_headers(
            ws_sched,
            [
                "Period",
                "Period End",
                "Completion %",
                f"Revenue {ccy}",
                f"Cumulative {ccy}",
                "FTA Filing",
            ],
        )
        for i, row in enumerate(report.get("period_schedule") or [], 2):
            ws_sched.cell(row=i, column=1, value=row.get("period"))
            ws_sched.cell(row=i, column=2, value=row.get("period_end"))
            ws_sched.cell(row=i, column=3, value=row.get("completion_pct"))
            ws_sched.cell(row=i, column=4, value=row.get("revenue_recognised"))
            ws_sched.cell(row=i, column=5, value=row.get("cumulative_revenue"))
            ws_sched.cell(row=i, column=6, value=row.get("fta_filing_period"))
        self._footer_note(ws_sched, report, len(report.get("period_schedule") or []) + 3)

        ws_vat = wb.create_sheet("VAT Alignment")
        self._write_headers(
            ws_vat,
            [f"Period", f"Revenue {ccy}", f"VAT @ 5% {ccy}", "FTA Filing"],
        )
        for i, row in enumerate(vat.get("alignment_table") or [], 2):
            ws_vat.cell(row=i, column=1, value=row.get("period"))
            ws_vat.cell(row=i, column=2, value=row.get("revenue_recognised"))
            ws_vat.cell(row=i, column=3, value=row.get("vat_5pct"))
            ws_vat.cell(row=i, column=4, value=row.get("fta_filing_period"))
        self._footer_note(ws_vat, report, len(vat.get("alignment_table") or []) + 3)

        score = float(report.get("disclosure_score") or 0)
        band, band_fill = self._disclosure_band(score)
        ws_disc = wb.create_sheet("Disclosure Quality")
        ws_disc["A1"] = "IFRS 15 Disclosure Quality Assessment — Real Estate UAE"
        ws_disc["A1"].font = self.title_font
        ws_disc["A3"] = "Overall Score"
        ws_disc["B3"] = f"{score}/100"
        ws_disc["B3"].fill = band_fill
        ws_disc["A4"] = "Score band"
        ws_disc["B4"] = band
        ws_disc["A6"] = "Criterion"
        ws_disc["B6"] = "Gap / issue"
        ws_disc["C6"] = "IFRS reference"
        for col in (1, 2, 3):
            ws_disc.cell(row=6, column=col).fill = self.header_fill
            ws_disc.cell(row=6, column=col).font = self.header_font
        dr = 7
        for gap in report.get("disclosure_gaps") or []:
            if isinstance(gap, dict):
                ws_disc.cell(row=dr, column=1, value=gap.get("criterion"))
                ws_disc.cell(row=dr, column=2, value=gap.get("gap"))
                ws_disc.cell(row=dr, column=3, value=gap.get("ifrs_reference"))
                dr += 1
        ws_disc.cell(row=dr + 1, column=1, value="Assessed by FinReportAI | IFRS 15 paras 110–129")
        self._footer_note(ws_disc, report, dr + 3)

        cancel = report.get("cancellation_refund")
        if cancel:
            ws_can = wb.create_sheet("Cancellation")
            ws_can["A1"] = "UAE Law No. 8 of 2007 — Article 11"
            ws_can["A1"].font = self.title_font
            can_rows = [
                ("Developer retention", cancel.get("developer_retention_amount")),
                ("Buyer refund", cancel.get("buyer_refund_amount")),
                ("Escrow to buyer", cancel.get("escrow_release_to_buyer")),
                ("Escrow to developer", cancel.get("escrow_release_to_developer")),
                ("Escrow shortfall", cancel.get("escrow_shortfall")),
                ("IFRS 15 revenue reversal", cancel.get("ifrs15_revenue_reversal")),
            ]
            for i, (lbl, val) in enumerate(can_rows, 3):
                ws_can.cell(row=i, column=1, value=lbl)
                ws_can.cell(row=i, column=2, value=val)
            self._footer_note(ws_can, report, 12)

        timeline = (report.get("escrow") or {}).get("timeline") or []
        if timeline:
            ws_esc = wb.create_sheet("RERA Escrow")
            self._write_headers(
                ws_esc,
                ["Type", "Date / Milestone", f"Amount {ccy}", "Escrow Balance", "Revenue TD"],
            )
            for i, ev in enumerate(timeline, 2):
                ws_esc.cell(row=i, column=1, value=ev.get("type"))
                ws_esc.cell(row=i, column=2, value=ev.get("date") or ev.get("milestone"))
                ws_esc.cell(
                    row=i,
                    column=3,
                    value=ev.get("amount") or ev.get("amount_released"),
                )
                ws_esc.cell(row=i, column=4, value=ev.get("escrow_balance"))
                ws_esc.cell(row=i, column=5, value=ev.get("revenue_recognised_to_date"))
            self._footer_note(ws_esc, report, len(timeline) + 3)

        oqoods = list(report.get("oqood_assessments") or [])
        if oqoods:
            ws_oq = wb.create_sheet("Oqood Amendments")
            ws_oq["A1"] = "Oqood Amendment Assessment — Dubai Land Department"
            ws_oq["A1"].font = self.title_font
            self._write_headers(
                ws_oq,
                ["Modification Type", "Requires Amendment", "Fee (AED)", "IFRS 15 Treatment", "Law Reference"],
                row=3,
            )
            row_i = 4
            for o in oqoods:
                ws_oq.cell(row=row_i, column=1, value=o.get("modification_type"))
                ws_oq.cell(row=row_i, column=2, value=bool(o.get("requires_oqood_amendment")))
                ws_oq.cell(row=row_i, column=3, value=o.get("amendment_fee_aed"))
                ws_oq.cell(row=row_i, column=4, value=o.get("ifrs15_modification_type"))
                ws_oq.cell(row=row_i, column=5, value=o.get("law_reference"))
                if bool(o.get("requires_oqood_amendment")):
                    for c in range(1, 6):
                        ws_oq.cell(row=row_i, column=c).fill = self.amber_fill
                row_i += 1
            ws_oq.cell(row=row_i + 1, column=1, value="Dubai Law No. 13 of 2008, Article 3")
            self._footer_note(ws_oq, report, row_i + 3)

        bundling = report.get("bundling_assessment")
        if isinstance(bundling, dict):
            ws_b = wb.create_sheet("Multi-Unit Bundling")
            ws_b["A1"] = "IFRS 15 Para 17 — Multi-Unit Contract Bundling Assessment"
            ws_b["A1"].font = self.title_font
            ws_b["A3"] = "Should bundle"
            ws_b["B3"] = bool(bundling.get("should_bundle"))
            ws_b["A4"] = "Audit risk"
            ws_b["B4"] = bundling.get("audit_risk_level")
            ws_b["A5"] = "Combined revenue (AED)"
            ws_b["B5"] = bundling.get("combined_revenue_recognised_aed")
            ws_b["A6"] = "Combined vs individual delta (AED)"
            ws_b["B6"] = bundling.get("discount_amount_aed")
            if bool(bundling.get("should_bundle")):
                ws_b["B3"].fill = self.amber_fill
            self._write_headers(
                ws_b,
                ["Contract ID", "Unit Number", "Unit Type", "Contract Price (AED)", "Completion %", "Revenue (AED)"],
                row=8,
            )
            rr = 9
            for unit in bundling.get("individual_schedules") or []:
                ws_b.cell(row=rr, column=1, value=unit.get("contract_id"))
                ws_b.cell(row=rr, column=2, value=unit.get("unit_number"))
                ws_b.cell(row=rr, column=3, value=unit.get("unit_type"))
                ws_b.cell(row=rr, column=4, value=unit.get("contract_price_aed"))
                ws_b.cell(row=rr, column=5, value=unit.get("completion_pct"))
                ws_b.cell(row=rr, column=6, value=unit.get("revenue_recognised_aed"))
                rr += 1
            ws_b.cell(row=rr + 1, column=1, value="IFRS 15 paragraph 17 — Combination of contracts")
            self._footer_note(ws_b, report, rr + 3)

        journals = off.get("journal_entries") or []
        if journals:
            ws_je = wb.create_sheet("Journal Entries")
            self._write_headers(
                ws_je, ["Dr", "Cr", f"Amount {ccy}", "Narrative"]
            )
            for i, je in enumerate(journals, 2):
                ws_je.cell(row=i, column=1, value=je.get("dr"))
                ws_je.cell(row=i, column=2, value=je.get("cr"))
                ws_je.cell(row=i, column=3, value=je.get("amount"))
                ws_je.cell(row=i, column=4, value=je.get("narrative"))
            self._footer_note(ws_je, report, len(journals) + 3)

        for ws in wb.worksheets:
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 36

        path = Path(filename)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return str(path)
