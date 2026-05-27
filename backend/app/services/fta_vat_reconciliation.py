"""
FTA VAT return reconciliation — IFRS 15 quarterly revenue vs FTA Box 1a/1b.
Federal Decree-Law No. 8 of 2017 (UAE VAT).
"""

from __future__ import annotations

import calendar
import io
import re
from datetime import date, datetime
from typing import Any, Dict, List, Literal, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field

VAT_RATE = 0.05
MATCH_TOLERANCE_AED = 1.0
TIMING_RISK_PCT = 20.0
UNEXPLAINED_PCT = 5.0


class FTAVATReturn(BaseModel):
    quarter: str
    period_start: str
    period_end: str
    box_1a_taxable_supplies: float = 0.0
    box_1b_vat_on_supplies: float = 0.0
    box_7_input_vat: float = 0.0
    box_8_net_vat_payable: float = 0.0
    fta_return_ref: Optional[str] = None
    filing_date: Optional[str] = None
    status: Literal["draft", "filed", "amended"] = "draft"


class VATReconciliationLine(BaseModel):
    quarter: str
    period_start: str
    period_end: str
    ifrs15_revenue_recognised: float
    ifrs15_vat_amount: float
    ifrs15_recognition_basis: str = "% completion over time"
    fta_taxable_supply: float = 0.0
    fta_vat_collected: float = 0.0
    fta_return_ref: Optional[str] = None
    fta_filing_date: Optional[str] = None
    fta_status: str = "draft"
    revenue_difference: float = 0.0
    vat_difference: float = 0.0
    difference_pct: float = 0.0
    reconciling_items: List[str] = Field(default_factory=list)
    status: Literal["matched", "timing_diff", "unexplained", "no_fta_data"] = "no_fta_data"
    risk_flag: bool = False
    auditor_note: str = ""


class FTAVATReconciliationReport(BaseModel):
    rera_registration_number: str
    project_name: str = ""
    developer_name: str = ""
    currency: str = "AED"
    vat_rate: float = VAT_RATE
    fta_filing_frequency: str = "quarterly"
    reconciliation_lines: List[VATReconciliationLine] = Field(default_factory=list)
    total_ifrs15_revenue: float = 0.0
    total_ifrs15_vat: float = 0.0
    total_fta_taxable_supply: float = 0.0
    total_fta_vat_collected: float = 0.0
    total_revenue_difference: float = 0.0
    total_vat_difference: float = 0.0
    quarters_matched: int = 0
    quarters_timing_diff: int = 0
    quarters_unexplained: int = 0
    quarters_no_fta_data: int = 0
    overall_risk: Literal["low", "medium", "high"] = "low"
    overall_risk_reason: str = ""
    reconciliation_summary: str = ""
    generated_at: str = ""


class FTAVATReconciliationRequest(BaseModel):
    rera_registration_number: str
    project_name: str = ""
    developer_name: str = ""
    currency: str = "AED"
    quarterly_schedule: List[Dict[str, Any]] = Field(default_factory=list)
    fta_returns: List[FTAVATReturn] = Field(default_factory=list)


def _parse_iso_date(raw: Optional[str]) -> Optional[date]:
    if not raw:
        return None
    try:
        return datetime.strptime(str(raw).strip()[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _period_bounds_from_end(period_end: str) -> Tuple[str, str]:
    d = _parse_iso_date(period_end)
    if not d:
        return "", str(period_end or "")[:10]
    q = (d.month - 1) // 3 + 1
    start_month = (q - 1) * 3 + 1
    start = date(d.year, start_month, 1)
    return start.isoformat(), d.isoformat()


def _quarter_key(label: str, period_start: Optional[str] = None) -> Optional[str]:
    """Normalize quarter labels to Q{n}{year} for matching."""
    text = (label or "").upper().strip()
    m = re.search(r"Q\s*([1-4])\s*[- ]?\s*(\d{4})", text)
    if m:
        return f"Q{m.group(1)}{m.group(2)}"

    year_m = re.search(r"(\d{4})", text)
    year = int(year_m.group(1)) if year_m else None

    if year and re.search(r"JAN\s*[-–]?\s*MAR", text):
        return f"Q1{year}"
    if year and re.search(r"APR\s*[-–]?\s*JUN", text):
        return f"Q2{year}"
    if year and re.search(r"JUL\s*[-–]?\s*SEP", text):
        return f"Q3{year}"
    if year and re.search(r"OCT\s*[-–]?\s*DEC", text):
        return f"Q4{year}"

    d = _parse_iso_date(period_start)
    if d:
        q = (d.month - 1) // 3 + 1
        return f"Q{q}{d.year}"
    return None


def _quarter_key_from_schedule_item(item: Dict[str, Any]) -> Optional[str]:
    label = str(item.get("quarter") or item.get("period") or "")
    period_start = item.get("period_start")
    if not period_start and item.get("period_end"):
        period_start, _ = _period_bounds_from_end(str(item["period_end"]))
    return _quarter_key(label, str(period_start) if period_start else None)


def _normalize_schedule_item(item: Dict[str, Any]) -> Dict[str, Any]:
    revenue = float(item.get("revenue") or item.get("revenue_recognised") or 0)
    quarter = str(item.get("quarter") or item.get("period") or "")
    period_end = str(item.get("period_end") or "")
    period_start = str(item.get("period_start") or "")
    if not period_start and period_end:
        period_start, period_end = _period_bounds_from_end(period_end)
    if not quarter and period_end:
        d = _parse_iso_date(period_end)
        if d:
            qn = (d.month - 1) // 3 + 1
            quarter = f"Q{qn} {d.year}"
    return {
        "quarter": quarter,
        "period_start": period_start,
        "period_end": period_end,
        "revenue": revenue,
    }


def _find_fta_return(
    schedule_item: Dict[str, Any], fta_returns: List[FTAVATReturn]
) -> Optional[FTAVATReturn]:
    sk = _quarter_key_from_schedule_item(schedule_item)
    if not sk:
        return None
    for fr in fta_returns:
        fk = _quarter_key(fr.quarter, fr.period_start)
        if fk and fk == sk:
            return fr
        # Fallback: period overlap
        s_start = _parse_iso_date(schedule_item.get("period_start"))
        s_end = _parse_iso_date(schedule_item.get("period_end"))
        f_start = _parse_iso_date(fr.period_start)
        f_end = _parse_iso_date(fr.period_end)
        if s_start and s_end and f_start and f_end:
            if s_start <= f_end and f_start <= s_end:
                return fr
    return None


def reconcile_vat_line(
    quarter_schedule_item: Dict[str, Any],
    fta_return: Optional[FTAVATReturn],
) -> VATReconciliationLine:
    norm = _normalize_schedule_item(quarter_schedule_item)
    ifrs15_revenue = round(float(norm["revenue"]), 2)
    ifrs15_vat = round(ifrs15_revenue * VAT_RATE, 2)

    base = VATReconciliationLine(
        quarter=norm["quarter"],
        period_start=norm["period_start"],
        period_end=norm["period_end"],
        ifrs15_revenue_recognised=ifrs15_revenue,
        ifrs15_vat_amount=ifrs15_vat,
    )

    if fta_return is None:
        base.status = "no_fta_data"
        base.auditor_note = (
            "FTA return not yet entered for this period. "
            "File VAT return with FTA and enter Box 1a/1b values."
        )
        base.risk_flag = False
        return base

    reconciling_items: List[str] = []
    fta_taxable = round(float(fta_return.box_1a_taxable_supplies), 2)
    fta_vat = round(float(fta_return.box_1b_vat_on_supplies), 2)

    expected_1b = round(fta_taxable * VAT_RATE, 2)
    if abs(fta_vat - expected_1b) > MATCH_TOLERANCE_AED:
        reconciling_items.append(
            f"WARNING: Box 1b ({fta_vat:,.2f}) does not equal 5% of "
            f"Box 1a ({fta_taxable:,.2f}). Expected: {expected_1b:,.2f}. "
            "Check FTA return for errors."
        )

    revenue_difference = round(fta_taxable - ifrs15_revenue, 2)
    vat_difference = round(fta_vat - ifrs15_vat, 2)
    difference_pct = round(abs(revenue_difference) / max(ifrs15_revenue, 1) * 100, 2)

    base.fta_taxable_supply = fta_taxable
    base.fta_vat_collected = fta_vat
    base.fta_return_ref = fta_return.fta_return_ref
    base.fta_filing_date = fta_return.filing_date
    base.fta_status = fta_return.status
    base.revenue_difference = revenue_difference
    base.vat_difference = vat_difference
    base.difference_pct = difference_pct
    base.reconciling_items = reconciling_items

    if abs(revenue_difference) < MATCH_TOLERANCE_AED:
        base.status = "matched"
        base.auditor_note = "IFRS 15 revenue and FTA taxable supply agree."
        base.risk_flag = False
    elif revenue_difference > 0:
        base.status = "timing_diff"
        reconciling_items.append(
            f"FTA taxable supply exceeds IFRS 15 recognised revenue by "
            f"AED {abs(revenue_difference):,.2f}. This is typical when "
            f"VAT invoices are raised on milestone payments before "
            f"completion-based recognition catches up (deferred revenue)."
        )
        base.auditor_note = (
            "Positive timing difference — VAT invoiced ahead of IFRS 15 "
            "recognition. Verify against deferred revenue balance. "
            "Common in off-plan UAE developments."
        )
        base.risk_flag = difference_pct > TIMING_RISK_PCT
    elif revenue_difference < 0:
        if difference_pct > UNEXPLAINED_PCT:
            base.status = "unexplained"
            reconciling_items.append(
                f"IFRS 15 recognised revenue exceeds FTA taxable supply by "
                f"AED {abs(revenue_difference):,.2f} ({difference_pct:.1f}%). "
                f"Revenue recognised without corresponding VAT invoice. "
                f"Review whether VAT invoices have been issued for all "
                f"recognised completion milestones."
            )
            base.auditor_note = (
                "POTENTIAL COMPLIANCE ISSUE: IFRS 15 revenue exceeds "
                "FTA taxable supply. VAT may be under-reported. "
                "Consult UAE tax advisor. "
                "Ref: Federal Decree-Law No. 8 of 2017, Article 25."
            )
            base.risk_flag = True
        else:
            base.status = "timing_diff"
            base.auditor_note = (
                "Minor negative timing difference within 5% tolerance. "
                "Monitor in subsequent quarters."
            )
            base.risk_flag = False

    base.reconciling_items = reconciling_items
    return base


def build_reconciliation_report(
    quarterly_schedule: List[Dict[str, Any]],
    fta_returns: List[FTAVATReturn],
    project_info: Dict[str, Any],
) -> FTAVATReconciliationReport:
    lines: List[VATReconciliationLine] = []
    for item in quarterly_schedule:
        norm = _normalize_schedule_item(item)
        matched = _find_fta_return(norm, fta_returns)
        lines.append(reconcile_vat_line(norm, matched))

    total_q = len(lines)
    quarters_matched = sum(1 for ln in lines if ln.status == "matched")
    quarters_timing_diff = sum(1 for ln in lines if ln.status == "timing_diff")
    quarters_unexplained = sum(1 for ln in lines if ln.status == "unexplained")
    quarters_no_fta_data = sum(1 for ln in lines if ln.status == "no_fta_data")

    total_ifrs15_revenue = round(sum(ln.ifrs15_revenue_recognised for ln in lines), 2)
    total_ifrs15_vat = round(sum(ln.ifrs15_vat_amount for ln in lines), 2)
    total_fta_taxable = round(
        sum(ln.fta_taxable_supply for ln in lines if ln.status != "no_fta_data"), 2
    )
    total_fta_vat = round(
        sum(ln.fta_vat_collected for ln in lines if ln.status != "no_fta_data"), 2
    )
    total_revenue_difference = round(sum(ln.revenue_difference for ln in lines), 2)
    total_vat_difference = round(sum(ln.vat_difference for ln in lines), 2)

    any_risk = any(ln.risk_flag for ln in lines)
    if any_risk or quarters_unexplained > 0:
        overall_risk: Literal["low", "medium", "high"] = "high"
        overall_risk_reason = (
            "Unexplained or high-risk timing differences require tax advisor review."
            if quarters_unexplained > 0
            else "One or more quarters exceed timing difference risk thresholds."
        )
    elif total_q > 0 and quarters_timing_diff > total_q / 2:
        overall_risk = "medium"
        overall_risk_reason = (
            f"Majority of quarters ({quarters_timing_diff}/{total_q}) show timing differences "
            "between IFRS 15 and FTA taxable supply."
        )
    else:
        overall_risk = "low"
        overall_risk_reason = "No material unexplained differences identified."

    reconciliation_summary = (
        f"{quarters_matched} quarter(s) matched, "
        f"{quarters_timing_diff} timing difference(s), "
        f"{quarters_unexplained} unexplained difference(s), "
        f"{quarters_no_fta_data} awaiting FTA data."
    )

    return FTAVATReconciliationReport(
        rera_registration_number=str(project_info.get("rera_registration_number") or ""),
        project_name=str(project_info.get("project_name") or ""),
        developer_name=str(project_info.get("developer_name") or ""),
        currency=str(project_info.get("currency") or "AED"),
        reconciliation_lines=lines,
        total_ifrs15_revenue=total_ifrs15_revenue,
        total_ifrs15_vat=total_ifrs15_vat,
        total_fta_taxable_supply=total_fta_taxable,
        total_fta_vat_collected=total_fta_vat,
        total_revenue_difference=total_revenue_difference,
        total_vat_difference=total_vat_difference,
        quarters_matched=quarters_matched,
        quarters_timing_diff=quarters_timing_diff,
        quarters_unexplained=quarters_unexplained,
        quarters_no_fta_data=quarters_no_fta_data,
        overall_risk=overall_risk,
        overall_risk_reason=overall_risk_reason,
        reconciliation_summary=reconciliation_summary,
        generated_at=datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
    )


def _status_fill(status: str) -> PatternFill:
    colors_map = {
        "matched": "D1FAE5",
        "timing_diff": "FEF9C3",
        "unexplained": "FEE2E2",
        "no_fta_data": "F3F4F6",
    }
    hex_c = colors_map.get(status, "FFFFFF")
    return PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")


def export_vat_reconciliation_excel(
    report: FTAVATReconciliationReport,
    fta_returns: List[FTAVATReturn],
) -> Tuple[bytes, str]:
    """Build Excel workbook; returns (bytes, filename)."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "VAT Reconciliation"

    headers1 = [
        "Quarter",
        "Period",
        "IFRS 15 Revenue",
        "IFRS 15 VAT",
        "FTA Box 1a",
        "FTA Box 1b",
        "Revenue Diff",
        "VAT Diff",
        "Diff %",
        "Status",
        "Risk",
        "Auditor Note",
    ]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

    for ln in report.reconciliation_lines:
        period = f"{ln.period_start} — {ln.period_end}"
        row_idx = ws1.max_row + 1
        ws1.append(
            [
                ln.quarter,
                period,
                ln.ifrs15_revenue_recognised,
                ln.ifrs15_vat_amount,
                ln.fta_taxable_supply if ln.status != "no_fta_data" else "",
                ln.fta_vat_collected if ln.status != "no_fta_data" else "",
                ln.revenue_difference,
                ln.vat_difference,
                ln.difference_pct,
                ln.status,
                "Yes" if ln.risk_flag else "No",
                ln.auditor_note,
            ]
        )
        fill = _status_fill(ln.status)
        for cell in ws1[row_idx]:
            cell.fill = fill

    totals_row = ws1.max_row + 1
    ws1.append(
        [
            "TOTAL",
            "",
            report.total_ifrs15_revenue,
            report.total_ifrs15_vat,
            report.total_fta_taxable_supply,
            report.total_fta_vat_collected,
            report.total_revenue_difference,
            report.total_vat_difference,
            "",
            "",
            "",
            "",
        ]
    )
    for cell in ws1[totals_row]:
        cell.font = Font(bold=True)

    ws2 = wb.create_sheet("FTA Return Data")
    ws2.append(
        [
            "Note: FTA return data as entered by user. "
            "Verify against official FTA portal records."
        ]
    )
    headers2 = [
        "Quarter",
        "Period Start",
        "Period End",
        "Box 1a",
        "Box 1b",
        "Box 7",
        "Box 8",
        "Return Ref",
        "Filing Date",
        "Status",
    ]
    ws2.append(headers2)
    for fr in fta_returns:
        ws2.append(
            [
                fr.quarter,
                fr.period_start,
                fr.period_end,
                fr.box_1a_taxable_supplies,
                fr.box_1b_vat_on_supplies,
                fr.box_7_input_vat,
                fr.box_8_net_vat_payable,
                fr.fta_return_ref or "",
                fr.filing_date or "",
                fr.status,
            ]
        )

    ws3 = wb.create_sheet("Reconciliation Summary")
    risk_colors = {"low": "DCFCE7", "medium": "FEF9C3", "high": "FEE2E2"}
    ws3["A1"] = "Overall Risk Rating"
    ws3["B1"] = report.overall_risk.upper()
    ws3["B1"].font = Font(bold=True, size=14)
    ws3["B1"].fill = PatternFill(
        start_color=risk_colors.get(report.overall_risk, "FFFFFF"),
        end_color=risk_colors.get(report.overall_risk, "FFFFFF"),
        fill_type="solid",
    )
    ws3["A2"] = report.overall_risk_reason
    ws3["A4"] = "Totals comparison"
    ws3.append(["", "IFRS 15", "FTA"])
    ws3.append(["Revenue / Taxable supply", report.total_ifrs15_revenue, report.total_fta_taxable_supply])
    ws3.append(["VAT", report.total_ifrs15_vat, report.total_fta_vat_collected])
    ws3.append(["Difference", report.total_revenue_difference, report.total_vat_difference])
    ws3.append([])
    ws3.append(["Reconciling items (by quarter)"])
    for ln in report.reconciliation_lines:
        for item in ln.reconciling_items:
            ws3.append([f"{ln.quarter}: {item}"])
    ws3.append([])
    ws3.append(
        [
            "Prepared under Federal Decree-Law No. 8 of 2017 "
            "and IFRS 15 (IASB 2014)"
        ]
    )
    ws3.append([])
    ws3.append(["Prepared by: _________________________"])
    ws3.append(["Reviewed by: _________________________"])
    ws3.append(["Date: _________________________"])

    buf = io.BytesIO()
    wb.save(buf)
    safe_rera = re.sub(r"[^\w\-]", "_", report.rera_registration_number or "RE")[:30]
    date_part = datetime.utcnow().strftime("%Y%m%d")
    filename = f"VAT_Rec_{safe_rera}_{date_part}.xlsx"
    return buf.getvalue(), filename
