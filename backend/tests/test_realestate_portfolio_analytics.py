"""Tests for UAE real estate portfolio analytics."""

import io

from openpyxl import load_workbook

from backend.app.services.realestate_portfolio_analytics import (
    build_portfolio_analytics,
    export_portfolio_analytics_excel,
    filter_projects,
)


def _sample_contracts():
    def c(cid, name, price, pct, escrow, health=None, oqood=0, bundle=False):
        snap = {
            "project_name": name,
            "rera_registration_number": f"RERA-{cid}",
            "developer_name": "Dev Co",
            "completion_pct": pct,
            "health_validation": health or {
                "check_a_pass": True,
                "check_b1_pass": True,
                "check_b2_pass": escrow != "violation",
                "check_c_pass": True,
                "check_d_pass": True,
                "check_e_pass": True,
            },
            "escrow_validation": {"is_violation": escrow == "violation"},
            "pending_oqood_filings": oqood,
            "bundling_alert": bundle,
        }
        return {
            "contract_type": "real_estate_off_plan",
            "contract_id": cid,
            "customer_name": name,
            "total_tp": price,
            "recognised_to_date": price * pct / 100,
            "deferred_balance": price * (100 - pct) / 100 * 0.1,
            "construction_completion_pct": pct,
            "risk": f"RERA-{cid}",
            "realestate_snapshot": snap,
        }

    return [
        c("A", "Project A", 3_500_000, 65, "compliant"),
        c("B", "Project B", 2_000_000, 40, "violation"),
        c("C", "Project C", 5_000_000, 80, "compliant"),
    ]


def test_analytics_three_projects_weighted_completion():
    analytics = build_portfolio_analytics(_sample_contracts())
    assert analytics.total_projects == 3
    assert analytics.total_contract_value_aed == 10_500_000
    assert analytics.escrow_violation_count == 1
    expected = (65 * 3.5 + 40 * 2.0 + 80 * 5.0) / (3.5 + 2.0 + 5.0)
    assert abs(analytics.portfolio_completion_pct - expected) < 0.2
    assert analytics.completion_distribution["26-50"] == 1
    assert analytics.completion_distribution["51-75"] == 1
    assert analytics.completion_distribution["76-100"] == 1


def test_filter_violations_only():
    base = build_portfolio_analytics(_sample_contracts())
    filtered = filter_projects(base, violations_only=True)
    assert filtered.total_projects == 1
    assert filtered.projects[0].contract_id == "B"


def test_excel_export_sheets():
    analytics = build_portfolio_analytics(_sample_contracts())
    data = export_portfolio_analytics_excel(analytics)
    wb = load_workbook(io.BytesIO(data))
    assert "Portfolio Summary" in wb.sheetnames
    assert "Project Detail" in wb.sheetnames
    assert "Compliance Matrix" in wb.sheetnames
