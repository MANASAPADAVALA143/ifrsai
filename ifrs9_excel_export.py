"""
IFRS 9 ECL Excel Export — five-sheet audit pack (aligned with IFRS 15 export styling).
"""

from __future__ import annotations

import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Colour scheme — match IFRS 15 / Prompt 2
DARK_BLUE = "003366"
MED_BLUE = "1F5C99"
GREEN = "1A7A4A"
RED = "CC0000"
ORANGE = "E36C09"
TEAL = "007B8A"
LIGHT_BG = "EEF4FB"
LIGHT_GREEN = "E8F5EE"
LIGHT_RED = "FFF0F0"
LIGHT_ORG = "FFF3E0"
GREY_BG = "F4F4F4"
WHITE = "FFFFFF"

STAGE1_FILL = "E8F5EE"
STAGE2_FILL = "FFF3E0"
STAGE3_FILL = "FFF0F0"
STAGE1_HDR = "1A7A4A"
STAGE2_HDR = "E36C09"
STAGE3_HDR = "CC0000"

OUTPUT_DIR = Path(__file__).resolve().parent / "outputs"


def thin_border() -> Border:
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def title_row(
    ws: Any,
    row: int,
    text: str,
    col_a: str = "A",
    col_b: str = "D",
    *,
    fill_hex: str = DARK_BLUE,
    font_size: int = 14,
    bold: bool = True,
    italic: bool = False,
    white: bool = True,
) -> None:
    ws.merge_cells(f"{col_a}{row}:{col_b}{row}")
    c = ws[f"{col_a}{row}"]
    c.value = text
    c.font = Font(bold=bold, size=font_size, color="FFFFFF" if white else "000000", italic=italic)
    c.fill = _fill(fill_hex)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def hdr_row(ws: Any, row: int, labels: List[str], start_col: int = 1, *, fill_hex: str = MED_BLUE) -> None:
    f = Font(bold=True, color="FFFFFF", size=11)
    fi = _fill(fill_hex)
    for i, lab in enumerate(labels):
        col = get_column_letter(start_col + i)
        cell = ws[f"{col}{row}"]
        cell.value = lab
        cell.font = f
        cell.fill = fi
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()


def data_cell(
    ws: Any,
    row: int,
    col: int,
    value: Any,
    *,
    fill_hex: Optional[str] = None,
    font_color: Optional[str] = None,
    bold: bool = False,
    number_format: Optional[str] = None,
    align: str = "left",
) -> None:
    letter = get_column_letter(col)
    cell = ws[f"{letter}{row}"]
    cell.value = value
    if fill_hex:
        cell.fill = _fill(fill_hex)
    if font_color:
        cell.font = Font(bold=bold, color=font_color, size=11)
    else:
        cell.font = Font(bold=bold, size=11)
    if number_format:
        cell.number_format = number_format
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = thin_border()


def spacer(ws: Any, row: int) -> None:
    ws.row_dimensions[row].height = 6


def section_hdr(ws: Any, row: int, text: str, col_a: str = "A", col_b: str = "D", *, fill_hex: str = ORANGE) -> None:
    ws.merge_cells(f"{col_a}{row}:{col_b}{row}")
    c = ws[f"{col_a}{row}"]
    c.value = text
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = _fill(fill_hex)
    c.alignment = Alignment(horizontal="left", vertical="center")


def _money_fmt() -> str:
    return '"$"#,##0.00'


def _pct_fmt() -> str:
    return "0.00%"


def _normalize_stage_summary(data: Dict[str, Any], loans: List[Dict[str, Any]]) -> Dict[str, Dict[str, float]]:
    ss = data.get("stage_summary")
    if isinstance(ss, dict) and "stage1" in ss:
        def one(k: str) -> Dict[str, float]:
            b = ss.get(k) or {}
            return {
                "count": float(b.get("count", 0) or 0),
                "ead": float(b.get("ead", 0) or 0),
                "ecl": float(b.get("ecl", 0) or 0),
            }

        return {"stage1": one("stage1"), "stage2": one("stage2"), "stage3": one("stage3")}

    if isinstance(ss, list):
        out: Dict[str, Dict[str, float]] = {}
        for idx, stn in enumerate([1, 2, 3]):
            block = ss[idx] if idx < len(ss) else {}
            out[f"stage{stn}"] = {
                "count": float(block.get("loan_count", block.get("count", 0)) or 0),
                "ead": float(block.get("ead", 0) or 0),
                "ecl": float(block.get("ecl", 0) or 0),
            }
        return out

    out = {"stage1": {"count": 0, "ead": 0, "ecl": 0}, "stage2": {"count": 0, "ead": 0, "ecl": 0}, "stage3": {"count": 0, "ead": 0, "ecl": 0}}
    for st in (1, 2, 3):
        sl = [x for x in loans if int(x.get("stage", 1) or 1) == st]
        key = f"stage{st}"
        out[key]["count"] = float(len(sl))
        out[key]["ead"] = sum(float(x.get("ead", 0) or 0) for x in sl)
        out[key]["ecl"] = sum(float(x.get("ecl", 0) or 0) for x in sl)
    return out


def _write_ifrs9_master_summary_sheet(ws: Any, master: Dict[str, Any]) -> None:
    """Sheet 0 — IFRS 9 Master Summary (matches IFRS 15 master sheet layout pattern)."""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = DARK_BLUE
    for col, w in ("A", 40), ("B", 22), ("C", 22), ("D", 36):
        ws.column_dimensions[col].width = w

    pn = master.get("portfolio_name") or "Portfolio"
    gen = master.get("generated_at") or ""
    rd = master.get("reporting_date") or ""
    en = master.get("entity_name") or ""
    title_row(ws, 1, "IFRS 9 — MASTER COMPLIANCE REPORT", "A", "D", fill_hex=DARK_BLUE, font_size=14)
    sub = f"{pn}"
    if en:
        sub += f" | {en}"
    sub += f" | Reporting: {rd} | Generated: {gen}"
    title_row(ws, 2, sub, "A", "D", fill_hex=MED_BLUE, font_size=10, italic=True)
    spacer(ws, 3)
    r = 4

    po = master.get("portfolio_overview") or {}
    section_hdr(ws, r, "PORTFOLIO OVERVIEW", "A", "D")
    r += 1
    for lab, val in [
        ("Portfolio", po.get("portfolio_name", "—")),
        ("Entity", po.get("entity_name") or "—"),
        ("Reporting date", po.get("reporting_date", "—")),
        ("Total EAD", po.get("total_exposure_ead", 0)),
        ("Total ECL provision", po.get("total_ecl_provision", 0)),
        ("Coverage ratio", (float(po.get("coverage_ratio", 0) or 0)) / 100.0),
        ("Weighted avg PD", (float(po.get("weighted_avg_pd", 0) or 0)) / 100.0 if float(po.get("weighted_avg_pd", 0) or 0) > 1 else float(po.get("weighted_avg_pd", 0) or 0)),
        ("Loan count", int(po.get("loan_count", 0) or 0)),
        ("Currency", po.get("currency", "USD")),
    ]:
        is_money = lab in ("Total EAD", "Total ECL provision")
        is_pct = lab == "Coverage ratio"
        is_pd = lab == "Weighted avg PD"
        data_cell(ws, r, 1, lab, fill_hex=GREY_BG, align="left")
        if is_money:
            data_cell(ws, r, 2, val, fill_hex=WHITE, number_format=_money_fmt(), align="right")
        elif is_pct:
            data_cell(ws, r, 2, val, fill_hex=WHITE, number_format=_pct_fmt(), align="right")
        elif is_pd:
            v = float(val or 0)
            cell_v = v / 100.0 if v > 1 else v
            data_cell(ws, r, 2, cell_v, fill_hex=WHITE, number_format=_pct_fmt(), align="right")
        else:
            data_cell(ws, r, 2, val, fill_hex=WHITE, align="left")
        r += 1
    r += 1

    es = master.get("ecl_summary") or {}
    section_hdr(ws, r, "ECL BY STAGE", "A", "D")
    r += 1
    hdr_row(ws, r, ["Stage", "Count", "EAD ($)", "ECL ($)"], 1)
    r += 1
    for key, lab, fill in [
        ("stage1", "Stage 1 (12m)", STAGE1_FILL),
        ("stage2", "Stage 2 (Lifetime)", STAGE2_FILL),
        ("stage3", "Stage 3 (Lifetime)", STAGE3_FILL),
    ]:
        b = es.get(key) or {}
        data_cell(ws, r, 1, lab, fill_hex=fill, bold=True)
        data_cell(ws, r, 2, int(b.get("count", 0) or 0), fill_hex=fill, align="right")
        data_cell(ws, r, 3, float(b.get("ead", 0) or 0), fill_hex=fill, number_format=_money_fmt(), align="right")
        data_cell(ws, r, 4, float(b.get("ecl", 0) or 0), fill_hex=fill, number_format=_money_fmt(), align="right")
        r += 1
    data_cell(ws, r, 1, "Total ECL", fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True)
    data_cell(ws, r, 2, "✓" if es.get("all_steps_complete") else "—", fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True, align="center")
    data_cell(ws, r, 3, "", fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True)
    data_cell(ws, r, 4, float(es.get("total_ecl", 0) or 0), fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True, number_format=_money_fmt(), align="right")
    r += 2

    asm = master.get("assessments") or {}
    section_hdr(ws, r, "ASSESSMENTS", "A", "D")
    r += 1
    hdr_row(ws, r, ["Module", "Assessed", "Key figures"], 1)
    r += 1
    ecl_a = asm.get("ecl_staging") or {}
    cl_a = asm.get("classification") or {}
    mo_a = asm.get("macro_overlay") or {}
    pm_a = asm.get("provision_matrix") or {}
    pwe = mo_a.get("probability_weighted_ecl")
    pwe_s = f"${float(pwe):,.0f}" if pwe is not None else "—"
    tpv = pm_a.get("total_provision")
    tpv_s = f"${float(tpv):,.0f}" if tpv is not None else "—"
    rows_asm = [
        ("ECL Staging", ecl_a.get("assessed"), f"ECL {ecl_a.get('total_ecl', 0):,.0f} | {ecl_a.get('method', '')}"),
        ("Classification", cl_a.get("assessed"), str(cl_a.get("measurement", "—"))),
        ("Macro overlay", mo_a.get("assessed"), f"PWE {pwe_s}"),
        ("Provision matrix", pm_a.get("assessed"), f"Provision {tpv_s}"),
    ]
    alt = False
    for name, ok, note in rows_asm:
        bg = LIGHT_BG if alt else WHITE
        alt = not alt
        data_cell(ws, r, 1, name, fill_hex=bg)
        data_cell(ws, r, 2, "Yes" if ok else "No", fill_hex=bg, align="center")
        data_cell(ws, r, 3, note, fill_hex=bg, align="left")
        r += 1
    r += 1

    section_hdr(ws, r, "RISK FLAGS", "A", "D")
    r += 1
    risks = list(master.get("risk_flags") or [])
    if not risks:
        ws.merge_cells(f"A{r}:D{r}")
        c = ws[f"A{r}"]
        c.value = "✓ No risks identified"
        c.font = Font(bold=True, color="006100", size=11)
        c.fill = _fill(LIGHT_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center")
        r += 1
    else:
        hdr_row(ws, r, ["Severity", "Module", "Message", "Action required"], 1)
        r += 1
        for rf in risks:
            sev = str(rf.get("severity", "LOW"))
            fill_hex = LIGHT_RED if sev == "HIGH" else LIGHT_ORG if sev == "MEDIUM" else "FFF2CC"
            data_cell(ws, r, 1, sev, fill_hex=fill_hex, bold=True)
            data_cell(ws, r, 2, rf.get("module", ""), fill_hex=fill_hex)
            data_cell(ws, r, 3, rf.get("message", ""), fill_hex=fill_hex)
            data_cell(ws, r, 4, rf.get("action_required", ""), fill_hex=fill_hex)
            r += 1
    r += 1

    ar = master.get("audit_readiness") or {}
    section_hdr(ws, r, "AUDIT READINESS", "A", "D")
    r += 1
    score = float(ar.get("score", 0) or 0)
    lvl = str(ar.get("level", ""))
    sc_fill = LIGHT_GREEN if lvl == "Ready" else LIGHT_ORG if lvl == "Needs Review" else LIGHT_RED
    ws.merge_cells(f"A{r}:D{r}")
    c = ws[f"A{r}"]
    c.value = f"Score: {score:.1f}% — {lvl}"
    c.font = Font(bold=True, size=14)
    c.fill = _fill(sc_fill)
    c.alignment = Alignment(horizontal="center", vertical="center")
    r += 1
    ws[f"A{r}"] = "Item"
    ws[f"B{r}"] = "Status"
    for col in ("A", "B"):
        ws[f"{col}{r}"].font = Font(bold=True, color="FFFFFF", size=10)
        ws[f"{col}{r}"].fill = _fill(MED_BLUE)
        ws[f"{col}{r}"].border = thin_border()
    r += 1
    for it in ar.get("checklist") or []:
        st = str(it.get("status", ""))
        colf = Font(color="006100") if st == "complete" else Font(color="C00000")
        data_cell(ws, r, 1, it.get("item", ""), align="left")
        ws[f"A{r}"].font = colf
        data_cell(ws, r, 2, st, align="left")
        ws[f"B{r}"].font = colf
        r += 1
    r += 1

    section_hdr(ws, r, "AI NARRATIVE", "A", "D")
    r += 1
    ws.merge_cells(f"A{r}:D{r + 18}")
    nar = ws[f"A{r}"]
    nar.value = master.get("ai_narrative") or "—"
    nar.font = Font(name="Times New Roman", size=11)
    nar.alignment = Alignment(wrap_text=True, vertical="top")
    nar.border = thin_border()


def _loan_rows(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    raw = data.get("loans") or []
    out: List[Dict[str, Any]] = []
    for i, l in enumerate(raw):
        if not isinstance(l, dict):
            continue
        lid = l.get("loan_id") or l.get("id") or f"LOAN-{i + 1}"
        pdv = float(l.get("pd") if l.get("pd") is not None else l.get("pd_pct", 0) or 0)
        if pdv > 1:
            pd_frac = pdv / 100.0
        elif pdv == 1.0:
            pd_frac = 0.01
        else:
            pd_frac = pdv
        lgv = float(l.get("lgd") if l.get("lgd") is not None else l.get("lgd_pct", 0) or 0)
        lgd_frac = lgv / 100.0 if lgv > 1 else lgv
        out.append(
            {
                "loan_id": lid,
                "stage": int(l.get("stage", 1) or 1),
                "ead": float(l.get("ead", 0) or 0),
                "pd_frac": pd_frac,
                "lgd_frac": lgd_frac,
                "ecl": float(l.get("ecl", 0) or 0),
                "days_past_due": int(l.get("days_past_due", 0) or 0),
                "sicr": bool(l.get("sicr", False)),
                "status": str(l.get("status", "Performing")),
            }
        )
    return out


def _stage_fill_alt(stage: int, alt: bool) -> str:
    base = STAGE1_FILL if stage == 1 else STAGE2_FILL if stage == 2 else STAGE3_FILL
    if not alt:
        return base
    return "F7FAFC" if stage == 1 else "FFFAF0" if stage == 2 else "FFF5F5"


def export_ifrs9_excel(data: Dict[str, Any], master_report: Optional[Dict[str, Any]] = None) -> str:
    """
    Build IFRS 9 ECL workbook (five audit sheets; optional sixth sheet — Master Summary — as sheet 0).
    Returns 8-char file_id (filename suffix).
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    portfolio = str(data.get("portfolio_name") or "Portfolio")
    entity = str(data.get("entity_name") or "")
    report_date = str(data.get("reporting_date") or datetime.now().strftime("%Y-%m-%d"))
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")

    applicable_ecl = float(data.get("applicable_ecl", 0) or 0)
    ecl_12m = float(data.get("ecl_12m", 0) or 0)
    total_ead = float(data.get("total_ead", 0) or 0)
    pd_used = float(data.get("pd_used", 0) or data.get("weighted_avg_pd", 0) or 0)
    lgd_used = float(data.get("lgd_used", 0) or 0)
    coverage_ratio = float(data.get("coverage_ratio", 0) or 0)

    loans = _loan_rows(data)
    stage_summary = _normalize_stage_summary(data, loans)

    s1e = stage_summary["stage1"]["ecl"]
    s2e = stage_summary["stage2"]["ecl"]
    s3e = stage_summary["stage3"]["ecl"]
    if ecl_12m <= 0 and s1e > 0:
        ecl_12m = s1e

    wb = Workbook()
    # Remove default sheet; rebuild with optional Master Summary first (same order as IFRS 15 pack).
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    insert_at = 0
    if master_report is not None:
        ws0 = wb.create_sheet("IFRS 9 Master Summary", insert_at)
        _write_ifrs9_master_summary_sheet(ws0, master_report)
        insert_at += 1

    # ─── Sheet 1 Summary ───
    ws1 = wb.create_sheet("Summary", insert_at)
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = DARK_BLUE
    for col, w in ("A", 34), ("B", 24), ("C", 20), ("D", 20):
        ws1.column_dimensions[col].width = w

    title_row(ws1, 1, "IFRS 9 — EXPECTED CREDIT LOSS", "A", "D", fill_hex=DARK_BLUE, font_size=14)
    subtitle = f"ECL Provision Report | {portfolio}"
    if entity:
        subtitle += f" | {entity}"
    subtitle += f" | Reporting Date: {report_date}"
    title_row(ws1, 2, subtitle, "A", "D", fill_hex=MED_BLUE, font_size=11, italic=True)
    spacer(ws1, 3)
    r = 4
    section_hdr(ws1, r, "KEY METRICS", "A", "D", fill_hex=ORANGE)
    r += 1

    # Payload uses percentage points (e.g. 2.5 = 2.5%) for PD/LGD/coverage from UI/API
    pd_cell = pd_used / 100.0
    lgd_cell = lgd_used / 100.0
    cov_cell = coverage_ratio / 100.0

    metrics: List[Tuple[str, Any, Optional[str], Optional[str]]] = [
        ("Portfolio Name", portfolio, None, None),
        ("Entity", entity or "—", None, None),
        ("Reporting Date", report_date, None, None),
        ("Total Exposure (EAD)", total_ead, _money_fmt(), None),
        ("Total ECL Provision", applicable_ecl, _money_fmt(), RED),
        ("Stage 1 ECL (12-month)", ecl_12m if ecl_12m > 0 else s1e, _money_fmt(), GREEN),
        ("Stage 2 ECL (Lifetime)", s2e, _money_fmt(), ORANGE),
        ("Stage 3 ECL (Lifetime)", s3e, _money_fmt(), RED),
        ("Weighted Average PD", pd_cell, _pct_fmt(), None),
        ("Weighted Average LGD", lgd_cell, _pct_fmt(), None),
        ("Coverage Ratio", cov_cell, _pct_fmt(), None),
        ("Report Generated", generated, None, None),
    ]

    alt = False
    for label, val, fmt, color in metrics:
        bg = GREY_BG if alt else WHITE
        alt = not alt
        data_cell(ws1, r, 1, label, fill_hex=bg, align="left")
        data_cell(ws1, r, 2, val, fill_hex=bg, font_color=color, number_format=fmt, align="right")
        r += 1

    r += 1
    section_hdr(ws1, r, "STAGE DISTRIBUTION", "A", "D", fill_hex=ORANGE)
    r += 1
    for col, (lab, hcol) in enumerate(
        zip(["Stage 1", "Stage 2", "Stage 3"], [STAGE1_HDR, STAGE2_HDR, STAGE3_HDR]),
        start=1,
    ):
        cell = ws1.cell(row=r, column=col)
        cell.value = lab
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = _fill(hcol)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    r += 1
    c1, c2, c3 = stage_summary["stage1"]["count"], stage_summary["stage2"]["count"], stage_summary["stage3"]["count"]
    data_cell(ws1, r, 1, f"{int(c1)} loans", fill_hex=STAGE1_FILL, align="center")
    data_cell(ws1, r, 2, f"{int(c2)} loans", fill_hex=STAGE2_FILL, align="center")
    data_cell(ws1, r, 3, f"{int(c3)} loans", fill_hex=STAGE3_FILL, align="center")
    r += 1
    e1, e2, e3 = stage_summary["stage1"]["ead"], stage_summary["stage2"]["ead"], stage_summary["stage3"]["ead"]
    data_cell(ws1, r, 1, e1, fill_hex=STAGE1_FILL, number_format=_money_fmt(), align="right")
    data_cell(ws1, r, 2, e2, fill_hex=STAGE2_FILL, number_format=_money_fmt(), align="right")
    data_cell(ws1, r, 3, e3, fill_hex=STAGE3_FILL, number_format=_money_fmt(), align="right")
    r += 1
    z1, z2, z3 = stage_summary["stage1"]["ecl"], stage_summary["stage2"]["ecl"], stage_summary["stage3"]["ecl"]
    data_cell(ws1, r, 1, z1, fill_hex=STAGE1_FILL, number_format=_money_fmt(), align="right")
    data_cell(ws1, r, 2, z2, fill_hex=STAGE2_FILL, number_format=_money_fmt(), align="right")
    data_cell(ws1, r, 3, z3, fill_hex=STAGE3_FILL, number_format=_money_fmt(), align="right")

    # ─── Sheet 2 Loan Staging ───
    ws2 = wb.create_sheet("Loan Staging Detail")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = STAGE1_HDR
    widths = [18, 12, 18, 12, 12, 18, 20, 22, 20]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    title_row(ws2, 1, f"LOAN STAGING DETAIL — {portfolio}", "A", "I", fill_hex=DARK_BLUE)
    title_row(
        ws2,
        2,
        "Stage 1: Performing (12-month ECL) | Stage 2: SICR (Lifetime ECL) | Stage 3: Impaired (Lifetime ECL)",
        "A",
        "I",
        fill_hex=MED_BLUE,
        font_size=10,
    )
    hdr_row(ws2, 3, ["Loan ID", "Stage", "EAD ($)", "PD (%)", "LGD (%)", "ECL ($)", "Days Past Due", "SICR Triggered", "Status"], 1)
    ws2.freeze_panes = "A4"

    if not loans:
        ws2.merge_cells("A4:I8")
        c = ws2["A4"]
        c.value = "No loan data available — upload portfolio and run calculation"
        c.font = Font(italic=True, size=12, color="64748B")
        c.fill = _fill(GREY_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        row = 4
        prev_stage = 0
        alt_idx = 0
        sum_ead = 0.0
        sum_ecl = 0.0
        for i, ln in enumerate(loans):
            st = int(ln["stage"])
            if st != prev_stage:
                alt_idx = 0
                prev_stage = st
            fill = _stage_fill_alt(st, alt_idx % 2 == 1)
            alt_idx += 1
            data_cell(ws2, row, 1, ln["loan_id"], fill_hex=fill, align="left")
            data_cell(ws2, row, 2, f"Stage {st}", fill_hex=fill, align="center")
            data_cell(ws2, row, 3, ln["ead"], fill_hex=fill, number_format=_money_fmt(), align="right")
            data_cell(ws2, row, 4, ln["pd_frac"], fill_hex=fill, number_format=_pct_fmt(), align="right")
            data_cell(ws2, row, 5, ln["lgd_frac"], fill_hex=fill, number_format=_pct_fmt(), align="right")
            data_cell(ws2, row, 6, ln["ecl"], fill_hex=fill, number_format=_money_fmt(), align="right")
            data_cell(ws2, row, 7, ln["days_past_due"], fill_hex=fill, align="right")
            data_cell(ws2, row, 8, "Yes" if ln["sicr"] else "No", fill_hex=fill, align="center")
            data_cell(ws2, row, 9, ln["status"], fill_hex=fill, align="left")
            sum_ead += ln["ead"]
            sum_ecl += ln["ecl"]
            row += 1
        data_cell(ws2, row, 1, "TOTAL", fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True, align="left")
        for c in range(2, 9):
            if c == 3:
                data_cell(ws2, row, c, sum_ead, fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True, number_format=_money_fmt(), align="right")
            elif c == 6:
                data_cell(ws2, row, c, sum_ecl, fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True, number_format=_money_fmt(), align="right")
            else:
                data_cell(ws2, row, c, "", fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True, align="center")

    # ─── Sheet 3 Movement ───
    ws3 = wb.create_sheet("ECL Movement")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = ORANGE
    for i, w in enumerate([24, 18, 18, 18, 18, 18, 18], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    title_row(ws3, 1, "ECL MOVEMENT ANALYSIS", "A", "G", fill_hex=DARK_BLUE)
    movement = data.get("ecl_movement") or {}
    r = 3
    if isinstance(movement, dict) and movement.get("stage1"):
        hdr_row(
            ws3,
            r,
            ["Stage", "Opening ECL", "New Loans", "Stage Transfers", "Write-offs", "Remeasurement", "Closing ECL"],
            1,
        )
        r += 1
        for st, fill in [("stage1", STAGE1_FILL), ("stage2", STAGE2_FILL), ("stage3", STAGE3_FILL)]:
            b = movement.get(st) or {}
            lbl = st.replace("stage", "Stage ")
            data_cell(ws3, r, 1, lbl, fill_hex=fill, bold=True)
            data_cell(ws3, r, 2, float(b.get("opening_ecl", 0) or 0), fill_hex=fill, number_format=_money_fmt(), align="right")
            data_cell(ws3, r, 3, float(b.get("new_loans", 0) or 0), fill_hex=fill, number_format=_money_fmt(), align="right")
            data_cell(ws3, r, 4, float(b.get("stage_transfers", 0) or 0), fill_hex=fill, number_format=_money_fmt(), align="right")
            data_cell(ws3, r, 5, float(b.get("write_offs", 0) or 0), fill_hex=fill, number_format=_money_fmt(), align="right")
            data_cell(ws3, r, 6, float(b.get("remeasurement", 0) or 0), fill_hex=fill, number_format=_money_fmt(), align="right")
            data_cell(ws3, r, 7, float(b.get("closing_ecl", 0) or 0), fill_hex=fill, number_format=_money_fmt(), align="right")
            r += 1
    else:
        ws3.merge_cells("A3:G10")
        c = ws3["A3"]
        c.value = (
            "ECL Movement analysis requires prior period data.\n\n"
            "Populate the following inputs to generate the movement schedule:\n"
            "• Opening ECL per stage\n"
            "• New loan originations\n"
            "• Stage transfer amounts\n"
            "• Write-off amounts\n"
            "• Remeasurement adjustments"
        )
        c.font = Font(italic=True, size=11, color="444444")
        c.fill = _fill(GREY_BG)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        r = 12

    r = max(r, 12)
    section_hdr(ws3, r, "SICR DETECTION CRITERIA", "A", "G", fill_hex=ORANGE)
    r += 1
    hdr_row(ws3, r, ["Indicator", "Threshold", "Description"], 1)
    r += 1
    sicr_rows = [
        ("Days Past Due", "> 30 days", "Backstop indicator"),
        ("Rating Downgrade", "> 2 notches", "Internal rating"),
        ("Watchlist", "Any", "Management overlay"),
        ("Macro Deterioration", "GDP < -1%", "Forward-looking"),
        ("Forbearance", "Any", "Restructuring flag"),
    ]
    alt = False
    for a, b, d in sicr_rows:
        bg = LIGHT_BG if alt else WHITE
        alt = not alt
        data_cell(ws3, r, 1, a, fill_hex=bg)
        data_cell(ws3, r, 2, b, fill_hex=bg)
        data_cell(ws3, r, 3, d, fill_hex=bg)
        r += 1

    # ─── Sheet 4 Journals ───
    ws4 = wb.create_sheet("Journal Entries")
    ws4.sheet_view.showGridLines = False
    ws4.sheet_properties.tabColor = MED_BLUE
    for i, w in enumerate([12, 28, 30, 18, 18], start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    title_row(ws4, 1, "IFRS 9 ECL JOURNAL ENTRIES", "A", "E", fill_hex=DARK_BLUE)
    r = 3
    hdr_row(ws4, r, ["Date", "Account", "Description", "Dr ($)", "Cr ($)"], 1)
    r += 1

    jrows: List[Dict[str, Any]] = list(data.get("journal_entries") or [])
    total_dr = 0.0
    if not jrows:
        blocks = []
        if ecl_12m > 0 or s1e > 0:
            amt = ecl_12m if ecl_12m > 0 else s1e
            blocks.append(("Initial Recognition — Stage 1", amt, "12-month ECL provision — Stage 1 performing loans"))
        if s2e > 0:
            blocks.append(("Stage 2 — SICR", s2e, "Lifetime ECL — SICR detected (Stage 2 transfer)"))
        if s3e > 0:
            blocks.append(("Stage 3 — Impaired", s3e, "Lifetime ECL — credit impaired (Stage 3)"))
        for _title, amt, desc in blocks:
            total_dr += amt
            data_cell(ws4, r, 1, report_date, fill_hex=LIGHT_BG, align="left")
            data_cell(ws4, r, 2, "ECL Expense (P&L)", fill_hex=LIGHT_BG, font_color=MED_BLUE, align="left")
            data_cell(ws4, r, 3, desc, fill_hex=LIGHT_BG, align="left")
            data_cell(ws4, r, 4, amt, fill_hex=LIGHT_BG, number_format=_money_fmt(), align="right")
            data_cell(ws4, r, 5, "", fill_hex=LIGHT_BG, align="right")
            r += 1
            data_cell(ws4, r, 1, report_date, fill_hex=LIGHT_GREEN, align="left")
            data_cell(ws4, r, 2, "    Loan Loss Allowance", fill_hex=LIGHT_GREEN, font_color=GREEN, align="left")
            data_cell(ws4, r, 3, desc, fill_hex=LIGHT_GREEN, align="left")
            data_cell(ws4, r, 4, "", fill_hex=LIGHT_GREEN, align="right")
            data_cell(ws4, r, 5, amt, fill_hex=LIGHT_GREEN, number_format=_money_fmt(), align="right")
            r += 1
        if applicable_ecl > 0 and not blocks:
            amt = applicable_ecl
            data_cell(ws4, r, 1, report_date, fill_hex=LIGHT_BG, align="left")
            data_cell(ws4, r, 2, "ECL Expense (P&L)", fill_hex=LIGHT_BG, font_color=MED_BLUE, align="left")
            data_cell(ws4, r, 3, f"Total ECL provision — {report_date}", fill_hex=LIGHT_BG, align="left")
            data_cell(ws4, r, 4, amt, fill_hex=LIGHT_BG, number_format=_money_fmt(), align="right")
            data_cell(ws4, r, 5, "", fill_hex=LIGHT_BG, align="right")
            r += 1
            data_cell(ws4, r, 1, report_date, fill_hex=LIGHT_GREEN, align="left")
            data_cell(ws4, r, 2, "    Loan Loss Allowance", fill_hex=LIGHT_GREEN, font_color=GREEN, align="left")
            data_cell(ws4, r, 3, f"Total ECL provision — {report_date}", fill_hex=LIGHT_GREEN, align="left")
            data_cell(ws4, r, 4, "", fill_hex=LIGHT_GREEN, align="right")
            data_cell(ws4, r, 5, amt, fill_hex=LIGHT_GREEN, number_format=_money_fmt(), align="right")
            r += 1
            total_dr = amt
    else:
        for je in jrows:
            amt = float(je.get("amount", 0) or 0)
            dr_ac = str(je.get("dr", "ECL Expense (P&L)"))
            cr_ac = str(je.get("cr", "Loan Loss Allowance"))
            desc = str(je.get("type", je.get("description", "ECL recognition")))
            total_dr += amt
            data_cell(ws4, r, 1, report_date, fill_hex=LIGHT_BG, align="left")
            data_cell(ws4, r, 2, dr_ac, fill_hex=LIGHT_BG, font_color=MED_BLUE, align="left")
            data_cell(ws4, r, 3, desc, fill_hex=LIGHT_BG, align="left")
            data_cell(ws4, r, 4, amt, fill_hex=LIGHT_BG, number_format=_money_fmt(), align="right")
            data_cell(ws4, r, 5, "", fill_hex=LIGHT_BG, align="right")
            r += 1
            data_cell(ws4, r, 1, report_date, fill_hex=LIGHT_GREEN, align="left")
            data_cell(ws4, r, 2, f"    {cr_ac}", fill_hex=LIGHT_GREEN, font_color=GREEN, align="left")
            data_cell(ws4, r, 3, desc, fill_hex=LIGHT_GREEN, align="left")
            data_cell(ws4, r, 4, "", fill_hex=LIGHT_GREEN, align="right")
            data_cell(ws4, r, 5, amt, fill_hex=LIGHT_GREEN, number_format=_money_fmt(), align="right")
            r += 1

    section_hdr(ws4, r, "Balance Sheet Presentation", "A", "E", fill_hex=MED_BLUE)
    r += 1
    ws4.merge_cells(f"A{r}:E{r + 3}")
    note = ws4[f"A{r}"]
    net_loans = max(total_ead - applicable_ecl, 0)
    note.value = (
        "Loan Loss Allowance is presented as a deduction from Gross Loans on the Balance Sheet per IFRS 9.5.5.1.\n\n"
        f"Net Loans = Gross Loans − Allowance\n"
        f"Net Loans = ${net_loans:,.2f}"
    )
    note.font = Font(size=10, color="333333")
    note.fill = _fill(GREY_BG)
    note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    r += 5
    te = applicable_ecl if applicable_ecl > 0 else total_dr
    data_cell(ws4, r, 1, "TOTAL ECL CHARGE", fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True)
    data_cell(ws4, r, 2, "", fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True)
    data_cell(ws4, r, 3, "", fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True)
    data_cell(ws4, r, 4, te, fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True, number_format=_money_fmt(), align="right")
    data_cell(ws4, r, 5, te, fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True, number_format=_money_fmt(), align="right")

    # ─── Sheet 5 Disclosure ───
    ws5 = wb.create_sheet("Disclosure Notes")
    ws5.sheet_view.showGridLines = False
    ws5.sheet_properties.tabColor = TEAL
    ws5.column_dimensions["A"].width = 20
    ws5.column_dimensions["B"].width = 80
    title_row(ws5, 1, "IFRS 9 DISCLOSURE NOTES", "A", "B", fill_hex=DARK_BLUE)
    title_row(ws5, 2, "Note [X] — Expected Credit Losses on Financial Assets", "A", "B", fill_hex=MED_BLUE, font_size=11)
    r = 4
    section_hdr(ws5, r, "DISCLOSURE 1 — ACCOUNTING POLICY", "A", "B", fill_hex=DARK_BLUE)
    r += 1
    dn = data.get("disclosure_notes")
    policy = ""
    if isinstance(dn, str):
        policy = dn
    elif isinstance(dn, dict):
        policy = str(dn.get("accounting_policy", "") or "")
    if not policy.strip():
        policy = (
            "The Group assesses on a forward-looking basis the ECL associated with its financial assets carried at amortised cost. "
            "The Group applies the three-stage model for recognising and measuring ECL:\n\n"
            "Stage 1 — Performing: 12-month ECL is recognised for financial instruments that have not had a significant increase in credit risk since origination.\n\n"
            "Stage 2 — Underperforming: Lifetime ECL is recognised when a significant increase in credit risk (SICR) has occurred since origination but the instrument is not yet credit-impaired.\n\n"
            "Stage 3 — Non-performing: Lifetime ECL is recognised for credit-impaired financial instruments."
        )
    ws5.merge_cells(f"B{r}:B{r + 8}")
    cell = ws5[f"B{r}"]
    cell.value = policy
    cell.font = Font(name="Times New Roman", size=11)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    r += 10

    section_hdr(ws5, r, "DISCLOSURE 2 — QUANTITATIVE ECL TABLE", "A", "B", fill_hex=ORANGE)
    r += 1
    ws5[f"A{r}"].value = f"ECL by Stage — {report_date}"
    ws5[f"A{r}"].font = Font(bold=True, size=11)
    r += 1
    hdr_row(ws5, r, ["Stage", "# Loans", "Gross EAD ($)", "ECL Allowance ($)", "Coverage %"], 1)
    r += 1
    tot_n = tot_ead = tot_ecl = 0.0
    for st in (1, 2, 3):
        key = f"stage{st}"
        sn = int(stage_summary[key]["count"])
        se = stage_summary[key]["ead"]
        sz = stage_summary[key]["ecl"]
        cov = (sz / se * 100) if se else 0.0
        fill = STAGE1_FILL if st == 1 else STAGE2_FILL if st == 2 else STAGE3_FILL
        data_cell(ws5, r, 1, f"Stage {st}", fill_hex=fill, bold=True)
        data_cell(ws5, r, 2, sn, fill_hex=fill, align="right")
        data_cell(ws5, r, 3, se, fill_hex=fill, number_format=_money_fmt(), align="right")
        data_cell(ws5, r, 4, sz, fill_hex=fill, number_format=_money_fmt(), align="right")
        data_cell(ws5, r, 5, cov / 100.0, fill_hex=fill, number_format=_pct_fmt(), align="right")
        tot_n += sn
        tot_ead += se
        tot_ecl += sz
        r += 1
    cov_t = (tot_ecl / tot_ead * 100) if tot_ead else 0.0
    data_cell(ws5, r, 1, "TOTAL", fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True)
    data_cell(ws5, r, 2, tot_n, fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True, align="right")
    data_cell(ws5, r, 3, tot_ead, fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True, number_format=_money_fmt(), align="right")
    data_cell(ws5, r, 4, tot_ecl, fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True, number_format=_money_fmt(), align="right")
    data_cell(ws5, r, 5, cov_t / 100.0, fill_hex=DARK_BLUE, font_color="FFFFFF", bold=True, number_format=_pct_fmt(), align="right")
    r += 2

    section_hdr(ws5, r, "DISCLOSURE 3 — SIGNIFICANT ASSUMPTIONS", "A", "B", fill_hex=DARK_BLUE)
    r += 1
    hdr_row(ws5, r, ["Assumption", "Value", "Methodology"], 1)
    r += 1
    macro = str(data.get("macro_sensitivity", "—"))
    disc_rate = data.get("discount_rate", "—")
    assumptions = [
        ("PD — Stage 1", f"{pd_used:.2f}%" if pd_used > 1 else f"{pd_used * 100:.2f}%", "Internal / external PD model"),
        ("LGD", f"{lgd_used:.2f}%" if lgd_used > 1 else f"{lgd_used * 100:.2f}%", "Collateral-based recovery"),
        ("EAD", f"${total_ead:,.2f}", "Outstanding balance"),
        ("Macro overlay", macro, "GDP / unemployment / rates sensitivity"),
        ("Discount rate", f"{disc_rate}%" if disc_rate != "—" else "—", "Effective interest"),
    ]
    alt = False
    for a, v, m in assumptions:
        bg = LIGHT_BG if alt else WHITE
        alt = not alt
        data_cell(ws5, r, 1, a, fill_hex=bg)
        data_cell(ws5, r, 2, v, fill_hex=bg)
        data_cell(ws5, r, 3, m, fill_hex=bg)
        r += 1

    r += 1
    section_hdr(ws5, r, "DISCLOSURE 4 — SENSITIVITY ANALYSIS", "A", "B", fill_hex=ORANGE)
    r += 1
    scen = data.get("scenario_results") or {}
    if isinstance(scen, dict) and scen.get("base") is not None:
        hdr_row(ws5, r, ["Scenario", "ECL ($)", "vs Base ($)"], 1)
        r += 1
        base_v = float(scen.get("base", 0) or 0)
        opt_v = float(scen.get("optimistic", 0) or 0)
        pess_v = float(scen.get("pessimistic", 0) or 0)
        rows = [
            ("Base", base_v, 0.0, LIGHT_BG),
            ("Optimistic", opt_v, opt_v - base_v, LIGHT_GREEN),
            ("Pessimistic", pess_v, pess_v - base_v, LIGHT_RED),
        ]
        for name, ev, diff, fill in rows:
            data_cell(ws5, r, 1, name, fill_hex=fill, bold=True)
            data_cell(ws5, r, 2, ev, fill_hex=fill, number_format=_money_fmt(), align="right")
            data_cell(ws5, r, 3, diff, fill_hex=fill, number_format=_money_fmt(), align="right")
            r += 1
    else:
        ws5.merge_cells(f"A{r}:C{r + 4}")
        c = ws5[f"A{r}"]
        c.value = (
            "Sensitivity analysis not available. Enable macro scenario modelling to generate optimistic/pessimistic ECL scenarios under IFRS 9."
        )
        c.font = Font(italic=True, size=11, color="666666")
        c.fill = _fill(GREY_BG)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    file_id = str(uuid.uuid4())[:8]
    safe_name = portfolio.replace(" ", "_").replace("/", "_")
    safe_date = report_date.replace("-", "")
    filename = f"IFRS9_ECL_{safe_name}_{safe_date}_{file_id}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    return file_id
