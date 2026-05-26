"""
UAE Real Estate portfolio — cross-project analytics and Excel export.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field

UAE_PEG = 3.6725


class ProjectSummary(BaseModel):
    contract_id: str
    project_name: str
    rera_registration_number: str
    developer_name: str
    contract_price_aed: float
    completion_pct: float
    revenue_recognised_aed: float
    deferred_revenue_aed: float
    vat_amount_aed: float
    escrow_balance_aed: float
    escrow_compliance: str
    health_score: int
    health_checks: Dict[str, Any] = Field(default_factory=dict)
    pending_oqood_filings: int = 0
    bundling_alert: bool = False
    rera_certificate_verified: bool = False
    disclosure_score: Optional[float] = None
    last_updated: str = ""
    completion_source: str = "manual_input"
    currency: str = "AED"
    deadline_overdue: int = 0
    deadline_due_soon: int = 0
    deadline_milestones: List[Dict[str, Any]] = Field(default_factory=list)


class PortfolioAnalytics(BaseModel):
    total_projects: int = 0
    total_contract_value_aed: float = 0.0
    total_revenue_recognised_aed: float = 0.0
    total_deferred_revenue_aed: float = 0.0
    total_vat_aed: float = 0.0
    total_remaining_po_aed: float = 0.0
    portfolio_completion_pct: float = 0.0
    escrow_compliant_count: int = 0
    escrow_violation_count: int = 0
    projects_with_oqood_pending: int = 0
    projects_with_bundling_alert: int = 0
    projects_with_health_issues: int = 0
    portfolio_deadlines_overdue: int = 0
    average_disclosure_score: float = 0.0
    completion_distribution: Dict[str, int] = Field(
        default_factory=lambda: {"0-25": 0, "26-50": 0, "51-75": 0, "76-100": 0}
    )
    projects: List[ProjectSummary] = Field(default_factory=list)
    bundling_alerts: List[Dict[str, Any]] = Field(default_factory=list)
    generated_at: str = ""
    currency: str = "AED"


def _completion_bucket(pct: float) -> str:
    if pct <= 25:
        return "0-25"
    if pct <= 50:
        return "26-50"
    if pct <= 75:
        return "51-75"
    return "76-100"


def _health_score_from_checks(hc: Dict[str, Any]) -> int:
    keys = (
        "check_a_pass",
        "check_b1_pass",
        "check_b2_pass",
        "check_c_pass",
        "check_d_pass",
        "check_e_pass",
    )
    if not hc:
        return 0
    score = 0
    for k in keys:
        if k not in hc:
            continue
        if bool(hc.get(k)):
            score += 1
    return score


def _escrow_compliance_for_contract(contract: Dict[str, Any]) -> str:
    snap = contract.get("realestate_snapshot") or {}
    ev = snap.get("escrow_validation") or contract.get("escrow_validation") or {}
    if ev.get("is_violation") is True:
        return "violation"
    if ev.get("is_violation") is False:
        return "compliant"
    try:
        from backend.app.services.rera_escrow_validator import validate_escrow_release

        receipts = list(contract.get("escrow_receipts") or [])
        releases = list(contract.get("escrow_releases") or [])
        tp = float(contract.get("total_tp") or 0)
        pct = contract.get("construction_completion_pct")
        if pct is None and tp > 0:
            pct = (float(contract.get("recognised_to_date") or 0) / tp) * 100.0
        result = validate_escrow_release(
            receipts, releases, float(pct or 0), tp
        )
        return "violation" if result.is_violation else "compliant"
    except Exception:
        return "unknown"


def _contract_in_bundling(contract_id: str, alerts: List[Dict[str, Any]]) -> bool:
    for alert in alerts:
        for u in alert.get("units") or []:
            if str(u.get("contract_id") or "") == contract_id:
                return True
        if str(alert.get("buyer_id") or "") and contract_id in str(alert):
            return True
    return False


def _map_contract_to_summary(
    contract: Dict[str, Any],
    bundling_alerts: List[Dict[str, Any]],
    display_currency: str = "AED",
) -> ProjectSummary:
    snap = contract.get("realestate_snapshot") or {}
    rate = UAE_PEG if display_currency.upper() == "USD" else 1.0

    contract_price = float(contract.get("total_tp") or 0) / rate
    revenue = float(
        contract.get("recognised_to_date")
        or snap.get("revenue_recognised")
        or 0
    ) / rate
    deferred = float(contract.get("deferred_balance") or snap.get("deferred_revenue") or 0) / rate
    vat = float(snap.get("vat_amount") or contract.get("vat_amount") or 0) / rate

    completion = contract.get("construction_completion_pct")
    if completion is None:
        completion = snap.get("completion_pct")
    if completion is None and contract_price > 0:
        completion = (revenue / contract_price) * 100.0
    completion = float(completion or 0)

    receipts = contract.get("escrow_receipts") or []
    released = sum(float(r.get("amount") or 0) for r in contract.get("escrow_releases") or [])
    received = sum(float(r.get("amount") or 0) for r in receipts)
    escrow_balance = (received - released) / rate

    hc = dict(snap.get("health_validation") or contract.get("health_validation") or {})
    health_score = int(snap.get("health_score") or _health_score_from_checks(hc))

    escrow_status = _escrow_compliance_for_contract(contract)
    cid = str(contract.get("contract_id") or "RE-UNKNOWN")

    from backend.app.services.rera_deadline_tracker import summarize_deadlines_for_contract

    deadline_summary = summarize_deadlines_for_contract(contract)

    return ProjectSummary(
        contract_id=cid,
        project_name=str(
            snap.get("project_name") or contract.get("project_name") or contract.get("customer_name") or "Unnamed Project"
        ),
        rera_registration_number=str(
            snap.get("rera_registration_number") or contract.get("risk") or contract.get("rera_registration_number") or "—"
        ),
        developer_name=str(snap.get("developer_name") or contract.get("developer_name") or "Developer"),
        contract_price_aed=round(contract_price, 2),
        completion_pct=round(completion, 1),
        revenue_recognised_aed=round(revenue, 2),
        deferred_revenue_aed=round(deferred, 2),
        vat_amount_aed=round(vat, 2),
        escrow_balance_aed=round(escrow_balance, 2),
        escrow_compliance=escrow_status,
        health_score=health_score,
        health_checks=hc,
        pending_oqood_filings=int(snap.get("pending_oqood_filings") or contract.get("pending_oqood_filings") or 0),
        bundling_alert=bool(
            snap.get("bundling_alert")
            or contract.get("bundling_alert")
            or _contract_in_bundling(cid, bundling_alerts)
        ),
        rera_certificate_verified=bool(
            snap.get("rera_certificate_verified") or contract.get("rera_certificate_verified")
        ),
        disclosure_score=(
            float(contract["disclosure_score"])
            if contract.get("disclosure_score") is not None
            else (float(snap["disclosure_score"]) if snap.get("disclosure_score") is not None else None)
        ),
        last_updated=str(
            contract.get("_updated_at") or snap.get("last_updated") or datetime.now(timezone.utc).isoformat()
        ),
        completion_source=str(
            snap.get("completion_source") or contract.get("completion_source") or "manual_input"
        ),
        currency=display_currency.upper(),
        deadline_overdue=int(deadline_summary.get("deadline_overdue") or 0),
        deadline_due_soon=int(deadline_summary.get("deadline_due_soon") or 0),
        deadline_milestones=list(deadline_summary.get("deadline_milestones") or []),
    )


def build_portfolio_analytics(
    portfolio_contracts: List[Dict[str, Any]],
    *,
    bundling_alerts: Optional[List[Dict[str, Any]]] = None,
    currency: str = "AED",
) -> PortfolioAnalytics:
    """Aggregate real_estate_off_plan portfolio contracts."""
    alerts = bundling_alerts or []
    re_contracts = [
        c
        for c in portfolio_contracts
        if str(c.get("contract_type") or "") == "real_estate_off_plan"
    ]

    projects = [_map_contract_to_summary(c, alerts, currency) for c in re_contracts]
    return _aggregate_from_projects(projects, alerts, currency)


def _aggregate_from_projects(
    projects: List[ProjectSummary],
    bundling_alerts: List[Dict[str, Any]],
    currency: str,
) -> PortfolioAnalytics:
    dist: Dict[str, int] = {"0-25": 0, "26-50": 0, "51-75": 0, "76-100": 0}
    for p in projects:
        dist[_completion_bucket(p.completion_pct)] = dist.get(_completion_bucket(p.completion_pct), 0) + 1
    total_cv = sum(p.contract_price_aed for p in projects)
    total_rev = sum(p.revenue_recognised_aed for p in projects)
    weighted = (
        sum(p.completion_pct * p.contract_price_aed for p in projects) / total_cv if total_cv > 0 else 0.0
    )
    scores = [p.disclosure_score for p in projects if p.disclosure_score is not None]
    return PortfolioAnalytics(
        total_projects=len(projects),
        total_contract_value_aed=round(total_cv, 2),
        total_revenue_recognised_aed=round(total_rev, 2),
        total_deferred_revenue_aed=round(sum(p.deferred_revenue_aed for p in projects), 2),
        total_vat_aed=round(sum(p.vat_amount_aed for p in projects), 2),
        total_remaining_po_aed=round(
            sum(max(0.0, p.contract_price_aed - p.revenue_recognised_aed) for p in projects), 2
        ),
        portfolio_completion_pct=round(weighted, 1),
        escrow_compliant_count=sum(1 for p in projects if p.escrow_compliance == "compliant"),
        escrow_violation_count=sum(1 for p in projects if p.escrow_compliance == "violation"),
        projects_with_oqood_pending=sum(1 for p in projects if p.pending_oqood_filings > 0),
        projects_with_bundling_alert=sum(1 for p in projects if p.bundling_alert),
        projects_with_health_issues=sum(1 for p in projects if p.health_score < 4),
        portfolio_deadlines_overdue=sum(p.deadline_overdue for p in projects),
        average_disclosure_score=round(sum(scores) / len(scores), 1) if scores else 0.0,
        completion_distribution=dist,
        projects=projects,
        bundling_alerts=bundling_alerts,
        generated_at=datetime.now(timezone.utc).isoformat(),
        currency=currency.upper(),
    )


def filter_projects(
    analytics: PortfolioAnalytics,
    *,
    developer_name: Optional[str] = None,
    min_completion: Optional[float] = None,
    max_completion: Optional[float] = None,
    health_issues_only: bool = False,
    violations_only: bool = False,
) -> PortfolioAnalytics:
    """Re-aggregate after filtering project list."""
    projects = list(analytics.projects)
    if developer_name:
        q = developer_name.strip().lower()
        projects = [p for p in projects if q in p.developer_name.lower()]
    if min_completion is not None:
        projects = [p for p in projects if p.completion_pct >= min_completion]
    if max_completion is not None:
        projects = [p for p in projects if p.completion_pct <= max_completion]
    if health_issues_only:
        projects = [p for p in projects if p.health_score < 4]
    if violations_only:
        projects = [p for p in projects if p.escrow_compliance == "violation"]
    return _aggregate_from_projects(projects, analytics.bundling_alerts, analytics.currency)


def export_portfolio_analytics_excel(analytics: PortfolioAnalytics) -> bytes:
    """Three-sheet portfolio workbook."""
    wb = Workbook()
    cur = analytics.currency
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    amber_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws1 = wb.active
    ws1.title = "Portfolio Summary"
    kpis = [
        ("Total projects", analytics.total_projects),
        (f"Total contract value ({cur})", analytics.total_contract_value_aed),
        (f"Revenue recognised ({cur})", analytics.total_revenue_recognised_aed),
        (f"Deferred revenue ({cur})", analytics.total_deferred_revenue_aed),
        (f"VAT ({cur})", analytics.total_vat_aed),
        (f"Remaining PO ({cur})", analytics.total_remaining_po_aed),
        ("Portfolio completion %", analytics.portfolio_completion_pct),
        ("Escrow compliant", analytics.escrow_compliant_count),
        ("Escrow violations", analytics.escrow_violation_count),
        ("Oqood pending projects", analytics.projects_with_oqood_pending),
        ("Bundling alert projects", analytics.projects_with_bundling_alert),
        ("Health issue projects", analytics.projects_with_health_issues),
        ("Avg disclosure score", analytics.average_disclosure_score),
    ]
    ws1.append(["KPI", "Value"])
    for c in ws1[1]:
        c.fill = header_fill
        c.font = header_font
    for label, val in kpis:
        ws1.append([label, val])

    ws2 = wb.create_sheet("Project Detail")
    headers = [
        "Project",
        "RERA",
        "Developer",
        "Completion %",
        f"Contract ({cur})",
        f"Revenue ({cur})",
        f"Deferred ({cur})",
        "Escrow",
        "Health",
        "Oqood Pending",
        "Bundling",
    ]
    ws2.append(headers)
    for c in ws2[1]:
        c.fill = header_fill
        c.font = header_font
    for p in analytics.projects:
        row_idx = ws2.max_row + 1
        ws2.append(
            [
                p.project_name,
                p.rera_registration_number,
                p.developer_name,
                p.completion_pct,
                p.contract_price_aed,
                p.revenue_recognised_aed,
                p.deferred_revenue_aed,
                p.escrow_compliance,
                f"{p.health_score}/5",
                p.pending_oqood_filings,
                "Yes" if p.bundling_alert else "No",
            ]
        )
        row_fill = None
        if p.escrow_compliance == "violation":
            row_fill = red_fill
        elif p.health_score < 3:
            row_fill = amber_fill
        if row_fill:
            for cell in ws2[row_idx]:
                cell.fill = row_fill
        oq_cell = ws2.cell(row=row_idx, column=10)
        if p.pending_oqood_filings > 0:
            oq_cell.fill = amber_fill
        bun_cell = ws2.cell(row=row_idx, column=11)
        if p.bundling_alert:
            bun_cell.fill = amber_fill

    ws3 = wb.create_sheet("Compliance Matrix")
    check_cols = ["A", "B1", "B2", "C", "D", "E"]
    ws3.append(["Project", "RERA"] + check_cols + ["Health"])
    for c in ws3[1]:
        c.fill = header_fill
        c.font = header_font
    key_map = {
        "A": "check_a_pass",
        "B1": "check_b1_pass",
        "B2": "check_b2_pass",
        "C": "check_c_pass",
        "D": "check_d_pass",
        "E": "check_e_pass",
    }
    for p in analytics.projects:
        row = [p.project_name, p.rera_registration_number]
        hc = p.health_checks or {}
        for col in check_cols:
            k = key_map[col]
            if k not in hc:
                row.append("—")
            elif hc.get(k):
                row.append("✓")
            else:
                row.append("✗")
        row.append(f"{p.health_score}/5")
        ws3.append(row)
        r = ws3.max_row
        for ci, col in enumerate(check_cols, start=3):
            cell = ws3.cell(row=r, column=ci)
            if cell.value == "✓":
                cell.fill = green_fill
            elif cell.value == "✗":
                cell.fill = red_fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
