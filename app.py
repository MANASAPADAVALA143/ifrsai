"""
IFRS 16 Lease Accounting Automation - FastAPI Application
Enterprise REST API for IFRS 16 calculations and reporting
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks, Query, Request, Form
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse, StreamingResponse, Response, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field, field_validator
from typing import Optional, List, Dict, Any
from datetime import datetime, timezone
from decimal import Decimal
import io
import hashlib
import os
import json
import uuid
from pathlib import Path
import shutil
from contextlib import asynccontextmanager
import sys
from dotenv import load_dotenv

# Avoid UnicodeEncodeError on Windows (cp1252) when logging emoji or other non-ASCII text.
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (OSError, ValueError):
        pass
from currency_format import format_currency_value  # noqa: F401 — INR vs international amount formatting
import gc
from macro_sensitivity_config import MACRO_SENSITIVITY_DEFAULTS, update_sensitivity, get_sensitivity, is_db_loaded, set_db_loaded
from macro_sensitivity_db import init_db, get_active_config, save_config, get_config_history

# Load environment variables from .env file
# Always load from project root so .env is found regardless of cwd
_project_root = Path(__file__).resolve().parent
load_dotenv(_project_root / ".env")
if not os.getenv("ANTHROPIC_API_KEY"):
    load_dotenv(_project_root / "frontend" / ".env")  # fallback if key is in frontend/.env

# Heavy imports (pandas, numpy, anthropic, openpyxl) moved into route handlers
# RAG engine loads on first use, not at startup

# Configuration
UPLOAD_DIR = Path("uploads")
OUTPUT_DIR = Path("outputs")
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

from backend.app.services.ifrs15_db import ifrs15_db


def _ifrs15_firm_id(request: Optional[Request] = None) -> str:
    """Tenant id for IFRS 15 persistence (header X-Firm-Id or IFRS15_FIRM_ID env)."""
    if request is not None:
        fid = request.headers.get("x-firm-id") or request.headers.get("X-Firm-Id")
        if fid and str(fid).strip():
            return str(fid).strip()
    return os.getenv("IFRS15_FIRM_ID", "default")


def _ifrs15_portfolio_contracts_from_db(firm_id: str) -> List[Dict[str, Any]]:
    rows = ifrs15_db.get_portfolio(firm_id)
    return [dict(row.get("contract_data") or {}) for row in rows]


def _ifrs15_portfolio_summary_data(contract_dict: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "arr": float(contract_dict.get("arr") or 0),
        "mrr": float(contract_dict.get("mrr") or 0),
        "total_tp": float(contract_dict.get("total_tp") or 0),
        "status": str(contract_dict.get("status") or "active"),
        "contract_type": str(contract_dict.get("contract_type") or "other"),
        "deferred_balance": float(contract_dict.get("deferred_balance") or 0),
        "rpo_amount": float(contract_dict.get("rpo_amount") or 0),
        "recognised_to_date": float(contract_dict.get("recognised_to_date") or 0),
        "risk": contract_dict.get("risk"),
        "disclosure_score": contract_dict.get("disclosure_score"),
    }


def _ifrs15_portfolio_saas_summary(firm_id: Optional[str] = None) -> Dict[str, Any]:
    """Roll-up metrics for IFRS 15 SaaS-style portfolio (Supabase)."""
    fid = firm_id or _ifrs15_firm_id()
    contracts = _ifrs15_portfolio_contracts_from_db(fid)
    if not contracts:
        return {"success": True, "summary": {}, "contracts": []}

    active = [c for c in contracts if c.get("status") != "churned"]
    churned = [c for c in contracts if c.get("status") == "churned"]
    at_risk = [c for c in contracts if c.get("status") == "at_risk"]

    total_arr = sum(float(c.get("arr") or 0) for c in active)
    total_mrr = sum(float(c.get("mrr") or 0) for c in active)
    total_deferred = sum(float(c.get("deferred_balance") or 0) for c in contracts)
    total_rpo = sum(float(c.get("rpo_amount") or 0) for c in contracts)
    total_recognised = sum(float(c.get("recognised_to_date") or 0) for c in contracts)

    n = len(contracts)
    churn_rate = (len(churned) / n * 100.0) if n else 0.0

    by_type: Dict[str, Any] = {}
    for c in contracts:
        ct = str(c.get("contract_type") or "other")
        if ct not in by_type:
            by_type[ct] = {"count": 0, "arr": 0.0, "deferred": 0.0}
        by_type[ct]["count"] += 1
        by_type[ct]["arr"] += float(c.get("arr") or 0)
        by_type[ct]["deferred"] += float(c.get("deferred_balance") or 0)

    bundling_alerts: List[Dict[str, Any]] = []
    try:
        from backend.app.services.multi_unit_bundling import UnitContract, assess_bundling

        by_buyer: Dict[str, List[Dict[str, Any]]] = {}
        for c in contracts:
            if str(c.get("contract_type") or "") != "real_estate_off_plan":
                continue
            bid = str(c.get("buyer_id_hash") or "").strip()
            if not bid:
                continue
            by_buyer.setdefault(bid, []).append(c)

        for _, rows in by_buyer.items():
            if len(rows) < 2:
                continue
            units: List[UnitContract] = []
            for i, row in enumerate(rows):
                units.append(
                    UnitContract(
                        contract_id=str(row.get("contract_id") or f"RE-{i+1}"),
                        unit_number=str(row.get("unit_number") or row.get("contract_id") or f"U-{i+1}"),
                        unit_type=str(row.get("unit_type") or "apartment"),
                        contract_price_aed=float(row.get("total_tp") or 0),
                        contract_date=str(row.get("start_date") or "2024-01-01"),
                        completion_pct=float(row.get("completion_pct") or 0),
                        costs_incurred_aed=float(row.get("costs_incurred_aed") or 0),
                        buyer_name=str(row.get("customer_name") or "Buyer"),
                        buyer_id=str(row.get("buyer_id_hash") or ""),
                    )
                )
            assessment = assess_bundling(units).model_dump()
            if assessment.get("should_bundle"):
                bundling_alerts.append(assessment)
    except Exception:
        bundling_alerts = []

    summary = {
        "total_contracts": n,
        "active_contracts": len(active),
        "at_risk_contracts": len(at_risk),
        "churned_contracts": len(churned),
        "total_arr": round(total_arr, 2),
        "total_mrr": round(total_mrr, 2),
        "total_deferred_revenue": round(total_deferred, 2),
        "total_rpo": round(total_rpo, 2),
        "total_recognised_to_date": round(total_recognised, 2),
        "churn_rate_pct": round(churn_rate, 2),
        "at_risk_rate_pct": round((len(at_risk) / n * 100.0) if n else 0.0, 2),
        "revenue_backlog": round(total_deferred + total_rpo, 2),
        "by_contract_type": by_type,
        "bundling_alerts_count": len(bundling_alerts),
    }
    return {"success": True, "summary": summary, "contracts": contracts, "bundling_alerts": bundling_alerts}


def _ifrs15_audit_append(
    action: str,
    contract_id: str,
    description: str,
    before_value: Dict[str, Any],
    after_value: Dict[str, Any],
    ifrs_reference: str,
    changed_amount: float = 0.0,
    user: str = "System",
    firm_id: Optional[str] = None,
    portfolio_row_id: Optional[str] = None,
) -> None:
    from ifrs15_calculator import IFRS15Calculator

    calc = IFRS15Calculator()
    entry = calc.create_audit_entry(
        action=action,
        contract_id=contract_id or "N/A",
        description=description,
        before_value=before_value,
        after_value=after_value,
        ifrs_reference=ifrs_reference,
        user=user,
        changed_amount=float(changed_amount or 0),
    )
    fid = firm_id or _ifrs15_firm_id()
    ifrs15_db.log_action(
        firm_id=fid,
        action=action,
        details=entry,
        contract_id=portfolio_row_id,
        user_id=user,
    )


def _ifrs15_audit_entries_from_db(
    firm_id: str,
    limit: int = 500,
    contract_id: str = "",
    action: str = "",
) -> List[Dict[str, Any]]:
    rows = ifrs15_db.get_audit_log(
        firm_id,
        limit=limit,
        business_contract_id=contract_id or None,
    )
    entries: List[Dict[str, Any]] = []
    for row in rows:
        details = row.get("details")
        if isinstance(details, dict) and details.get("entry_id"):
            entries.append(details)
        else:
            entries.append(
                {
                    "entry_id": str(row.get("id", ""))[:8].upper(),
                    "timestamp": row.get("created_at", ""),
                    "user": row.get("user_id") or "System",
                    "action": row.get("action", ""),
                    "contract_id": (details or {}).get("contract_id", "N/A")
                    if isinstance(details, dict)
                    else "N/A",
                    "description": str((details or {}).get("description", row.get("action", ""))),
                    "before_value": (details or {}).get("before_value", {})
                    if isinstance(details, dict)
                    else {},
                    "after_value": (details or {}).get("after_value", {})
                    if isinstance(details, dict)
                    else {},
                    "ifrs_reference": (details or {}).get("ifrs_reference", "")
                    if isinstance(details, dict)
                    else "",
                    "sign_off_required": False,
                    "signed_off_by": "",
                    "signed_off_at": "",
                    "notes": "",
                }
            )
    if action:
        entries = [e for e in entries if str(e.get("action", "")) == action]
    entries.sort(key=lambda x: str(x.get("timestamp", "")), reverse=True)
    return entries

# Initialize services
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")


def get_public_api_base() -> str:
    """Base URL for printed links (docs, health). Cloud: RENDER_EXTERNAL_URL / PUBLIC_API_URL. Local: _IFRS_BIND_PORT."""
    r = (os.getenv("RENDER_EXTERNAL_URL") or "").strip().rstrip("/")
    if r:
        return r
    p = (os.getenv("PUBLIC_API_URL") or "").strip().rstrip("/")
    if p:
        return p
    port = (os.getenv("_IFRS_BIND_PORT") or "9000").strip()
    return f"http://127.0.0.1:{port}"

# Global RAG engine instance (lazy-loaded on first use)
rag_engine = None

def get_rag_engine():
    """Initialize RAG engine on first call. Loads ChromaDB + SentenceTransformer (~400MB)."""
    global rag_engine
    if rag_engine is not None:
        return rag_engine
    import traceback
    try:
        from rag_engine import IFRSRagEngine
        rag_engine = IFRSRagEngine(anthropic_api_key=ANTHROPIC_API_KEY)
        print("RAG engine initialized successfully")
        gc.collect()
        return rag_engine
    except Exception as e:
        print(f"RAG engine failed: {e}")
        traceback.print_exc()
        return None


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Lifespan: startup then yield (shutdown runs after yield)."""
    init_db()
    try:
        from entity_consolidation import init_consolidation_db
        init_consolidation_db()
    except Exception as e:
        print(f"Consolidation DB init warning: {e}")
    row = get_active_config()
    if row:
        update_sensitivity("gdp_sensitivity", row[0])
        update_sensitivity("unemployment_sensitivity", row[1])
        update_sensitivity("interest_rate_sensitivity", row[2])
        print(f"Macro sensitivity loaded from DB: GDP={row[0]}, Unemp={row[1]}, Rate={row[2]}")
    else:
        set_db_loaded(False)
        print("Macro sensitivity using hardcoded defaults")
    print("="*70)
    print("IFRS AI AUTOMATION API - SERVER STARTED")
    print("="*70)
    _base = get_public_api_base()
    print(f"API Documentation: {_base}/api/docs")
    print(f"ReDoc: {_base}/api/redoc")
    print(f"Health Check: {_base}/health")
    print(f"Claude API: {'Configured' if ANTHROPIC_API_KEY else 'Not configured'}")
    print("RAG engine: loads on first /api/chat or embed request")
    print("="*70)
    yield


# Initialize FastAPI with lifespan (replaces deprecated on_event)
app = FastAPI(
    title="IFRS 16 Lease Accounting Automation API",
    description="AI-powered IFRS 16 lease accounting automation using Claude API",
    version="1.0.0",
    docs_url="/api/docs",
    redoc_url="/api/redoc",
    lifespan=lifespan,
)

# CORS middleware - MUST be added before route definitions
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:3002",
        "http://localhost:3003",
        "http://localhost:3004",
        "http://localhost:3005",
        "http://localhost:5173",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:3002",
        "http://127.0.0.1:3003",
        "http://127.0.0.1:3004",
        "http://127.0.0.1:3005",
        "http://127.0.0.1:5173",
        "https://ifrsai.vercel.app",
        "https://ifrs-ai.vercel.app",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

@app.get("/health")
def health():
    """Simple health check for Render/load balancers"""
    return {"status": "ok"}

# Pydantic Models
class LeaseRequest(BaseModel):
    """Manual lease input request"""
    lease_id: str = Field(..., description="Unique lease identifier")
    company_id: str = Field(default="", description="Company identifier for RAG data isolation")
    asset_description: str = Field(..., description="Description of leased asset")
    lessee_name: str = Field(default="", description="Lessee company name")
    lessor_name: str = Field(default="", description="Lessor name")
    commencement_date: str = Field(..., description="Lease commencement date (YYYY-MM-DD)")
    lease_term_months: int = Field(..., gt=0, description="Lease term in months")
    monthly_payment: float = Field(..., gt=0, description="Monthly payment amount")
    non_lease_component: float = Field(default=0, ge=0, description="Monthly non-lease portion (e.g. service/maintenance) excluded from IFRS 16 liability")
    non_lease_description: str = Field(default="", description="What the non-lease component covers")
    practical_expedient_elected: bool = Field(default=False, description="IFRS 16 §15: elect not to separate components")
    annual_discount_rate: float = Field(..., gt=0.0001, le=1, description="Annual discount rate (e.g., 0.085 for 8.5%). IBR must be > 0%.")
    initial_direct_costs: float = Field(default=0, ge=0, description="Initial direct costs (legacy; use legal+brokerage+other if provided)")
    legal_fees: float = Field(default=0, ge=0, description="Legal fees")
    brokerage_fees: float = Field(default=0, ge=0, description="Brokerage / agent fees")
    other_initial_direct_costs: float = Field(default=0, ge=0, description="Other initial direct costs")
    initial_direct_costs_description: str = Field(default="", description="e.g. Legal fees for lease negotiation, agent commission")
    escalation_rate: float = Field(default=0, ge=0, description="Annual escalation rate (e.g., 0.05 for 5%)")
    cpi_index_base: float = Field(default=0, ge=0, description="CPI index at lease commencement (e.g. 100)")
    cpi_index_current: float = Field(default=0, ge=0, description="Latest CPI index value (e.g. 107.5)")
    cpi_adjustment_frequency_months: int = Field(default=12, ge=1, description="Months between CPI payment reviews (typically 12)")
    currency: str = Field(default="INR", description="Currency code")
    payment_type: str = Field(default="Arrears", description="Payment timing: 'Arrears' (end of period) or 'Advance' (beginning of period)")
    rent_free_months: int = Field(default=0, ge=0, description="Rent-free period in months")
    cash_incentive: float = Field(default=0, ge=0, description="Cash incentive received (reduces ROU asset)")
    lease_incentive_description: str = Field(default="", description="e.g. '2 months rent free'")
    rvg_amount: float = Field(default=0, ge=0, description="Residual value guarantee (guaranteed amount)")
    rvg_guaranteed_by: str = Field(default="None", description="Lessee | Third party | None")
    rvg_expected_payment: float = Field(default=0, ge=0, description="Expected payment at end of lease")
    
    class Config:
        json_schema_extra = {
            "example": {
                "lease_id": "LEASE-2024-001",
                "company_id": "COMP-ABC-001",
                "asset_description": "Commercial Office - 5,000 sq ft",
                "lessee_name": "TechCorp India Pvt. Ltd.",
                "lessor_name": "Prime Properties Ltd.",
                "commencement_date": "2024-01-01",
                "lease_term_months": 36,
                "monthly_payment": 50000,
                "annual_discount_rate": 0.085,
                "initial_direct_costs": 40000,
                "escalation_rate": 0.05,
                "currency": "INR"
            }
        }


class CalculationResponse(BaseModel):
    """Calculation response model"""
    status: str
    lease_id: str
    results: dict
    excel_file_id: Optional[str] = None


class ExtractionResponse(BaseModel):
    """Contract extraction response"""
    status: str
    extraction_id: str
    extracted_data: dict
    validation: dict


class ExtractionRequest(BaseModel):
    """Contract extraction request"""
    contract_text: str = Field(..., description="Lease contract text to extract")


class ChatRequest(BaseModel):
    """RAG chat request"""
    company_id: str = Field(..., description="Company identifier for data isolation")
    question: str = Field(..., description="Question to ask about company's IFRS documents")
    document_type: Optional[str] = Field(None, description="Filter by document type (lease, variance, contract, revenue, ecl)")
    top_k: int = Field(default=5, ge=1, le=20, description="Number of context chunks to retrieve")
    
    class Config:
        json_schema_extra = {
            "example": {
                "company_id": "COMP-ABC-001",
                "question": "What is the total lease liability for all office leases?",
                "document_type": "lease",
                "top_k": 5
            }
        }


class ChatResponse(BaseModel):
    """RAG chat response"""
    status: str
    answer: str
    sources: List[dict]
    context_count: int


class CFOInsightsRequest(BaseModel):
    """CFO strategic insights request - prompt built client-side from lease portfolio"""
    prompt: str = Field(..., description="Full prompt for Claude with lease portfolio JSON")


class IFRS16BulkCalculateRequest(BaseModel):
    """Bulk IFRS 16 calculation — same lease objects as /api/calculate"""
    leases: List[LeaseRequest] = Field(..., description="Leases to calculate")


class IFRS16ExportExcelRequest(BaseModel):
    """Build IFRS 16 .xlsx from calculator JSON (e.g. bulk-calculate) — single response, no file_id / disk coupling."""

    lease_id: str = Field(default="lease", description="Sanitized into download filename")
    calculation_results: Dict[str, Any] = Field(..., description="Full IFRS16Calculator output (JSON shape)")


class IFRS16CFOInsightsRequest(BaseModel):
    """CFO strategic insights — lease portfolio from client repository."""

    leases: List[Dict[str, Any]] = Field(default_factory=list)
    total_assets: float = 0
    annual_revenue: float = 0
    budget_lease_cost: float = 0


class ModificationAdviceRequest(BaseModel):
    """IFRS 16 §44 vs §45 modification advisor — extractor hints + live modification inputs."""
    extractor_hints: Dict[str, Any] = Field(default_factory=dict)
    modification_inputs: Dict[str, Any] = Field(default_factory=dict)


class IbrSuggestRequest(BaseModel):
    country: str = Field(default="India")
    currency: str = Field(default="INR")
    lease_term_months: int = Field(..., gt=0)
    asset_type: Optional[str] = Field(default="")
    lessee_type: str = Field(default="Corporate")


# IFRS 15 Models
class PerformanceObligationRequest(BaseModel):
    obligation_id: str
    description: str
    standalone_selling_price: float
    recognition_method: str  # "over_time" or "point_in_time"
    duration_months: int = 0
    transfer_date: Optional[str] = None  # YYYY-MM-DD


class IFRS15CalculateRequest(BaseModel):
    contract_id: str
    customer_name: str
    vendor_name: str = ""
    effective_date: str  # YYYY-MM-DD
    contract_term_months: int
    fixed_consideration: float
    variable_consideration: float = 0
    vc_constraint_factors: dict = Field(
        default_factory=lambda: {
            "susceptible_to_external": False,
            "long_resolution_period": False,
            "wide_range_of_outcomes": False,
            "limited_experience": False,
            "broad_price_concession_practice": False,
        }
    )
    # IFRS 15 §56-58 variable consideration constraint
    variable_consideration_constrained: float = Field(default=0, ge=0, description="Explicitly constrained VC amount (used when constraint_method='amount')")
    constraint_percentage: float = Field(default=100, ge=0, le=100, description="% of variable consideration that is highly probable (0–100). Default 100 = no constraint.")
    constraint_method: str = Field(default="percentage", description="'percentage' or 'amount'")
    discounts: float = 0
    rebates: float = 0
    financing_adjustment: float = 0
    currency: str = "USD"
    cash_received: float = 0
    contract_type: str = "fixed_price"
    hourly_rate: float = 0
    hours_worked: float = 0
    tm_cap: float = 0
    cumulative_billed: float = 0
    total_estimated_cost: float = 0
    actual_cost_to_date: float = 0
    prior_revenue_recognised: float = 0
    maintenance_term_months: int = 0
    volume_slabs: list = Field(default_factory=list)
    estimated_annual_volume: float = 0
    can_estimate_volume: bool = True
    sla_items: list = Field(default_factory=list)
    performance_obligations: List[PerformanceObligationRequest]
    payment_terms: str = ""


class IFRS15ClassifyContractRequest(BaseModel):
    contract_text: str = Field(..., min_length=1)


class IFRS15ModificationRequest(BaseModel):
    original_contract_id: str
    modification_date: str
    modification_description: str = ""
    new_goods_services: List[str] = Field(default_factory=list)
    price_change: float
    remaining_transaction_price: float = 0.0
    remaining_performance_obligations: List[str] = Field(default_factory=list)
    original_ssps: Dict[str, float] = Field(default_factory=dict)


class DeferredRevenueRequest(BaseModel):
    period: str
    opening_balance: float
    new_bookings: float
    revenue_released: float
    cancellations: float = 0.0
    modifications_impact: float = 0.0
    fx_impact: float = 0.0
    gl_closing_balance: float
    currency: str = "USD"
class IFRS15DownloadExcelRequest(BaseModel):
    contract_id: str
    customer_name: Optional[str] = ""
    effective_date: Optional[str] = None
    contract_term_months: Optional[int] = 0
    currency: Optional[str] = "USD"
    results: Dict[str, Any]
    master_report_data: Optional[Dict[str, Any]] = None


class VariableScenario(BaseModel):
    outcome: str
    amount: float
    probability: float = 0.0  # 0.0 to 1.0


class IFRS15VariableConsiderationRequest(BaseModel):
    method: str  # "expected_value" | "scenario_weighted" | "most_likely"
    scenarios: List[VariableScenario]
    constraint_factors: List[bool]  # exactly 5 items
    contract_id: Optional[str] = None
    total_contract_value: Optional[float] = None


class IFRS15ReversalRiskRequest(BaseModel):
    contract_id: Optional[str] = None
    constraint_level: str
    contract_term_months: int
    customer_type: str
    variable_consideration: float
    total_contract_value: float
    refund_type: str = "none"  # none | partial | full
    recognition_type: str  # over_time | point_in_time
    historical_attainment_pct: Optional[float] = None
    has_external_dependency: bool = False
    dependency_level: str = "low"  # low | medium | high


class RPOObligation(BaseModel):
    obligation_name: str
    allocated_amount: float
    recognised_to_date: float
    expected_end_date: str  # ISO date string
    original_expected_duration_months: Optional[float] = None
    is_right_to_invoice: bool = False


class RPOPerformanceObligation(BaseModel):
    name: str
    allocated_amount: float
    recognised_to_date: float
    expected_recognition_pattern: str = "within_1_year"
    recognition_type: str = "over_time"


class RPOContractRequest(BaseModel):
    contract_id: str
    customer_name: str
    contract_start: str
    contract_end: str
    total_transaction_price: float
    revenue_recognised_to_date: float
    performance_obligations: List[RPOPerformanceObligation]
    practical_expedient_applied: bool = False


class IFRS15RPORequest(BaseModel):
    obligations: List[RPOObligation] = Field(default_factory=list)
    contract_id: Optional[str] = None
    contracts: List[RPOContractRequest] = Field(default_factory=list)


class ContractCostLineRequest(BaseModel):
    cost_id: str
    contract_id: str
    description: str
    cost_type: str = "incremental_obtaining"
    cost_amount: float
    incurred_date: str
    contract_start: str
    contract_end: str
    expected_renewal: bool = False
    expected_renewal_months: int = 0
    currency: str = "USD"


class IFRS15ContractCostsRequest(BaseModel):
    """Legacy commission asset calculator OR batch IFRS 15.91–95 contract cost lines."""

    commission_amount: Optional[float] = None
    contract_term_months: Optional[int] = None
    contract_total_value: Optional[float] = None
    contract_id: Optional[str] = None
    costs: List[ContractCostLineRequest] = Field(default_factory=list)


class LicenseIPItemRequest(BaseModel):
    license_id: str
    product_name: str
    license_description: str = ""
    license_fee: float
    license_start: str
    license_end: str = ""
    is_perpetual: bool = False
    entity_activities_affect_ip: bool
    customer_exposed_to_effect: bool
    no_separate_functional_utility: bool
    currency: str = "USD"


class LicensesIPRequest(BaseModel):
    licenses: List[LicenseIPItemRequest]


class CustomerOptionRequest(BaseModel):
    option_id: str
    contract_id: str
    description: str
    option_type: str = "renewal_discount"
    original_contract_value: float
    original_ssp: float
    option_price: float
    option_ssp: float
    exercise_probability: float
    points_granted: float = 0.0
    point_value: float = 0.0
    currency: str = "USD"


class CustomerOptionsRequest(BaseModel):
    options: List[CustomerOptionRequest]


class WarrantyRequest(BaseModel):
    warranty_id: str
    contract_id: str
    product_description: str = ""
    warranty_description: str
    warranty_period_months: int
    warranty_value: float
    required_by_law: bool = False
    covers_specs_only: bool
    customer_can_purchase_separately: bool
    provides_additional_service: bool
    allocated_fee: float = 0
    currency: str = "USD"


class WarrantiesRequest(BaseModel):
    warranties: List[WarrantyRequest]


class BillAndHoldRequest(BaseModel):
    arrangement_id: str
    contract_id: str
    customer_name: str
    product_description: str
    contract_value: float
    expected_delivery_date: str
    billing_date: str
    reason_is_substantive: bool
    product_separately_identified: bool
    product_ready_for_transfer: bool
    entity_cannot_redirect: bool
    currency: str = "USD"


class BillAndHoldsRequest(BaseModel):
    arrangements: List[BillAndHoldRequest]


class FinancingComponentRequest(BaseModel):
    contract_id: str
    description: str = ""
    contract_value: float
    payment_date: str
    transfer_date: str
    payment_timing: str  # "advance" | "deferred"
    discount_rate: float
    currency: str = "USD"


class FinancingComponentsRequest(BaseModel):
    contracts: List[FinancingComponentRequest]


class NonCashItem(BaseModel):
    item_id: str
    contract_id: str
    description: str = ""
    consideration_type: str = "goods"
    fair_value_determinable: bool = True
    fair_value: float = 0
    fallback_ssp: float = 0
    currency: str = "USD"


class ConsiderationPayableItem(BaseModel):
    item_id: str
    contract_id: str
    description: str = ""
    payment_type: str = "cash"
    amount: float
    distinct_benefit_received: bool = False
    fair_value_of_benefit: float = 0
    currency: str = "USD"


class TransactionPriceAdjustmentsRequest(BaseModel):
    non_cash_items: List[NonCashItem] = Field(default_factory=list)
    consideration_payable_items: List[ConsiderationPayableItem] = Field(default_factory=list)


class EscrowReleaseEntry(BaseModel):
    date: str = ""
    amount: float = 0.0
    release_pct: float = 0.0
    milestone_description: str = ""
    rera_approval_ref: Optional[str] = None


class PortfolioContractRequest(BaseModel):
    contract_id: str
    customer_name: str
    contract_type: str = "subscription"
    arr: float = 0
    mrr: float = 0
    start_date: str
    end_date: str
    total_tp: float = 0
    recognised_to_date: float = 0
    deferred_balance: float = 0
    rpo_amount: float = 0
    status: str = "active"
    currency: str = "USD"
    disclosure_score: Optional[float] = None
    risk: Optional[str] = None
    buyer_id: Optional[str] = None
    escrow_receipts: List[Dict[str, Any]] = Field(default_factory=list)
    escrow_releases: List[EscrowReleaseEntry] = Field(default_factory=list)
    construction_completion_pct: Optional[float] = None


class AuditSignOffRequest(BaseModel):
    entry_id: str
    reviewer: str
    notes: str = ""


class IFRS15PrincipalAgentRequest(BaseModel):
    contract_id: Optional[str] = None
    # Legacy IFRS15PrincipalAgentEngine payload
    transaction_price: Optional[float] = None
    cost_paid_to_supplier: Optional[float] = None
    obtains_before_transfer: Optional[bool] = None
    sets_price_independently: Optional[bool] = None
    primarily_responsible: Optional[bool] = None
    # IFRS 15.B34–B38 extended assessment (IFRS15Calculator)
    arrangement_id: Optional[str] = None
    description: str = ""
    third_party_involved: bool = True
    gross_contract_value: Optional[float] = None
    third_party_cost: Optional[float] = None
    controls_before_transfer: Optional[bool] = None
    primary_obligor: Optional[bool] = None
    inventory_risk: Optional[bool] = None
    pricing_discretion: Optional[bool] = None
    credit_risk: Optional[bool] = None


class IFRS15LicenseRequest(BaseModel):
    transaction_price: float
    licence_term_months: int
    licence_start_date: str
    significantly_affects_ip: bool
    customer_exposed_as_occurs: bool
    activities_not_separate_good: bool
    includes_usage_royalties: bool
    contract_id: Optional[str] = None


class IFRS15MasterReportRequest(BaseModel):
    contract_id: str
    customer_name: str = ""
    core_results: Dict[str, Any]
    modification_result: Optional[Dict[str, Any]] = None
    variable_consideration_result: Optional[Dict[str, Any]] = None
    rpo_result: Optional[Dict[str, Any]] = None
    contract_costs_result: Optional[Dict[str, Any]] = None
    principal_agent_result: Optional[Dict[str, Any]] = None
    license_result: Optional[Dict[str, Any]] = None


class IFRS15MasterReportExcelRequest(BaseModel):
    master_report: Dict[str, Any]


class IFRS15ClientReportRequest(BaseModel):
    contract_id: str
    customer_name: str
    calculation_results: Dict[str, Any]
    master_report_data: Optional[Dict[str, Any]] = None
    include_auditor_qa: bool = True
    prepared_by: str = "IFRS AI"


class IFRS15DisclosureScorerRequest(BaseModel):
    disclosure_text: str
    calculation_results: Optional[Dict[str, Any]] = None


class RealEstateOffPlanRequest(BaseModel):
    contract_value: float
    construction_start: str
    expected_handover: str
    current_date: str
    costs_incurred_to_date: float
    total_estimated_costs: float
    escrow_receipts: List[Dict[str, Any]] = Field(default_factory=list)
    revenue_prior_period: float = 0.0
    billings: Optional[List[Dict[str, Any]]] = None
    contract_id: Optional[str] = None


class RealEstateEscrowRequest(BaseModel):
    contract_value: float
    costs_incurred_to_date: float
    total_estimated_costs: float
    escrow_receipts: List[Dict[str, Any]] = Field(default_factory=list)
    milestone_releases: List[Dict[str, Any]] = Field(default_factory=list)
    revenue_prior_period: float = 0.0


class RealEstateModificationRequest(BaseModel):
    original_contract: Dict[str, Any]
    modification_type: str  # price_change | unit_swap | cancellation | extension
    modification_details: Dict[str, Any] = Field(default_factory=dict)
    contract_id: Optional[str] = None
    oqood_filed: bool = False
    modification_date: Optional[str] = None


class RealEstateContractCostsRequest(BaseModel):
    commission_paid: float
    commission_date: Optional[str] = None
    contract_value: float
    expected_amortisation_period: int = 24
    contract_id: Optional[str] = None


class RealEstatePrincipalAgentRequest(BaseModel):
    gross_contract_value: float = 0.0
    contract_value: Optional[float] = None
    agent_commission: float = 0.0
    controls_before_transfer: str = "developer"
    inventory_risk: str = "developer"
    pricing_discretion: str = "developer"
    credit_risk: str = "developer"
    contract_id: Optional[str] = None


class RealEstateVatRequest(BaseModel):
    revenue_schedule: List[Dict[str, Any]] = Field(default_factory=list)
    revenue_current_period: Optional[float] = None
    period: Optional[str] = None
    fta_filing_period: Optional[str] = None
    currency: str = "AED"


class RealEstateFullRequest(BaseModel):
    """Run off-plan, escrow, VAT, and optional modules in one call."""
    off_plan: RealEstateOffPlanRequest
    escrow: Optional[RealEstateEscrowRequest] = None
    vat_schedule: List[Dict[str, Any]] = Field(default_factory=list)
    spa_mapped: Optional[Dict[str, Any]] = None


class RealEstateToCalculateRequest(BaseModel):
    """Build main IFRS 15 /calculate payload from real estate outputs."""
    off_plan: Dict[str, Any]
    spa: Optional[Dict[str, Any]] = None
    spa_mapped: Optional[Dict[str, Any]] = None
    construction_start: str = ""
    expected_handover: str = ""
    contract_value: float = 0.0
    costs_incurred_to_date: float = 0.0
    total_estimated_costs: float = 0.0
    revenue_prior_period: float = 0.0
    escrow_total: float = 0.0
    escrow_receipts: List[Dict[str, Any]] = Field(default_factory=list)
    escrow_releases: List[EscrowReleaseEntry] = Field(default_factory=list)


class RealEstateContractInput(BaseModel):
    """Shared UAE real estate contract fields."""

    rera_registration_number: str = Field(..., min_length=4, max_length=20)
    project_name: str = ""
    currency: str = Field(default="AED", pattern="^(AED|USD)$")
    exchange_rate: float = Field(default=3.6725, gt=0)
    revenue_recognition_trigger: str = Field(
        default="earlier_of_both",
        pattern="^(rera_completion_certificate|spa_handover_date|earlier_of_both)$",
    )
    rera_completion_date: Optional[str] = None
    spa_handover_date: Optional[str] = None

    @field_validator("rera_registration_number")
    @classmethod
    def validate_rera_field(cls, v: str) -> str:
        from backend.app.services.ifrs15_realestate import validate_rera_registration_number

        return validate_rera_registration_number(v)


class RealEstateReportRequest(RealEstateContractInput):
    """Full UAE real estate report with quarterly schedule and optional modules."""

    rera_certificate_ref: Optional[str] = None
    rera_certificate_date: Optional[str] = None
    rera_certificate_verified_pct: Optional[float] = Field(default=None, ge=0, le=100)

    contract_value: float
    construction_start: str
    expected_handover: str
    current_date: str
    costs_incurred_to_date: float
    total_estimated_costs: float
    escrow_receipts: List[Dict[str, Any]] = Field(default_factory=list)
    escrow_releases: List[EscrowReleaseEntry] = Field(default_factory=list)
    milestone_releases: List[Dict[str, Any]] = Field(default_factory=list)
    revenue_prior_period: float = 0.0
    commission_paid: Optional[float] = None
    expected_amortisation_period: int = 24
    assess_principal_agent: bool = False
    gross_contract_value: Optional[float] = None
    controls_before_transfer: str = "developer"
    inventory_risk: str = "developer"
    pricing_discretion: str = "developer"
    credit_risk: str = "developer"
    spa: Optional[Dict[str, Any]] = None
    spa_mapped: Optional[Dict[str, Any]] = None
    contract_id: Optional[str] = None
    cancellation_refund: Optional[Dict[str, Any]] = None
    modifications: List[Dict[str, Any]] = Field(default_factory=list)
    units: List[Dict[str, Any]] = Field(default_factory=list)


class CancellationRefundInput(BaseModel):
    contract_price: float
    amount_paid_by_buyer: float
    construction_completion_pct: float = Field(ge=0, le=100)
    rera_registration_number: str = Field(..., min_length=4, max_length=20)
    cancellation_reason: str = Field(
        default="buyer_default",
        pattern="^(buyer_default|developer_default|mutual_agreement)$",
    )
    escrow_balance: float = 0.0
    escrow_receipts: List[Dict[str, Any]] = Field(default_factory=list)
    escrow_releases: List[EscrowReleaseEntry] = Field(default_factory=list)

    @field_validator("rera_registration_number")
    @classmethod
    def validate_rera_cancel(cls, v: str) -> str:
        from backend.app.services.ifrs15_realestate import validate_rera_registration_number

        return validate_rera_registration_number(v)


class RealEstateExportExcelRequest(BaseModel):
    report: Dict[str, Any]
    contract_id: Optional[str] = None
    escrow_receipts: List[Dict[str, Any]] = Field(default_factory=list)
    escrow_releases: List[EscrowReleaseEntry] = Field(default_factory=list)
    construction_completion_pct: Optional[float] = None


class RealEstateBundlingCheckRequest(BaseModel):
    units: List[Dict[str, Any]] = Field(default_factory=list)
    currency: str = "AED"


class RealEstateOqoodFiledPatchRequest(BaseModel):
    modification_id: str
    oqood_filed: bool = True


def _escrow_entries_to_dicts(entries: Any) -> List[Dict[str, Any]]:
    if not entries:
        return []
    out: List[Dict[str, Any]] = []
    for e in entries:
        if isinstance(e, dict):
            out.append(dict(e))
        elif hasattr(e, "model_dump"):
            out.append(e.model_dump(mode="python"))
        else:
            out.append(dict(e))
    return out


def _rera_escrow_gate(
    escrow_receipts: Any,
    escrow_releases: Any,
    construction_completion_pct: float,
    contract_price_aed: float,
) -> Optional[JSONResponse]:
    from backend.app.services.rera_escrow_validator import (
        validate_escrow_release,
        rera_escrow_violation_response_body,
    )

    receipts = _escrow_entries_to_dicts(escrow_receipts)
    releases = _escrow_entries_to_dicts(escrow_releases)
    ev = validate_escrow_release(receipts, releases, construction_completion_pct, contract_price_aed)
    if ev.is_violation:
        return JSONResponse(status_code=422, content=rera_escrow_violation_response_body(ev))
    return None


# Helper Functions
def convert_lease_request_to_input(request: LeaseRequest):
    """Convert API request to LeaseInput dataclass"""
    from ifrs16_calculator import LeaseInput
    raw_date = (request.commencement_date or "").strip()[:10]
    try:
        commencement_dt = datetime.strptime(raw_date, "%Y-%m-%d")
    except ValueError as e:
        raise ValueError(
            f"commencement_date must be YYYY-MM-DD (got {request.commencement_date!r})"
        ) from e
    return LeaseInput(
        lease_id=request.lease_id,
        asset_description=request.asset_description,
        lessee_name=request.lessee_name,
        lessor_name=request.lessor_name,
        commencement_date=commencement_dt,
        lease_term_months=request.lease_term_months,
        monthly_payment=Decimal(str(request.monthly_payment)),
        non_lease_component=Decimal(str(getattr(request, 'non_lease_component', 0) or 0)),
        non_lease_description=getattr(request, 'non_lease_description', '') or '',
        practical_expedient_elected=bool(getattr(request, 'practical_expedient_elected', False)),
        annual_discount_rate=Decimal(str(request.annual_discount_rate)),
        initial_direct_costs=Decimal(str(request.initial_direct_costs)),
        legal_fees=Decimal(str(getattr(request, "legal_fees", 0) or 0)),
        brokerage_fees=Decimal(str(getattr(request, "brokerage_fees", 0) or 0)),
        other_initial_direct_costs=Decimal(str(getattr(request, "other_initial_direct_costs", 0) or 0)),
        initial_direct_costs_description=getattr(request, "initial_direct_costs_description", "") or "",
        escalation_rate=Decimal(str(request.escalation_rate)),
        cpi_index_base=Decimal(str(getattr(request, "cpi_index_base", 0) or 0)),
        cpi_index_current=Decimal(str(getattr(request, "cpi_index_current", 0) or 0)),
        cpi_adjustment_frequency_months=int(getattr(request, "cpi_adjustment_frequency_months", 12) or 12),
        currency=request.currency,
        payment_type=getattr(request, "payment_type", "Arrears") or "Arrears",
        rent_free_months=getattr(request, "rent_free_months", 0) or 0,
        cash_incentive=Decimal(str(getattr(request, "cash_incentive", 0) or 0)),
        lease_incentive_description=getattr(request, "lease_incentive_description", "") or "",
        rvg_amount=Decimal(str(getattr(request, "rvg_amount", 0) or 0)),
        rvg_guaranteed_by=getattr(request, "rvg_guaranteed_by", "None") or "None",
        rvg_expected_payment=Decimal(str(getattr(request, "rvg_expected_payment", 0) or 0)),
    )


# API Endpoints

@app.get("/", response_class=HTMLResponse)
async def root():
    """Root endpoint - Frontend page"""
    html_file = Path("templates/index.html")
    if html_file.exists():
        return html_file.read_text(encoding='utf-8')
    else:
        return """
        <html>
            <head><title>IFRS AI Platform</title></head>
            <body style="font-family: Arial; text-align: center; padding: 50px;">
                <h1>🚀 IFRS AI Automation Platform</h1>
                <p>Server is running!</p>
                <a href="/api/docs" style="display: inline-block; background: #667eea; color: white; padding: 15px 30px; text-decoration: none; border-radius: 8px; margin: 20px;">
                    Open API Documentation
                </a>
            </body>
        </html>
        """

@app.get("/ifrs16", response_class=HTMLResponse)
async def ifrs16_page():
    """IFRS 16 Lease Accounting Form"""
    html_file = Path("templates/ifrs16.html")
    if html_file.exists():
        return html_file.read_text(encoding='utf-8')
    else:
        raise HTTPException(status_code=404, detail="IFRS 16 page not found")

@app.get("/ifrs15")
async def ifrs15_redirect():
    """Legacy HTML route → Next.js dashboard."""
    return RedirectResponse(url="/dashboard/ifrs15", status_code=301)


@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "anthropic_configured": bool(ANTHROPIC_API_KEY)
    }


@app.post("/api/health/alert")
async def send_health_alert(alert_data: dict):
    """Send email alert when health checks fail"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    try:
        subject = alert_data.get("subject", "IFRS AI Health Alert")
        message = alert_data.get("message", "")
        failures = alert_data.get("failures", [])

        body = f"""
IFRS AI Health Monitor Alert
============================
Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

{message}

DETAILS:
"""
        for f in failures:
            body += f"\n❌ {f.get('name', '')}"
            body += f"\n   Error: {f.get('error', '')}"
            body += f"\n   Fix: {f.get('fix', 'Check logs')}\n"

        body += """
---
This is an automated alert from IFRS AI Health Monitor.
Fix issues before client demos.
        """

        sender_email = os.getenv("ALERT_EMAIL_FROM", "")
        sender_password = os.getenv("ALERT_EMAIL_PASSWORD", "")
        receiver_email = os.getenv("ALERT_EMAIL_TO", "manasa@gnanova.pro")

        if sender_email and sender_password:
            msg = MIMEMultipart()
            msg["From"] = sender_email
            msg["To"] = receiver_email
            msg["Subject"] = subject
            msg.attach(MIMEText(body, "plain"))

            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                server.login(sender_email, sender_password)
                server.sendmail(sender_email, receiver_email, msg.as_string())

            return {"status": "alert_sent", "email": receiver_email}
        else:
            print(f"\n{'='*50}")
            print("HEALTH ALERT (email not configured):")
            print(body)
            print("="*50)
            return {"status": "logged", "message": "Configure ALERT_EMAIL_FROM and ALERT_EMAIL_PASSWORD in .env"}
    except Exception as e:
        print(f"Alert error: {e}")
        return {"status": "error", "message": str(e)}


class QAValidateRequest(BaseModel):
    """QA validation request with IFRS calculation results"""
    ifrs16_results: Optional[Dict] = None
    ifrs15_results: Optional[Dict] = None
    ifrs9_results: Optional[Dict] = None


@app.post("/api/health/qa-validate")
async def qa_validate(request: QAValidateRequest):
    """Validate IFRS calculation results using Claude (Big4 auditor style)"""
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY not configured")
    import anthropic
    try:
        payload = json.dumps({
            "ifrs16": request.ifrs16_results,
            "ifrs15": request.ifrs15_results,
            "ifrs9": request.ifrs9_results,
        }, indent=2, default=str)[:12000]
        prompt = f"""You are a Big4 auditor validating IFRS calculations.

Check these results against IFRS standards:

IFRS 16: Is lease_liability = PV of lease payments? Is ROU asset = lease_liability + initial direct costs?
IFRS 15: Does total recognised + deferred = contract value? 5-step model applied?
IFRS 9: Is ECL = PD × LGD × EAD for each loan? (If no IFRS 9 data, set pass: false, issues: ["IFRS 9 API not implemented"])

RESULTS TO VALIDATE:
{payload}

Return ONLY valid JSON in this exact format:
{{"pass": true/false, "issues": ["issue1", "issue2"], "confidence": 0-100, "ifrs16_notes": "string", "ifrs15_notes": "string", "ifrs9_notes": "string"}}"""

        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        msg = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1500,
            temperature=0,
            messages=[{"role": "user", "content": prompt}]
        )
        text = msg.content[0].text.strip()
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0].strip()
        elif "```" in text:
            text = text.split("```")[1].split("```")[0].strip()
        data = json.loads(text)
        gc.collect()
        return {"status": "success", "validation": data}
    except json.JSONDecodeError as e:
        return {"status": "error", "validation": {"pass": False, "issues": [f"Parse error: {e}"], "confidence": 0}}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class QAReportEmailRequest(BaseModel):
    report_text: str
    subject: str = "IFRS AI — QA Validation Report"


@app.post("/api/health/qa-report-email")
async def qa_report_email(request: QAReportEmailRequest):
    """Email the QA report to configured recipient"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    sender_email = os.getenv("ALERT_EMAIL_FROM", "")
    sender_password = os.getenv("ALERT_EMAIL_PASSWORD", "")
    receiver_email = os.getenv("ALERT_EMAIL_TO", "manasa@gnanova.pro")
    if not sender_email or not sender_password:
        return {"status": "logged", "message": "Email not configured. Set ALERT_EMAIL_FROM and ALERT_EMAIL_PASSWORD."}
    try:
        msg = MIMEMultipart()
        msg["From"] = sender_email
        msg["To"] = receiver_email
        msg["Subject"] = request.subject
        msg.attach(MIMEText(request.report_text, "plain"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, receiver_email, msg.as_string())
        return {"status": "sent", "email": receiver_email}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/api/calculate", response_model=CalculationResponse)
async def calculate_lease(request: LeaseRequest):
    """
    Calculate IFRS 16 lease accounting metrics
    
    Performs complete IFRS 16 calculation including:
    - Lease liability (PV calculation)
    - ROU asset
    - Amortization schedule
    - Journal entries
    - Maturity analysis
    - P&L impact
    """
    try:
        from ifrs16_calculator import IFRS16Calculator
        from ifrs16_excel_export import IFRS16ExcelExporter
        # Convert request to LeaseInput
        lease_input = convert_lease_request_to_input(request)
        
        # Perform calculation
        calculator = IFRS16Calculator()
        results = calculator.calculate_full_ifrs16(lease_input)
        
        # Convert DataFrame to dict for JSON serialization
        results_json = results.copy()
        if 'amortization_schedule' in results_json:
            results_json['amortization_schedule'] = results_json['amortization_schedule'].to_dict(orient='records')
        
        # Generate Excel file (calculation still succeeds if export fails)
        file_id = str(uuid.uuid4())
        excel_filename = OUTPUT_DIR / f"IFRS16_{request.lease_id}_{file_id}.xlsx"
        try:
            exporter = IFRS16ExcelExporter()
            exporter.export_ifrs16_workbook(results, str(excel_filename))
        except Exception as excel_err:
            print(f"Excel export failed (non-critical): {excel_err}")
            file_id = None
        
        # Auto-trigger RAG embedding only when company_id is provided.
        engine = get_rag_engine() if request.company_id else None
        if engine:
            try:
                # Prepare content for embedding (use original results with DataFrames converted)
                embed_content = {
                    "lease_id": request.lease_id,
                    "company_id": request.company_id,
                    "asset_description": request.asset_description,
                    "lessee_name": request.lessee_name,
                    "lessor_name": request.lessor_name,
                    "commencement_date": request.commencement_date,
                    "lease_term_months": request.lease_term_months,
                    "monthly_payment": request.monthly_payment,
                    "annual_discount_rate": request.annual_discount_rate,
                    "currency": request.currency,
                    "lease_liability": float(results['lease_liability']),
                    "rou_asset": float(results['rou_asset']),
                    "monthly_depreciation": float(results['monthly_depreciation']),
                    "total_interest": float(results['total_interest']),
                    "year_1_impact": results_json['year_1_impact']
                }
                
                # Store in RAG engine
                rag_result = engine.embed_and_store(
                    company_id=request.company_id,
                    document_type="lease",
                    content=embed_content,
                    document_id=request.lease_id
                )
                print(f"RAG embedding completed: {rag_result.get('status')}")
                try:
                    from ifrs16_rag_leases import index_lease
                    index_lease(
                        engine,
                        request.lease_id,
                        "",
                        {
                            "property_name": request.asset_description,
                            "start_date": request.commencement_date,
                            "monthly_payment": request.monthly_payment,
                            "currency": request.currency,
                            "ibr": request.annual_discount_rate,
                            "tenant_name": request.lessee_name,
                            "lease_liability": float(results["lease_liability"]),
                            "rou_asset": float(results["rou_asset"]),
                        },
                        company_id=request.company_id or "default",
                    )
                except Exception as lease_idx_err:
                    print(f"Lease RAG index (non-critical): {lease_idx_err}")
                
            except Exception as rag_error:
                print(f"RAG embedding failed (non-critical): {rag_error}")
                # Don't fail the request if RAG embedding fails
        
        gc.collect()
        return CalculationResponse(
            status="success",
            lease_id=request.lease_id,
            results=results_json,
            excel_file_id=file_id
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Calculation error: {str(e)}")


@app.post("/api/extract", response_model=ExtractionResponse)
async def extract_contract(request: ExtractionRequest):
    """
    Extract lease terms from contract text using Claude API
    
    Uses Claude Sonnet 4.5 to extract:
    - Lease dates and terms
    - Payment amounts and schedules
    - Initial costs
    - Options and clauses
    - IFRS 16 classification
    """
    if not ANTHROPIC_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="Claude API not configured. Set ANTHROPIC_API_KEY environment variable."
        )
    
    try:
        from ifrs16_extractor import IFRS16LeaseExtractor
        extractor = IFRS16LeaseExtractor(api_key=ANTHROPIC_API_KEY)
        
        # Extract lease terms
        extracted_data = extractor.extract_lease_terms(request.contract_text)
        
        # Validate extraction
        validation = extractor.validate_extraction(extracted_data)
        
        # Save extraction
        extraction_id = str(uuid.uuid4())
        extraction_file = OUTPUT_DIR / f"extraction_{extraction_id}.json"
        extractor.save_extraction(extracted_data, str(extraction_file))
        
        gc.collect()
        return ExtractionResponse(
            status="success",
            extraction_id=extraction_id,
            extracted_data=extracted_data,
            validation=validation
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Extraction error: {str(e)}")


@app.get("/api/upload-contract")
async def upload_contract_get():
    """GET not allowed; use POST with a file. See /api/docs for usage."""
    return JSONResponse(
        status_code=200,
        content={
            "message": "This endpoint accepts POST only. Upload a file from the app or use the API docs.",
            "docs": f"{get_public_api_base()}/api/docs",
            "method": "POST",
        },
    )


@app.post("/api/upload-contract")
async def upload_contract(file: UploadFile = File(...)):
    """
    Upload lease contract file (PDF, DOCX, TXT, Excel) and extract terms
    
    Supports:
    - PDF files (with OCR if needed)
    - Word documents (.docx)
    - Text files (.txt)
    - Excel files (.xlsx, .xls)
    """
    import traceback
    safe_name = (file.filename or "upload").replace("\\", "_").replace("/", "_")
    print(f"📥 Upload received: {safe_name} ({file.content_type})")

    if not ANTHROPIC_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="Claude API not configured. Set ANTHROPIC_API_KEY environment variable."
        )

    allowed_extensions = ['.pdf', '.docx', '.txt', '.xlsx', '.xls']
    file_ext = Path(safe_name).suffix.lower()

    if file_ext not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type. Allowed: {', '.join(allowed_extensions)}"
        )

    file_id = str(uuid.uuid4())
    upload_path = UPLOAD_DIR / f"{file_id}_{safe_name}"
    
    try:
        # Save uploaded file
        with open(upload_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        print(f"❌ Save error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to save file: {str(e)}")
    
    try:
        # Extract from file
        from ifrs16_extractor import IFRS16LeaseExtractor
        extractor = IFRS16LeaseExtractor(api_key=ANTHROPIC_API_KEY)
        extracted_data = extractor.extract_from_file(str(upload_path))
        
        # Validate extraction (don't let validation failure break the response)
        try:
            validation = extractor.validate_extraction(extracted_data)
        except Exception as val_err:
            print(f"⚠️ Validation warning: {val_err}")
            validation = {
                'is_valid': False,
                'errors': [str(val_err)],
                'warnings': [],
                'requires_review': True,
                'low_confidence_count': 0,
                'error_count': 1,
                'warning_count': 0
            }
        
        # Save extraction
        try:
            extraction_file = OUTPUT_DIR / f"extraction_{file_id}.json"
            extractor.save_extraction(extracted_data, str(extraction_file))
        except Exception as save_err:
            print(f"⚠️ Could not save extraction JSON: {save_err}")
        
        del extractor  # free ref so gc can reclaim
        print(f"✅ Extraction complete: {safe_name}")
        gc.collect()
        
        # Ensure JSON-serializable (datetime, Decimal, etc.)
        def make_serializable(obj):
            if obj is None:
                return None
            if isinstance(obj, dict):
                return {str(k): make_serializable(v) for k, v in obj.items()}
            if isinstance(obj, list):
                return [make_serializable(v) for v in obj]
            if isinstance(obj, (str, int, float, bool)):
                return obj
            if hasattr(obj, 'isoformat'):
                return obj.isoformat()
            if hasattr(obj, '__float__') and not isinstance(obj, (int, float)):
                try:
                    return float(obj)
                except (TypeError, ValueError):
                    return str(obj)
            return str(obj)
        
        extracted_data = make_serializable(extracted_data)
        validation = make_serializable(validation)
        
        return {
            "status": "success",
            "file_id": file_id,
            "filename": safe_name,
            "extracted_data": extracted_data,
            "validation": validation
        }
        
    except Exception as e:
        err_msg = str(e)
        print(f"❌ Upload error: {err_msg}")
        traceback.print_exc()
        # Common causes: Claude rate limit, memory after multiple uploads, temp file / file handle
        if "api_key" in err_msg.lower() or "authentication" in err_msg.lower():
            raise HTTPException(status_code=503, detail="Invalid Claude API key. Check ANTHROPIC_API_KEY in .env")
        if "rate" in err_msg.lower() or "overloaded" in err_msg.lower():
            raise HTTPException(status_code=503, detail="Claude API rate limit. Please try again in a moment.")
        if "File not found" in err_msg or "No such file" in err_msg:
            raise HTTPException(status_code=500, detail="File could not be read. Try a different file.")
        if "JSON" in err_msg or "json" in err_msg:
            raise HTTPException(status_code=500, detail="AI extraction returned invalid format. Try again or use a clearer contract.")
        raise HTTPException(status_code=500, detail=f"Extraction failed: {err_msg[:200]}")
    finally:
        # Always delete temp upload file to avoid disk buildup and possible lock/memory issues
        try:
            if upload_path.exists():
                upload_path.unlink()
                print(f"🗑️ Cleaned up temp file: {upload_path.name}")
        except Exception as cleanup_err:
            print(f"⚠️ Could not delete temp file {upload_path}: {cleanup_err}")
        gc.collect()


@app.post("/api/ifrs16/modification-advice")
async def ifrs16_modification_advice(body: ModificationAdviceRequest):
    """
    Suggest IFRS 16 modification treatment (§44 separate lease vs §45 remeasurement)
    from extractor-style JSON plus current modification form values.
    """
    from modification_advisor import advise_modification

    result = advise_modification(body.extractor_hints or {}, body.modification_inputs or {})
    return result


@app.get("/api/download/{file_id}")
async def download_file(file_id: str):
    """
    Download generated Excel report
    
    Args:
        file_id: File ID returned from /api/calculate endpoint
    """
    # Find file with this ID
    matching_files = list(OUTPUT_DIR.glob(f"*{file_id}*.xlsx"))
    
    if not matching_files:
        raise HTTPException(status_code=404, detail="File not found")
    
    file_path = matching_files[0]
    
    return FileResponse(
        path=str(file_path),
        filename=file_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.delete("/api/cleanup")
async def cleanup_files(older_than_days: int = 7):
    """
    Cleanup old files from uploads and outputs directories
    
    Args:
        older_than_days: Delete files older than this many days
    """
    import time
    
    deleted_count = 0
    cutoff_time = time.time() - (older_than_days * 86400)
    
    for directory in [UPLOAD_DIR, OUTPUT_DIR]:
        for file_path in directory.iterdir():
            if file_path.is_file() and file_path.stat().st_mtime < cutoff_time:
                file_path.unlink()
                deleted_count += 1
    
    return {
        "status": "success",
        "deleted_files": deleted_count,
        "older_than_days": older_than_days
    }


@app.post("/api/batch-calculate")
async def batch_calculate(leases: List[LeaseRequest]):
    """
    Batch calculate multiple leases
    
    Calculates IFRS 16 metrics for multiple leases in one request
    """
    from ifrs16_calculator import IFRS16Calculator
    results = []
    
    for lease in leases:
        try:
            lease_input = convert_lease_request_to_input(lease)
            calculator = IFRS16Calculator()
            result = calculator.calculate_full_ifrs16(lease_input)
            
            # Convert DataFrame
            result_json = result.copy()
            if 'amortization_schedule' in result_json:
                result_json['amortization_schedule'] = result_json['amortization_schedule'].to_dict(orient='records')
            
            results.append({
                "lease_id": lease.lease_id,
                "status": "success",
                "results": result_json
            })
            
        except Exception as e:
            results.append({
                "lease_id": lease.lease_id,
                "status": "error",
                "error": str(e)
            })
    
    gc.collect()
    return {
        "status": "completed",
        "total_leases": len(leases),
        "successful": sum(1 for r in results if r["status"] == "success"),
        "failed": sum(1 for r in results if r["status"] == "error"),
        "results": results
    }


@app.post("/api/ifrs16/bulk-calculate")
async def ifrs16_bulk_calculate(body: IFRS16BulkCalculateRequest):
    """
    Calculate IFRS 16 for many leases in one request.
    One failure does not stop the rest. Returns summary metrics per lease + portfolio roll-up.
    """
    from ifrs16_calculator import IFRS16Calculator

    out_results: List[Dict[str, Any]] = []
    total = len(body.leases)
    successful = 0
    failed = 0
    total_lease_liability = 0.0
    total_rou_asset = 0.0
    ibr_sum = 0.0
    currency_counts: Dict[str, int] = {}

    for lease in body.leases:
        try:
            lease_input = convert_lease_request_to_input(lease)
            calculator = IFRS16Calculator()
            result = calculator.calculate_full_ifrs16(lease_input)
            result_json = result.copy()
            if "amortization_schedule" in result_json and hasattr(result_json["amortization_schedule"], "to_dict"):
                result_json["amortization_schedule"] = result_json["amortization_schedule"].to_dict(orient="records")

            ll = float(result.get("lease_liability", 0) or 0)
            rou = float(result.get("rou_asset", 0) or 0)
            md = float(result.get("monthly_depreciation", 0) or 0)
            ti = float(result.get("total_interest", 0) or 0)

            total_lease_liability += ll
            total_rou_asset += rou
            ibr_sum += float(lease.annual_discount_rate or 0)
            ccy = (lease.currency or "INR").upper()
            currency_counts[ccy] = currency_counts.get(ccy, 0) + 1
            successful += 1

            out_results.append(
                {
                    "lease_id": lease.lease_id,
                    "status": "success",
                    "error": None,
                    "lease_liability": ll,
                    "rou_asset": rou,
                    "monthly_depreciation": md,
                    "total_interest": ti,
                    "calculation_results": result_json,
                }
            )
        except Exception as e:
            failed += 1
            out_results.append(
                {
                    "lease_id": getattr(lease, "lease_id", "") or "",
                    "status": "error",
                    "error": str(e),
                    "lease_liability": 0.0,
                    "rou_asset": 0.0,
                    "monthly_depreciation": 0.0,
                    "total_interest": 0.0,
                    "calculation_results": None,
                }
            )

    gc.collect()
    avg_ibr = (ibr_sum / successful) if successful else 0.0
    return {
        "total": total,
        "successful": successful,
        "failed": failed,
        "results": out_results,
        "portfolio_summary": {
            "total_lease_liability": total_lease_liability,
            "total_rou_asset": total_rou_asset,
            "avg_ibr": avg_ibr,
            "currency_breakdown": currency_counts,
        },
    }


@app.post("/api/ifrs16/export-excel")
async def ifrs16_export_excel(body: IFRS16ExportExcelRequest):
    """
    Return IFRS 16 workbook bytes from existing calculation_results.
    Use after bulk-calculate so Quick Analysis can download without a second /api/calculate
    (avoids Excel+RAG path failures and ephemeral file IDs on multi-instance hosts).
    """
    try:
        from ifrs16_excel_export import IFRS16ExcelExporter

        exporter = IFRS16ExcelExporter()
        data = exporter.export_ifrs16_workbook_bytes(body.calculation_results)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel export failed: {str(e)}")
    raw = (body.lease_id or "lease").strip()
    safe = "".join(c for c in raw if c.isalnum() or c in "._-")[:80] or "lease"
    fname = f"IFRS16_{safe}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/api/ifrs16/bulk-template")
async def ifrs16_bulk_template():
    """Download Excel template for portfolio bulk upload (3 sample rows)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Leases"

    headers = [
        "lease_id",
        "asset_description",
        "lessee_name",
        "lessor_name",
        "commencement_date",
        "lease_term_months",
        "monthly_payment",
        "annual_discount_rate",
        "currency",
        "payment_type",
        "rent_free_months",
        "escalation_rate",
        "legal_fees",
        "brokerage_fees",
        "other_initial_direct_costs",
        "cash_incentive",
        "rvg_amount",
        "rvg_guaranteed_by",
        "rvg_expected_payment",
        "cpi_index_base",
        "cpi_index_current",
        "cpi_adjustment_frequency_months",
        "non_lease_component",
        "non_lease_description",
        "practical_expedient_elected",
    ]

    yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    instr = (
        "Fill in your lease data below. Do not change column headers. "
        "Date format: YYYY-MM-DD. Rates as decimals: 8.5% = 0.085"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c1 = ws.cell(row=1, column=1, value=instr)
    c1.fill = yellow
    c1.font = Font(bold=True)
    c1.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 36

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)

    samples = [
        [
            "OFFICE-SAMPLE-001",
            "Corporate Office — Banjara Hills, Hyderabad",
            "Gnanova Technologies Pvt Ltd",
            "Prestige Estates Pvt Ltd",
            "2024-04-01",
            36,
            85000,
            0.085,
            "INR",
            "Arrears",
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            "None",
            0,
            0,
            0,
            12,
            0,
            "",
            "FALSE",
        ],
        [
            "DC-SAMPLE-001",
            "Data Centre Space — HITEC City, Hyderabad",
            "Gnanova Technologies Pvt Ltd",
            "CtrlS Datacenters Ltd",
            "2024-04-01",
            60,
            500000,
            0.105,
            "INR",
            "Arrears",
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            "None",
            0,
            100,
            107.5,
            12,
            75000,
            "Data centre cooling, physical security, UPS power",
            "FALSE",
        ],
        [
            "IBM-SAMPLE-001",
            "IBM Power Server Rack — HITEC City Data Centre",
            "Gnanova Technologies Pvt Ltd",
            "IBM India Pvt Ltd",
            "2024-07-01",
            60,
            250000,
            0.105,
            "INR",
            "Arrears",
            2,
            0.05,
            40000,
            35000,
            25000,
            50000,
            300000,
            "Lessee",
            150000,
            100,
            103,
            12,
            30000,
            "IBM hardware maintenance included in rent",
            "FALSE",
        ],
    ]

    for r_idx, row in enumerate(samples, start=3):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    gc.collect()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ifrs16_bulk_lease_template.xlsx"'},
    )


def _ifrs16_lease_liability(lease: Dict[str, Any]) -> float:
    res = lease.get("results") if isinstance(lease.get("results"), dict) else {}
    return float(
        lease.get("lease_liability")
        or lease.get("liability")
        or (res or {}).get("lease_liability")
        or 0
    )


def _ifrs16_rou_asset(lease: Dict[str, Any]) -> float:
    res = lease.get("results") if isinstance(lease.get("results"), dict) else {}
    return float(lease.get("rou_asset") or lease.get("rou") or (res or {}).get("rou_asset") or 0)


def _ifrs16_monthly_payment(lease: Dict[str, Any]) -> float:
    payments = lease.get("payments") if isinstance(lease.get("payments"), dict) else {}
    return float(lease.get("monthly_payment") or payments.get("monthly") or 0)


def _ifrs16_end_date(lease: Dict[str, Any]) -> str:
    dates = lease.get("dates") if isinstance(lease.get("dates"), dict) else {}
    return str(lease.get("end_date") or dates.get("end") or "")


def _ifrs16_lease_active(lease: Dict[str, Any]) -> bool:
    status = str(lease.get("status", "active")).lower()
    return status not in ("expired", "draft", "terminated", "churned")


@app.post("/api/ifrs16/cfo-insights")
async def generate_cfo_insights(req: IFRS16CFOInsightsRequest):
    """IFRS 16 CFO strategic insights — portfolio metrics + Claude analysis."""
    try:
        from datetime import date

        leases = req.leases
        today = date.today()

        if not leases:
            return {
                "success": True,
                "insights": [],
                "metrics": {},
                "risk_score": 0,
            }

        if not ANTHROPIC_API_KEY:
            raise HTTPException(
                status_code=503,
                detail="Claude API not configured. Set ANTHROPIC_API_KEY in .env",
            )

        total_liability = sum(_ifrs16_lease_liability(l) for l in leases)
        total_rou = sum(_ifrs16_rou_asset(l) for l in leases)
        total_annual_payments = sum(_ifrs16_monthly_payment(l) * 12 for l in leases)
        active_leases = [l for l in leases if _ifrs16_lease_active(l)]

        leverage_ratio = (
            total_liability / req.total_assets * 100 if req.total_assets > 0 else 0
        )
        lease_intensity = (
            total_annual_payments / req.annual_revenue * 100 if req.annual_revenue > 0 else 0
        )
        budget_variance = (
            total_annual_payments - req.budget_lease_cost if req.budget_lease_cost > 0 else 0
        )

        expiring_90: List[Dict[str, Any]] = []
        expiring_365: List[Dict[str, Any]] = []
        for l in active_leases:
            end_str = _ifrs16_end_date(l)
            if not end_str:
                continue
            try:
                end = datetime.strptime(end_str[:10], "%Y-%m-%d").date()
                days_left = (end - today).days
                if 0 < days_left <= 90:
                    expiring_90.append(l)
                elif 90 < days_left <= 365:
                    expiring_365.append(l)
            except Exception:
                continue

        top_leases = sorted(
            active_leases,
            key=lambda x: _ifrs16_lease_liability(x),
            reverse=True,
        )[:5]

        import anthropic

        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

        portfolio_summary = f"""
Lease Portfolio for CFO Analysis:
- Total Leases: {len(leases)} ({len(active_leases)} active)
- Total Lease Liability: ${total_liability:,.0f}
- Total ROU Assets: ${total_rou:,.0f}
- Annual Lease Payments: ${total_annual_payments:,.0f}
- Leases Expiring in 90 days: {len(expiring_90)}
- Leases Expiring in 12 months: {len(expiring_365)}
- Leverage Ratio (liability/assets): {leverage_ratio:.1f}%
- Lease Intensity (payments/revenue): {lease_intensity:.1f}%
- Budget Variance: ${budget_variance:+,.0f}

Top 5 Leases by Liability:
{chr(10).join([
    f"  - {l.get('title', l.get('contract_id', l.get('id', 'Unknown')))}: "
    f"${_ifrs16_lease_liability(l):,.0f} liability, "
    f"ends {_ifrs16_end_date(l) or 'N/A'}"
    for l in top_leases
])}

Expiring in 90 days:
{chr(10).join([
    f"  - {l.get('title', l.get('asset', 'Unknown'))}: "
    f"${_ifrs16_monthly_payment(l):,.0f}/month"
    for l in expiring_90
]) or "  None"}
"""

        ai_response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1500,
            messages=[
                {
                    "role": "user",
                    "content": f"""You are a CFO advisor analysing a company's lease portfolio under IFRS 16.

{portfolio_summary}

Generate exactly 4 strategic insights for the CFO.
Each insight must be specific, actionable, and reference actual numbers.

Format your response as JSON only — no other text:
{{
  "insights": [
    {{
      "id": "1",
      "category": "LEVERAGE" | "RENEWAL_RISK" | "COST_OPTIMISATION" | "BUDGET" | "CONCENTRATION_RISK",
      "severity": "HIGH" | "MEDIUM" | "LOW",
      "title": "Short title (max 8 words)",
      "finding": "One sentence with specific numbers",
      "recommendation": "One specific actionable recommendation",
      "impact": "Estimated financial impact or risk amount"
    }}
  ],
  "overall_health": "STRONG" | "ADEQUATE" | "AT_RISK" | "CRITICAL",
  "health_score": 0-100,
  "one_line_summary": "One sentence CFO-level portfolio summary"
}}""",
                }
            ],
        )

        ai_text = "".join(
            b.text for b in ai_response.content if getattr(b, "type", None) == "text"
        ).strip()
        if ai_text.startswith("```"):
            ai_text = ai_text.split("```")[1]
            if ai_text.startswith("json"):
                ai_text = ai_text[4:]
        ai_data = json.loads(ai_text.strip())

        metrics = {
            "total_liability": round(total_liability, 2),
            "total_rou_assets": round(total_rou, 2),
            "total_annual_payments": round(total_annual_payments, 2),
            "active_lease_count": len(active_leases),
            "expiring_90_days": len(expiring_90),
            "expiring_12_months": len(expiring_365),
            "leverage_ratio_pct": round(leverage_ratio, 2),
            "lease_intensity_pct": round(lease_intensity, 2),
            "budget_variance": round(budget_variance, 2),
            "top_leases_by_liability": [
                {
                    "title": l.get("title", l.get("asset", "Unknown")),
                    "liability": _ifrs16_lease_liability(l),
                    "monthly_payment": _ifrs16_monthly_payment(l),
                    "end_date": _ifrs16_end_date(l) or "N/A",
                }
                for l in top_leases
            ],
            "expiring_soon": [
                {
                    "title": l.get("title", l.get("asset", "Unknown")),
                    "end_date": _ifrs16_end_date(l) or "N/A",
                    "monthly_payment": _ifrs16_monthly_payment(l),
                }
                for l in expiring_90
            ],
        }

        gc.collect()
        return {
            "success": True,
            "insights": ai_data.get("insights", []),
            "overall_health": ai_data.get("overall_health", "ADEQUATE"),
            "health_score": ai_data.get("health_score", 50),
            "one_line_summary": ai_data.get("one_line_summary", ""),
            "metrics": metrics,
        }

    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse AI response: {str(e)}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== IFRS 15 Endpoints ====================

class IFRS15ExtractRequest(BaseModel):
    contract_text: str = Field(..., description="Contract text to extract")


class IFRS15ClauseDetectionRequest(BaseModel):
    contract_text: str


@app.post("/api/ifrs15/detect-clauses")
async def ifrs15_detect_clauses(request: IFRS15ClauseDetectionRequest):
    """Standalone IFRS 15 non-standard clause scan (same engine as upload)."""
    try:
        from ifrs15_extractor import detect_nonstandard_clauses

        return detect_nonstandard_clauses(request.contract_text, api_key=ANTHROPIC_API_KEY)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/extract")
async def ifrs15_extract(request: IFRS15ExtractRequest):
    """Extract IFRS 15 terms from pasted contract text"""
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=503, detail="Claude API not configured.")
    try:
        from ifrs15_extractor import IFRS15ContractExtractor, detect_nonstandard_clauses

        extractor = IFRS15ContractExtractor(api_key=ANTHROPIC_API_KEY)
        extracted_data = extractor.extract_contract_terms(request.contract_text)
        validation = extractor.validate_ifrs15_extraction(extracted_data)
        clause_detection = detect_nonstandard_clauses(request.contract_text, api_key=ANTHROPIC_API_KEY)
        gc.collect()
        return {
            "status": "success",
            "extracted_data": extracted_data,
            "validation": validation,
            "clause_detection": clause_detection,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/upload-contract")
async def ifrs15_upload_contract(file: UploadFile = File(...)):
    """
    Upload revenue contract (PDF, DOCX, TXT, XLSX) and extract IFRS 15 terms
    """
    if not ANTHROPIC_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="Claude API not configured. Set ANTHROPIC_API_KEY environment variable."
        )
    allowed_extensions = ['.pdf', '.docx', '.txt', '.xlsx', '.xls']
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"Unsupported file type. Allowed: {', '.join(allowed_extensions)}")
    try:
        file_id = str(uuid.uuid4())
        upload_path = UPLOAD_DIR / f"ifrs15_{file_id}_{file.filename}"
        with open(upload_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        from ifrs15_extractor import IFRS15ContractExtractor, detect_nonstandard_clauses, read_contract_file_to_text

        extractor = IFRS15ContractExtractor(api_key=ANTHROPIC_API_KEY)
        extracted_data = extractor.extract_from_file(str(upload_path))
        validation = extractor.validate_ifrs15_extraction(extracted_data)
        extraction_file = OUTPUT_DIR / f"ifrs15_extraction_{file_id}.json"
        extractor.save_extraction(extracted_data, str(extraction_file))
        try:
            contract_text = read_contract_file_to_text(str(upload_path))
        except Exception:
            contract_text = ""
        clause_detection = detect_nonstandard_clauses(contract_text, api_key=ANTHROPIC_API_KEY)
        gc.collect()
        return {
            "status": "success",
            "file_id": file_id,
            "filename": file.filename,
            "extracted_data": extracted_data,
            "validation": validation,
            "clause_detection": clause_detection,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/calculate")
async def ifrs15_calculate(request: IFRS15CalculateRequest):
    """
    Calculate IFRS 15 revenue recognition
    """
    try:
        from ifrs15_calculator import IFRS15Calculator, IFRS15Input, PerformanceObligation
        obligations = []
        for ob in request.performance_obligations:
            transfer_dt = None
            if ob.transfer_date:
                try:
                    transfer_dt = datetime.strptime(ob.transfer_date, "%Y-%m-%d")
                except ValueError:
                    pass
            obligations.append(PerformanceObligation(
                obligation_id=ob.obligation_id,
                description=ob.description,
                standalone_selling_price=Decimal(str(ob.standalone_selling_price)),
                recognition_method=ob.recognition_method,
                duration_months=ob.duration_months,
                transfer_date=transfer_dt
            ))
        contract = IFRS15Input(
            contract_id=request.contract_id,
            customer_name=request.customer_name,
            vendor_name=request.vendor_name,
            effective_date=datetime.strptime(request.effective_date, "%Y-%m-%d"),
            contract_term_months=request.contract_term_months,
            fixed_consideration=Decimal(str(request.fixed_consideration)),
            variable_consideration=Decimal(str(request.variable_consideration)),
            variable_consideration_constrained=Decimal(str(request.variable_consideration_constrained)),
            constraint_percentage=Decimal(str(request.constraint_percentage)),
            constraint_method=request.constraint_method,
            vc_constraint_factors=dict(request.vc_constraint_factors or {}),
            discounts=Decimal(str(request.discounts)),
            rebates=Decimal(str(request.rebates)),
            financing_adjustment=Decimal(str(request.financing_adjustment)),
            currency=request.currency,
            contract_type=request.contract_type,
            hourly_rate=Decimal(str(request.hourly_rate)),
            hours_worked=Decimal(str(request.hours_worked)),
            tm_cap=Decimal(str(request.tm_cap)),
            cumulative_billed=Decimal(str(request.cumulative_billed)),
            total_estimated_cost=Decimal(str(request.total_estimated_cost)),
            actual_cost_to_date=Decimal(str(request.actual_cost_to_date)),
            prior_revenue_recognised=Decimal(str(request.prior_revenue_recognised)),
            maintenance_term_months=request.maintenance_term_months,
            volume_slabs=request.volume_slabs,
            estimated_annual_volume=Decimal(str(request.estimated_annual_volume)),
            can_estimate_volume=request.can_estimate_volume,
            sla_items=request.sla_items,
            performance_obligations=obligations,
            payment_terms=(request.payment_terms or "").strip(),
        )
        calc = IFRS15Calculator()
        results = calc.calculate_full_ifrs15(contract, cash_received=Decimal(str(request.cash_received)))
        results_json = results.copy()
        if 'recognition_schedule' in results_json and hasattr(results_json['recognition_schedule'], 'to_dict'):
            results_json['recognition_schedule'] = results_json['recognition_schedule'].to_dict(orient='records')
        file_id = str(uuid.uuid4())
        safe_id = "".join(c for c in request.contract_id if c.isalnum() or c in "-_")[:30]
        excel_path = OUTPUT_DIR / f"IFRS15_{safe_id}_{file_id}.xlsx"
        try:
            from ifrs15_excel_export import IFRS15ExcelExporter
            exporter = IFRS15ExcelExporter()
            exporter.export_ifrs15_workbook(
                results=results_json,
                contract_meta={
                    "contract_id": request.contract_id,
                    "customer_name": request.customer_name,
                    "effective_date": request.effective_date,
                    "contract_term_months": request.contract_term_months,
                    "currency": request.currency,
                },
                filename=str(excel_path),
            )
        except Exception:
            pass
        tr = float(results_json.get("total_recognised") or 0)
        td = float(results_json.get("total_deferred") or 0)
        tpv = float(results_json.get("transaction_price") or results_json.get("total_contract_value") or 0)
        _ifrs15_audit_append(
            "CALCULATE",
            request.contract_id,
            f"Revenue calculation — ${tr:,.2f} recognised, ${td:,.2f} deferred",
            {},
            {"total_recognised": tr, "total_deferred": td, "transaction_price": tpv},
            "IFRS 15 — 5-Step Model",
            changed_amount=tr,
        )
        gc.collect()
        return {
            "status": "success",
            "contract_id": request.contract_id,
            "results": results_json,
            "excel_file_id": file_id
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/download-excel")
async def ifrs15_download_excel(request: IFRS15DownloadExcelRequest):
    """Generate IFRS 15 six-sheet audit workbook (master summary + five core sheets) from calculation results."""
    try:
        from ifrs15_excel_export import IFRS15ExcelExporter

        file_id = str(uuid.uuid4())
        safe_id = "".join(c for c in request.contract_id if c.isalnum() or c in "-_")[:30]
        excel_path = OUTPUT_DIR / f"IFRS15_{safe_id}_{file_id}.xlsx"

        exporter = IFRS15ExcelExporter()
        res = request.results or {}
        exporter.export_ifrs15_workbook(
            results=res,
            contract_meta={
                "contract_id": request.contract_id,
                "customer_name": request.customer_name or "",
                "effective_date": request.effective_date or "",
                "contract_term_months": request.contract_term_months or 0,
                "currency": request.currency or "USD",
                "variable_consideration_assessment": res.get("variable_consideration_assessment"),
                "rpo_disclosure": res.get("rpo_disclosure"),
                "contract_costs_assessment": res.get("contract_costs_assessment"),
                "principal_agent_assessment": res.get("principal_agent_assessment"),
                "license_classification": res.get("license_classification"),
                "deferred_revenue_rollforward": res.get("deferred_revenue_rollforward"),
                "rpo_disclosure_ifrs120": res.get("rpo_disclosure_ifrs120"),
                "principal_agent_history": res.get("principal_agent_history"),
                "contract_costs_ifrs9194": res.get("contract_costs_ifrs9194"),
                "licenses_ip_export": res.get("licenses_ip_export"),
                "audit_entries": res.get("audit_entries"),
                "material_rights_ifrs1540": res.get("material_rights_ifrs1540"),
                "warranties_ifrs1528": res.get("warranties_ifrs1528"),
                "bill_and_hold_ifrs1579": res.get("bill_and_hold_ifrs1579"),
                "financing_component_ifrs1560": res.get("financing_component_ifrs1560"),
                "tp_adjustments_ifrs1566": res.get("tp_adjustments_ifrs1566"),
            },
            filename=str(excel_path),
            master_report=request.master_report_data,
        )

        gc.collect()
        return {"status": "success", "file_id": file_id, "filename": excel_path.name}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/modification")
async def ifrs15_modification(request: IFRS15ModificationRequest):
    """Assess IFRS 15 contract modification type (IFRS 15.18–21)."""
    try:
        from ifrs15_calculator import ContractModification, IFRS15Calculator

        mod = ContractModification(
            original_contract_id=request.original_contract_id,
            modification_date=request.modification_date,
            modification_description=request.modification_description,
            new_goods_services=list(request.new_goods_services or []),
            price_change=float(request.price_change),
            remaining_transaction_price=float(request.remaining_transaction_price),
            remaining_performance_obligations=list(request.remaining_performance_obligations or []),
            original_ssps=dict(request.original_ssps or {}),
        )
        calculator = IFRS15Calculator()
        result = calculator.assess_modification(mod)
        _ifrs15_audit_append(
            "MODIFICATION",
            request.original_contract_id,
            f"Modification assessed — {result.get('modification_type_name') or result.get('modification_type_label') or ''}",
            {},
            {
                "modification_type": result.get("modification_type"),
                "catch_up_amount": float(result.get("catch_up_amount") or 0),
            },
            "IFRS 15.18-21",
            changed_amount=float(result.get("catch_up_amount") or 0),
        )
        return {"success": True, "modification": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/deferred-revenue-rollforward")
async def ifrs15_deferred_revenue_rollforward(req: DeferredRevenueRequest):
    try:
        from ifrs15_calculator import DeferredRevenueInput, IFRS15Calculator

        data = DeferredRevenueInput(
            period=req.period,
            opening_balance=req.opening_balance,
            new_bookings=req.new_bookings,
            revenue_released=req.revenue_released,
            cancellations=req.cancellations,
            modifications_impact=req.modifications_impact,
            fx_impact=req.fx_impact,
            gl_closing_balance=req.gl_closing_balance,
            currency=req.currency,
        )
        calculator = IFRS15Calculator()
        result = calculator.deferred_revenue_rollforward(data)
        cid = f"DEFERRED-{req.period}"
        _ifrs15_audit_append(
            "DEFERRED_REC",
            cid,
            f"Deferred revenue roll-forward — period {req.period}, reconciled={result.get('reconciled')}",
            {},
            {"variance": float(result.get("variance") or 0), "gl_closing_balance": float(req.gl_closing_balance)},
            "IFRS 15.116",
            changed_amount=float(result.get("variance") or 0),
        )
        return {"success": True, "rollforward": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/variable-consideration")
async def ifrs15_variable_consideration(request: IFRS15VariableConsiderationRequest):
    """Estimate and constrain variable consideration (IFRS 15.50–58)."""
    if len(request.constraint_factors) != 5:
        raise HTTPException(status_code=400, detail="constraint_factors must have exactly 5 values")
    try:
        from ifrs15_calculator import IFRS15VariableConsiderationEngine

        body: Dict[str, Any] = {
            "method": request.method,
            "scenarios": [s.model_dump() for s in request.scenarios],
            "constraint_factors": list(request.constraint_factors),
        }
        if request.total_contract_value is not None:
            body["total_contract_value"] = request.total_contract_value
        engine = IFRS15VariableConsiderationEngine()
        return engine.estimate(body)
    except ValueError as e:
        msg = str(e)
        if "Probabilities must sum" in msg or "100%" in msg:
            raise HTTPException(status_code=400, detail="Probabilities must sum to 100%")
        raise HTTPException(status_code=400, detail=msg)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/reversal-risk")
async def ifrs15_reversal_risk(request: IFRS15ReversalRiskRequest):
    """Revenue reversal risk score from VC constraint and contract profile."""
    try:
        from ifrs15_calculator import IFRS15ReversalRiskEngine

        engine = IFRS15ReversalRiskEngine()
        return engine.score(request.model_dump())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/rpo")
async def ifrs15_rpo(request: IFRS15RPORequest):
    """Remaining performance obligations disclosure (IFRS 15.120–122)."""
    try:
        if request.contracts:
            from ifrs15_calculator import IFRS15Calculator, RPOContract

            contracts = [
                RPOContract(
                    contract_id=c.contract_id,
                    customer_name=c.customer_name,
                    contract_start=c.contract_start,
                    contract_end=c.contract_end,
                    total_transaction_price=c.total_transaction_price,
                    revenue_recognised_to_date=c.revenue_recognised_to_date,
                    performance_obligations=[po.model_dump() for po in c.performance_obligations],
                    practical_expedient_applied=c.practical_expedient_applied,
                )
                for c in request.contracts
            ]
            calculator = IFRS15Calculator()
            result = calculator.calculate_rpo(contracts)
            ac = (
                request.contracts[0].contract_id
                if request.contracts
                else (request.contract_id or "MULTI")
            )
            _ifrs15_audit_append(
                "RPO",
                ac,
                f"RPO disclosure (IFRS 15.120–122) — total RPO ${float(result.get('total_rpo') or 0):,.2f}",
                {},
                {"total_rpo": result.get("total_rpo"), "contract_count": result.get("contract_count")},
                "IFRS 15.120-122",
                changed_amount=float(result.get("total_rpo") or 0),
            )
            return {"success": True, "rpo": result}

        if not request.obligations:
            raise HTTPException(
                status_code=400,
                detail="Provide either contracts (IFRS 15.120–122) or obligations (legacy RPO engine).",
            )

        from ifrs15_calculator import IFRS15RPOEngine

        body = {"obligations": [o.model_dump() for o in request.obligations]}
        leg = IFRS15RPOEngine().calculate_rpo(body)
        _ifrs15_audit_append(
            "RPO",
            request.contract_id or "N/A",
            "RPO disclosure (legacy obligations engine)",
            {},
            {"total_rpo": leg.get("total_rpo"), "keys": list(leg.keys())[:12]},
            "IFRS 15.120-122",
            changed_amount=float(leg.get("total_rpo") or 0),
        )
        return leg
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/contract-costs")
async def ifrs15_contract_costs(request: IFRS15ContractCostsRequest):
    """Costs to obtain / fulfil contracts (IFRS 15.91–95) — batch lines or legacy commission engine."""
    try:
        if request.costs:
            from ifrs15_calculator import IFRS15Calculator, ContractCostInput

            inputs = [
                ContractCostInput(
                    cost_id=c.cost_id,
                    contract_id=c.contract_id,
                    description=c.description,
                    cost_type=c.cost_type,
                    cost_amount=float(c.cost_amount),
                    incurred_date=c.incurred_date,
                    contract_start=c.contract_start,
                    contract_end=c.contract_end,
                    expected_renewal=bool(c.expected_renewal),
                    expected_renewal_months=int(c.expected_renewal_months or 0),
                    currency=c.currency or "USD",
                )
                for c in request.costs
            ]
            calculator = IFRS15Calculator()
            result = calculator.calculate_contract_costs(inputs)
            cid = request.costs[0].contract_id if request.costs else "MULTI"
            summ = result.get("summary") or {}
            _ifrs15_audit_append(
                "CONTRACT_COSTS",
                cid,
                (
                    f"Contract costs assessed — capitalised ${float(summ.get('total_capitalised', 0) or 0):,.2f}, "
                    f"expensed ${float(summ.get('total_expensed_immediately', 0) or 0):,.2f}"
                ),
                {},
                dict(summ),
                "IFRS 15.91-94",
                changed_amount=float(summ.get("total_capitalised", 0) or 0),
            )
            return {"success": True, "contract_costs": result}

        if (
            request.commission_amount is None
            or request.contract_term_months is None
            or request.contract_total_value is None
        ):
            raise HTTPException(
                status_code=400,
                detail="Provide either costs[] (IFRS 15.91–95 batch) or commission_amount, contract_term_months, and contract_total_value (legacy).",
            )

        from ifrs15_calculator import IFRS15ContractCostsEngine

        body = {
            "commission_amount": request.commission_amount,
            "contract_term_months": request.contract_term_months,
            "contract_total_value": request.contract_total_value,
        }
        legacy = IFRS15ContractCostsEngine().calculate(body)
        _ifrs15_audit_append(
            "CONTRACT_COSTS",
            request.contract_id or "N/A",
            (
                f"Legacy commission asset — capitalise={legacy.get('capitalise')}, "
                f"amount ${float(legacy.get('commission_amount', 0) or 0):,.2f}"
            ),
            {},
            {"capitalise": legacy.get("capitalise"), "commission_amount": legacy.get("commission_amount")},
            "IFRS 15.91-94",
            changed_amount=float(legacy.get("commission_amount") or 0),
        )
        return legacy
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/principal-agent")
async def ifrs15_principal_agent(request: IFRS15PrincipalAgentRequest):
    """Principal vs agent (gross vs net) assessment (IFRS 15.B34–B38)."""
    try:
        ext_id = (request.arrangement_id or "").strip()
        if ext_id:
            missing = []
            if request.gross_contract_value is None:
                missing.append("gross_contract_value")
            if request.third_party_cost is None:
                missing.append("third_party_cost")
            for fld, val in [
                ("controls_before_transfer", request.controls_before_transfer),
                ("primary_obligor", request.primary_obligor),
                ("inventory_risk", request.inventory_risk),
                ("pricing_discretion", request.pricing_discretion),
                ("credit_risk", request.credit_risk),
            ]:
                if val is None:
                    missing.append(fld)
            if missing:
                raise HTTPException(
                    status_code=400,
                    detail=f"Extended principal-agent assessment missing: {', '.join(missing)}",
                )
            from ifrs15_calculator import IFRS15Calculator, PrincipalAgentInput

            data = PrincipalAgentInput(
                arrangement_id=ext_id,
                description=request.description or "",
                third_party_involved=bool(request.third_party_involved),
                gross_contract_value=float(request.gross_contract_value),
                third_party_cost=float(request.third_party_cost),
                controls_before_transfer=bool(request.controls_before_transfer),
                primary_obligor=bool(request.primary_obligor),
                inventory_risk=bool(request.inventory_risk),
                pricing_discretion=bool(request.pricing_discretion),
                credit_risk=bool(request.credit_risk),
            )
            calculator = IFRS15Calculator()
            result = calculator.assess_principal_agent(data)
            _ifrs15_audit_append(
                "PRINCIPAL_AGENT",
                ext_id,
                f"Principal vs agent — {result.get('conclusion')} ({result.get('revenue_treatment')})",
                {},
                {"conclusion": result.get("conclusion"), "revenue_treatment": result.get("revenue_treatment")},
                "IFRS 15.B34-B38",
                changed_amount=float(result.get("gross_contract_value") or 0),
            )
            return {"success": True, "assessment": result}

        if (
            request.transaction_price is None
            or request.cost_paid_to_supplier is None
            or request.obtains_before_transfer is None
            or request.sets_price_independently is None
            or request.primarily_responsible is None
        ):
            raise HTTPException(
                status_code=400,
                detail="Legacy principal-agent requires transaction_price, cost_paid_to_supplier, and three boolean indicators.",
            )

        from ifrs15_calculator import IFRS15PrincipalAgentEngine

        legacy_body = {
            "transaction_price": request.transaction_price,
            "cost_paid_to_supplier": request.cost_paid_to_supplier,
            "obtains_before_transfer": request.obtains_before_transfer,
            "sets_price_independently": request.sets_price_independently,
            "primarily_responsible": request.primarily_responsible,
            "contract_id": request.contract_id,
        }
        leg = IFRS15PrincipalAgentEngine().assess(legacy_body)
        _ifrs15_audit_append(
            "PRINCIPAL_AGENT",
            request.contract_id or "N/A",
            f"Principal vs agent (legacy) — {leg.get('role') or leg.get('conclusion') or leg.get('basis', '')}",
            {},
            {"transaction_price": request.transaction_price, "payload_keys": list(leg.keys())[:12]},
            "IFRS 15.B34-B38",
            changed_amount=float(request.transaction_price or 0),
        )
        return leg
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/license-classification")
async def ifrs15_license_classification(request: IFRS15LicenseRequest):
    """Licence of IP: right to access vs right to use (IFRS 15.B52–B63)."""
    try:
        from ifrs15_calculator import IFRS15LicenseEngine

        return IFRS15LicenseEngine().classify(request.model_dump())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/licenses-ip")
async def ifrs15_licenses_ip(req: LicensesIPRequest):
    """IFRS 15.B52–B63 — licence of IP (calculator assessment; batch)."""
    try:
        from ifrs15_calculator import IFRS15Calculator, LicenseIPInput

        calculator = IFRS15Calculator()
        results: List[Dict[str, Any]] = []
        for lic in req.licenses:
            data = LicenseIPInput(
                license_id=lic.license_id,
                product_name=lic.product_name,
                license_description=lic.license_description or "",
                license_fee=float(lic.license_fee),
                license_start=lic.license_start,
                license_end=lic.license_end or "",
                is_perpetual=bool(lic.is_perpetual),
                entity_activities_affect_ip=bool(lic.entity_activities_affect_ip),
                customer_exposed_to_effect=bool(lic.customer_exposed_to_effect),
                no_separate_functional_utility=bool(lic.no_separate_functional_utility),
                currency=lic.currency or "USD",
            )
            results.append(calculator.assess_license_ip(data))
        summary = {
            "total_licenses": len(results),
            "right_to_access": sum(1 for r in results if r.get("license_type") == "RIGHT_TO_ACCESS"),
            "right_to_use": sum(1 for r in results if r.get("license_type") == "RIGHT_TO_USE"),
            "judgement_required": sum(1 for r in results if r.get("license_type") == "JUDGEMENT_REQUIRED"),
        }
        cid = req.licenses[0].license_id if req.licenses else "N/A"
        _ifrs15_audit_append(
            "LICENSE_IP",
            cid,
            f"Licences of IP assessed — {summary['total_licenses']} licence(s), "
            f"RTA {summary['right_to_access']}, RTU {summary['right_to_use']}, judgement {summary['judgement_required']}",
            {},
            summary,
            "IFRS 15.B52-B63",
            changed_amount=float(sum(float(r.get("license_fee", 0) or 0) for r in results)),
        )
        return {"success": True, "licenses": results, "summary": summary}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/material-rights")
async def ifrs15_material_rights(req: CustomerOptionsRequest):
    """IFRS 15.B40–B43 — customer options / material rights (batch)."""
    try:
        from ifrs15_calculator import IFRS15Calculator, CustomerOptionInput

        calculator = IFRS15Calculator()
        results: List[Dict[str, Any]] = []
        for opt in req.options:
            data = CustomerOptionInput(
                option_id=opt.option_id,
                contract_id=opt.contract_id,
                description=opt.description,
                option_type=opt.option_type,
                original_contract_value=float(opt.original_contract_value),
                original_ssp=float(opt.original_ssp),
                option_price=float(opt.option_price),
                option_ssp=float(opt.option_ssp),
                exercise_probability=float(opt.exercise_probability),
                points_granted=float(opt.points_granted or 0),
                point_value=float(opt.point_value or 0),
                currency=opt.currency or "USD",
            )
            result = calculator.assess_material_right(data)
            results.append(result)
            _ifrs15_audit_append(
                "MATERIAL_RIGHTS",
                opt.contract_id or "N/A",
                f"Material rights — {opt.description}: "
                f"{'SEPARATE PO identified' if result.get('material_right_exists') else 'No material right'}",
                {},
                {
                    "material_right": result.get("material_right_exists"),
                    "allocated_to_option": result.get("allocated_to_option", 0),
                    "option_id": opt.option_id,
                },
                "IFRS 15.B40-B43",
                changed_amount=float(result.get("allocated_to_option") or 0),
            )
        material_right_count = sum(1 for r in results if r.get("material_right_exists"))
        total_deferred = sum(float(r.get("allocated_to_option") or 0) for r in results)
        summary = {
            "total_options_assessed": len(results),
            "material_rights_found": material_right_count,
            "no_material_right": len(results) - material_right_count,
            "total_deferred_to_options": round(total_deferred, 2),
        }
        return {"success": True, "options": results, "summary": summary}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/warranties")
async def classify_warranties(req: WarrantiesRequest):
    """IFRS 15.B28–B33 / IAS 37 — warranty classification (batch)."""
    try:
        from ifrs15_calculator import IFRS15Calculator, WarrantyInput

        calculator = IFRS15Calculator()
        results: List[Dict[str, Any]] = []
        for w in req.warranties:
            data = WarrantyInput(
                warranty_id=w.warranty_id,
                contract_id=w.contract_id,
                product_description=w.product_description or "",
                warranty_description=w.warranty_description,
                warranty_period_months=int(w.warranty_period_months),
                warranty_value=float(w.warranty_value),
                required_by_law=bool(w.required_by_law),
                covers_specs_only=bool(w.covers_specs_only),
                customer_can_purchase_separately=bool(w.customer_can_purchase_separately),
                provides_additional_service=bool(w.provides_additional_service),
                allocated_fee=float(w.allocated_fee or 0),
                currency=w.currency or "USD",
            )
            result = calculator.classify_warranty(data)
            results.append(result)
            _ifrs15_audit_append(
                "WARRANTY",
                w.contract_id or "N/A",
                (
                    f"Warranty classification — {w.warranty_description}: "
                    f"{result['warranty_type']} ({result['accounting_standard']})"
                ),
                {},
                {
                    "type": result["warranty_type"],
                    "standard": result["accounting_standard"],
                    "deferred": result["deferred_revenue_amount"],
                },
                "IFRS 15.B28-B33",
                changed_amount=float(result["deferred_revenue_amount"] or 0),
            )
        by_id = {str(r.get("warranty_id")): r for r in results}
        total_ias37 = sum(
            float(w.warranty_value)
            for w in req.warranties
            if by_id.get(w.warranty_id, {}).get("warranty_type") == "ASSURANCE"
        )
        return {
            "success": True,
            "warranties": results,
            "summary": {
                "total": len(results),
                "assurance_type": sum(1 for r in results if r.get("warranty_type") == "ASSURANCE"),
                "service_type": sum(1 for r in results if r.get("warranty_type") == "SERVICE"),
                "judgement_required": sum(
                    1 for r in results if r.get("warranty_type") == "JUDGEMENT_REQUIRED"
                ),
                "total_ias37_provision": round(total_ias37, 2),
                "total_ifrs15_deferred": round(
                    sum(float(r.get("deferred_revenue_amount") or 0) for r in results), 2
                ),
            },
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/bill-and-hold")
async def assess_bill_and_hold(req: BillAndHoldsRequest):
    """IFRS 15.B79–B82 — bill-and-hold arrangements (batch)."""
    try:
        from ifrs15_calculator import IFRS15Calculator, BillAndHoldInput

        calculator = IFRS15Calculator()
        results: List[Dict[str, Any]] = []
        for arr in req.arrangements:
            data = BillAndHoldInput(
                arrangement_id=arr.arrangement_id,
                contract_id=arr.contract_id,
                customer_name=arr.customer_name,
                product_description=arr.product_description,
                contract_value=float(arr.contract_value),
                expected_delivery_date=arr.expected_delivery_date,
                billing_date=arr.billing_date,
                reason_is_substantive=bool(arr.reason_is_substantive),
                product_separately_identified=bool(arr.product_separately_identified),
                product_ready_for_transfer=bool(arr.product_ready_for_transfer),
                entity_cannot_redirect=bool(arr.entity_cannot_redirect),
                currency=arr.currency or "USD",
            )
            result = calculator.assess_bill_and_hold(data)
            results.append(result)
            _ifrs15_audit_append(
                "BILL_AND_HOLD",
                arr.contract_id or "N/A",
                (
                    f"Bill-and-hold assessment — {arr.product_description}: "
                    f"{result['conclusion']} ({result['criteria_met_count']}/4 criteria met)"
                ),
                {},
                {
                    "conclusion": result["conclusion"],
                    "criteria_met": result["criteria_met_count"],
                    "deferred": result["revenue_deferred"],
                },
                "IFRS 15.B79-B82",
                changed_amount=float(result["revenue_deferred"] or 0),
            )
        return {
            "success": True,
            "arrangements": results,
            "summary": {
                "total": len(results),
                "recognisable_now": sum(
                    1 for r in results if r.get("conclusion") == "REVENUE_RECOGNISABLE"
                ),
                "deferred_until_delivery": sum(
                    1 for r in results if r.get("conclusion") == "DEFER_UNTIL_DELIVERY"
                ),
                "total_revenue_deferred": round(
                    sum(float(r.get("revenue_deferred") or 0) for r in results), 2
                ),
            },
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/financing-component")
async def financing_component_calculate(req: FinancingComponentsRequest):
    """IFRS 15.60–65 — significant financing component (batch)."""
    try:
        from ifrs15_calculator import IFRS15Calculator, FinancingComponentInput

        calculator = IFRS15Calculator()
        results: List[Dict[str, Any]] = []
        for c in req.contracts:
            data = FinancingComponentInput(
                contract_id=c.contract_id,
                description=c.description or "",
                contract_value=float(c.contract_value),
                payment_date=c.payment_date,
                transfer_date=c.transfer_date,
                payment_timing=c.payment_timing,
                discount_rate=float(c.discount_rate),
                currency=c.currency or "USD",
            )
            result = calculator.calculate_financing_component(data)
            results.append(result)
            _ifrs15_audit_append(
                "FINANCING_COMPONENT",
                c.contract_id or "N/A",
                (
                    f"Financing component ({c.payment_timing}) — "
                    f"PV ${float(result.get('pv_of_payment') or 0):,.2f} vs nominal "
                    f"${float(c.contract_value):,.2f}; "
                    f"financing ${float(result.get('financing_amount') or 0):,.2f}"
                ),
                {"nominal": float(c.contract_value)},
                {
                    "revenue": result.get("revenue_amount"),
                    "financing": result.get("financing_amount"),
                    "type": result.get("financing_type"),
                },
                "IFRS 15.60-65",
                changed_amount=float(result.get("financing_amount") or 0),
            )
        expedient_applied = sum(1 for r in results if r.get("practical_expedient_applied"))
        return {
            "success": True,
            "contracts": results,
            "summary": {
                "total": len(results),
                "expedient_applied": expedient_applied,
                "financing_adjusted": len(results) - expedient_applied,
                "total_financing_amount": round(
                    sum(float(r.get("financing_amount") or 0) for r in results), 2
                ),
                "interest_income_total": round(
                    sum(
                        float(r.get("financing_amount") or 0)
                        for r in results
                        if r.get("financing_type") == "INTEREST_INCOME"
                    ),
                    2,
                ),
                "interest_expense_total": round(
                    sum(
                        float(r.get("financing_amount") or 0)
                        for r in results
                        if r.get("financing_type") == "INTEREST_EXPENSE"
                    ),
                    2,
                ),
            },
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/transaction-price-adjustments")
async def transaction_price_adjustments(req: TransactionPriceAdjustmentsRequest):
    """IFRS 15.66–72 — non-cash consideration and consideration payable (transaction price)."""
    try:
        from ifrs15_calculator import (
            IFRS15Calculator,
            NonCashConsiderationInput,
            ConsiderationPayableInput,
        )

        calculator = IFRS15Calculator()
        result: Dict[str, Any] = {}

        if req.non_cash_items:
            nc_inputs = [
                NonCashConsiderationInput(**i.model_dump()) for i in req.non_cash_items
            ]
            nc_result = calculator.calculate_non_cash_consideration(nc_inputs)
            result["non_cash"] = nc_result
            cid0 = req.non_cash_items[0].contract_id if req.non_cash_items else "N/A"
            tp_nc = float(nc_result.get("total_tp_from_non_cash") or 0)
            _ifrs15_audit_append(
                "NON_CASH_CONSIDERATION",
                cid0,
                f"Non-cash consideration: ${tp_nc:,.2f} added to transaction price",
                {},
                {"total_tp_from_non_cash": tp_nc, "items": len(nc_result.get("items") or [])},
                "IFRS 15.66-69",
                changed_amount=tp_nc,
            )

        if req.consideration_payable_items:
            cp_inputs = [
                ConsiderationPayableInput(**i.model_dump()) for i in req.consideration_payable_items
            ]
            cp_result = calculator.calculate_consideration_payable(cp_inputs)
            result["consideration_payable"] = cp_result
            cid1 = req.consideration_payable_items[0].contract_id if req.consideration_payable_items else "N/A"
            tr = float(cp_result.get("total_revenue_reduction") or 0)
            tc = float(cp_result.get("total_cost_recognition") or 0)
            _ifrs15_audit_append(
                "CONSIDERATION_PAYABLE",
                cid1,
                f"Consideration payable: ${tr:,.2f} reduces revenue; ${tc:,.2f} as cost",
                {},
                {
                    "total_revenue_reduction": tr,
                    "total_cost_recognition": tc,
                    "items": len(cp_result.get("items") or []),
                },
                "IFRS 15.70-72",
                changed_amount=tr,
            )

        return {"success": True, "result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class TPChangeRequest(BaseModel):
    contract_id: str = ""
    adjustment_reason: str = "variable_consideration"
    original_transaction_price: float
    new_transaction_price: float
    revenue_recognised_to_date: float = 0
    remaining_performance_obligations: int = 1
    adjustment_method: str = "cumulative_catchup"
    adjustment_date: str = ""


@app.post("/api/ifrs15/tp-adjustments")
async def tp_adjustments_change(req: TPChangeRequest):
    """IFRS 15 — cumulative catch-up vs prospective transaction price changes."""
    try:
        from ifrs15_tp_change import calculate_tp_adjustment

        result = calculate_tp_adjustment(req.model_dump())
        _ifrs15_audit_append(
            "TP_ADJUSTMENT",
            req.contract_id or "N/A",
            f"TP adjustment {result['adjustment_amount']:,.2f} ({result['method_used']})",
            {},
            result,
            "IFRS 15.87-90",
            changed_amount=float(result.get("current_period_impact") or 0),
        )
        return {"success": True, **result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/portfolio/add")
async def add_portfolio_contract(req: PortfolioContractRequest, request: Request):
    """Upsert a contract into the IFRS 15 portfolio store."""
    firm_id = _ifrs15_firm_id(request)
    try:
        if req.contract_type == "real_estate_off_plan":
            tp = float(req.total_tp or 0)
            cmp_pct = req.construction_completion_pct
            if cmp_pct is None and tp > 0:
                cmp_pct = (float(req.recognised_to_date or 0) / tp) * 100.0
            gate = _rera_escrow_gate(
                req.escrow_receipts,
                req.escrow_releases,
                float(cmp_pct if cmp_pct is not None else 0),
                tp,
            )
            if gate:
                return gate
        contract_dict = req.model_dump()
        buyer_id_raw = str(contract_dict.get("buyer_id") or "").strip()
        if buyer_id_raw:
            contract_dict["buyer_id_hash"] = hashlib.sha256(buyer_id_raw.encode("utf-8")).hexdigest()
            contract_dict["buyer_id_masked"] = f"EID-****-{buyer_id_raw[-4:]}" if len(buyer_id_raw) >= 4 else "EID-****-0000"
            contract_dict.pop("buyer_id", None)
        summary = _ifrs15_portfolio_summary_data(contract_dict)
        contract_name = req.customer_name or req.contract_id
        existing_row = ifrs15_db.find_portfolio_by_business_contract_id(firm_id, req.contract_id)
        portfolio_row_id: Optional[str] = None
        if existing_row:
            portfolio_row_id = str(existing_row["id"])
            ifrs15_db.update_contract(
                firm_id,
                portfolio_row_id,
                contract_data=contract_dict,
                summary_data=summary,
                contract_name=contract_name,
            )
        else:
            row = ifrs15_db.add_contract(firm_id, contract_name, contract_dict, summary)
            portfolio_row_id = str(row.get("id", ""))
        portfolio_size = len(_ifrs15_portfolio_contracts_from_db(firm_id))
        _ifrs15_audit_append(
            "PORTFOLIO_ADD",
            req.contract_id,
            f"Portfolio upsert — {req.customer_name} ({req.contract_type})",
            {},
            {"portfolio_size": portfolio_size, "arr": float(req.arr or 0)},
            "IFRS 15 Portfolio",
            changed_amount=float(req.total_tp or 0),
            firm_id=firm_id,
            portfolio_row_id=portfolio_row_id,
        )
        return {"success": True, "portfolio_size": portfolio_size}
    except Exception as e:
        raise HTTPException(status_code=500, detail={"error": "Database error", "detail": str(e)})


@app.post("/api/ifrs15/realestate/portfolio")
async def ifrs15_realestate_portfolio(req: PortfolioContractRequest, http_request: Request):
    """Real estate portfolio — RERA escrow Art. 8 hard block; same persistence as /portfolio/add."""
    return await add_portfolio_contract(req, http_request)


@app.get("/api/ifrs15/portfolio/summary")
async def get_portfolio_summary(request: Request):
    """Aggregated SaaS / revenue metrics across portfolio contracts."""
    try:
        return _ifrs15_portfolio_saas_summary(_ifrs15_firm_id(request))
    except Exception as e:
        raise HTTPException(status_code=500, detail={"error": "Database error", "detail": str(e)})


@app.delete("/api/ifrs15/portfolio/{contract_id}")
async def remove_portfolio_contract(contract_id: str, request: Request):
    firm_id = _ifrs15_firm_id(request)
    try:
        before = len(_ifrs15_portfolio_contracts_from_db(firm_id))
        deleted = ifrs15_db.delete_contract_by_business_id(firm_id, contract_id)
        if not deleted:
            raise HTTPException(status_code=404, detail=f"{contract_id} not found")
        portfolio_size = len(_ifrs15_portfolio_contracts_from_db(firm_id))
        _ifrs15_audit_append(
            "PORTFOLIO_REMOVE",
            contract_id,
            "Removed contract from portfolio",
            {"portfolio_size": before},
            {"portfolio_size": portfolio_size},
            "IFRS 15 Portfolio",
            changed_amount=0.0,
            firm_id=firm_id,
        )
        return {"success": True, "removed": contract_id, "portfolio_size": portfolio_size}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail={"error": "Database error", "detail": str(e)})


@app.get("/api/ifrs15/portfolio/export-excel")
async def export_portfolio_excel(request: Request):
    """Download portfolio workbook (summary KPIs + contract detail)."""
    from ifrs15_excel_export import IFRS15ExcelExporter

    try:
        payload = _ifrs15_portfolio_saas_summary(_ifrs15_firm_id(request))
    except Exception as e:
        raise HTTPException(status_code=500, detail={"error": "Database error", "detail": str(e)})
    contracts = payload.get("contracts") or []
    if not contracts:
        raise HTTPException(status_code=400, detail="Portfolio is empty")
    summary = payload.get("summary") or {}
    buf = IFRS15ExcelExporter().export_portfolio_saas_to_bytes(summary, contracts)
    fn = f"IFRS15_Portfolio_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


@app.get("/api/ifrs15/audit-log")
async def get_ifrs15_audit_log(
    request: Request,
    contract_id: str = Query("", description="Filter by contract id"),
    action: str = Query("", description="Filter by action"),
    limit: int = Query(50, ge=1, le=500),
):
    try:
        firm_id = _ifrs15_firm_id(request)
        filtered = _ifrs15_audit_entries_from_db(
            firm_id, limit=limit, contract_id=contract_id, action=action
        )[:limit]
        pending = sum(
            1
            for e in filtered
            if e.get("sign_off_required") and not (e.get("signed_off_by") or "").strip()
        )
        return {
            "success": True,
            "entries": filtered,
            "total": len(filtered),
            "pending_sign_off": pending,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail={"error": "Database error", "detail": str(e)})


@app.post("/api/ifrs15/audit-log/sign-off")
async def sign_off_ifrs15_audit_entry(req: AuditSignOffRequest, request: Request):
    firm_id = _ifrs15_firm_id(request)
    try:
        rows = ifrs15_db.get_audit_log(firm_id, limit=500)
        for row in rows:
            details = row.get("details") or {}
            if str(details.get("entry_id")) == req.entry_id:
                from ifrs15_calculator import IFRS15Calculator

                updated = IFRS15Calculator().sign_off_entry(details, req.reviewer, req.notes)
                ifrs15_db.update_audit_log_entry(firm_id, str(row["id"]), updated)
                return {"success": True, "entry": updated}
        raise HTTPException(status_code=404, detail=f"Entry {req.entry_id} not found")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail={"error": "Database error", "detail": str(e)})


@app.post("/api/ifrs15/audit-log/export-excel")
async def export_ifrs15_audit_log_excel(
    request: Request,
    contract_id: str = Query("", description="Filter by contract id"),
    action: str = Query("", description="Filter by action"),
):
    """Minimal workbook: audit log only (same filters as GET audit-log, no row limit)."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import timezone

    try:
        firm_id = _ifrs15_firm_id(request)
        filtered = _ifrs15_audit_entries_from_db(
            firm_id, limit=500, contract_id=contract_id, action=action
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail={"error": "Database error", "detail": str(e)})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IFRS 15 Audit Log"

    orange = PatternFill("solid", fgColor="F97316")
    amber = PatternFill("solid", fgColor="FEF3C7")
    green = PatternFill("solid", fgColor="D1FAE5")

    headers = [
        "Entry ID",
        "Timestamp",
        "Action",
        "Contract ID",
        "Description",
        "IFRS Reference",
        "Sign-Off Required",
        "Signed Off By",
        "Signed Off At",
        "Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = orange
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for e in filtered:
        row = [
            e.get("entry_id", ""),
            e.get("timestamp", ""),
            e.get("action", ""),
            e.get("contract_id", ""),
            e.get("description", ""),
            e.get("ifrs_reference", ""),
            "YES" if e.get("sign_off_required") else "NO",
            e.get("signed_off_by", ""),
            e.get("signed_off_at", ""),
            e.get("notes", ""),
        ]
        ws.append(row)
        last = ws.max_row
        req = bool(e.get("sign_off_required"))
        done = bool(str(e.get("signed_off_by", "") or "").strip())
        fill = green if req and done else amber if req and not done else None
        if fill:
            for cell in ws[last]:
                cell.fill = fill

    max_col = len(headers)
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(v or "")))
        ws.column_dimensions[letter].width = min(max_len + 4, 50)

    last_row = max(ws.max_row, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{last_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.getvalue()

    filename = f"IFRS15_AuditLog_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        iter([data]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/ifrs15/master-report")
async def ifrs15_master_report(request: IFRS15MasterReportRequest):
    """Aggregate IFRS 15 core + optional modules into a master compliance report."""
    try:
        from ifrs15_calculator import IFRS15MasterSummaryEngine

        body = request.model_dump()
        rep = IFRS15MasterSummaryEngine().generate(body)

        narrative = ""
        if ANTHROPIC_API_KEY:
            try:
                import anthropic

                ctx = json.dumps(
                    {
                        "contract_overview": rep.get("contract_overview"),
                        "financial_summary": rep.get("financial_summary"),
                        "five_step_status": rep.get("five_step_status"),
                        "assessments": rep.get("assessments"),
                        "risk_flags": rep.get("risk_flags"),
                        "audit_readiness": rep.get("audit_readiness"),
                    },
                    indent=2,
                    default=str,
                )[:12000]
                prompt = (
                    "You are an IFRS 15 technical accounting expert. Write a 3-paragraph executive summary of this "
                    "contract's revenue recognition treatment (max 400 words). Use the actual numbers from the JSON.\n"
                    "Paragraph 1: Contract overview and revenue recognised.\n"
                    "Paragraph 2: Key judgements made and their impact.\n"
                    "Paragraph 3: Outstanding risks and recommended actions.\n"
                    "Write in the style of a Big 4 technical memo. Plain text only, no markdown.\n\n"
                    f"Data:\n{ctx}"
                )
                client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
                response = client.messages.create(
                    model="claude-sonnet-4-20250514",
                    max_tokens=1500,
                    messages=[{"role": "user", "content": prompt}],
                )
                narrative = (response.content[0].text if response.content else "").strip()
            except Exception:
                narrative = (
                    "AI narrative could not be generated (Claude API error). "
                    "Review the structured sections in this report and draft the executive summary manually."
                )
        else:
            narrative = (
                "AI narrative is not available (ANTHROPIC_API_KEY not configured). "
                "Configure the API key to generate a three-paragraph executive summary automatically."
            )
        rep["ai_narrative"] = narrative
        gc.collect()
        return rep
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/master-report/download-excel")
async def ifrs15_master_report_download_excel(request: IFRS15MasterReportExcelRequest):
    """Generate standalone IFRS 15 Master Summary Excel workbook."""
    try:
        from ifrs15_excel_export import IFRS15ExcelExporter

        file_id = str(uuid.uuid4())
        excel_path = OUTPUT_DIR / f"IFRS15_Master_{file_id}.xlsx"
        exporter = IFRS15ExcelExporter()
        exporter.export_ifrs15_master_report(request.master_report, str(excel_path))
        gc.collect()
        return {"status": "success", "file_id": file_id, "filename": excel_path.name}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/score-disclosure")
async def ifrs15_score_disclosure(request: IFRS15DisclosureScorerRequest):
    """Score IFRS 15 narrative disclosure quality (keywords + optional Claude suggestions)."""
    try:
        from ifrs15_calculator import IFRS15DisclosureScorer

        scorer = IFRS15DisclosureScorer()
        return scorer.score(
            request.disclosure_text,
            request.calculation_results or {},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/client-report")
async def ifrs15_client_report(request: IFRS15ClientReportRequest):
    """Generate client-facing IFRS 15 PDF report (10 pages)."""
    try:
        from ifrs15_client_report import generate_client_report

        payload = request.model_dump()
        out = generate_client_report(payload, api_key=ANTHROPIC_API_KEY)
        gc.collect()
        return {
            "status": "success",
            "file_id": out.get("file_id"),
            "filename": out.get("filename"),
            "pages": out.get("pages", 10),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ifrs15/client-report/{file_id}")
async def ifrs15_client_report_download(file_id: str):
    """Download generated client-facing IFRS 15 PDF report."""
    matching = list(OUTPUT_DIR.glob(f"{file_id}_*.pdf"))
    if not matching:
        raise HTTPException(status_code=404, detail="File not found")
    path = matching[0]
    return FileResponse(
        path=str(path),
        filename=path.name,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{path.name}"'},
    )


@app.post("/api/ifrs15/realestate/off-plan")
async def ifrs15_realestate_off_plan(request: RealEstateOffPlanRequest):
    """UAE off-plan sales — over-time revenue via input method (% completion)."""
    try:
        from backend.app.services.ifrs15_realestate import OffPlanSalesEngine

        result = OffPlanSalesEngine().calculate(request.model_dump())
        cid = request.contract_id or "RE-OFFPLAN"
        _ifrs15_audit_append(
            "RE_OFFPLAN",
            cid,
            f"Off-plan revenue — {result.get('completion_pct')}% complete, "
            f"AED {float(result.get('revenue_recognised_to_date', 0) or 0):,.0f} to date",
            {},
            result,
            "IFRS 15.35(c) — UAE Real Estate",
            changed_amount=float(result.get("revenue_current_period") or 0),
        )
        return {"success": True, "result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/realestate/escrow")
async def ifrs15_realestate_escrow(request: RealEstateEscrowRequest):
    """RERA escrow timeline — revenue at later of % completion or milestone release."""
    try:
        from backend.app.services.ifrs15_realestate import ReraEscrowTracker

        result = ReraEscrowTracker().analyse(request.model_dump())
        return {"success": True, "result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/realestate/modification")
async def ifrs15_realestate_modification(request: RealEstateModificationRequest):
    """UAE real estate contract modifications — prospective vs cumulative catch-up."""
    try:
        from backend.app.services.ifrs15_realestate import RealEstateModificationEngine
        from backend.app.services.oqood_validator import assess_oqood_requirement

        payload = request.model_dump()
        result = RealEstateModificationEngine().assess(payload)
        oqood = assess_oqood_requirement(
            {
                "modification_type": request.modification_type,
                "old_value": (request.original_contract or {}).get("value"),
                "new_value": (request.modification_details or {}).get("new_price"),
                "modification_date": request.modification_date,
                "currency": "AED",
                "exchange_rate": 3.6725,
            }
        ).model_dump()
        oqood["oqood_filed"] = bool(request.oqood_filed)
        result["oqood_assessment"] = oqood
        cid = request.contract_id or "RE-MOD"
        _ifrs15_audit_append(
            "RE_MODIFICATION",
            cid,
            f"Modification {request.modification_type} — {result.get('treatment')}",
            {},
            result,
            "IFRS 15.18-21 — UAE Real Estate",
            changed_amount=float(result.get("revenue_adjustment") or 0),
        )
        return {"success": True, "result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/realestate/bundling-check")
async def ifrs15_realestate_bundling_check(request: RealEstateBundlingCheckRequest):
    """IFRS 15 para 17 multi-unit bundling assessment."""
    try:
        from backend.app.services.multi_unit_bundling import UnitContract, assess_bundling

        units = [UnitContract(**u) for u in request.units]
        assessment = assess_bundling(units).model_dump()
        return {"success": True, "assessment": assessment}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.patch("/api/ifrs15/realestate/modification/oqood-filed")
async def ifrs15_realestate_patch_oqood_filed(request: RealEstateOqoodFiledPatchRequest):
    """Mark Oqood amendment filing status for a modification."""
    return {"success": True, "modification_id": request.modification_id, "oqood_filed": request.oqood_filed}


@app.post("/api/ifrs15/realestate/contract-costs")
async def ifrs15_realestate_contract_costs(request: RealEstateContractCostsRequest):
    """Sales commissions — capitalise when amortisation > 12 months (IFRS 15.91)."""
    try:
        from backend.app.services.ifrs15_realestate import RealEstateContractCostsEngine

        result = RealEstateContractCostsEngine().calculate(request.model_dump())
        return {"success": True, "result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/realestate/principal-agent")
async def ifrs15_realestate_principal_agent(request: RealEstatePrincipalAgentRequest):
    """Developer principal vs agent — gross vs net revenue (IFRS 15.B34–B38)."""
    try:
        from backend.app.services.ifrs15_realestate import RealEstatePrincipalAgentChecker

        body = request.model_dump()
        if not body.get("gross_contract_value") and body.get("contract_value"):
            body["gross_contract_value"] = body["contract_value"]
        result = RealEstatePrincipalAgentChecker().assess(body)
        return {"success": True, "result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/realestate/vat")
async def ifrs15_realestate_vat(request: RealEstateVatRequest):
    """UAE 5% VAT alignment with IFRS 15 revenue recognition schedule."""
    try:
        from backend.app.services.ifrs15_realestate import UaeVatTimingEngine

        result = UaeVatTimingEngine().align(request.model_dump())
        return {"success": True, "result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/realestate/calculate")
async def ifrs15_realestate_calculate(request: RealEstateFullRequest):
    """Combined UAE real estate calculation — off-plan, escrow, VAT."""
    try:
        from backend.app.services.ifrs15_realestate import (
            OffPlanSalesEngine,
            ReraEscrowTracker,
            UaeVatTimingEngine,
        )

        off_plan = OffPlanSalesEngine().calculate(request.off_plan.model_dump())
        escrow_result = None
        if request.escrow:
            escrow_result = ReraEscrowTracker().analyse(request.escrow.model_dump())
        vat_rows = list(request.vat_schedule or [])
        if not vat_rows and off_plan.get("revenue_current_period"):
            vat_rows = [
                {
                    "period": off_plan.get("estimated_handover", "Current"),
                    "revenue_recognised": off_plan["revenue_current_period"],
                }
            ]
        vat_result = UaeVatTimingEngine().align(
            {"revenue_schedule": vat_rows, "currency": "AED"}
        )
        return {
            "success": True,
            "off_plan": off_plan,
            "escrow": escrow_result,
            "vat": vat_result,
            "spa_mapped": request.spa_mapped,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/realestate/cancellation-refund")
async def ifrs15_realestate_cancellation_refund(request: CancellationRefundInput):
    """UAE Law No. 8 of 2007 Article 11 — cancellation refund waterfall."""
    try:
        gate = _rera_escrow_gate(
            request.escrow_receipts,
            request.escrow_releases,
            float(request.construction_completion_pct),
            float(request.contract_price),
        )
        if gate:
            return gate
        from backend.app.services.ifrs15_realestate import CancellationRefundEngine

        result = CancellationRefundEngine().calculate(request.model_dump())
        return {"success": True, "result": result}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/realestate/report")
async def ifrs15_realestate_report(request: RealEstateReportRequest):
    """Full UAE real estate IFRS 15 report — off-plan, escrow, quarterly VAT schedule."""
    try:
        from backend.app.services.ifrs15_realestate import (
            RealEstateReportEngine,
            validate_realestate_health,
            effective_completion_pct,
        )
        from backend.app.services.oqood_validator import assess_oqood_requirement
        from backend.app.services.multi_unit_bundling import UnitContract, assess_bundling

        body = request.model_dump()
        if request.assess_principal_agent:
            body["gross_contract_value"] = request.gross_contract_value or request.contract_value
        if not body.get("spa_handover_date"):
            body["spa_handover_date"] = request.expected_handover

        completion_pct_gate = effective_completion_pct(body)
        gate_rep = _rera_escrow_gate(
            request.escrow_receipts,
            request.escrow_releases,
            completion_pct_gate,
            float(request.contract_value),
        )
        if gate_rep:
            return gate_rep

        report = RealEstateReportEngine().build(body)
        oqood_assessments: List[Dict[str, Any]] = []
        for mod in (request.modifications or []):
            oq = assess_oqood_requirement(mod).model_dump()
            oq["modification_date"] = mod.get("modification_date")
            oq["modification_type"] = mod.get("modification_type") or mod.get("type")
            oq["oqood_filed"] = bool(mod.get("oqood_filed", False))
            oqood_assessments.append(oq)
        if oqood_assessments:
            report["oqood_assessments"] = oqood_assessments

        bundling_assessment = None
        if request.units:
            units = [UnitContract(**u) for u in request.units]
            bundling_assessment = assess_bundling(units).model_dump()
            report["bundling_assessment"] = bundling_assessment

        hv = dict(report.get("health_validation") or validate_realestate_health(report))
        pending_oqood = 0
        for oq in oqood_assessments:
            if not bool(oq.get("requires_oqood_amendment")):
                continue
            mod_dt = oq.get("modification_date")
            dt = datetime.strptime(mod_dt[:10], "%Y-%m-%d").date() if isinstance(mod_dt, str) and len(mod_dt) >= 10 else None
            if dt and dt <= datetime.now().date() and not bool(oq.get("oqood_filed")):
                pending_oqood += 1
        hv["check_d_pass"] = pending_oqood == 0
        hv["check_e_pass"] = bool(bundling_assessment) or not request.units
        details = list(hv.get("details") or [])
        details.append(
            "D: Oqood Amendment Check — "
            + (
                f"WARN {pending_oqood} modification(s) may require Oqood amendment filing"
                if pending_oqood
                else "PASS"
            )
        )
        details.append(
            "E: Multi-Unit Bundling Check — "
            + (
                "PASS"
                if hv["check_e_pass"]
                else "WARN potential IFRS 15 para 17 bundling issue — run bundling check"
            )
        )
        hv["details"] = details
        hv["overall_pass"] = bool(hv.get("check_a_pass")) and bool(hv.get("check_b_pass")) and bool(hv.get("check_c_pass")) and bool(hv["check_d_pass"]) and bool(hv["check_e_pass"])
        report["health_validation"] = hv

        if request.cancellation_refund:
            report["cancellation_refund"] = request.cancellation_refund
        cid = request.contract_id or "RE-REPORT"
        off = report.get("off_plan") or {}
        _ifrs15_audit_append(
            "RE_REPORT",
            cid,
            f"Real estate full report — {off.get('completion_pct')}% complete",
            {},
            {
                "periods": len(report.get("period_schedule") or []),
                "rera": request.rera_registration_number,
                "health_pass": (report.get("health_validation") or {}).get("overall_pass"),
            },
            "IFRS 15 — UAE Real Estate",
            changed_amount=float(off.get("revenue_current_period") or 0),
        )
        return {"success": True, "report": report}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/realestate/export-excel")
async def ifrs15_realestate_export_excel(request: RealEstateExportExcelRequest):
    """Export UAE real estate IFRS 15 report to Excel."""
    try:
        rpt = dict(request.report)
        receipts = list(request.escrow_receipts or rpt.get("escrow_receipts") or [])
        releases = list(request.escrow_releases or rpt.get("escrow_releases") or [])
        cmp_pct_x = request.construction_completion_pct
        if cmp_pct_x is None:
            cmp_pct_x = float((rpt.get("off_plan") or {}).get("completion_pct") or 0)
        cv_x = float(rpt.get("contract_value") or 0)
        gate_x = _rera_escrow_gate(receipts, releases, float(cmp_pct_x), cv_x)
        if gate_x:
            return gate_x

        from ifrs15_realestate_excel import IFRS15RealEstateExcelExporter

        cid = (request.contract_id or "RE").replace(" ", "_")[:40]
        file_id = str(uuid.uuid4())
        excel_path = OUTPUT_DIR / f"IFRS15_RE_{cid}_{file_id}.xlsx"
        exporter = IFRS15RealEstateExcelExporter()
        exporter.export_workbook(request.report, str(excel_path))
        return {
            "status": "success",
            "file_id": file_id,
            "filename": excel_path.name,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/realestate/to-calculate-payload")
async def ifrs15_realestate_to_calculate(request: RealEstateToCalculateRequest):
    """Map UAE real estate recognition into main IFRS 15 calculate request."""
    try:
        gate_tc = _rera_escrow_gate(
            request.escrow_receipts,
            request.escrow_releases,
            float(request.off_plan.get("completion_pct") or 0),
            float(request.contract_value),
        )
        if gate_tc:
            return gate_tc
        from backend.app.services.ifrs15_realestate import build_ifrs15_calculate_payload

        payload = build_ifrs15_calculate_payload(
            off_plan=request.off_plan,
            spa=request.spa,
            spa_mapped=request.spa_mapped,
            construction_start=request.construction_start,
            expected_handover=request.expected_handover,
            contract_value=request.contract_value,
            costs_incurred=request.costs_incurred_to_date,
            total_costs=request.total_estimated_costs,
            revenue_prior=request.revenue_prior_period,
            escrow_total=request.escrow_total,
        )
        return {"success": True, "calculate_payload": payload}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/realestate/upload-spa")
async def ifrs15_realestate_upload_spa(file: UploadFile = File(...)):
    """Upload UAE SPA (PDF/DOCX) — extract fields and map to IFRS 15 inputs."""
    if not ANTHROPIC_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="Claude API not configured. Set ANTHROPIC_API_KEY environment variable.",
        )
    try:
        import uuid
        from backend.app.services.arabic_pdf_handler import (
            extract_spa_fields,
            extraction_meta_dict,
            legacy_extracted_from_result,
        )
        from backend.app.services.ifrs15_realestate import map_spa_to_ifrs15_inputs
        from backend.app.services.spa_parser import SPAParser

        file_id = str(uuid.uuid4())
        upload_path = UPLOAD_DIR / f"spa_{file_id}_{file.filename}"
        content = await file.read()
        upload_path.write_bytes(content)
        parser = SPAParser(api_key=ANTHROPIC_API_KEY)
        extraction = extract_spa_fields(
            content,
            parser._client,
            filename=file.filename or "upload.pdf",
            model=parser._model,
        )
        extracted = legacy_extracted_from_result(extraction)
        ifrs15_inputs = (
            {} if extraction.fallback_triggered else map_spa_to_ifrs15_inputs(extracted)
        )
        meta = extraction_meta_dict(extraction)
        extraction_file = OUTPUT_DIR / f"spa_extraction_{file_id}.json"
        extraction_file.write_text(
            json.dumps(
                {"extracted": extracted, "ifrs15_inputs": ifrs15_inputs, "extraction_meta": meta},
                indent=2,
            ),
            encoding="utf-8",
        )
        _ifrs15_audit_append(
            action="SPA_UPLOAD",
            contract_id=str(extracted.get("property_unit_number") or "RE-SPA"),
            description=(
                f"SPA uploaded — language: {meta['language_detected']}, "
                f"method: {meta['extraction_method']}, "
                f"confidence: {meta['confidence_score']:.0%}"
            ),
            before_value={},
            after_value={"filename": file.filename, "extraction_meta": meta},
            ifrs_reference="IFRS 15 — UAE SPA extraction",
        )
        return {
            "status": "success",
            "file_id": file_id,
            "filename": file.filename,
            "extracted": extracted,
            "ifrs15_inputs": ifrs15_inputs,
            "extraction_meta": meta,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/realestate/upload-rera-certificate")
async def ifrs15_realestate_upload_rera_certificate(
    file: UploadFile = File(...),
    rera_registration_number: Optional[str] = Form(None),
    form_completion_pct: Optional[float] = Form(None),
    currency: str = Form("AED"),
):
    """Upload RERA / DLD completion certificate PDF — extract verified completion %."""
    if not ANTHROPIC_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="Claude API not configured. Set ANTHROPIC_API_KEY environment variable.",
        )
    max_bytes = 10 * 1024 * 1024
    try:
        import uuid
        from backend.app.services.rera_certificate_handler import (
            cross_check_rera_number,
            extract_rera_certificate,
        )
        from backend.app.services.spa_parser import SPAParser

        content = await file.read()
        if len(content) > max_bytes:
            raise HTTPException(status_code=400, detail="Certificate PDF must be 10MB or smaller.")
        fname = (file.filename or "").lower()
        if not fname.endswith(".pdf"):
            raise HTTPException(status_code=400, detail="Only PDF certificates are accepted.")

        parser = SPAParser(api_key=ANTHROPIC_API_KEY)
        result = extract_rera_certificate(
            content,
            parser._client,
            form_completion_pct=form_completion_pct,
            model=parser._model,
        )
        result = cross_check_rera_number(result, rera_registration_number)

        file_id = str(uuid.uuid4())
        cert_path = UPLOAD_DIR / f"rera_cert_{file_id}_{file.filename}"
        cert_path.write_bytes(content)

        fields = result.fields.model_dump()
        payload = {
            "success": result.success,
            "language_detected": result.language_detected,
            "confidence_score": result.confidence_score,
            "fields": fields,
            "low_confidence_fields": result.low_confidence_fields,
            "warnings": result.warnings,
            "extraction_method": result.extraction_method,
            "mismatch_detected": result.mismatch_detected,
            "mismatch_detail": result.mismatch_detail,
            "file_id": file_id,
            "filename": file.filename,
            "currency": currency,
        }
        if result.success:
            cert_ref = fields.get("certificate_ref") or "N/A"
            cert_date = fields.get("certificate_date") or "N/A"
            pct = fields.get("completion_pct")
            _ifrs15_audit_append(
                action="RERA_CERTIFICATE_UPLOAD",
                contract_id=str(fields.get("rera_registration_number") or rera_registration_number or "RE-CERT"),
                description=(
                    f"RERA certificate uploaded — ref: {cert_ref}, date: {cert_date}, "
                    f"verified completion: {pct}%, method: {result.extraction_method}, "
                    f"confidence: {result.confidence_score:.0%}"
                ),
                before_value={},
                after_value=payload,
                ifrs_reference="IFRS 15 — RERA completion certificate",
            )
        return payload
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs15/classify-contract")
async def ifrs15_classify_contract(request: IFRS15ClassifyContractRequest):
    """Classify IT services contract type for IFRS 15 revenue method selection."""
    if not ANTHROPIC_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="Claude API not configured. Set ANTHROPIC_API_KEY environment variable."
        )
    try:
        import anthropic
        import json
        import re

        prompt = (
            "Read this IT services contract and classify it. Return ONLY JSON:\n"
            "{\n"
            "  'contract_type':\n"
            "    'time_and_material' |\n"
            "    'fixed_price' |\n"
            "    'capped_tm' |\n"
            "    'maintenance',\n"
            "  'confidence': 0-100,\n"
            "  'reason': 'one sentence explanation',\n"
            "  'key_indicators': [\n"
            "    'indicator 1',\n"
            "    'indicator 2'\n"
            "  ],\n"
            "  'hourly_rate': number or null,\n"
            "  'tm_cap': number or null,\n"
            "  'estimated_cost': number or null\n"
            "}\n\n"
            f"Contract:\n{request.contract_text}"
        )

        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1200,
            messages=[{"role": "user", "content": prompt}]
        )
        text = response.content[0].text if response.content else ""
        cleaned = text.strip()
        if cleaned.startswith("```"):
            cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.I)
            cleaned = re.sub(r"\s*```$", "", cleaned)
        parsed = json.loads(cleaned.replace("'", '"'))
        return {"status": "success", "classification": parsed}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Contract classification failed: {str(e)}")


@app.get("/api/ifrs15/download/{file_id}")
async def ifrs15_download(file_id: str):
    """Download IFRS 15 Excel report"""
    matching = list(OUTPUT_DIR.glob(f"*{file_id}*.xlsx"))
    if not matching:
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(
        path=str(matching[0]),
        filename=matching[0].name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.post("/api/chat", response_model=ChatResponse)
async def chat_with_rag(request: ChatRequest):
    """
    Chat with company's IFRS documents using RAG
    
    Uses Retrieval-Augmented Generation to answer questions about:
    - Lease calculations (IFRS 16)
    - Revenue contracts (IFRS 15)
    - Expected credit loss (IFRS 9)
    - Variances and analyses
    
    CRITICAL: Data is filtered by company_id to ensure isolation
    """
    engine = get_rag_engine()
    if not engine:
        raise HTTPException(
            status_code=503,
            detail="RAG engine not initialized. Check server logs."
        )
    
    try:
        # Ask question with context
        result = engine.ask_with_context(
            company_id=request.company_id,
            question=request.question,
            document_type=request.document_type,
            top_k=request.top_k
        )
        
        if result['status'] == 'error':
            raise HTTPException(
                status_code=500,
                detail=f"RAG query failed: {result.get('error', 'Unknown error')}"
            )
        
        gc.collect()
        return ChatResponse(
            status="success",
            answer=result['answer'],
            sources=result.get('sources', []),
            context_count=result.get('context_count', 0)
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Chat error: {str(e)}")


@app.post("/api/cfo-insights")
async def cfo_insights(request: CFOInsightsRequest):
    """
    CFO Strategic Insights - AI analysis of lease portfolio
    
    Accepts a prompt (built client-side from lease_repository) and returns structured
    insights: risk flags, cost-saving opportunities, renewal alerts, efficiency recommendations.
    Uses Claude API; requires ANTHROPIC_API_KEY.
    """
    if not ANTHROPIC_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="Claude API not configured. Set ANTHROPIC_API_KEY in .env"
        )
    try:
        import anthropic
        import re
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4000,
            messages=[{"role": "user", "content": request.prompt}],
        )
        raw = "".join(
            b.text for b in response.content if getattr(b, "type", None) == "text"
        )
        clean = re.sub(r"```json|```", "", raw).strip()
        parsed = json.loads(clean)
        gc.collect()
        return {"status": "success", "result": parsed}
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse AI response: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/rag/stats/{company_id}")
async def get_rag_stats(company_id: str):
    """
    Get RAG statistics for a company
    
    Shows:
    - Total documents indexed
    - Document counts by type
    - Total chunks stored
    """
    engine = get_rag_engine()
    if not engine:
        raise HTTPException(
            status_code=503,
            detail="RAG engine not initialized"
        )
    
    try:
        stats = engine.get_company_stats(company_id)
        return {
            "status": "success",
            "stats": stats
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Stats error: {str(e)}")


# IFRS 16 Smart Alerts
@app.get("/api/ifrs16/alerts/defaults")
async def get_alert_defaults():
    """Return default email from env for pre-fill"""
    return {"email": os.getenv("ALERT_EMAIL_TO", "")}


@app.post("/api/ifrs16/alerts/configure")
async def configure_alerts(config: dict):
    """Save alert configuration to JSON file"""
    try:
        with open("alert_config.json", "w") as f:
            json.dump(config, f, indent=2)
        return {"status": "saved"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs16/alerts/send-test")
async def send_test_alert(config: dict):
    """Send a test alert email"""
    import smtplib
    from email.mime.text import MIMEText
    email = config.get("email", "")
    if not email:
        return {"status": "error", "message": "Email required"}
    message = "IFRS AI Test Alert - your alert system is working!"
    try:
        sender = os.getenv("ALERT_EMAIL_FROM", "")
        password = os.getenv("ALERT_EMAIL_PASSWORD", "")
        if sender and password:
            msg = MIMEText(message)
            msg["Subject"] = "IFRS AI Test Alert"
            msg["From"] = sender
            msg["To"] = email
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
                s.login(sender, password)
                s.sendmail(sender, email, msg.as_string())
            return {"status": "sent", "email": email}
        return {"status": "logged", "message": "Configure ALERT_EMAIL_FROM and ALERT_EMAIL_PASSWORD in .env"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


class AlertsCheckRequest(BaseModel):
    leases: List[dict] = Field(default_factory=list, description="Lease repository data from frontend")


@app.post("/api/ifrs16/alerts/check")
async def check_alerts(request: AlertsCheckRequest):
    """Check lease expiry and return alerts list"""
    from datetime import datetime, timedelta
    alerts = []
    today = datetime.now().date()
    for lease in request.leases:
        end_str = None
        if isinstance(lease, dict):
            dates = lease.get("dates") or {}
            end_str = dates.get("end")
            asset = lease.get("asset") or lease.get("lease_id") or lease.get("id", "Unknown")
        else:
            continue
        if not end_str:
            continue
        try:
            end_date = datetime.strptime(end_str, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            continue
        diff = (end_date - today).days
        if diff < 0:
            alerts.append({
                "id": str(lease.get("id", "")),
                "type": "expired",
                "severity": "red",
                "title": f"{asset} has expired",
                "message": f"Lease expired {abs(diff)} days ago",
            })
        elif diff <= 7:
            alerts.append({
                "id": str(lease.get("id", "")),
                "type": "expiring_7",
                "severity": "red",
                "title": f"{asset} expires in {diff} days",
                "message": f"Renew or terminate by {end_str}",
            })
        elif diff <= 30:
            alerts.append({
                "id": str(lease.get("id", "")),
                "type": "expiring_30",
                "severity": "amber",
                "title": f"{asset} expires in {diff} days",
                "message": f"Renew or terminate by {end_str}",
            })
        elif diff <= 90:
            alerts.append({
                "id": str(lease.get("id", "")),
                "type": "expiring_90",
                "severity": "amber",
                "title": f"{asset} expires in {diff} days",
                "message": f"Renew or terminate by {end_str}",
            })
    return {"alerts": alerts}


@app.post("/api/ifrs16/suggest-ibr")
async def suggest_ibr(payload: IbrSuggestRequest):
    fallback_by_currency = {
        "INR": {"ibr_low": 7.5, "ibr_mid": 9.0, "ibr_high": 12.0},
        "USD": {"ibr_low": 5.5, "ibr_mid": 7.0, "ibr_high": 9.0},
        "EUR": {"ibr_low": 3.5, "ibr_mid": 5.0, "ibr_high": 7.0},
    }
    fallback = fallback_by_currency.get((payload.currency or "").upper(), {"ibr_low": 6.0, "ibr_mid": 8.5, "ibr_high": 11.0})
    fallback_response = {
        **fallback,
        "rationale": "Fallback estimate based on currency-level market conditions and typical borrowing spreads.",
        "market_references": ["Policy rate", "Commercial lending benchmarks", "Typical credit spread"],
    }
    if not ANTHROPIC_API_KEY:
        return fallback_response

    try:
        import anthropic

        system_prompt = """You are an IFRS 16 expert. Given a lease's country, currency, tenor, asset type,
and lessee type, suggest a realistic Incremental Borrowing Rate (IBR) range.

Respond ONLY with a JSON object. No preamble, no markdown, no explanation outside
the JSON. Format:
{
  "ibr_low": 7.5,
  "ibr_mid": 8.5,
  "ibr_high": 10.0,
  "rationale": "2-3 sentence explanation referencing current market context,
                 RBI/central bank rates, and IFRS 16 para 26 guidance",
  "market_references": ["RBI repo rate", "SBI MCLR", "typical corporate bond spread"]
}"""

        user_prompt = (
            f"Country: {payload.country}, Currency: {payload.currency}, "
            f"Lease term: {payload.lease_term_months} months, "
            f"Asset type: {payload.asset_type or 'N/A'}, Lessee type: {payload.lessee_type}. "
            "Suggest IBR range."
        )

        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        msg = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=800,
            temperature=0,
            system=system_prompt,
            messages=[{"role": "user", "content": user_prompt}],
        )
        text = msg.content[0].text.strip()
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0].strip()
        elif "```" in text:
            text = text.split("```")[1].split("```")[0].strip()
        data = json.loads(text)
        return {
            "ibr_low": float(data.get("ibr_low")),
            "ibr_mid": float(data.get("ibr_mid")),
            "ibr_high": float(data.get("ibr_high")),
            "rationale": str(data.get("rationale", "")),
            "market_references": data.get("market_references", []) or [],
        }
    except Exception:
        return fallback_response


@app.post("/api/ifrs16/consolidate")
async def consolidate_ifrs16(request: dict):
    """
    Consolidate IFRS 16 across multiple entities.

    Body:
    {
      "group_currency": "USD",
      "entities": [
        {
          "entity_name": "Subsidiary A",
          "entity_currency": "INR",
          "fx_rate_to_group": 0.012,
          "leases": [ ...lease results... ]
        }
      ]
    }
    """
    try:
        from ifrs16_calculator import consolidate_leases

        entities = request.get("entities", [])
        if not entities:
            raise HTTPException(400, "At least one entity required")
        result = consolidate_leases(entities)
        result["group_currency"] = request.get("group_currency", "USD")
        return result
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/consolidation/entities")
async def consolidation_entities(parent_id: str = ""):
    from entity_consolidation import get_entities
    return get_entities(parent_id or None)


@app.post("/api/consolidation/entities")
async def consolidation_add_entity(payload: dict):
    from entity_consolidation import save_entity
    if not payload.get("entity_id") or not payload.get("entity_name"):
        raise HTTPException(400, "entity_id and entity_name are required")
    save_entity(
        entity_id=payload["entity_id"],
        entity_name=payload["entity_name"],
        parent_id=payload.get("parent_id"),
        currency=payload.get("currency", "INR"),
        fx_rate=float(payload.get("fx_rate_to_group", 1.0)),
    )
    return {"status": "saved"}


@app.post("/api/consolidation/intercompany")
async def consolidation_add_intercompany(payload: dict):
    from entity_consolidation import save_intercompany
    required = ["lessor_entity", "lessee_entity", "lease_id", "monthly_amount"]
    missing = [k for k in required if k not in payload]
    if missing:
        raise HTTPException(400, f"Missing fields: {', '.join(missing)}")
    save_intercompany(
        lessor=payload["lessor_entity"],
        lessee=payload["lessee_entity"],
        lease_id=payload["lease_id"],
        monthly_amount=float(payload["monthly_amount"]),
    )
    return {"status": "saved"}


@app.post("/api/consolidation/run")
async def consolidation_run(payload: dict):
    from entity_consolidation import consolidate
    entity_ids = payload.get("entity_ids", [])
    if not entity_ids:
        raise HTTPException(400, "At least one entity required")
    return consolidate(entity_ids=entity_ids, group_currency=payload.get("group_currency", "INR"))


# ==================== IFRS 9 ECL Endpoints ====================

# Standard PD by credit rating (IFRS 9 para 5.5.17)
PD_BY_RATING = {
    "AAA": 0.0001, "AA+": 0.0002, "AA": 0.0002, "AA-": 0.0003,
    "A+": 0.0004, "A": 0.0005, "A-": 0.001,
    "BBB+": 0.0015, "BBB": 0.002, "BBB-": 0.003,
    "BB+": 0.005, "BB": 0.01, "BB-": 0.02,
    "B+": 0.03, "B": 0.05, "B-": 0.08,
    "CCC": 0.10, "CC": 0.20, "C": 0.30, "D": 1.0,
}


class IFRS9ClassificationRequest(BaseModel):
    instrument_name: str
    instrument_type: str
    business_model_indicators: List[str]
    sppi_features: List[str]
    prepayment_penalty_reasonable: bool = True
    fair_value_option_elected: bool = False
    fvo_reason: Optional[str] = None
    business_model_changed: bool = False
    nominal_rate: Optional[float] = None
    issue_price: Optional[float] = None
    face_value: Optional[float] = None
    term_months: Optional[int] = None


class MacroScenario(BaseModel):
    gdp_growth: float
    unemployment_rate: float
    interest_rate: float
    property_price_change: float = 0.0
    credit_spread: float = 0.0
    probability: float


class IFRS9MacroOverlayRequest(BaseModel):
    portfolio_name: str
    base_pd: float
    lgd: float
    ead: float
    base_scenario: MacroScenario
    optimistic_scenario: MacroScenario
    pessimistic_scenario: MacroScenario
    loans: Optional[List[Any]] = None


class ReceivableItem(BaseModel):
    invoice_id: str
    customer: str
    gross_amount: float
    days_past_due: int
    currency: str = "USD"


class BucketTotal(BaseModel):
    label: str
    gross_amount: float
    historical_loss_rate: Optional[float] = None


class HistoricalData(BaseModel):
    bucket_label: str
    historical_balance: float
    historical_writeoffs: float


class IFRS9ProvisionMatrixRequest(BaseModel):
    portfolio_name: str
    reporting_date: Optional[str] = None
    receivable_type: str = "trade_receivables"
    receivables: Optional[List[ReceivableItem]] = None
    bucket_totals: Optional[List[BucketTotal]] = None
    loss_rates: Optional[Dict[str, Any]] = None
    historical_data: Optional[List[HistoricalData]] = None
    macro_adjustment_factor: Optional[float] = 0.0
    writeoffs_this_period: Optional[float] = 0.0
    custom_buckets: Optional[List[Dict[str, Any]]] = None


class IFRS9DownloadExcelRequest(BaseModel):
    portfolio_name: str = "Portfolio"
    entity_name: Optional[str] = ""
    reporting_date: Optional[str] = None
    applicable_ecl: Optional[float] = 0
    ecl_12m: Optional[float] = 0
    ecl_lifetime: Optional[float] = 0
    total_ead: Optional[float] = 0
    pd_used: Optional[float] = 0
    lgd_used: Optional[float] = 0
    coverage_ratio: Optional[float] = 0
    weighted_avg_pd: Optional[float] = 0
    stage_summary: Optional[Dict[str, Any]] = None
    loans: Optional[List[Any]] = None
    ecl_movement: Optional[Dict[str, Any]] = None
    journal_entries: Optional[List[Any]] = None
    disclosure_notes: Optional[Any] = None
    scenario_results: Optional[Dict[str, Any]] = None
    bucket_results: Optional[List[Any]] = None
    macro_sensitivity: Optional[str] = None
    discount_rate: Optional[float] = None
    master_report_data: Optional[Dict[str, Any]] = None


class IFRS9MasterReportRequest(BaseModel):
    portfolio_name: str
    entity_name: Optional[str] = ""
    reporting_date: Optional[str] = None
    core_results: Optional[Dict[str, Any]] = None
    classification_result: Optional[Dict[str, Any]] = None
    macro_overlay_result: Optional[Dict[str, Any]] = None
    provision_matrix_result: Optional[Dict[str, Any]] = None


@app.get("/api/ifrs9/pd-rates")
async def get_ifrs9_pd_rates():
    """Return standard PD rates by credit rating (AAA to D)."""
    return {"pd_rates": PD_BY_RATING}


@app.post("/api/ifrs9/upload-portfolio")
async def upload_portfolio(file: UploadFile = File(...)):
    """
    Read uploaded Excel/CSV debtor ageing or loan schedule.
    Returns structured data for form auto-fill (counterparty, amounts, ageing buckets).
    """
    allowed = [".xlsx", ".xls", ".csv"]
    ext = Path(file.filename).suffix.lower()
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"Allowed: {', '.join(allowed)}")
    try:
        import pandas as pd
        contents = await file.read()
        if ext == ".csv":
            df = pd.read_csv(io.BytesIO(contents), encoding="utf-8", on_bad_lines="skip")
        else:
            df = pd.read_excel(io.BytesIO(contents), sheet_name=0)
        # Normalize column names (strip, lower, replace spaces)
        df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
        # Build ageing-style buckets if we have amount columns
        rows = df.head(100).to_dict(orient="records")
        # Try to infer: outstanding_balance, days_past_due, counterparty, etc.
        extracted = {
            "rows": rows,
            "columns": list(df.columns),
            "row_count": len(df),
        }
        gc.collect()
        return {"status": "success", "filename": file.filename, "extracted_data": extracted}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload error: {str(e)}")


@app.post("/api/ifrs9/calculate")
async def calculate_ecl(data: dict):
    """
    Core ECL calculation engine.
    Input: approach (simplified/general), stage (1/2/3), pd_12m, pd_lifetime, lgd, ead,
           ageing_buckets (for simplified), scenarios (base/optimistic/pessimistic with weights).
    Returns: ecl_12m, ecl_lifetime, applicable_ecl, coverage_ratio, journal_entries, disclosure_notes, scenario_results.
    """
    try:
        def apply_macro_overlay_pd_pct(base_pd_pct: float, macro: dict | None) -> float:
            """Apply macro overlay to PD in percentage points."""
            if not macro:
                return base_pd_pct
            gdp_growth = float((macro or {}).get("gdp_growth", 0) or 0)
            unemployment = float((macro or {}).get("unemployment", 0) or 0)
            interest_rate = float((macro or {}).get("interest_rate", 0) or 0)
            gdp_s = get_sensitivity("gdp_sensitivity")
            unemp_s = get_sensitivity("unemployment_sensitivity")
            rate_s = get_sensitivity("interest_rate_sensitivity")
            gdp_adj = max(0.20, 1 + ((-gdp_growth) * gdp_s))
            unemp_adj = max(0.20, 1 + (unemployment * unemp_s))
            rate_adj = max(0.20, 1 + (interest_rate * rate_s))
            return max(0.0, min(100.0, base_pd_pct * gdp_adj * unemp_adj * rate_adj))

        approach = data.get("approach", "general")
        stage = int(data.get("stage", 1))
        pd_12m = float(data.get("pd_12m", 0.01))
        pd_lifetime = float(data.get("pd_lifetime", 0.05))
        lgd = float(data.get("lgd", 0.45))
        ead = float(data.get("ead", 0))
        discount_rate = float(data.get("discount_rate", 0.08))

        scenarios = data.get("scenarios") or {}
        base_macro = scenarios.get("base_macro") or {}
        optimistic_macro = scenarios.get("optimistic_macro") or {}
        pessimistic_macro = scenarios.get("pessimistic_macro") or {}
        pd_12m_base = apply_macro_overlay_pd_pct(pd_12m, base_macro)
        pd_lifetime_base = apply_macro_overlay_pd_pct(pd_lifetime, base_macro)

        if approach == "simplified" and data.get("ageing_buckets"):
            # Provision matrix: sum(amount * rate) per bucket
            total_ecl = 0
            bucket_results = []
            for b in data["ageing_buckets"]:
                amt = float(b.get("amount", 0) or 0)
                rate = float(b.get("rate", 0) or 0) / 100
                ecl = amt * rate
                total_ecl += ecl
                bucket_results.append({"bucket": b.get("bucket", ""), "amount": amt, "rate_pct": rate * 100, "ecl": ecl})
            applicable_ecl = total_ecl
        else:
            # General: ECL = PD × LGD × EAD (with optional discount)
            dfactor = 1 / (1 + discount_rate)
            ecl_12m = ead * (pd_12m_base / 100) * (lgd / 100) * dfactor
            ecl_lifetime = ead * (pd_lifetime_base / 100) * (lgd / 100) * dfactor
            applicable_ecl = float(ecl_12m if stage == 1 else ecl_lifetime)
            bucket_results = []

        coverage = (applicable_ecl / ead * 100) if ead else 0

        # Scenario-weighted ECL (optional)
        base_w = float(scenarios.get("base_weight", 50)) / 100
        opt_w = float(scenarios.get("optimistic_weight", 30)) / 100
        pess_w = float(scenarios.get("pessimistic_weight", 20)) / 100
        if approach == "general":
            dfactor = 1 / (1 + discount_rate)
            pd_12m_opt = apply_macro_overlay_pd_pct(pd_12m, optimistic_macro)
            pd_12m_pess = apply_macro_overlay_pd_pct(pd_12m, pessimistic_macro)
            pd_lifetime_opt = apply_macro_overlay_pd_pct(pd_lifetime, optimistic_macro)
            pd_lifetime_pess = apply_macro_overlay_pd_pct(pd_lifetime, pessimistic_macro)
            base_ecl_calc = ead * ((pd_12m_base if stage == 1 else pd_lifetime_base) / 100) * (lgd / 100) * dfactor
            opt_ecl_calc = ead * ((pd_12m_opt if stage == 1 else pd_lifetime_opt) / 100) * (lgd / 100) * dfactor
            pess_ecl_calc = ead * ((pd_12m_pess if stage == 1 else pd_lifetime_pess) / 100) * (lgd / 100) * dfactor
            base_ecl = float(scenarios.get("base_ecl", base_ecl_calc))
            opt_ecl = float(scenarios.get("optimistic_ecl", opt_ecl_calc))
            pess_ecl = float(scenarios.get("pessimistic_ecl", pess_ecl_calc))
        else:
            base_ecl = float(scenarios.get("base_ecl", applicable_ecl))
            opt_ecl = float(scenarios.get("optimistic_ecl", applicable_ecl * 0.8))
            pess_ecl = float(scenarios.get("pessimistic_ecl", applicable_ecl * 1.2))
        ecl_weighted = base_ecl * base_w + opt_ecl * opt_w + pess_ecl * pess_w

        journal_entries = [
            {"type": "ECL Recognition", "dr": "Impairment Loss (P&L)", "cr": "Loss Allowance (BS)", "amount": applicable_ecl},
        ]

        disclosure_notes = (
            f"Note: Expected Credit Losses (IFRS 9)\n\n"
            f"The Company applies the {approach} approach. ECL = PD × LGD × EAD.\n"
            f"PD (12M): {pd_12m}% (macro-adjusted: {pd_12m_base:.2f}%) | "
            f"PD (Lifetime): {pd_lifetime}% (macro-adjusted: {pd_lifetime_base:.2f}%) | LGD: {lgd}% | EAD: {ead:,.0f}\n"
            f"Applicable ECL: {applicable_ecl:,.0f} | Coverage: {coverage:.2f}%\n"
            f"Probability-weighted ECL (scenarios): {ecl_weighted:,.0f}"
        )

        return {
            "ecl_12m": ead * (pd_12m_base / 100) * (lgd / 100) if approach == "general" else None,
            "ecl_lifetime": ead * (pd_lifetime_base / 100) * (lgd / 100) if approach == "general" else None,
            "ecl_simplified": applicable_ecl if approach == "simplified" else None,
            "applicable_ecl": applicable_ecl,
            "ecl_weighted": ecl_weighted,
            "coverage_ratio": round(coverage, 2),
            "pd_used": pd_12m_base if stage == 1 else pd_lifetime_base,
            "macro_pd": {
                "base_pd_12m": pd_12m_base,
                "base_pd_lifetime": pd_lifetime_base,
                "optimistic_pd_12m": apply_macro_overlay_pd_pct(pd_12m, optimistic_macro),
                "optimistic_pd_lifetime": apply_macro_overlay_pd_pct(pd_lifetime, optimistic_macro),
                "pessimistic_pd_12m": apply_macro_overlay_pd_pct(pd_12m, pessimistic_macro),
                "pessimistic_pd_lifetime": apply_macro_overlay_pd_pct(pd_lifetime, pessimistic_macro),
            },
            "macro_sensitivity_used": {
                "gdp_sensitivity": get_sensitivity("gdp_sensitivity"),
                "unemployment_sensitivity": get_sensitivity("unemployment_sensitivity"),
                "interest_rate_sensitivity": get_sensitivity("interest_rate_sensitivity"),
                "config_source": "db" if is_db_loaded() else "default",
            },
            "lgd_used": lgd,
            "ead_used": ead,
            "journal_entries": journal_entries,
            "disclosure_notes": disclosure_notes,
            "bucket_results": bucket_results,
            "scenario_results": {
                "base": base_ecl,
                "optimistic": opt_ecl,
                "pessimistic": pess_ecl,
                "weighted": ecl_weighted,
            },
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/macro-sensitivity")
async def get_macro_sensitivity(tenant_id: str = "default", portfolio_type: str = "all"):
    row = get_active_config(tenant_id, portfolio_type)
    if row:
        return {
            "source": "database",
            "tenant_id": tenant_id,
            "portfolio_type": portfolio_type,
            "gdp_sensitivity": row[0],
            "unemployment_sensitivity": row[1],
            "interest_rate_sensitivity": row[2],
            "effective_from": row[3],
            "approved_by": row[4],
            "approval_notes": row[5],
            "created_at": row[6],
        }
    return {"source": "defaults", **MACRO_SENSITIVITY_DEFAULTS}


@app.post("/api/admin/macro-sensitivity")
async def update_macro_sensitivity(payload: dict):
    required = ["gdp_sensitivity", "unemployment_sensitivity", "interest_rate_sensitivity"]
    for key in required:
        if key not in payload:
            raise HTTPException(400, f"Missing field: {key}")
    save_config(
        tenant_id=payload.get("tenant_id", "default"),
        portfolio_type=payload.get("portfolio_type", "all"),
        gdp=float(payload["gdp_sensitivity"]),
        unemp=float(payload["unemployment_sensitivity"]),
        rate=float(payload["interest_rate_sensitivity"]),
        approved_by=payload.get("approved_by", ""),
        notes=payload.get("approval_notes", ""),
    )
    update_sensitivity("gdp_sensitivity", float(payload["gdp_sensitivity"]))
    update_sensitivity("unemployment_sensitivity", float(payload["unemployment_sensitivity"]))
    update_sensitivity("interest_rate_sensitivity", float(payload["interest_rate_sensitivity"]))
    return {"status": "updated", "effective": "immediately"}


@app.get("/api/admin/macro-sensitivity/history")
async def macro_sensitivity_history(tenant_id: str = "default", portfolio_type: str = "all"):
    rows = get_config_history(tenant_id, portfolio_type)
    cols = [
        "id",
        "tenant_id",
        "portfolio_type",
        "gdp_sensitivity",
        "unemployment_sensitivity",
        "interest_rate_sensitivity",
        "effective_from",
        "approved_by",
        "approval_notes",
        "is_active",
        "created_at",
    ]
    return [dict(zip(cols, r)) for r in rows]


@app.post("/api/ifrs9/classify")
async def classify_instrument(request: IFRS9ClassificationRequest):
    """IFRS 9 classification & measurement (business model, SPPI, optional EIR)."""
    from ifrs9_ecl_calculator import IFRS9ClassificationEngine

    engine = IFRS9ClassificationEngine()
    return engine.classify(request.model_dump())


@app.post("/api/ifrs9/macro-overlay")
async def ifrs9_macro_overlay(request: IFRS9MacroOverlayRequest):
    """Forward-looking macro scenario overlay for ECL (IFRS 9.5.5.17)."""
    from ifrs9_ecl_calculator import IFRS9MacroOverlayEngine

    try:
        engine = IFRS9MacroOverlayEngine()
        payload = request.model_dump()
        if payload.get("loans") is None:
            payload["loans"] = []
        return engine.calculate(payload)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/ifrs9/provision-matrix")
async def ifrs9_provision_matrix(request: IFRS9ProvisionMatrixRequest):
    """IFRS 9.5.5.15 simplified provision matrix (ageing buckets × loss rates)."""
    from ifrs9_ecl_calculator import IFRS9ProvisionMatrixEngine

    try:
        engine = IFRS9ProvisionMatrixEngine()
        payload = request.model_dump()
        if payload.get("receivables") is None:
            payload["receivables"] = []
        if payload.get("bucket_totals") is None:
            payload["bucket_totals"] = []
        if payload.get("historical_data") is None:
            payload["historical_data"] = []
        return engine.calculate(payload)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/ifrs9/master-report")
async def ifrs9_master_report(request: IFRS9MasterReportRequest):
    """Aggregate IFRS 9 ECL, classification, macro overlay, and provision matrix into one master report."""
    from ifrs9_ecl_calculator import IFRS9MasterSummaryEngine

    engine = IFRS9MasterSummaryEngine()
    return engine.generate(request.model_dump())


@app.post("/api/ifrs9/download-report")
async def download_ecl_report(data: dict):
    """Generate Excel ECL report. Returns file_id for frontend download."""
    try:
        import pandas as pd
        file_id = str(uuid.uuid4())
        path = OUTPUT_DIR / f"ecl_report_{file_id}.xlsx"
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            pd.DataFrame([{"ECL Summary": data.get("applicable_ecl", 0), "Coverage %": data.get("coverage_ratio", 0)}]).to_excel(w, sheet_name="ECL Summary", index=False)
            if data.get("bucket_results"):
                pd.DataFrame(data["bucket_results"]).to_excel(w, sheet_name="Provision Matrix", index=False)
            if data.get("journal_entries"):
                pd.DataFrame(data["journal_entries"]).to_excel(w, sheet_name="Journal Entries", index=False)
        gc.collect()
        return {"file_id": file_id, "filename": path.name}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ifrs9/download-excel")
async def ifrs9_download_excel(request: IFRS9DownloadExcelRequest):
    """IFRS 9 ECL audit pack (openpyxl); optional first sheet when master_report_data is supplied."""
    try:
        from ifrs9_excel_export import export_ifrs9_excel

        payload = request.model_dump()
        master = payload.pop("master_report_data", None)
        file_id = export_ifrs9_excel(payload, master_report=master)
        safe_name = (request.portfolio_name or "Portfolio").replace(" ", "_").replace("/", "_")
        date_str = (request.reporting_date or datetime.now().strftime("%Y-%m-%d")).replace("-", "")
        filename = f"IFRS9_ECL_{safe_name}_{date_str}_{file_id}.xlsx"
        gc.collect()
        n_sheets = 6 if master else 5
        return {"file_id": file_id, "filename": filename, "sheets": n_sheets}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ifrs9/download/{file_id}")
async def download_ifrs9_file(file_id: str):
    """Download IFRS 9 Excel (audit pack or legacy ecl_report)."""
    matching = list(OUTPUT_DIR.glob(f"*{file_id}*.xlsx"))
    if not matching:
        raise HTTPException(status_code=404, detail="File not found")
    path = sorted(matching, key=lambda p: p.stat().st_mtime, reverse=True)[0]
    return FileResponse(
        path=str(path),
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{path.name}"'},
    )


@app.delete("/api/rag/document/{company_id}/{document_id}")
async def delete_rag_document(company_id: str, document_id: str):
    """
    Delete a document from RAG storage
    
    Args:
        company_id: Company identifier
        document_id: Document identifier to delete
    """
    engine = get_rag_engine()
    if not engine:
        raise HTTPException(
            status_code=503,
            detail="RAG engine not initialized"
        )
    
    try:
        result = engine.delete_document(company_id, document_id)
        if result['status'] == 'error':
            raise HTTPException(status_code=500, detail=result.get('error'))
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Delete error: {str(e)}")


try:
    from backend.app.routers.ifrs16_extensions import router as _ifrs16_ext_router

    app.include_router(_ifrs16_ext_router)
except ImportError as _ifrs16_ext_err:
    print(f"WARNING: IFRS 16 extensions router not loaded: {_ifrs16_ext_err}")

try:
    from backend.app.routers.rev_rec_recon import router as _rev_rec_recon_router

    app.include_router(_rev_rec_recon_router)
except ImportError as _rev_rec_import_err:
    print(f"WARNING: Rev-rec reconciliation router not loaded: {_rev_rec_import_err}")


# Run application (local dev). Tries PORT/API_PORT then scans for a free port if busy (Windows 10048, etc.).
if __name__ == "__main__":
    import socket
    import uvicorn

    def _pick_bind_port(start: int, span: int = 40) -> int:
        for p in range(start, start + span):
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                try:
                    s.bind(("127.0.0.1", p))
                except OSError:
                    continue
                return p
        raise RuntimeError(f"No free TCP port between {start} and {start + span - 1}")

    base = int(os.environ.get("PORT", os.environ.get("API_PORT", "9000")))
    chosen = _pick_bind_port(base)
    os.environ["_IFRS_BIND_PORT"] = str(chosen)
    os.environ["PUBLIC_API_URL"] = f"http://127.0.0.1:{chosen}"

    port_file = _project_root / "api_dev_port.txt"
    try:
        port_file.write_text(str(chosen), encoding="utf-8")
    except OSError as werr:
        print(f"WARNING: Could not write {port_file.name}: {werr}")

    if chosen != base:
        print(f"WARNING: Port {base} is already in use; using {chosen} instead.")
    print(f"API base: http://127.0.0.1:{chosen}")
    print(f"Wrote {port_file.name} for Next.js proxy (restart npm run dev if the UI was already running).")

    uvicorn.run(app, host="127.0.0.1", port=chosen, reload=False)




