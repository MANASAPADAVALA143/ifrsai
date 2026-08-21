"""IFRS 16 lease commercial-terms flattening, deadline scan, and Excel export.

Used by:
  POST /api/export-lease-terms
  POST /api/ifrs16/alerts/scan-deadlines  (n8n / cron)
"""

from __future__ import annotations

import io
import re
from calendar import monthrange
from datetime import date, datetime, timedelta
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _v(obj: Any) -> Any:
    if obj is None:
        return None
    if isinstance(obj, dict) and "value" in obj:
        return obj.get("value")
    return obj


def _dig(data: Any, *path: str) -> Any:
    cur = data
    for p in path:
        if not isinstance(cur, dict):
            return None
        cur = cur.get(p)
    return cur


def _as_date(raw: Any) -> Optional[date]:
    if raw is None or raw == "":
        return None
    if isinstance(raw, date) and not isinstance(raw, datetime):
        return raw
    if isinstance(raw, datetime):
        return raw.date()
    s = str(raw).strip()[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def _as_int(raw: Any) -> Optional[int]:
    if raw is None or raw == "":
        return None
    try:
        return int(float(raw))
    except (TypeError, ValueError):
        return None


def _as_float(raw: Any) -> Optional[float]:
    if raw is None or raw == "":
        return None
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def _add_months(d: date, months: int) -> date:
    if months <= 0:
        return d
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    last = monthrange(y, m)[1]
    return date(y, m, min(d.day, last))


def _escalation_pct(text: Any) -> Optional[float]:
    if not text:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)\s*%", str(text))
    return float(m.group(1)) if m else None


def unwrap_extraction(payload: Any) -> dict[str, Any]:
    """Accept raw extractor JSON or wrappers with extracted_data / contract_data."""
    if not isinstance(payload, dict):
        return {}
    if "extracted_data" in payload and isinstance(payload["extracted_data"], dict):
        return payload["extracted_data"]
    if "contract_data" in payload and isinstance(payload["contract_data"], dict):
        return payload["contract_data"]
    # lease_data shape from ifrs16_leases
    if "lease_data" in payload and isinstance(payload["lease_data"], dict):
        ld = payload["lease_data"]
        cd = ld.get("contract_data")
        if isinstance(cd, dict) and cd:
            return cd
        return ld
    return payload


def flatten_lease_terms(payload: Any, lease_id: Optional[str] = None) -> dict[str, Any]:
    """Flatten one extraction (or stored lease) into a single export/scan row."""
    top = payload if isinstance(payload, dict) else {}
    data = unwrap_extraction(top)
    ld = top.get("lease_data") if isinstance(top.get("lease_data"), dict) else {}
    if not ld and isinstance(top.get("id"), str) and top.get("dates"):
        ld = top

    lessor = _v(_dig(data, "basic_info", "lessor_name")) or ld.get("lessor") or ""
    lessee = _v(_dig(data, "basic_info", "lessee_name")) or ld.get("lessee") or ""
    asset = _v(_dig(data, "basic_info", "asset_description")) or ld.get("asset") or ld.get("title") or ""

    start = _as_date(_v(_dig(data, "dates", "commencement_date"))) or _as_date(
        (ld.get("dates") or {}).get("commencement") or ld.get("start_date")
    )
    end = _as_date(_v(_dig(data, "dates", "end_date"))) or _as_date(
        (ld.get("dates") or {}).get("end") or ld.get("end_date")
    )
    term = _as_int(_v(_dig(data, "dates", "lease_term_months"))) or _as_int(ld.get("term_months"))

    monthly = _as_float(_v(_dig(data, "payments", "monthly_amount")))
    if monthly is None:
        monthly = _as_float(ld.get("monthly_payment") or ld.get("base_rent_amount"))
    freq = _v(_dig(data, "payments", "payment_frequency")) or ld.get("payment_frequency") or "Monthly"
    currency = _v(_dig(data, "payments", "currency")) or ld.get("currency") or ""
    escalation_text = _v(_dig(data, "payments", "escalation_clause")) or ""
    rent_free = _as_int(_v(_dig(data, "payments", "rent_free_months")))
    if rent_free is None:
        rent_free = _as_int(ld.get("rent_free_months")) or 0
    fit_out = _as_int(_v(_dig(data, "payments", "fit_out_period_months")))
    if fit_out is None:
        fit_out = _as_int(ld.get("fit_out_period_months")) or 0
    service = _as_float(_v(_dig(data, "payments", "non_lease_component")))
    if service is None:
        service = _as_float(ld.get("non_lease_component"))

    deposit = _as_float(_v(_dig(data, "deposits", "security_deposit_amount")))
    if deposit is None:
        deposit = _as_float(ld.get("security_deposit_amount"))
    deposit_refund_days = _as_int(_v(_dig(data, "deposits", "security_deposit_refund_days")))
    if deposit_refund_days is None:
        deposit_refund_days = _as_int(ld.get("security_deposit_refund_days"))

    renewal_type = str(_v(_dig(data, "options", "renewal_type")) or ld.get("renewal_type") or "").strip().lower()
    if renewal_type not in {"auto", "manual", "none"}:
        renewal_type = renewal_type or ""
    notice_days = _as_int(_v(_dig(data, "options", "renewal_notice_period_days")))
    if notice_days is None:
        notice_days = _as_int(ld.get("renewal_notice_period_days"))
    renewal_text = _v(_dig(data, "options", "renewal_options")) or ld.get("renewal_options") or ""
    break_text = _v(_dig(data, "options", "termination_clause")) or ld.get("termination_clauses") or ""
    break_penalty = _as_float(_v(_dig(data, "options", "break_clause_penalty_amount")))
    if break_penalty is None:
        break_penalty = _as_float(ld.get("break_clause_penalty_amount"))
    break_after = _as_int(_v(_dig(data, "options", "break_clause_eligible_after_months")))
    if break_after is None:
        break_after = _as_int(ld.get("break_clause_eligible_after_months"))

    esc_pct = _escalation_pct(escalation_text)
    if esc_pct is None and ld.get("escalation_value") is not None:
        esc_pct = _as_float(ld.get("escalation_value"))

    lid = (
        lease_id
        or top.get("lease_id")
        or ld.get("lease_id")
        or ld.get("id")
        or top.get("id")
        or ""
    )

    return {
        "lease_id": str(lid),
        "landlord": str(lessor or ""),
        "tenant": str(lessee or ""),
        "property_unit": str(asset or ""),
        "lease_term_months": term,
        "commencement_date": start.isoformat() if start else "",
        "expiry_date": end.isoformat() if end else "",
        "monthly_rent": monthly,
        "annual_rent": round(monthly * 12, 4) if monthly is not None else None,
        "payment_frequency": str(freq or ""),
        "currency": str(currency or ""),
        "rent_escalation_pct": esc_pct,
        "escalation_clause": str(escalation_text or ""),
        "security_deposit_amount": deposit,
        "security_deposit_refund_days": deposit_refund_days,
        "renewal_type": renewal_type,
        "renewal_notice_period_days": notice_days,
        "renewal_options": str(renewal_text or ""),
        "break_clause_terms": str(break_text or ""),
        "break_clause_penalty_amount": break_penalty,
        "break_clause_eligible_after_months": break_after,
        "rent_free_months": rent_free or 0,
        "fit_out_period_months": fit_out or 0,
        "service_charges": service,
        "_start": start,
        "_end": end,
        "_notice_days": notice_days,
        "_break_after": break_after,
        "_escalation_pct": esc_pct,
        "_escalation_start": _as_date(ld.get("escalation_start_date")),
    }


def compute_deadline_dates(row: dict[str, Any], today: Optional[date] = None) -> dict[str, Any]:
    """Derive notice / escalation / break / expiry deadlines and days-until-next."""
    today = today or date.today()
    start: Optional[date] = row.get("_start")
    end: Optional[date] = row.get("_end")
    notice_days = row.get("_notice_days")
    break_after = row.get("_break_after")
    esc_start: Optional[date] = row.get("_escalation_start")
    esc_pct = row.get("_escalation_pct")

    notice_deadline: Optional[date] = None
    if end is not None and notice_days is not None and notice_days >= 0:
        notice_deadline = end - timedelta(days=int(notice_days))

    break_window_open: Optional[date] = None
    if start is not None and break_after is not None and break_after >= 0:
        break_window_open = _add_months(start, int(break_after))

    next_escalation: Optional[date] = None
    if esc_pct is not None and esc_pct > 0 and start is not None:
        base = esc_start or _add_months(start, 12)
        cand = base
        # Walk annual anniversaries until on/after today
        guard = 0
        while cand < today and guard < 40:
            cand = _add_months(cand, 12)
            guard += 1
        if end is None or cand <= end:
            next_escalation = cand

    candidates: list[tuple[str, date]] = []
    if end is not None:
        candidates.append(("expiry", end))
    if notice_deadline is not None:
        candidates.append(("renewal_notice", notice_deadline))
    if next_escalation is not None:
        candidates.append(("escalation", next_escalation))
    if break_window_open is not None:
        candidates.append(("break_clause_window", break_window_open))

    upcoming = [(k, d, (d - today).days) for k, d in candidates]
    future = [x for x in upcoming if x[2] >= 0]
    past = [x for x in upcoming if x[2] < 0]

    if future:
        next_type, next_date, days = min(future, key=lambda x: x[2])
    elif past:
        # Most recently missed / overdue
        next_type, next_date, days = max(past, key=lambda x: x[2])
    else:
        next_type, next_date, days = None, None, None

    return {
        "notice_deadline": notice_deadline.isoformat() if notice_deadline else "",
        "next_escalation_date": next_escalation.isoformat() if next_escalation else "",
        "break_clause_window_opens": break_window_open.isoformat() if break_window_open else "",
        "next_deadline_type": next_type or "",
        "next_deadline_date": next_date.isoformat() if next_date else "",
        "days_until_next_deadline": days,
        "_deadlines": upcoming,
    }


def build_export_row(payload: Any, lease_id: Optional[str] = None, today: Optional[date] = None) -> dict[str, Any]:
    flat = flatten_lease_terms(payload, lease_id=lease_id)
    deadlines = compute_deadline_dates(flat, today=today)
    out = {k: v for k, v in flat.items() if not k.startswith("_")}
    out.update({k: v for k, v in deadlines.items() if not k.startswith("_")})
    return out


EXPORT_COLUMNS = [
    "lease_id",
    "landlord",
    "tenant",
    "property_unit",
    "lease_term_months",
    "commencement_date",
    "expiry_date",
    "monthly_rent",
    "annual_rent",
    "payment_frequency",
    "currency",
    "rent_escalation_pct",
    "escalation_clause",
    "security_deposit_amount",
    "security_deposit_refund_days",
    "renewal_type",
    "renewal_notice_period_days",
    "renewal_options",
    "break_clause_terms",
    "break_clause_penalty_amount",
    "break_clause_eligible_after_months",
    "rent_free_months",
    "fit_out_period_months",
    "service_charges",
    "notice_deadline",
    "next_escalation_date",
    "break_clause_window_opens",
    "next_deadline_type",
    "next_deadline_date",
    "days_until_next_deadline",
]


def export_lease_terms_excel(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lease Terms"
    header_fill = PatternFill("solid", fgColor="0D9488")
    header_font = Font(color="FFFFFF", bold=True)
    for col, name in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(1, col, name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, name in enumerate(EXPORT_COLUMNS, start=1):
            val = row.get(name)
            ws.cell(r_idx, c_idx, "" if val is None else val)
    for col in ws.columns:
        letter = col[0].column_letter
        width = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 40)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _bucket(days: int) -> str:
    if days < 0:
        return "overdue"
    if days <= 30:
        return "within_30"
    if days <= 60:
        return "within_60"
    if days <= 90:
        return "within_90"
    return "beyond_90"


def scan_lease_deadlines(
    lease_rows: list[dict[str, Any]],
    *,
    today: Optional[date] = None,
    horizon_days: int = 90,
) -> list[dict[str, Any]]:
    """Scan stored ifrs16_leases (or extraction payloads) for actionable deadlines."""
    today = today or date.today()
    alerts: list[dict[str, Any]] = []
    for raw in lease_rows:
        flat = flatten_lease_terms(raw)
        deadlines = compute_deadline_dates(flat, today=today)
        lid = flat.get("lease_id") or ""
        asset = flat.get("property_unit") or lid or "Lease"
        for dtype, ddate, days in deadlines.get("_deadlines") or []:
            if days > horizon_days:
                continue
            severity = "red" if days <= 30 or days < 0 else "amber"
            label = {
                "expiry": "Lease expiry",
                "renewal_notice": "Renewal notice deadline",
                "escalation": "Rent escalation anniversary",
                "break_clause_window": "Break-clause window opens",
            }.get(dtype, dtype)
            if days < 0:
                title = f"{asset}: {label} passed {abs(days)} days ago"
                message = f"{label} was {ddate.isoformat()}"
            else:
                title = f"{asset}: {label} in {days} days"
                message = f"{label} on {ddate.isoformat()}"
            alerts.append(
                {
                    "id": str(lid),
                    "lease_id": str(lid),
                    "type": dtype,
                    "bucket": _bucket(days),
                    "severity": severity,
                    "title": title,
                    "message": message,
                    "deadline_date": ddate.isoformat(),
                    "days_until": days,
                    "landlord": flat.get("landlord"),
                    "tenant": flat.get("tenant"),
                    "property_unit": flat.get("property_unit"),
                }
            )
    alerts.sort(key=lambda a: (a.get("days_until") is None, a.get("days_until", 9999)))
    return alerts
